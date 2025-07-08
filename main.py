
from reportlab.pdfbase.pdfmetrics import registerFontFamily
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.styles import ParagraphStyle as PS
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.colors import *
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import TableStyle, Paragraph, SimpleDocTemplate
from reportlab.pdfgen.canvas import Canvas 
from PyPDF2 import PdfMerger
from main_window import Ui_MainWindow
from start_form import Ui_Form
from datetime import *
from PyQt5 import *
from PyQt5.QtCore import QAbstractTableModel, QThread, pyqtSignal
from PyQt5.QtGui import QIcon, QBrush, QColor, QFont, QPalette
from PyQt5.QtWidgets import QPushButton, QRadioButton, QHeaderView, QComboBox, QListWidgetItem, QItemDelegate, QStyledItemDelegate
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox, QMenu, QInputDialog, QTableWidgetItem, QLineEdit, QLabel
from PyQt5.QtWidgets import QAbstractItemView, QFileDialog, QProgressDialog, QAction, QDesktopWidget, QTableView, QColorDialog, QMessageBox
from PyQt5 import QtGui, QtWidgets, QtCore

import itertools
from models import *
from collections import Counter
from itertools import *
import os
import openpyxl as op
import pandas as pd
import numpy as np
import contextlib
import sys

import pymysql
import subprocess
#=============
import pathlib
from pathlib import Path
from dateutil.relativedelta import relativedelta
import random
import math
from sys import platform
import time
os.environ['QT_AUTO_SCREEN_SCALE_FACTOR'] = '1'
# WindowsArguments = dpiawareness = 1

# app = QtGui.QApplication(sys.argv)
# screen_rect = app.desktop().screenGeometry()
# width, height = screen_rect.width(), screen_rect.height()
# import collections
from playhouse.migrate import * # для удаления, редактирования таблиц DB

if not os.path.isdir("table_pdf"):  # создает папку 
    os.mkdir("table_pdf")
if not os.path.isdir("competition_pdf"):  # создает папку 
    os.mkdir("competition_pdf")
if not os.path.isdir("sign"):  # создает папку 
    os.mkdir("sign")


def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press ⌘F8 to toggle the breakpoint.

if __name__ == '__main__':
    print_hi('PyCharm and Alex')

# from playhouse.sqlite_ext import SqliteExtDatabase, backup_to_file, backup

registerFontFamily('DejaVuSerif', normal='DejaVuSerif',
                   bold='DejaVuSerif-Bold', italic='DejaVuSerif-Italic')
outpath = os.path.join(os.getcwd(), 'font')
pdfmetrics.registerFont(TTFont('DejaVuSans', os.path.join(outpath, 'DejaVuSans.ttf')))
pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', os.path.join(outpath, 'DejaVuSans-Bold.ttf')))
pdfmetrics.registerFont(TTFont('DejaVuSerif', os.path.join(outpath, 'DejaVuSerif.ttf')))
pdfmetrics.registerFont(TTFont('DejaVuSerif-Bold', os.path.join(outpath, 'DejaVuSerif-Bold.ttf')))
pdfmetrics.registerFont(TTFont('DejaVuSerif-Italic', os.path.join(outpath, 'DejaVuSerif-Italic.ttf')))
# ============== рабочий вариант
class MyTableModel(QAbstractTableModel):
    def __init__(self, data):
        super().__init__()
        self._data = data
        self.horizontalHeaderLabels = []

    def setHorizontalHeaderLabels(self, horizontalHeaderLabels):
        self.horizontalHeaderLabels = horizontalHeaderLabels
 
    def headerData(self, section: int, orientation: QtCore.Qt.Orientation, role: QtCore.Qt.ItemDataRole):
        if (orientation == QtCore.Qt.Horizontal and role == QtCore.Qt.DisplayRole and len(self.horizontalHeaderLabels) == self.columnCount(None)):
            return self.horizontalHeaderLabels[section]
            # return self._data.columns[section]
        return super().headerData(section, orientation, role)
    
    def rowCount(self, parent):
        return len(self._data)
    
    def columnCount(self, parent):
        if len(self._data) > 0:
            return len(self._data[0])
        else:
            return 0

    def data(self, index, role):
        coach_tmp_list = []
        tb = my_win.tabWidget.currentIndex()
        if role == QtCore.Qt.ItemDataRole.DisplayRole:
            return str(self._data[index.row()][index.column()])
        elif role == QtCore.Qt.ForegroundRole: # выделяет фамилию красным цветом
            val = str(self._data[index.row()][index.column()])
            if index.column() == 1 and tb == 1: # создание списка должников и если находит окрашывает фамилии красным   
                dolg_R_list = dolg_R() 
                if val in dolg_R_list:     
                    return QtGui.QBrush(QtCore.Qt.red)
            elif index.column() == 4 and tb == 1: # создание списка должников и если находит окрашывает фамилии красным 
                city_dict = unconfirmed_city()
                val_player = self._data[index.row()][1]
                if val_player in city_dict.keys():     
                    return QtGui.QBrush(QtCore.Qt.blue)
            elif index.column() == 3 and tb == 2: # выделяет повторяющиеся фамилии тренеров
                if my_win.checkBox_repeat_regions.isChecked(): # отмечен чекбокс проверки повтора регионов в группе
                    ind = my_win.comboBox_filter_number_group_final.currentIndex()
                    if ind > 0: # значит выбран номер группы
                        n_gr = my_win.comboBox_filter_number_group_final.currentText()
                        group_coach_list = dupl_coach(n_gr) # список всех тренеров группы 
                        count_frequency = filter(lambda x: group_coach_list.count(x) > 1, group_coach_list)
                        double_coach_list = list(set(count_frequency))                       
                        znak = val.find(",")
                        if znak == -1: # один тренер
                            coach_tmp_list.append(val)
                        else: # у игрока не один тренер и делает из них список тренеров
                            coach_1 = val[:znak]
                            coach_tmp_list.append(coach_1)
                            if val.find(",", znak) == -1:
                                znak_1 = val.find(",", znak + 1)
                                coach_2 = val[znak: znak_1]
                                coach_tmp_list.append(coach_2)
                            else:
                                coach_2 = val[znak + 2:]
                                znak_1 = val.find(",", znak + 1)
                                if val.find(",", znak_1) == -1:
                                    coach_tmp_list.append(coach_2)
                                else:
                                    coach_2 = val[znak + 2:znak_1]
                                    coach_tmp_list.append(coach_2)
                                    coach_3 = val[znak_1 + 2:]
                                    coach_tmp_list.append(coach_3)
                        for dc in coach_tmp_list:
                            if dc in double_coach_list:    
                                return QtGui.QBrush(QtCore.Qt.blue)
                            else:
                                return QtGui.QBrush(QtCore.Qt.black)
            elif index.column() == 2 and tb == 2: # выделяет совпадающие регионы
                if my_win.checkBox_repeat_regions.isChecked(): # отмечен чекбокс проверки повтора регионов в группе
                    ind = my_win.comboBox_filter_number_group_final.currentIndex()
                    if ind > 0: # значит выбран номер группы
                        n_gr = my_win.comboBox_filter_number_group_final.currentText()
                        region_group_list = dupl_regions(n_gr)
                        p = 0
                        for l in region_group_list:
                            if val == l:
                                p += 1
                        if p > 1:    
                            return QtGui.QBrush(QtCore.Qt.darkGreen)
                        else:
                            return QtGui.QBrush(QtCore.Qt.black)
            # elif index.column() == 2 and tb == 3: # выделяет совпадающие регионы при жеребьевки групп
            #     # if my_win.checkBox_repeat_regions.isChecked(): # отмечен чекбокс проверки повтора регионов в группе
            #     # ind = my_win.comboBox_filter_number_group_final.currentIndex()
            #     if ind > 0: # значит выбран номер группы
            #         n_gr = my_win.comboBox_filter_number_group_final.currentText()
            #         region_group_list = dupl_regions(n_gr)
            #         p = 0
            #         for l in region_group_list:
            #             if val == l:
            #                 p += 1
            #         if p > 1:    
            #             return QtGui.QBrush(QtCore.Qt.darkGreen)
            #         else:
            #             return QtGui.QBrush(QtCore.Qt.black)



class _MyTableModel(QAbstractTableModel): # === вариант эксперементальный ============
    def __init__(self, data):
        super().__init__()
        self._data = data
        self.horizontalHeaderLabels = []

    def setHorizontalHeaderLabels(self, horizontalHeaderLabels):
        self.horizontalHeaderLabels = horizontalHeaderLabels
 
    def headerData(self, section: int, orientation: QtCore.Qt.Orientation, role: QtCore.Qt.ItemDataRole):
        if (orientation == QtCore.Qt.Horizontal and role == QtCore.Qt.DisplayRole and len(self.horizontalHeaderLabels) == self.columnCount(None)):
        # if (orientation == QtCore.Qt.Horizontal and role == QtCore.Qt.DisplayRole  and len(self.horizontalHeaderLabels) == 10):
            return self.horizontalHeaderLabels[section]
            # return self._data.columns[section]
        return super().headerData(section, orientation, role)
    
    def rowCount(self, parent):
        return len(self._data)
    
    def columnCount(self, parent):
        if len(self._data) > 0:
            return len(self._data[0])
            # return len(self._data.columns)
        else:
            return 0

    def data(self, index, role):
        coach_tmp_list = []
        tb = my_win.tabWidget.currentIndex()
        if role == QtCore.Qt.ItemDataRole.DisplayRole:
            return str(self._data.iloc[index.row()][index.column()])
        elif role == QtCore.Qt.ForegroundRole: # выделяет фамилию красным цветом
            val = str(self._data.iloc[index.row()][index.column()])
            # ===== new ===
            if tb == 1:
                if index.column() == 1:
                    dolg_R_list = dolg_R() 
                    if val in dolg_R_list:     
                        return QtGui.QBrush(QtCore.Qt.red)
                elif index.column() == 4:
                    city_dict = unconfirmed_city()
                    val_player = self._data.iloc[index.row()][1]
                    if val_player in city_dict.keys():     
                        return QtGui.QBrush(QtCore.Qt.blue)
            elif tb == 2:
                if index.column() == 3:
                    if my_win.checkBox_repeat_regions.isChecked(): # отмечен чекбокс проверки повтора регионов в группе
                        ind = my_win.comboBox_filter_number_group_final.currentIndex()
                    else:
                        ind = 0
                    if ind > 0: # значит выбран номер группы
                        n_gr = my_win.comboBox_filter_number_group_final.currentText()
                        group_coach_list = dupl_coach(n_gr) # список всех тренеров группы 
                        count_frequency = filter(lambda x: group_coach_list.count(x) > 1, group_coach_list)
                        double_coach_list = list(set(count_frequency))                       
                        znak = val.find(",")
                        if znak == -1: # один тренер
                            coach_tmp_list.append(val)
                        else: # у игрока не один тренер и делает из них список тренеров
                            coach_1 = val[:znak]
                            coach_tmp_list.append(coach_1)
                            if val.find(",", znak) == -1:
                                znak_1 = val.find(",", znak + 1)
                                coach_2 = val[znak: znak_1]
                                coach_tmp_list.append(coach_2)
                            else:
                                coach_2 = val[znak + 2:]
                                znak_1 = val.find(",", znak + 1)
                                if val.find(",", znak_1) == -1:
                                    coach_tmp_list.append(coach_2)
                                else:
                                    coach_2 = val[znak + 2:znak_1]
                                    coach_tmp_list.append(coach_2)
                                    coach_3 = val[znak_1 + 2:]
                                    coach_tmp_list.append(coach_3)
                        for dc in coach_tmp_list:
                            if dc in double_coach_list:    
                                return QtGui.QBrush(QtCore.Qt.blue)
                            else:
                                return QtGui.QBrush(QtCore.Qt.black)
                elif index.column() == 2:
                    if my_win.checkBox_repeat_regions.isChecked(): # отмечен чекбокс проверки повтора регионов в группе
                        ind = my_win.comboBox_filter_number_group_final.currentIndex()
                    else:
                        ind = 0
                    if ind > 0: # значит выбран номер группы
                        n_gr = my_win.comboBox_filter_number_group_final.currentText()
                        region_group_list = dupl_regions(n_gr)
                        p = 0
                        for l in region_group_list:
                            if val == l:
                                p += 1
                        if p > 1:    
                            return QtGui.QBrush(QtCore.Qt.darkGreen)
                        else:
                            return QtGui.QBrush(QtCore.Qt.black)


class MainWindow(QMainWindow, Ui_MainWindow):

    def __init__(self, parent=None, *args, **kwargs) -> object:
        QMainWindow.__init__(self)
        self.setupUi(self)

        w = self.size().width()     # "определение ширины"
        h = self.size().height()    # "определение высоты"

        # app = QApplication(sys.argv)
        # screen_rect = app.desktop().screenGeometry()
        # width, height = screen_rect.width(), screen_rect.height()
        # if height < h:
        #     form_heidht = height
        # else:
        #     form_heidht = height
        # # print(self.w, self.h) # получение разрешения экрана
        # my_win.setGeometry(w, form_heidht)
        self._createAction()
        self._createMenuBar()
        self._connectActions()

        self.menuBar()
        # centralWidget = QMainWindow()
        Button_turnir_1 = QPushButton("Proba", self) # (в каком виджете размещена)
        Button_turnir_1.resize(40, 10) # размеры кнопки (длина 120, ширина 50)
        Button_turnir_1.move(10, 5) # разммещение кнопки (от левого края 900, от верхнего 0) от виджета в котором размещен
        font = QFont('Times New Roman', 8)
        Button_turnir_1.setFont(font)
        # Button_turnir_1.setText("Proba")
 
    
        Button_turnir_1.setFlat(True)
        Button_turnir_1.show()
        # Button_turnir_1.clicked.connect(view)

        self.Button_title_made.setEnabled(False)
        self.Button_system_made.setEnabled(False)
        self.tabWidget.setCurrentIndex(0)  # включает вкладку титул
        # ++ отключение страниц
        self.tabWidget.setTabEnabled(1, True)
        self.tabWidget.setTabEnabled(2, False)
        self.tabWidget.setTabEnabled(3, False)
        self.tabWidget.setTabEnabled(4, False)
        self.tabWidget.setTabEnabled(5, False)
        self.tabWidget.setTabEnabled(6, True)
        # self.tabWidget.setTabEnabled(7, True)

        self.tabWidget_stage.setTabEnabled(0, False)
        self.tabWidget_stage.setTabEnabled(1, False)
        self.tabWidget_stage.setTabEnabled(2, False)


        self.toolBox.setItemEnabled(0, False)
        self.toolBox.setItemEnabled(1, False)
        self.toolBox.setItemEnabled(2, False)
        self.toolBox.setItemEnabled(3, False)
        self.toolBox.setItemEnabled(4, False)
        self.toolBox.setItemEnabled(5, False)
        self.toolBox.setItemEnabled(6, True)
        # self.toolBox.setItemEnabled(7, True)
 
        self.tabWidget_2.setTabVisible(0, False) # скрывает ярлыки у tabWidget
        self.tabWidget_2.setTabVisible(1, False)
        self.tabWidget_2.setTabVisible(2, False)
        self.tabWidget_2.setTabVisible(3, False)
    def closeEvent(self, event):
        # Создание бэкап DB при закрытии формы -main- по нажатию на крестик
        sender = my_win.sender()
        if sender != self.exitAction:
            reply = QMessageBox.question\
                    (self, 'Вы нажали на крестик',
                        "Вы уверены, что хотите уйти?\n"
                        "если сделать копию DB то нажмите -Yes-\n",              
                QMessageBox.Yes,
                QMessageBox.No)
            if reply == QMessageBox.Yes:
                flag  = 1
                exit_comp(flag)
            # else:
            #     event.ignore()  
    # ====== создание строки меню ===========
    def _createMenuBar(self):
        menuBar = self.menuBar()
        menuBar.setNativeMenuBar(False)  # разрешает показ менюбара

        # меню Соревнования
        fileMenu = QMenu("Соревнования", self)  # основное
        menuBar.addMenu(fileMenu)
        # подменю с выбором (addMenu добавляет к пункту возможность выбора)
        new_comp = fileMenu.addMenu("Новые")
        fileMenu.addSeparator()  # вставляет разделительную черту
        go_to = fileMenu.addMenu("Перейти к")
        fileMenu.addSeparator()  # вставляет разделительную черту
        # подменю без выбора (addAction создает сразу действие)
        system = fileMenu.addMenu("Система")
        choice = fileMenu.addMenu("Жеребьевка")
        # saveList = fileMenu.addMenu("Сохранить")
        fileMenu.addSeparator()
        last_comp = fileMenu.addMenu("Последние")
        fileMenu.addSeparator()
        fileMenu.addAction(self.exitAction)
        # меню Редактировать
        editMenu = menuBar.addMenu("Редактировать")  # основное
        # меню Печать
        printMenu = menuBar.addMenu("Печать") # основное
        printMenu.addAction(self.print_double_family_Action)
        # ============ создание подменю
        new_comp.addAction(self.new_comp_Action) # создание титула для повтора мальчики или девочки если одни уже созданы
        go_to.addAction(self.go_to_Action)  # подменю выбора соревнования
        system.addAction(self.system_made_Action)  # подменю создание системы
        system.addAction(self.system_edit_Action)  # подменю редактирование системы
        system.addAction(self.system_clear_Action)  # подменю редактирование системы
        choice.addAction(self.choice_one_table_Action) # подменю одна таблица
        choice.addAction(self.choice_gr_Action)  # подменю группы
        choice.addAction(self.choice_pf_Action)  # подменю полуфиналы
        choice.addAction(self.choice_fin_Action)  # подменю финалы
    
        last_comp.addAction(self.first_comp_Action)
        last_comp.addAction(self.second_comp_Action)
        last_comp.addAction(self.third_comp_Action)
        last_comp.addAction(self.fourth_comp_Action)

        ed_Menu = editMenu.addMenu("Жеребьевка")
        ed_Menu.addAction(self.ed_one_table_Action)
        ed_Menu.addAction(self.ed_etap_Action)

        editMenu.addAction(self.vid_edit_Action)  #в осн меню -Редактировать- добавлен пункт сразу с акцией -Вид страницы этапов
        editMenu.addAction(self.indent_edit_Action)
        # меню Рейтинг
        rank_Menu = menuBar.addMenu("Рейтинг")  # основное
        rank_Menu.addAction(self.rAction)
        rank_Menu.addAction(self.r1Action)
        # меню печать
        print_Menu = printMenu.addMenu("Чистые таблицы") 
        print_Menu.addAction(self.clear_s8_full_Action)  
        print_Menu.addAction(self.clear_s8_2_Action)       
        print_Menu.addAction(self.clear_s16_Action)
        print_Menu.addAction(self.clear_s16_2_Action)
        print_Menu.addAction(self.clear_s32_Action)
        print_Menu.addAction(self.clear_s32_full_Action)
        print_Menu.addAction(self.clear_s32_2_Action)
        print_Menu = printMenu.addMenu("Должники за R")
        print_Menu.addAction(self.print_list_nopay_R_Action)
        print_Menu.addAction(self.print_list_pay_R_Action)


        # меню просмотр (последовательность вида в меню)
        view_Menu = menuBar.addMenu("Просмотр")
        view_Menu.addAction(self.view_all_comp_Action)
        view_Menu.addSeparator()
        view_Menu.addAction(self.view_title_Action)
        view_Menu.addAction(self.view_regions_list_Action)
        view_Menu.addAction(self.view_referee_list_Action)
        view_Menu.addAction(self.view_winners_list_Action)
        view_Menu.addAction(self.view_list_Action)
        view_Menu.addSeparator()
        view_Menu.addAction(self.view_gr_Action)
        pf_view_Menu = view_Menu.addMenu("Полуфиналы")
        pf_view_Menu.addAction(self.view_pf1_Action)
        pf_view_Menu.addAction(self.view_pf2_Action)
        view_Menu.addAction(self.view_one_table_Action)
        v_Menu = view_Menu.addMenu("Финалы")
        v_Menu.addAction(self.view_fin1_Action)
        v_Menu.addAction(self.view_fin2_Action)
        v_Menu.addAction(self.view_fin3_Action)
        v_Menu.addAction(self.view_fin4_Action)
        v_Menu.addAction(self.view_fin5_Action)
        v_Menu.addAction(self.view_fin6_Action)
        v_Menu.addAction(self.view_fin7_Action)
        v_Menu.addAction(self.view_fin8_Action)
        v_Menu.addAction(self.view_fin9_Action)
        v_Menu.addAction(self.view_fin10_Action)
        v_Menu.addAction(self.view_superfin_Action)

        # меню помощь
        help_Menu = menuBar.addMenu("Помощь")  # основное
        help_Menu.addAction(self.copy_db_Action)
        help_Menu.addAction(self.delete_copy_db_Action)
        help_Menu.addSeparator()
        help_Menu.addAction(self.stat_Action)
        help_Menu.addAction(self.player_stat_Action)
    #  создание действий меню
    def _createAction(self):
        self.helpAction = QAction(self)
        self.system_edit_Action = QAction("Редактировать")
        self.system_made_Action = QAction("Создать")
        self.system_clear_Action = QAction("Очистить")
        self.exitAction = QAction("Выход")
        self.rAction = QAction("Текущий рейтинг")
        self.r1Action = QAction("Рейтинг за январь")
        self.first_comp_Action = QAction("пусто")
        self.second_comp_Action = QAction("пусто")
        self.third_comp_Action = QAction("пусто")
        self.fourth_comp_Action = QAction("пусто")
        self.ed_one_table_Action = QAction("Редакитровать таблицу")

        self.print_list_nopay_R_Action = QAction("Список, неоплативших R")
        self.print_list_pay_R_Action = QAction("Список, оплативших R")
        self.print_double_family_Action = QAction("Двойные фамилии")

        self.ed_etap_Action = QAction("Редактирование этапов")  # подменю редактор
        self.vid_edit_Action = QAction("Вид страницы этапов")
        self.indent_edit_Action = QAction("Изменить отступ в PDF")

        self.choice_one_table_Action = QAction("Одна таблица")
        # подменю жеребьевка -группы-
        self.choice_gr_Action = QAction("Группы")
        # подменю жеребьевка -полуфиналы-
        self.choice_pf_Action = QAction("Полуфиналы")
        self.choice_fin_Action = QAction("Финалы")  # подменю жеребьевка -финалы-
        self.view_all_comp_Action = QAction("Полные соревнования")
        self.view_title_Action = QAction("Титульный лист")
        self.view_referee_list_Action = QAction("Список ГСК")
        self.view_regions_list_Action = QAction("Список субъектов РФ")
        self.view_winners_list_Action = QAction("Список победителей")
        self.view_list_Action = QAction("Список участников")
        self.view_gr_Action = QAction("Группы")
        self.view_pf1_Action = QAction("1-й полуфинал")
        self.view_pf2_Action = QAction("2-й полуфинал")

        self.view_one_table_Action = QAction("Одна таблица")
        self.go_to_Action = QAction("пусто")
        self.new_comp_Action = QAction("Клонирование титула")
        # подменю -печать-
        self.clear_s8_full_Action = QAction("Сетка 8")
        self.clear_s8_2_Action = QAction("Сетка 8 минус 2")
        self.clear_s16_Action = QAction("Сетка 16")
        self.clear_s16_2_Action = QAction("Сетка 16 минус 2")
        self.clear_s32_2_Action = QAction("Сетка 32 минус 2")
        self.clear_s32_full_Action = QAction("Сетка 32 прогрессивная")
        self.clear_s32_Action = QAction("Сетка 32 (1-3 места)")
        # ======== подменю финалы ============= сделать в зависимости от кол-во финалов остальные невидимые
        self.view_fin1_Action = QAction("1-финал")
        self.view_fin2_Action = QAction("2-финал")
        self.view_fin3_Action = QAction("3-финал")
        self.view_fin4_Action = QAction("4-финал")
        self.view_fin5_Action = QAction("5-финал")
        self.view_fin6_Action = QAction("6-финал")
        self.view_fin7_Action = QAction("7-финал")
        self.view_fin8_Action = QAction("8-финал")
        self.view_fin9_Action = QAction("9-финал")
        self.view_fin10_Action = QAction("10-финал")
        self.view_superfin_Action = QAction("Суперфинал")

        # выключает пункты меню пока не создана система
        self.choice_one_table_Action.setEnabled(False)
        self.choice_gr_Action.setEnabled(False)
        self.choice_pf_Action.setEnabled(False)
        self.choice_fin_Action.setEnabled(False)

        self.view_one_table_Action.setEnabled(False)
        self.view_gr_Action.setEnabled(False)
        self.view_pf1_Action.setEnabled(False)
        self.view_pf2_Action.setEnabled(False)
 
        self.view_fin1_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin2_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin3_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin4_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin5_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin6_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin7_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin8_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin9_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_fin10_Action.setEnabled(False)  # делает пункт меню не видимым
        self.view_superfin_Action.setEnabled(False)  # делает пункт меню не видимым
        # пункты меню редактирование жеребьевки
        self.ed_one_table_Action.setEnabled(False)  # делает пункт меню не видимым
        self.ed_etap_Action.setEnabled(False)  # делает пункт меню не видимым
         # пункты меню помощь
        self.copy_db_Action = QAction("Импорт из базы данных")
        self.delete_copy_db_Action = QAction("Удаление старых копий DB")
        self.stat_Action = QAction("Число встреч для R отчета")
        self.player_stat_Action = QAction("Статистика игрока")

    def _connectActions(self):
        self.system_made_Action.triggered.connect(self.system_made)
        self.system_edit_Action.triggered.connect(self.system_made)
        self.system_clear_Action.triggered.connect(self.system_clear)
        self.vid_edit_Action.triggered.connect(self.vid_edit)
        self.indent_edit_Action.triggered.connect(self.indent_edit)
        self.exitAction.triggered.connect(self.exit)
        self.choice_one_table_Action.triggered.connect(self.choice)
        self.choice_gr_Action.triggered.connect(self.choice)
        self.choice_pf_Action.triggered.connect(self.choice)
        self.choice_fin_Action.triggered.connect(self.choice)
        self.view_all_comp_Action.triggered.connect(self.view)
        self.view_title_Action.triggered.connect(self.view)
        self.view_referee_list_Action.triggered.connect(self.view)
        self.view_regions_list_Action.triggered.connect(self.view)
        self.view_winners_list_Action.triggered.connect(self.view)
        self.view_list_Action.triggered.connect(self.view)
        self.view_one_table_Action.triggered.connect(self.view)
        self.view_gr_Action.triggered.connect(self.view)
        self.view_pf1_Action.triggered.connect(self.view)
        self.view_pf2_Action.triggered.connect(self.view)
        self.view_fin1_Action.triggered.connect(self.view)
        self.view_fin2_Action.triggered.connect(self.view)
        self.view_fin3_Action.triggered.connect(self.view)
        self.view_fin4_Action.triggered.connect(self.view)
        self.view_fin5_Action.triggered.connect(self.view)
        self.view_fin6_Action.triggered.connect(self.view)
        self.view_fin7_Action.triggered.connect(self.view)
        self.view_fin8_Action.triggered.connect(self.view)
        self.view_fin9_Action.triggered.connect(self.view)
        self.view_fin10_Action.triggered.connect(self.view)
        self.view_superfin_Action.triggered.connect(self.view)
        self.clear_s8_full_Action.triggered.connect(self.print_clear)
        self.clear_s8_2_Action.triggered.connect(self.print_clear)
        self.clear_s16_Action.triggered.connect(self.print_clear)
        self.clear_s16_2_Action.triggered.connect(self.print_clear)
        self.clear_s32_full_Action.triggered.connect(self.print_clear)
        self.clear_s32_Action.triggered.connect(self.print_clear)
        self.clear_s32_2_Action.triggered.connect(self.print_clear)

        self.first_comp_Action.triggered.connect(self.last)
        self.second_comp_Action.triggered.connect(self.last)
        self.third_comp_Action.triggered.connect(self.last)
        self.fourth_comp_Action.triggered.connect(self.last)

        self.ed_etap_Action.triggered.connect(self.edit_etap)
        self.new_comp_Action.triggered.connect(self.clon_titul)
        self.go_to_Action.triggered.connect(self.open)
        # Connect Рейтинг actions
        self.rAction.triggered.connect(self.r_File)
        self.r1Action.triggered.connect(self.r1_File)

        self.print_list_nopay_R_Action.triggered.connect(self.check_debitor_R)
        self.print_list_pay_R_Action.triggered.connect(self.check_debitor_R)
        self.print_double_family_Action.triggered.connect(self.duplicate_family)

        self.copy_db_Action.triggered.connect(self.import_db)
        # self.delete_copy_db_Action.triggered.connect(self.delete_db_copy)
        self.stat_Action.triggered.connect(self.statistika)
        self.player_stat_Action.triggered.connect(self.player_stat)

    def clon_titul(self):
        """Клонирование титула при создании новых соревновании если уже созданы мальчики или девочки"""
        t_id = title_id()
        titles = Title.select().where(Title.id == t_id).get()

        pl_gamer = titles.gamer
        full_name = titles.full_name_comp
        short_name = titles.short_name_comp
        count = len(short_name)
        s_name = short_name[:count - 1]
        if pl_gamer == 'Девочки':
            gm = 'Мальчики'
            pol = 'M'
        elif pl_gamer == 'Девушки':
            gm = 'Юноши'
            pol = 'M'
        elif pl_gamer == 'Юниорки':
            gm = 'Юниоры'
            pol = 'M'
        elif pl_gamer == 'Женщины':
            gm = 'Мужчины'
            pol = 'M'
        elif pl_gamer == 'Мальчики':
            gm = 'Девочки'
            pol = 'D'
        elif pl_gamer == 'Юноши':
            gm = 'Девушки' 
            pol = 'D'
        elif pl_gamer == 'Юниоры':
            gm = 'Юниорки' 
            pol = 'D'
        elif pl_gamer == 'Мужчины':
            gm = 'Женщины' 
            pol = 'D'     

        full_name_comp = full_name.replace(pl_gamer, gm)
        short_name_comp = s_name + pol

        title = Title(name=titles.name,
                    sredi=titles.sredi, 
                    vozrast=titles.vozrast,
                    data_start=titles.data_start,
                    data_end=titles.data_end,
                    mesto=titles.mesto, 
                    referee=titles.referee,
                    kat_ref=titles.kat_ref, 
                    secretary=titles.secretary, 
                    kat_sec=titles.kat_sec, 
                    gamer=gm, 
                    full_name_comp=full_name_comp, 
                    pdf_comp="",
                    short_name_comp=short_name_comp, 
                    tab_enabled="Титул Участники", 
                    multiregion=titles.multiregion).save()

            # получение последней записи в таблице
        t_id_last = Title.select().order_by(Title.id.desc()).get()
        system = System(title_id=t_id_last, total_athletes=0, total_group=0, max_player=0, stage="", type_table="",
                            page_vid="", label_string="", kol_game_string="", choice_flag=False, score_flag=5,
                            visible_game=False, stage_exit="", mesta_exit=0, no_game="").save()
        my_win.tabWidget.setCurrentIndex(0)
        db_r(gamer=gm)
        db_select_title()

    def check_debitor_R(self):
        check_player_whitout_R()

    def newFile(self):
        # Logic for creating a new file goes here...
        my_win.textEdit.setText("Нажата кнопка меню соревнования")
        gamer = db_select_title()

    def r_File(self):
        # Logic for creating a new file goes here...
        self.statusbar.showMessage("Загружен рейтинг-лист на текущий месяц")
        my_win.tabWidget.setTabEnabled(5, True)
        my_win.tabWidget.setCurrentIndex(5)
        fill_table_R_list()
        my_win.comboBox_choice_R.setCurrentIndex(0)
        my_win.lineEdit_find_player_in_R.setFocus()

    def r1_File(self):
        self.statusbar.showMessage("Загружен рейтинг-лист на январь месяц")
        my_win.tabWidget.setTabEnabled(5, True)
        my_win.tabWidget.setCurrentIndex(5)
        fill_table_R1_list()
        my_win.comboBox_choice_R.setCurrentIndex(1)
        my_win.lineEdit_find_player_in_R.setFocus()

    def import_db(self):
        """Импорт из бэкап в базу данных"""
            # Connect to the MySQL database
        cnx = pymysql.connect(user='root', password='db_pass', host='localhost')
        username='root'
        database='mysql_db'
        password='db_pass'
        host='localhost'

        cursor = cnx.cursor()
        fname = QFileDialog.getOpenFileName(my_win, "Выбрать файл базы данных", "backup_db", "*.sql")
        filepath = str(fname[0])
        backup_file = filepath.replace("/", "\\")
        restore_command = f"mysql --host={host} --user={username} --password={password} {database} < {backup_file}"
        subprocess.run(restore_command, shell=True)
        cnx.commit()

        print("Database restored successfully!")
        my_win.statusbar.showMessage("Импорт базы данных завершен успешно", 5000)
        cursor.close()
        cnx.close()
        my_win.tabWidget.setCurrentIndex(1)

    def statistika(self):
        """статистика встреч для точного обсчета рейтинга"""
        my_win.tableWidget.clear()
        my_win.tableWidget.show()
        my_win.tabWidget.setCurrentIndex(7)
        sf_list = ["1-й полуфинал", "2-й полуфинал"]
        sf_game_list = []
        my_win.tableWidget.setColumnCount(2) # устанавливает колво столбцов
        my_win.tableWidget.setRowCount(5)
        systems = System.select().where(System.title_id == title_id())
        for i in systems:
            if i.stage in sf_list:
                pl_exit = i.mesta_exit
                group_sf = i.total_group
                game_in_group = (pl_exit * (pl_exit - 1))
                sf_game = game_in_group * group_sf
                sf_game_list.append(sf_game)
        sum_game_sf = sum(sf_game_list)
        column_label = ["Стадия", "Колчество встреч"]
        my_win.tableWidget.resizeColumnsToContents()
        for i in range(0, 2):  # закрашивает заголовки таблиц  рейтинга зеленым цветом
            my_win.tableWidget.showColumn(i)
            item = QtWidgets.QTableWidgetItem()
            brush = QtGui.QBrush(QtGui.QColor(76, 100, 255))
            brush.setStyle(QtCore.Qt.SolidPattern)
            item.setForeground(brush)
            my_win.tableWidget.setHorizontalHeaderItem(i, item)
        my_win.tableWidget.setHorizontalHeaderLabels(column_label) # заголовки столбцов в tableWidget
        
 
        my_win.tableWidget.setItem(0, 0, QTableWidgetItem("Общее число игр в турнире"))
        my_win.tableWidget.setItem(1, 0, QTableWidgetItem("Встречи по неявке"))
        my_win.tableWidget.setItem(2, 0, QTableWidgetItem("Встречи в сетке с X"))
        my_win.tableWidget.setItem(3, 0, QTableWidgetItem("Встречи в ПФ, попавшие в финалы"))
        my_win.tableWidget.setItem(4, 0, QTableWidgetItem("Всего встреч для обсчета рейтинга"))

        all_game_result = Result.select().where(Result.title_id == title_id())               
        game_no_playing = all_game_result.select().where(Result.score_in_game == "В : П")        
        game_no = all_game_result.select().where(Result.loser == "X")


        all_game = len(all_game_result)
        no_playing = len( game_no_playing)
        no_game = len( game_no)
        sum_game_no_playing = no_playing + no_game + sum_game_sf
        sum_game_rejting = all_game - sum_game_no_playing

        my_win.tableWidget.setItem(0, 1, QTableWidgetItem(str(all_game)))
        my_win.tableWidget.setItem(1, 1, QTableWidgetItem(str(no_playing)))
        my_win.tableWidget.setItem(2, 1, QTableWidgetItem(str(no_game)))
        my_win.tableWidget.setItem(3, 1, QTableWidgetItem(str(sum_game_sf)))
        my_win.tableWidget.setItem(4, 1, QTableWidgetItem(str(sum_game_rejting)))
   
        my_win.tableWidget.resizeColumnsToContents() 

    def player_stat(self):
        """статистика встреч игрока"""
        my_win.tabWidget.setCurrentIndex(7)
        my_win.groupBox_4.hide()
        my_win.tableView.show()

    def duplicate_family(self):
        """печать списка двойныъх фамилий"""
        double_family()
    

    def exit(self):
        flag = 0
        exit_comp(flag)

    def choice(self):
        msg = QMessageBox
        sender = self.sender()
        system = System.select().where(System.title_id == title_id())
        if sender == self.choice_one_table_Action: # одна таблица
            sys = system.select().where(System.stage == "Одна таблица").get()
            type = sys.type_table
            group = sys.total_group
            fin = "Одна таблица"
            check_flag = check_choice(fin) # проверка на жеребьевку True - значит сделана
            if check_flag  is True:
                reply = msg.information(my_win, 'Уведомление', f"Жеребъевка {fin} была произведена,"
                                                                            f"\nесли хотите сделать "
                                                                            "повторно\nнажмите-ОК-, "
                                                                            "если нет то - Cancel-",
                                                msg.Ok,
                                                msg.Cancel)
                if reply == msg.Ok:
                    if type == "круг":
                        id_system = system_id(stage=fin)
                        clear_db_before_choice_final(fin)
                        System.update(choice_flag=0).where(System.id == id_system).execute()
                        # player_fin_on_circle(fin) # создание жеребьевки по кругу
                        one_table(fin, group)
                        # player_in_one_table(fin)
                        # player_in_table_group_and_write_Game_list_Result(stage=fin)
                    else:
                        clear_db_before_choice_final(fin)
                        posev_data = player_choice_in_setka(fin)
                        player_in_setka_and_write_Game_list_and_Result(fin, posev_data)
                        load_combobox_filter_final()
                else:
                    return
            else:
                if type == "круг":
                    player_fin_on_circle(fin) # создание жеребьевки по кругу
                else:
                    posev_data = player_choice_in_setka(fin)
                    player_in_setka_and_write_Game_list_and_Result(fin, posev_data)
                    load_combobox_filter_final()
            add_open_tab(tab_page="Результаты")
        elif sender == self.choice_gr_Action:  # нажат подменю жеребъевка групп
            flag_checking = checking_before_the_draw() # проверка что все игроки зарегистрировались
            if flag_checking is False:
                return             
            for stage_sys in system:
                stage = stage_sys.stage
                if stage == "Предварительный":
                    fin = stage
                    check_flag = check_choice(fin)
                    if check_flag == True:
                        reply = msg.information(my_win, 'Уведомление',
                                                        "Жеребъевка была произведена,\nесли хотите сделать "
                                                        "повторно\nнажмите -ОК-, если нет то - Cancel-",
                                                        msg.Ok, msg.Cancel)

                        if reply == msg.Ok:
                            my_win.tabWidget.setCurrentIndex(2)
                            clear_db_before_choice(stage)
                            choice_gr_automat()
                            add_open_tab(tab_page="Результаты")
                            my_win.tabWidget.setCurrentIndex(3)
                            my_win.ed_etap_Action.setEnabled(True) # включает меню - редактирование жеребьеввки групп
                            return
                        else:
                            return
                    else:
                        my_win.tabWidget.setCurrentIndex(2)
                        choice_gr_automat()
                        add_open_tab(tab_page="Результаты")
                        my_win.tabWidget.setCurrentIndex(3)
                        my_win.ed_etap_Action.setEnabled(True) # включает меню - редактирование жеребьеввки групп
                        enabled_menu_after_choice()
                        return
        elif sender == self.choice_pf_Action: # подменю полуфиналы            
            stage = select_choice_semifinal()
            # +++ new
            id_system = system_id(stage)
            system_stage = system.select().where(System.id == id_system).get()
            # ======
            choice_flag = system_stage.choice_flag
            if stage is None: # если отмена при выборе жеребьевки
                return
            if choice_flag is True:
                reply = msg.information(my_win, 'Уведомление',
                                                "Жеребъевка была произведена,\nесли хотите сделать "
                                                "повторно\nнажмите -ОК-, если нет то - Cancel-",
                                                msg.Ok, msg.Cancel)

                if reply == msg.Ok:
                    clear_db_before_choice_semifinal(stage)
                    # === вставить ручной вид жеребьевки
                    choice_semifinal_automat(stage)
# ======= заполнение сыграныыми играми в группах
                    reply = msg.information(my_win, 'Уведомление', f"Хотите заполнить {stage} результатами "
                                                                            f"встреч, сыгранных в группах.",                                                                            
                                            msg.Ok,
                                            msg.Cancel)
                    if reply == msg.Ok:
                        load_playing_game_in_table_for_semifinal(stage)
                    else:
                        return
                    add_open_tab(tab_page="Результаты")
                    my_win.tabWidget.setCurrentIndex(4)
                    my_win.ed_etap_Action.setEnabled(True) # включает меню - редактирование жеребьеввки групп
                    return
                else:
                    return
            else:
                # проверяет все или игры в группе сыграны
                remains = 0
                if remains == 0:
                    choice_semifinal_automat(stage)
                    reply = msg.information(my_win, 'Уведомление', f"Хотите заполнить {stage} результатами "
                                                                            f"встреч, сыгранных в группах.",                                                                            
                                            msg.Ok,
                                            msg.Cancel)
                    if reply == msg.Ok:
                        load_playing_game_in_table_for_semifinal(stage)
                    else:
                        return
                    add_open_tab(tab_page="Результаты")
                    my_win.tabWidget.setCurrentIndex(4)
                    my_win.ed_etap_Action.setEnabled(True) # включает меню - редактирование жеребьеввки групп
        elif sender == self.choice_fin_Action:  # нажат подменю жеребьевка финалов
            fin = select_choice_final()
            if fin is None: # если отмена при выборе жеребьевки
                return
            fin_list = []
            stage_list = ["Одна таблица", "Предварительный", "1-й полуфинал", "2-й полуфинал"]
            for k in system:
                stage = k.stage
                if stage not in stage_list:
                    fin_list.append(stage)
            stage = fin
            id_system = system_id(stage)
            sys = system.select().where(System.id == id_system).get()
            type = sys.type_table
            kol_player_exit = sys.mesta_exit
            etap_exit = sys.stage_exit
            if etap_exit == "Предварительный":
                etap_replacing = etap_exit.replace("ый", "ом")
            elif etap_exit == "1-й полуфинал" or etap_exit == "2-й полуфинал":
                etap_replacing = etap_exit + "е"
            fin_replacing = fin.replace("й", "ого") + "а"
            if fin is not None:
                checking_flag = checking_possibility_choice(stage) # флаг жеребьевки этапа, если True значит все игры предварительного или полуфиналов сыграны
                if checking_flag is False:
                    return
                check_flag = check_choice(fin) # была ли сделана жеребьевка
                if check_flag is True:
                    reply = msg.information(my_win, 'Уведомление', f"Жеребъевка {fin} была произведена,"
                                                                        f"\nесли хотите сделать "
                                                                        "повторно\nнажмите-ОК-, "
                                                                        "если нет то - Cancel-",
                                                    msg.Ok,
                                                    msg.Cancel)
                    if reply == msg.Ok:
                        if type == "круг":
                            clear_db_before_choice_final(fin)
                            player_fin_on_circle(fin)
                            if kol_player_exit > 1:
                                reply = msg.information(my_win, 'Уведомление', f"Хотите заполнить игры {fin_replacing} результатами "
                                                                        f"встреч, сыгранных в {etap_replacing} этапе.",                                                                            
                                                    msg.Ok,
                                                    msg.Cancel)
                                if reply == msg.Ok:
                                    load_playing_game_in_table_for_final(fin)
                                else:
                                    return
                            add_open_tab(tab_page="Результаты")
                        else:
                            my_win.tabWidget.setCurrentIndex(3) 
                            clear_db_before_choice_final(fin)
                            posev_data = player_choice_in_setka(fin)
                            player_in_setka_and_write_Game_list_and_Result(fin, posev_data)
                            load_combobox_filter_final()
                            add_open_tab(tab_page="Результаты")
                            load_name_net_after_choice_for_wiev(fin)

                            my_win.statusbar.showMessage(
                                            f"Жеребъевка {fin} завершена успешно", 5000)
                            check_choice_net(fin) # == проверка жеребьевки сетки на 1-ю встречу игроков одного региона или одного тренера
                    else:
                        return
                else:
                    if type == "круг":
                        player_fin_on_circle(fin)
                        if kol_player_exit > 1:
                            reply = msg.information(my_win, 'Уведомление', f"Хотите заполнить игры {fin_replacing} результатами "
                                                                        f"встреч, сыгранных в {etap_replacing} этапе.",
                                                                            
                                                    msg.Ok,
                                                    msg.Cancel)
                            if reply == msg.Ok:
                                load_playing_game_in_table_for_final(fin)
                            else:
                                return                   
                    else:
                        posev_data = player_choice_in_setka(fin)
                        player_in_setka_and_write_Game_list_and_Result(fin, posev_data)
                        load_combobox_filter_final()
                        add_open_tab(tab_page="Результаты")
                        load_name_net_after_choice_for_wiev(fin)
                    my_win.tabWidget.setCurrentIndex(5) 
            else:
                return
        enabled_menu_after_choice()

    def system_made(self):
        system_competition()

    def system_clear(self):
        system_clear()
        my_win.tabWidget.setCurrentIndex(1)

    def help(self):
        pass

    def edit_etap(self):
        """редактирование жеребьевки этапов соревнования"""
        my_win.tabWidget.setCurrentIndex(6)
        my_win.tabWidget_2.setCurrentIndex(1)
        my_win.comboBox_first_group.clear()
        my_win.comboBox_second_group.clear()
        # my_win.tableView.hide()
        my_win.widget.show()

    def open(self):
        go_to()

    def view(self):
        view()

    def vid_edit(self):
        change_page_vid()

    
    def indent_edit(self):
        change_indent_page()

    def print_clear(self):
        """Печать чистых таблиц"""
        sender = self.sender()

        if sender == self.clear_s32_Action:
            setka_32_made(fin="1-й финал")
        elif sender == self.clear_s32_full_Action:
            setka_32_full_made(fin="1-й финал")
        elif sender == self.clear_s32_2_Action:
            setka_32_2_made(fin="1-й финал")
        elif sender == self.clear_s16_Action:
            setka_16_full_made(fin="1-й финал")
        elif sender == self.clear_s16_2_Action:
            setka_16_2_made(fin="1-й финал")
        elif sender == self.clear_s8_full_Action:
            setka_8_full_made(fin="1-й финал")
        elif sender == self.clear_s8_2_Action:
            setka_8_2_made(fin="1-й финал")
        view()

    def last(self):
        """открыте соревнований из пункта меню - последние-"""
        sender = self.sender()
        if sender == self.first_comp_Action:
            go_to()
        elif sender == self.second_comp_Action:
            go_to()
        elif sender == self.third_comp_Action:
            go_to()
        elif sender == self.fourth_comp_Action:
            go_to()
 
    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

['Breeze', 'Oxygen', 'QtCurve', 'Windows', 'Fusion']
app = QApplication(sys.argv)
my_win = MainWindow()
app.setStyle('Fusion')
my_win.setWindowTitle("Соревнования по настольному теннису")
my_win.setWindowIcon(QIcon("CTT.png"))
# my_win.resize(1390, 804)
my_win.resize(1390, 780)
my_win.center()

def dolg_R():
    dolg_R_list = []
    dolg_player = Player.select().where((Player.title_id == title_id()) & (Player.pay_rejting == "долг"))
    for k in dolg_player:
        family = k.player
        dolg_R_list.append(family)
    return dolg_R_list  


def unconfirmed_city():
    """список городов которые еще не подтвердились"""
    player_city_dict = {}
    city_player = Player.select().where((Player.title_id == title_id()) & (Player.application == "предварительная"))
    for k in city_player:
        city = k.city
        player = k.player
        player_city_dict[player] = city
    return player_city_dict 

class StartWindow(QMainWindow, Ui_Form):
    """Стартовое окно приветствия"""
    def __init__(self):
        super(StartWindow, self).__init__()
        self.setupUi(self)  # загружает настройки формы(окна) из QT
        self.setWindowTitle('Добро пожаловать в COMPETITIONS_TT')
        self.setWindowIcon(QIcon("CTT.png"))
        self.Button_open.clicked.connect(self.open)
        self.Button_new.clicked.connect(self.new)
        self.Button_view_pdf.clicked.connect(self.view_competition_on_arhive)
        self.Button_old.clicked.connect(self.last_competition)
        self.Button_R.clicked.connect(self.r_load)
        self.LinkButton.clicked.connect(self.last_comp)

        self.Button_open.setEnabled(False)
        self.Button_view_pdf.setEnabled(False)
        self.comboBox_arhive_year.setEnabled(False)
        self.comboBox_arhive_year.currentTextChanged.connect(self.choice_competition)
        # ========== проверяет создана ли база, если нет то создает
        conn = pymysql.connect(host='localhost', user='root', password='db_pass')
        md = conn.cursor()
        md.execute("SHOW DATABASES")
        databases = md.fetchall()
        database_exists = False
        for database in databases:
            if 'mysql_db' in database:
                database_exists = True
                break
        if database_exists:
            pass
        else:
            dbase()

        count = len(Title.select())
        if count != 0:
            # получение последней записи в таблице
            t_id = Title.select().order_by(Title.id.desc()).get()
            id = t_id.id
            old_title = Title.get(Title.id == id)
            last_comp = old_title.full_name_comp
            self.LinkButton.setText(f"{last_comp}")
        else:
            self.LinkButton.setText("Список прошедших соревнований пуст")
            self.LinkButton.setEnabled(False)
            self.Button_open.setEnabled(False)
            self.Button_old.setEnabled(False)   
    

    def last_comp(self):
        """открытие последних соревнований"""
        sex = ["Девочки", "Девушки", "Юниорки", "Женщины"]
        id_title = db_select_title()
        tab_enabled(id_title)
        title_new = Title.select().where(Title.id == id_title).get()
        gamer = title_new.gamer
        self.close()
        if gamer in sex:
            my_win.setStyleSheet("#MainWindow{background-color:lightpink}")
        else:
            my_win.setStyleSheet("#MainWindow{background-color:lightblue}")
        # === вставить  проверку DB ======      
        flag = check_delete_db()
        if flag == 0 or flag == 1:
            return
        else:
            delete_db_copy(del_files_list=flag)


    def open(self):
        """открытие соревнований из архива"""
        self.close() 
        my_win.resize(1110, 750)
        flag = check_delete_db()
        if isinstance(flag, list): # узнает принадлежит переменная к типу
           delete_db_copy(del_files_list=flag) 
        go_to() 
        my_win.show()


    def new(self):
        """запускает новые соревнования"""
        msgBox = QMessageBox
        result = msgBox.question(my_win, "", "Вы действительно хотите создать новые соревнования?",
                                 msgBox.Ok, msgBox.Cancel)
        if result == msgBox.Ok:
            gamer = ("Мальчики", "Девочки", "Юноши",
                     "Девушки", "Юниоры", "Юниорки", "Мужчины", "Женщины")
            gamer, ok = QInputDialog.getItem(
                my_win, "Участники", "Выберите категорию спортсменов", gamer, 0, False)

            title = Title(name="", sredi="", vozrast="", data_start="", data_end="", mesto="", referee="",
                          kat_ref="", secretary="", kat_sec="", gamer=gamer, full_name_comp="", pdf_comp="",
                          short_name_comp="", tab_enabled="Титул", multiregion="").save()
            # получение последней записи в таблице
            t_id = Title.select().order_by(Title.id.desc()).get()
            id_title = t_id.id
            my_win.lineEdit_title_gamer.setText(gamer)
            db_r(gamer)
            system = System(title_id=id_title, total_athletes=0, total_group=0, max_player=0, stage="", type_table="",
                            page_vid="", label_string="", kol_game_string="", choice_flag=False, score_flag=5,
                            visible_game=False, stage_exit="", mesta_exit=0, no_game="").save()
            self.close()
            tab_enabled(id_title)
            my_win.show()
        else:
            return

    def last_competition(self):
        """заполняет меню -последние- прошедшими соревнованиями 5 штук"""
        data_list = []
        fir_window.comboBox_arhive_year.clear()
        fir_window.comboBox_arhive_year.setEnabled(True)
        titles = Title.select()
        for t in titles:
            date_start = t.data_start
            str_data = date_start.strftime("%Y-%B")
            if str_data not in data_list:
                data_list.append(str_data)
        data_list.sort(reverse=True)
        data_list.insert(0, "-выберите дату-")
        fir_window.comboBox_arhive_year.addItems(data_list)

    def choice_competition(self):
        """выбор соревнования из архива"""
        full_name_list = []
        fir_window.comboBox.clear()
        index = fir_window.comboBox_arhive_year.currentIndex()
        data_text = fir_window.comboBox_arhive_year.currentText()
        if index > 0:
            date_object = datetime.strptime(data_text, '%Y-%B').date()
            end_date = date_object + relativedelta(months=1)
            title = Title.select().where((Title.data_start >= date_object) & (Title.data_start < end_date))
            for m in title:
                full_comp = m.full_name_comp
                age = m.vozrast
                full_name_list.append(f"{full_comp} {age}")
            fir_window.comboBox.addItems(full_name_list)
        fir_window.Button_open.setEnabled(True)
        fir_window.Button_view_pdf.setEnabled(True)

    def r_load(self):
        pass

    def load_old(self):
        """загружает в комбобокс архивные соревнования"""
        self.label_4.show()
        comp_list = []
        # ==== получение записи текущего соревнования
        id_current = Title.select().where(Title.id == title_id()).get()
        full_name_current = f"{id_current.full_name_comp} {id_current.vozrast}" # текущие соревнования
        # получение последней записи в таблице
        t_id = Title.select().order_by(Title.id.desc())
        count = len(t_id)
        n = 4
        if count != 0: 
            for i in t_id:
                if n == 0:
                    break
                else:                  
                    old_comp = i.name
                    gamer = i.gamer
                    age = i.vozrast
                    full_name = f"{i.full_name_comp} {age}" # соревнования, на которые переходит из меню

                    if old_comp != "":
                        name_comp = f"{old_comp}.{gamer} {age}"
                        self.comboBox.addItem(name_comp)
                        if full_name_current == full_name:
                            full_name = "Активно"
                    else:
                        full_name = "пусто"
                    comp_list.append(full_name)
                n -= 1
 
        else:       
            print("нет соревнований")
        if count > 3:
            my_win.first_comp_Action.setText(comp_list[0]) 
            my_win.second_comp_Action.setText(comp_list[1])
            my_win.third_comp_Action.setText(comp_list[2])
            my_win.fourth_comp_Action.setText(comp_list[3])
        elif count == 3:
            my_win.first_comp_Action.setText(comp_list[0]) 
            my_win.second_comp_Action.setText(comp_list[1])
            my_win.third_comp_Action.setText(comp_list[2])
        elif count == 2:
            my_win.first_comp_Action.setText(comp_list[0]) 
            my_win.second_comp_Action.setText(comp_list[1])
        elif count == 1:
            my_win.first_comp_Action.setText(comp_list[0])
               
        if fir_window.comboBox.currentText() != "":
            fir_window.Button_open.setEnabled(True)

    def view_competition_on_arhive(self):
        """Просмотр полного соревнования из архива"""
        msgBox = QMessageBox()
        full_name = fir_window.comboBox.currentText()
        t_id = Title.select().where(Title.full_name_comp == full_name).get()
        catalog = 2
        change_dir(catalog)
        view_file = t_id.pdf_comp
        flag = open_close_file(view_file)
        if flag is False:
            result = msgBox.information(my_win, "", "Такой файл не существует.\n"
                                                    "необходимо его создать!",
                                        msgBox.Ok)
            return
        else:
            if platform == "linux" or platform == "linux2":  # linux
                pass
            elif platform == "darwin":  # OS X
                os.system(f"open {view_file}")
            elif platform == "win32":  # Windows...
                os.system(f"{view_file}")
            os.chdir("..")
         

class ToolTip(): # создание всплывающих подсказок
    my_win.Button_made_R_file.setToolTip("Создание файла Excel для обсчета рейтинга")
    my_win.Button_made_one_file_pdf.setToolTip("Перед созданием одного файла, передвиньте строки с названием этапаов в необходимом порядке")

class ProgressBarThread(QThread):
    def __init__(self, fir_window, parent=None):
        super().__init__()
        countChanged = pyqtSignal(int)
        self.fir_window = fir_window
        self.ProgressBarThread_instance = ProgressBarThread(fir_window=self)

    def run(value):
        if value < 100:
            fir_window.progressBar.setValue(value)
        else:           
           fir_window.progressBar.setValue(0)


def check_delete_db():
    """Проверка сроков на удаления бэкап DB"""
    msgBox = QMessageBox
    current_date = datetime.now().strftime('%Y-%m-%d') # текущая дата в формате 01_01_2000
    del_files_list = []
    cur_date = datetime.strptime(current_date, '%Y-%m-%d')
    dir_path = pathlib.Path.cwd()
    parent_dir = str(f"{dir_path}\\backup_db")
    files = os.listdir(parent_dir)
    for f in files:
        znak = f.find("db")
        date_file = f[znak + 3:znak + 13]
        Year = date_file[6:]
        mon = date_file[3:5]
        day = date_file[:2]
        date_str = f"{Year}-{mon}-{day}"
        df = datetime.strptime(date_str, '%Y-%m-%d')
        time_difference = cur_date - df
        if time_difference > timedelta(days=3):
            del_files_list.append(f) 
    count = len(del_files_list)
    my_win.show()
    if count > 1:
        result = msgBox.information(my_win, "", "Есть бэкап DB срок создания более 3-х дней назад.\n"
                                                    "необходимо их удалить.",
                                    msgBox.Ok, msgBox.Cancel)
        if result == msgBox.Ok:
            flag = del_files_list
        else:
            flag = 1 # отмена
            return flag
    else:
        flag = 0 # нет старых баз
    return flag


def delete_db_copy(del_files_list):
    """Удаление копий базы данных старше 3-х дней"""
    from datetime import timedelta
    msgBox = QMessageBox
    txt_tmp = []
    for t in del_files_list:
        txt_date = t[14:24]
        db_txt = f"Бэкап DB от {txt_date}"
        txt_tmp.append(db_txt)
    text_str = (',\n'.join(txt_tmp)) # создание текстового файла по строчно из списка
    result = msgBox.question(my_win, "Бэкап DB", f"Вы действительно хотите удалить копии базы данных,\nкоторые были созданы более 3-х дней назад?\n\n{text_str}",
                                msgBox.Ok, msgBox.Cancel)
    if result == msgBox.Ok:
        dir_path = pathlib.Path.cwd()
        parent_dir = str(f"{dir_path}\\backup_db")
        for f in del_files_list:
            del_file = f"{parent_dir}\\{f}"
            os.remove(del_file)
    else:
        return


def dbase():
    """Создание DB и таблиц"""
    conn = pymysql.connect(host='localhost', user='root', password='db_pass')
    conn.cursor().execute('CREATE DATABASE mysql_db')
    conn.close()

    with db:
        db.create_tables([Title, R_list_m, R_list_d, Region, City, Player, R1_list_m, R1_list_d, Coach, System,
                          Result, Game_list, Choice, Delete_player, Referee, Patronymic])


def db_r(gamer):  # table_db присваивает по умолчанию значение R_list
    """переходит на функцию выбора файла рейтинга в зависимости от текущего или январского,
     а потом загружает список регионов базу данных"""
    msgbox = QMessageBox
    gamer_list = ["Мальчики", "Юноши", "Юниоры", "Мужчины"]
    if gamer in gamer_list:
        table_db = R_list_m
    else:
        table_db = R_list_d
    reply = msgbox.information(my_win, 'Уведомление', "Выберите файл с текущим рейтингом, \nзатем файл рейтинга за январь месяц.",
                                                  msgbox.Ok)
 
    fname = QFileDialog.getOpenFileName(
        my_win, "Выбрать файл R-листа", "", "Excel files(*.xls *.xlsx)")
    if fname == ("", ""):
        # получение последней записи в таблице
        title = Title.select().order_by(Title.id.desc()).get()
        system = System.get(Title.id == title)
        system.delete_instance()
        title.delete_instance()
        return
    control_R_list(fname, gamer)
    load_listR_in_db(fname, table_db)
    my_win.statusbar.showMessage("Текущий рейтинг загружен")
    if gamer in gamer_list:
        table_db = R1_list_m
        ext = "(*01_m.xlsx *01_m.xls)"
    else:
        table_db = R1_list_d
        ext = "(*01_w.xlsx *01_w.xls)"
    fname = QFileDialog.getOpenFileName(
        my_win, "Выбрать файл R-листа", "", f"Excels files {ext}")

    load_listR_in_db(fname, table_db)
    my_win.statusbar.showMessage("Январский рейтинг загружен")
    # добавляет в таблицу регионы
    # получение последней записи в таблице
    # === вариант если title id не номер один но соревнования первые
    titles = Title.select()
    count = len(titles)
    if count == 1:
        wb = op.load_workbook("regions.xlsx")
        s = wb.sheetnames[0]
        sheet = wb[s]
        reg = []
        for i in range(1, 86):
            a = sheet['B%s' % i].value
            region_mod = a.strip() # удаляет лишние пробелы в регионах из excel файла
            reg.append([region_mod])
        with db:
            Region.insert_many(reg).execute()
    region()
    # показывает статус бар на 5 секунд
    my_win.statusbar.showMessage("Список регионов загружен", 5000)
    my_win.lineEdit_title_nazvanie.hasFocus()


def control_R_list(fname, gamer):
    """проверка рейтинга текущему месяцу"""
    filepatch = str(fname[0])
    znak = filepatch.rfind("/")
    month_vybor = filepatch[znak + 6:znak + 8]
    d = date.today()
    current_month = d.strftime("%m")
    if current_month != month_vybor:
        message = "Вы выбрали файл с не актуальным рейтингом!\nесли все равно хотите его использовать, нажмите <Ок>\nесли хотите вернуться, нажмите <Cancel>"
        reply = QtWidgets.QMessageBox.information(my_win, 'Уведомление', message,
                                                  QtWidgets.QMessageBox.Ok,
                                                  QtWidgets.QMessageBox.Cancel)
        if reply == QMessageBox.Ok:
            return
        else:
            db_r(gamer)
    else:
        return


def load_listR_in_db(fname, table_db):
    """при отсутствии выбора файла рейтинга, позволяет выбрать вторично или выйти из диалога
    если выбор был сделан загружает в базу данных"""
    msgBox = QMessageBox
    step = 0
    filepatch = str(fname[0])
    if table_db == R_list_m or table_db == R_list_d:
        r = "текущим"
    elif table_db == R1_list_m or table_db == R1_list_d:
        r = "январским"
    if filepatch == "":
        message = f"Вы не выбрали файл с {r} рейтингом!\nесли хотите выйти, нажмите <Ок>\nесли хотите вернуться, нажмите <Cancel>"
        reply = msgBox.information(my_win, 'Уведомление', message,
                                                  msgBox.Ok,
                                                  msgBox.Cancel)
        if reply == msgBox.Ok:
            return
        else:
            db_r(table_db)
    else:
        data = []
        data_tmp = []

        rlist = table_db.delete().execute()
       
        excel_data = pd.read_excel(filepatch)  # читает  excel файл Pandas
        data_pandas = pd.DataFrame(excel_data)  # получает Dataframe
        # создает список заголовков столбцов
        column = data_pandas.columns.ravel().tolist()

        count = len(data_pandas)  # кол-во строк в excel файле

        count_column = len(column)
        if count_column == 5:
            data_list_new = []
            data_list = [""]
            data_list_new = data_list * count
            data_pandas["Субъект РФ"] = data_list_new
            data_pandas["Федеральный округ"] = data_list_new
            column = data_pandas.columns.ravel().tolist()

        for i in range(0, count):  # цикл по строкам
            pr = 100 * i / count
            rpr = math.ceil(pr)
            ProgressBarThread.run(value=rpr)
            for col in column:  # цикл по столбцам
                player_data = data_pandas.iloc[i][col]
                # заменяет пустые строки рейтинга на ноль и преобразовывает в тип int
                data_pandas['Рейтинг'] = data_pandas['Рейтинг'].fillna (0)
                data_pandas['Рейтинг'] = data_pandas['Рейтинг'].astype(int)
                data_tmp.append(player_data)  # получает временный список строки
            data.append(data_tmp.copy())  # добавляет в список Data
            data_tmp.clear()  # очищает временный список
        with db.atomic():
            for idx in range(0, len(data), 100):
                table_db.insert_many(data[idx:idx+100]).execute()


def region():
    """добавляет из таблицы в комбобокс регионы"""
    count = len(Region.select())
    if my_win.comboBox_region.currentIndex() > 0:  # проверка на заполненность комбобокса данными
        return
    else:
        with db:
            for r in range(1, count + 1):
                reg = Region.get(Region.id == r)
                my_win.comboBox_region.addItem(reg.region)


fir_window = StartWindow()  # Создаём объект класса ExampleApp
fir_window.show()  # Показываем окно
#  ==== наполнение комбобоксов ==========
page_orient = ("альбомная", "книжная")
kategoria_list = ("-выбор категории-", "2-я кат.", "1-я кат.", " ССВК")
mylist = ('мальчиков и девочек', 'юношей и девушек', 'мужчин и женщин')
raz = ("б/р", "3-юн", "2-юн", "1-юн", "3-р",
       "2-р", "1-р", "КМС", "МС", "МСМК", "ЗМС")
res = ("все игры", "завершенные", "не сыгранные")
vid_setki_one_table = ("-выбор типа таблицы-", "Сетка (-2)", "Сетка (с розыгрышем всех мест)",
             "Сетка (за 1-3 место)", "Круговая система")

my_win.comboBox_page_vid.addItems(page_orient)
my_win.comboBox_kategor_ref.addItems(kategoria_list)
my_win.comboBox_kategor_sec.addItems(kategoria_list)
my_win.comboBox_sredi.addItems(mylist)
my_win.comboBox_razryad.addItems(raz)
my_win.comboBox_filter_played.addItems(res)
my_win.comboBox_filter_played_sf.addItems(res)
my_win.comboBox_filter_played_fin.addItems(res)

my_win.comboBox_table_1.addItems(vid_setki_one_table)
my_win.comboBox_table_2.addItems(vid_setki_one_table)
my_win.comboBox_table_3.addItems(vid_setki_one_table)
my_win.comboBox_table_4.addItems(vid_setki_one_table)
my_win.comboBox_table_5.addItems(vid_setki_one_table)
my_win.comboBox_table_6.addItems(vid_setki_one_table)
my_win.comboBox_table_7.addItems(vid_setki_one_table)
my_win.comboBox_table_8.addItems(vid_setki_one_table)
my_win.comboBox_table_9.addItems(vid_setki_one_table)
my_win.comboBox_table_10.addItems(vid_setki_one_table)
my_win.comboBox_table_11.addItems(vid_setki_one_table)
my_win.comboBox_table_12.addItems(vid_setki_one_table)
# ставит сегодняшнюю дату в виджете календарь
my_win.dateEdit_start.setDate(date.today())
my_win.dateEdit_end.setDate(date.today())


def tab_enabled(id_title):
    """Включает вкладки в зависимости от создании системы и жеребьевки"""
    # включает вкладки меню системы
    title_list = []
    my_win.system_edit_Action.setEnabled(True) # делает меню  -редактировать- видиммым
    my_win.system_made_Action.setEnabled(True) # делает меню  -редактировать- видиммым

    sender = my_win.sender()
    tab_index = ["Титул", "Участники", "Система", "Результаты"]
    titles = Title.select().order_by(Title.id.desc())  # получает все title.id по убыванию
    title_new = Title.select().where(Title.id == id_title).get()
    vozrast = title_new.vozrast
    gamer = title_new.gamer
    name = title_new.name
    date_comp = title_new.data_start
    #=== сделать вариант с списком соревнований дев и юн
    n = 2
    t_name = titles.select().where((Title.name == name) & (Title.data_start == date_comp))
    t_age = t_name.select().where(Title.vozrast == vozrast)
    count = len(t_age)
    if count == 2:
        for k in t_age:
            if n != 0:  
                if gamer == k.gamer:
                    title_list.insert(0, k.id)
                else:
                    title_list.insert(1, k.id)
                n -= 1
            else:
                break
    else:
        for k in titles:
            if n != 0:  
                title_list.append(k.id)
                n -= 1
            else:
                break

    count_title = len(Title.select())
    title_id_current = title_list[0] # текущие соревнования
    title_id_last = title_list[1] if count_title > 1 else title_id_current

    if count_title > 1: # если соревнования не первые
        my_win.setWindowTitle(f"Соревнования по настольному теннису. {gamer} {vozrast}")
        # === new ===
        title_current = id_title
        tit_id = Title.get(Title.id == title_id_last) if title_current == title_id_current else Title.get(Title.id == title_id_current)
        old_comp = tit_id.name
        old_data = tit_id.data_start
        old_gamer = tit_id.gamer
        old_age = tit_id.vozrast
        comp = f"{old_comp}.{old_data}.{old_gamer} {old_age}" # соревнования предыдущие
        my_win.go_to_Action.setText(comp) # пункт меню -перейти к- соревнования предыдущие
        fir_window.load_old() # загружает в меню -последние- пять

    my_win.tabWidget.setTabEnabled(1, False)        
    my_win.tabWidget.setTabEnabled(2, False)
    my_win.tabWidget.setTabEnabled(3, False)
    my_win.tabWidget.setTabEnabled(4, False)
    my_win.tabWidget.setTabEnabled(5, False)
    my_win.tabWidget.setTabEnabled(6, True)
    # my_win.tabWidget.setTabEnabled(7, True)
# включает вкладки записаные в Титул
    tab_str = title_new.tab_enabled
    tab_list = tab_str.split(" ")
    for k in tab_list:
        ind = tab_index.index(k)
        my_win.tabWidget.setTabEnabled(ind, True)
        my_win.toolBox.setItemEnabled(ind, True)
    if gamer == "":
        gamer = my_win.lineEdit_title_gamer.text()
    my_win.toolBox.setCurrentIndex(0) # включает toolbox вкладку титул
    # Скрывает подменю системы в зависимости от созданной системы или нет
    if "Система" not in tab_list:
        my_win.system_edit_Action.setEnabled(False) # делает меню  -редактировать- не видиммым
    else:
        my_win.system_made_Action.setEnabled(False) # делает меню - создать- не видиммым
    enabled_menu_after_choice()


def add_open_tab(tab_page):
    """добавляет в таблицу -Title- список открытых вкладок"""
    tab_index = ["Титул", "Участники", "Система", "Результаты"]
    titles = Title.select().where(Title.id == title_id()).get()

    if tab_page != "":
        tab_str = titles.tab_enabled
        tab_list = tab_str.split(" ")

        if tab_page not in tab_list:
            tab_list.append(tab_page)        

        for k in tab_list:
            ind = tab_index.index(k)
            my_win.tabWidget.setTabEnabled(ind, True)
        tab_str = (' '.join(tab_list))
        #=====
        Title.update(tab_enabled = tab_str).where(Title.id == title_id()).execute()
        #======


def enabled_menu_after_choice():
    """Скрывает меню если еще не сделана жеребьевка"""
    systems = System.select().where(System.title_id == title_id())
    for k in systems:
        choice = k.choice_flag
        if choice is True:
            stage = k.stage
            if stage == "Одна таблица":
                my_win.view_one_table_Action.setEnabled(True)
            elif stage == "Предварительный":
                my_win.view_gr_Action.setEnabled(True)
            elif stage == "1-й полуфинал":
                my_win.view_pf1_Action.setEnabled(True)
            elif stage == "2-й полуфинал":
                my_win.view_pf2_Action.setEnabled(True)
            elif stage == "1-й финал":
                my_win.view_fin1_Action.setEnabled(True)
            elif stage == "2-й финал":
                my_win.view_fin2_Action.setEnabled(True)
            elif stage == "3-й финал":
                my_win.view_fin3_Action.setEnabled(True)
            elif stage == "4-й финал":
                my_win.view_fin4_Action.setEnabled(True)
            elif stage == "5-й финал":
                my_win.view_fin5_Action.setEnabled(True)
            elif stage == "6-й финал":
                my_win.view_fin6_Action.setEnabled(True)
            elif stage == "7-й финал":
                my_win.view_fin7_Action.setEnabled(True)
            elif stage == "8-й финал":
                my_win.view_fin8_Action.setEnabled(True)
            elif stage == "9-й финал":
                my_win.view_fin9_Action.setEnabled(True)
            elif stage == "10-й финал":
                my_win.view_fin10_Action.setEnabled(True)
            elif stage == "Суперфинал":
                my_win.view_superfin_Action.setEnabled(True)
            my_win.ed_etap_Action.setEnabled(True)
        stage = k.stage

        # for i in range(3):
        #     my_win.tabWidget_stage.setTabEnabled(i, False) # выключает вкладки фильтров на вкладки -Результаты-

        if stage == "Одна таблица":
            my_win.choice_one_table_Action.setEnabled(True)
            # my_win.tabWidget_stage.setTabEnabled(2, True) # включает вкладки фильтров на вкладки -Результаты-
        elif stage == "Предварительный":
            my_win.choice_gr_Action.setEnabled(True)
            # my_win.tabWidget_stage.setTabEnabled(0, True) # включает вкладки фильтров на вкладки -Результаты-
        elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
            my_win.choice_pf_Action.setEnabled(True)
            # my_win.tabWidget_stage.setTabEnabled(1, True) # включает вкладки фильтров на вкладки -Результаты-
        else:
            my_win.choice_fin_Action.setEnabled(True)
            # my_win.tabWidget_stage.setTabEnabled(2, True) # включает вкладки фильтров на вкладки -Результаты-


def db_insert_title(title_str):
    """Вставляем запись в таблицу титул"""
    msgBox = QMessageBox()
    nm = title_str[0]
    sr = title_str[1]
    vz = title_str[2]
    ds = title_str[3]
    de = title_str[4]
    ms = title_str[5]
    ref = title_str[6]
    sek = title_str[7]
    kr = title_str[8]
    ks = title_str[9]
    gm = title_str[10]
    fn = title_str[11]
    ind = title_str[12]

    reply = msgBox.question(my_win, "Уведомление", "Если соревнования межрегиональные\n нажмите -YES-, \n"
                                            "Если принимают участие спортсмены одного реиона\n нажмите -NO-", msgBox.Yes, msgBox.No)
    if msgBox.Yes:
        mr = 1
    else:
        mr = 0                                        
    reply = msgBox.question(my_win, "Уведомление", "Будет ли в списках участников\n присутствовать отчество? \n"
                                            "Если да нажмите -YES-\n нажмите -NO-", msgBox.Yes, msgBox.No)
    if msgBox.Yes:
        otc = 1
    else:
        otc = 0  

    short_name, ok = QInputDialog.getText(my_win, "Краткое имя соревнования", "Создайте краткое имя соревнования,\nдля"
                                          " отбражения в названии файла при "
                                          "сохранении,\nиспользуете латинские буквы"
                                          " без пробелов.\n"
                                          "В формате: название, возраст участников_дата,"
                                          " месяц, год и кто "
                                          "играет.")
    if ok:
        # получение последней записи в таблице
        # == создание папки соревнования === новое ===
        # os.mkdir(f"{short_name}")
        # =====
        t = Title.select().order_by(Title.id.desc()).get()
        title = Title(id=t, name=nm, sredi=sr, vozrast=vz, data_start=ds, data_end=de, mesto=ms, referee=ref,
                     kat_ref=kr, secretary=sek, kat_sec=ks, gamer=gm, full_name_comp=fn, pdf_comp="",
                     short_name_comp=short_name, multiregion=mr, perenos=ind, otchestvo=otc).save()
    else:
        return


def go_to():
    """переход на предыдущие соревнования и обратно при нажатии меню -перейти к- или из меню -последние-"""
    sender = my_win.sender()
    sex = ["Девочки", "Девушки", "Юниорки", "Женщины"]

    if sender == fir_window.Button_open:
        full_name_with_age = fir_window.comboBox.currentText()
    elif sender == my_win.first_comp_Action:
        full_name_with_age = my_win.first_comp_Action.text()
    elif sender == my_win.second_comp_Action:
        full_name_with_age = my_win.second_comp_Action.text()
    elif sender == my_win.third_comp_Action:
        full_name_with_age = my_win.third_comp_Action.text()
    elif sender == my_win.fourth_comp_Action:
        full_name_with_age = my_win.fourth_comp_Action.text()
    elif sender == my_win.go_to_Action:
        full_name_with_age = my_win.go_to_Action.text()  # полное название к которым переходим
      # ==== смена названия в меню -перейти к-
    t = Title.select().where(Title.id == title_id()).get()
    full_name_current = t.full_name_comp
    age_current = t.vozrast
    my_win.go_to_Action.setText(f"{full_name_current} {age_current}") # надпись на меню -перейти к- соревнования которые были

    mark = full_name_with_age.find("до") 
    if mark > 0: 
        full_name = full_name_with_age[:mark - 1]
        age = full_name_with_age[mark:]
    else:
        full_name = full_name_with_age
        age = ""

    titles = Title.get((Title.full_name_comp == full_name) & (Title.vozrast == age)) 
    id_title = titles.id # id соревнования на которое переходим
    gamer = titles.gamer
    # смена цвета фона формы в зависимости от пола играющих
    if gamer in sex:
        my_win.setStyleSheet("#MainWindow{background-color:lightpink}")
    else:
        my_win.setStyleSheet("#MainWindow{background-color:lightblue}")
        
    my_win.lineEdit_title_nazvanie.setText(titles.name)
    my_win.lineEdit_title_vozrast.setText(titles.vozrast)
    my_win.dateEdit_start.setDate(titles.data_start)
    my_win.dateEdit_end.setDate(titles.data_end)
    my_win.lineEdit_city_title.setText(titles.mesto)
    my_win.comboBox_referee.setCurrentText(titles.referee)
    my_win.comboBox_kategor_ref.setCurrentText(titles.kat_ref)
    my_win.comboBox_secretary.setCurrentText(titles.secretary)
    my_win.comboBox_kategor_sec.setCurrentText(titles.kat_sec)
    my_win.lineEdit_title_gamer.setText(titles.gamer)
    my_win.tabWidget.setCurrentIndex(0)  # открывает вкладку списки
     #===== new
    tab_enabled(id_title)
    player_list = Player.select().where(Player.title_id == id_title)
    count_player = len(player_list)
    my_win.label_46.setText(f"Всего: {count_player} участников")
    
    list_player_pdf(player_list)
    fir_window.load_old()


def db_select_title():
    """извлекаем из таблицы данные и заполняем поля титула для редактирования или просмотра"""
    sender = fir_window.sender()  # от какой кнопки сигнал
    if sender == my_win.go_to_Action:  # переход к соревнованиям из меню основного окна
        title = Title.get(Title.id == title_id())
        name = title.name
        data = title.data_start
        gamer_current = title.gamer
        # === вариант с ид титула =====
        id_title = title_id()
        # полное название текущих соревнований
        full_name_current = f"{name}.{data}.{gamer_current}"
        # присваиваем новый текст соревнований в меню -перейти к-
        my_win.go_to_Action.setText(full_name_current)
        gamer = title.gamer
    elif sender == my_win.toolBox or sender.text() != "Открыть":# переход от последнего соревнования в окне приветствия
        title = Title.get(Title.id == title_id())
        name = title.name
        gamer = title.gamer
        # === вариант с ид титула =====
        id_title = title_id()
    elif sender == my_win.new_comp_Action:
        title = Title.select().order_by(Title.id.desc()).get()
        name = title.name
        gamer = title.gamer
        id_title = title_id()
    # сигнал от кнопки с текстом -открыть- соревнования из архива (стартовое окно)
    else:
        txt = fir_window.comboBox.currentText()
        key = txt.rindex(".")
        gamer = txt[key +  1:]
        name = txt[:key - 11]
        # sroki = fir_window.label_4.text()
        data = txt[key - 10: key]
        titles = Title.select()
        for title in titles:
            name_title = title.name
            gamer_title = title.gamer
            data_title = str(title.data_start)
            if name == name_title and gamer == gamer_title:
                if data == data_title:
                    break
    if name != "":
        my_win.lineEdit_title_nazvanie.setText(title.name)
        my_win.lineEdit_title_vozrast.setText(title.vozrast)
        my_win.dateEdit_start.setDate(title.data_start)
        my_win.dateEdit_end.setDate(title.data_end)
        my_win.lineEdit_city_title.setText(title.mesto)
        my_win.comboBox_sredi.setCurrentText(title.sredi)
        my_win.comboBox_referee.setCurrentText(title.referee)
        my_win.comboBox_kategor_ref.setCurrentText(title.kat_ref)
        my_win.comboBox_secretary.setCurrentText(title.secretary)
        my_win.comboBox_kategor_sec.setCurrentText(title.kat_sec)
        my_win.lineEdit_title_gamer.setText(title.gamer)
    else:
        load_comboBox_referee()
    tab_enabled(id_title)
    return id_title


def system_made():
    """Заполняет таблицу система кол-во игроков, кол-во групп и прочее"""
    systems  = System.select().where(System.title_id == title_id()).get()
    count_system = len(systems)  # получение количества записей (этапов) в системе
    sg = my_win.comboBox_table.currentText()
    page_v = my_win.comboBox_page_1.currentText()
    total_group = systems.total_group
    total_athletes = systems.total_athletes
    max_player = systems.max_player
    if sg == "одна таблица":
        system = System(id=systems, title_id=title_id(), total_athletes=total_athletes, total_group=0,
                        max_player=0, stage=sg, page_vid=page_v, label_string="", kol_game_string="",
                        choice_flag=False, score_flag=5, visible_game=True, no_game="").save()
    else:  # предварительный этап
        for i in range(1, count_system + 1):
            system = System(id=systems, title_id=title_id(), total_athletes=total_athletes, total_group=total_group,
                            max_player=max_player, stage=sg, page_vid=page_v, label_string="", kol_game_string="",
                            choice_flag=False, score_flag=5, visible_game=True, no_game="").save()
    player_in_table_group_and_write_Game_list_Result()
    my_win.label_33.setText("Всего: 0 игр.")
    my_win.checkBox_2.setChecked(False)
    my_win.checkBox_3.setChecked(False)
    my_win.Button_system_made.setEnabled(False)


def r_list_load_tableView():
    my_win.lineEdit_find_player_in_R.clear()
    r_combo_index = my_win.comboBox_choice_R.currentIndex()  
    if r_combo_index == 0:
        fill_table_R_list() 
    else:
        fill_table_R1_list() 


def title_string():
    """ переменные строк титульного листа """
    title_str = []
    # получение последней записи в таблице
    title = Title.select().order_by(Title.id.desc()).get()

    nm = my_win.lineEdit_title_nazvanie.text()
     # ======== если длинное название перенос на две строки
    total_mark = len(nm)
    if total_mark > 60:
        nm_list = nm.split()
        word, ok = QInputDialog.getItem(my_win, "Название", "Выберите после какого слова\n"
                                        "перенести на другую строку", nm_list)
        ind = nm_list.index(word)
    # ====== 
    sr = my_win.comboBox_sredi.currentText()
    vz = my_win.lineEdit_title_vozrast.text()
    ds = my_win.dateEdit_start.text()
    de = my_win.dateEdit_end.text()
    ms = my_win.lineEdit_city_title.text()
    ref = my_win.comboBox_referee.currentText()
    sek = my_win.comboBox_secretary.currentText()
    kr = my_win.comboBox_kategor_ref.currentText()
    ks = my_win.comboBox_kategor_sec.currentText()
    gm = title.gamer
    fn = f"{nm}.{ds}.{gm}"

    title_str = [nm, sr, vz, ds, de, ms, ref, sek, kr, ks, gm, fn, ind]
    return title_str


def title_pdf():
    """сохранение в PDF формате титульной страницы"""
    msgBox = QMessageBox
    string_data = data_title_string()
    nz = my_win.lineEdit_title_nazvanie.text()
    sr = my_win.comboBox_sredi.currentText()
    vz = my_win.lineEdit_title_vozrast.text()
    ct = my_win.lineEdit_city_title.text()

    message = "Хотите добавить изображение в титульный лист?"
    reply = msgBox.question(my_win, 'Уведомление', message,
                                           msgBox.Yes,
                                           msgBox.No)
    if reply == msgBox.Yes:
        fname = QFileDialog.getOpenFileName(
            my_win, "Выбрать изображение", "/desktop", "Image files (*.jpg, *.png)")
        if fname[0] == "":
            return
        filepatch = str(fname[0])
    else:
        filepatch = None

    tit_id = Title.select().where(Title.id == title_id()).get()
    short_name = tit_id.short_name_comp
    canvas = Canvas(f"{short_name}_title.pdf", pagesize=A4)

    if filepatch == None:
        canvas.setFont("DejaVuSerif-Italic", 14)
        canvas.drawString(5 * cm, 28 * cm, "Федерация настольного тенниса России")
        canvas.drawString(3 * cm, 27 * cm, "Федерация настольного тенниса Нижегородской области")
        canvas.setFont("DejaVuSerif-Italic", 20)
        canvas.drawString(2 * cm, 23 * cm, nz)
        canvas.setFont("DejaVuSerif-Italic", 16)
        canvas.drawString(2.5 * cm, 22 * cm, f"среди {sr} {vz}")
        canvas.setFont("DejaVuSerif-Italic", 14)
        canvas.drawString(5.5 * cm, 5 * cm, f"г. {ct} Нижегородская область")
        canvas.drawString(7.5 * cm, 4 * cm, string_data)
    else:
        canvas.drawImage(filepatch, 7 * cm, 12 * cm, 6.9 * cm, 4.9 * cm,
                         mask=[0, 2, 0, 2, 0, 2])  # делает фон прозрачным
        canvas.setFont("DejaVuSerif-Italic", 14)
        canvas.drawString(5 * cm, 28 * cm, "Федерация настольного тенниса России")
        canvas.drawString(3 * cm, 27 * cm, "Федерация настольного тенниса Нижегородской области")
        canvas.setFont("DejaVuSerif-Italic", 20)
        canvas.drawString(2 * cm, 23 * cm, nz) # попробовать выравнить титул
        canvas.setFont("DejaVuSerif-Italic", 16)
        canvas.drawString(2.5 * cm, 22 * cm, f"среди {sr} {vz}")
        canvas.setFont("DejaVuSerif-Italic", 14)
        canvas.drawString(5.5 * cm, 5 * cm, f"г. {ct} Нижегородская область")
        canvas.drawString(7.5 * cm, 4 * cm, string_data)
    catalog = 1 # файл сохраяняется в каталоге /table_pdf
    change_dir(catalog)
    canvas.save()
    os.chdir("..")


def title_made():
    """создание титульного листа соревнования"""
    title_str = title_string()
    if my_win.Button_title_made.text() == "Редактировать":
        title_update()
        my_win.checkBox.setChecked(False)
        add_open_tab(tab_page="Участники")
        return
    else:
        db_insert_title(title_str)
    title_pdf()
    # после заполнения титула выключает чекбокс
    my_win.checkBox.setChecked(False)
    my_win.Button_title_made.setText("Создать")
    region()
    # получение последней записи в таблице
    title = Title.select().order_by(Title.id.desc()).get()

    # получение последнего id системы соревнования
    s = System.select().order_by(System.id.desc()).get()
    add_open_tab(tab_page="Участники")
    with db:
        System.create_table()
        sys = System(id=s, title_id=title, total_athletes=0, total_group=0, max_player=0, stage="", page_vid="",
                     label_string="", kol_game_string="", choice_flag=False, score_flag=5, visible_game=True, stage_exit="", mesta_exit="", no_game="").save()


def data_title_string():
    """получение строки начало и конец соревнований для вставки в титульный лист"""
    months_list = ("января", "февраля", "марта", "апреля", "мая", "июня", "июля",
                   "августа", "сентября", "октября", "ноября", "декабря")
    # получение последней записи в таблице
    title = Title.select().order_by(Title.id.desc()).get()
    datastart = str(title.data_start)
    dataend = str(title.data_end)
    ds = datastart[8:10]  # получаем число день из календаря
    ms = datastart[5:7]  # получаем число месяц из календаря
    ys = datastart[0:4]  # получаем число год из календаря
    me = dataend[5:7]
    de = dataend[8:10]
    month_st = months_list[int(ms) - 1]
    if de > ds:  # получаем строку начало и конец соревнования в
        # одном месяце или два месяца если начало и конец в разных месяцах
        return f"{ds}-{de} {month_st} {ys} г."
    elif de == ds:
        return f"{ds} {month_st} {ys} г."
    else:
        month_end = months_list[int(me) - 1]
        return f"{ds} {month_st}-{de} {month_end} {ys} г."


def title_update():
    """обновляет запись титула, если был он изменен"""
    title_str = title_string()
    nm = title_str[0]
    vz = title_str[2]
    ds = title_str[3]
    de = title_str[4]
    ms = title_str[5]
    ref = title_str[6]
    sek = title_str[7]
    kr = title_str[8]
    ks = title_str[9]
 
    Title.update(name=nm, vozrast=vz, data_start=ds, data_end=de, mesto=ms, referee=ref,
                 kat_ref=kr, secretary=sek, kat_sec=ks).where(Title.id == title_id()).execute()

    title_pdf()


def clear_filter_rejting_list():
    """сбрасывает данные фильтра на вкладкк -рейтинг-"""
    my_win.lineEdit_find_player_in_R.clear()
    my_win.comboBox_filter_region_in_R.setCurrentIndex(0)
    my_win.comboBox_filter_city_in_R.setCurrentIndex(0)
    my_win.comboBox_filter_date_in_R.setCurrentIndex(0)
    filter_rejting_list()


def find_in_rlist():
    """при создании списка участников ищет спортсмена в текущем R-листе"""
    tb = my_win.tabWidget.currentIndex()
    if my_win.checkBox_find_player.isChecked():
        find_in_player_list()
    else:
        r_data_m = [R_list_m, R1_list_m]
        r_data_w = [R_list_d, R1_list_d]
        t_id = Title.get(Title.id == title_id())
        gamer = t_id.gamer
        my_win.listWidget.clear()
        if tb == 6:
            cur_index = my_win.comboBox_choice_R.currentIndex()
            txt = my_win.lineEdit_find_player_in_R.text()
        else:
            my_win.textEdit.clear()
            txt = my_win.lineEdit_Family_name.text()

        zn = txt.find(" ")
        if zn != -1:
            family = txt[:zn]
            name = txt[zn + 1:]
            if name != "":
                family = family.capitalize()
                name = name.capitalize()  # Переводит первую букву в заглавную
                txt = f"{family} {name}"
        else:
            txt = txt.capitalize()  # Переводит первую букву в заглавную
        if gamer == "Девочки" or gamer == "Девушки" or gamer == "Юниорки"or gamer == "Женщины":
            if tb == 6 and cur_index == 0:
                r_data = r_data_w[0] # текущий рейтинг
            elif tb == 6 and cur_index == 1:
                r_data = r_data_w[1] # январский рейтинг
            else:
                r_data = r_data_w
        else:
            if tb == 6 and cur_index == 0:
                r_data = r_data_m[0]
            elif tb == 6 and cur_index == 1:
                r_data = r_data_m[1]
            else:
                r_data = r_data_m
        
        r = 0
        if tb == 6: # вкладка рейтинг
            if cur_index == 0:
                player_list = r_data.select().where(r_data.r_fname ** f'{txt}%')  # like поиск в текущем рейтинге
            else:
                player_list = r_data.select().where(r_data.r1_fname ** f'{txt}%')  # like поиск в январском рейтинге
        else:
            for r_list in r_data:
                p = r_list.select()
                if r == 0 :
                    my_win.label_63.setText("Поиск в текущем рейтинг листе.")
                    p = p.where(r_list.r_fname ** f'{txt}%')  # like поиск в текущем рейтинге
                    if r == 0  and len(p) != 0:
                        for pl in p:
                            full_stroka = f"{pl.r_fname}, {str(pl.r_list)}, {pl.r_bithday}, {pl.r_city}"
                            my_win.listWidget.addItem(full_stroka) # заполняет лист виджет спортсменами
                        return
                    elif r == 0:
                        r = 1
                        continue
                else:
                    my_win.label_63.setText("Поиск в январском рейтинге.")
                    p = p.where(r_list.r1_fname ** f'{txt}%')  # like поиск в январском рейтинге
                    if len(p) > 0:
                        for pl in p:
                            full_stroka = f"{pl.r1_fname}, {str(pl.r1_list)}, {pl.r1_bithday}, {pl.r1_city}"
                            my_win.listWidget.addItem(full_stroka) # заполняет лист виджет спортсменами
                    else:
                        full_stroka = ""
                        my_win.listWidget.addItem(full_stroka) # заполняет лист виджет спортсменами
                    return
      

def input_player():
    """Ввод нового игрока если его нет в рейтинг листе текущем и январском"""
    flaf_otc = 0
    msgBox = QMessageBox()
    text = my_win.lineEdit_Family_name.text()
    zn = text.find(" ")
    family = text[:zn]
    zn_1 = text.rfind(" ")
    if zn == zn_1: # значит нет отчество 
        flag_otc = 0       
        name = text[zn + 1:]        
    else:
        flag_otc = 1
        name = text[zn + 1:zn_1]
        otc = text[zn_1 + 1:]
        otc = otc.capitalize()  # Переводит первую букву в заглавную
    # family = family.capitalize()
    family = family.upper()
    name = name.capitalize() # Переводит первую букву в заглавную    
    f_name = f"{family} {name} {otc}" if flag_otc == 1 else f"{family} {name}"
    # повторная проверка игрока в январском рейтинге если два однофамильца и одинаковые имена 
    titles = Title.select().where(Title.id == title_id()).get()
    sex = titles.gamer
    woman_list = ["Девочки", "Девушки", "Юниорки", "Женщины"]
    r_data = R1_list_d if sex in woman_list else R1_list_m
    p = r_data.select().where(r_data.r1_fname == f_name)  # like поиск в январском рейтинге
    if len(p) > 0:
        my_win.listWidget.clear()
        my_win.label_63.setText("Поиск в январском рейтинге.")
        for pl in p:
            full_stroka = f"{pl.r1_fname}, {str(pl.r1_list)}, {pl.r1_bithday}, {pl.r1_city}"
            my_win.listWidget.addItem(full_stroka) # заполняет лист виджет спортсменами
        reply = msgBox.question(my_win, 'Уведомление', "Если это новый игрок\nнажмите -Yes-\n"
        "и заполните его данные,\nа если он есть\nв январском рейтинге\n нажмите -Cancel-",
                                                  msgBox.Yes,
                                                  msgBox.Cancel)
        if reply == msgBox.Yes:
            full_stroka = ""
            my_win.listWidget.addItem(full_stroka) # заполняет лист виджет спортсменами
            family = family.upper()
            my_win.lineEdit_Family_name.setText(f"{family} {name}")
            my_win.lineEdit_bday.setFocus()
            my_win.lineEdit_bday.setInputMask('00.00.0000')    
    else:
        full_stroka = ""
        my_win.listWidget.addItem(full_stroka) # заполняет лист виджет спортсменами
        family = family.upper()
        my_win.lineEdit_Family_name.setText(f"{f_name}")
        my_win.lineEdit_bday.setFocus()
        my_win.lineEdit_bday.setInputMask('00.00.0000')


def next_field():
    """переход к следующему полю ввода спортсмена"""
    my_win.lineEdit_R.setText('0')
    pl = my_win.lineEdit_Family_name.text()
    check_rejting_pay(pl)
    my_win.label_63.setText("Список городов.")
    my_win.lineEdit_city_list.setFocus()


def find_city():
    """Поиск городов и область"""
    city_list = []
    sender = my_win.sender()
    my_win.listWidget.clear()
    txt = my_win.label_63.text()
    city_field = my_win.lineEdit_city_list.text()
    if txt == "Список городов.":
        city_field = city_field.capitalize()  # Переводит первую букву в заглавную
        index = city_field.find(".")
        if index != -1:
            second_word = city_field[index + 1:]
            second_word = second_word.capitalize()
            city_field = city_field[:index + 1] + second_word
    
    c = City.select()
    c = c.where(City.city ** f'{city_field}%')  # like
    if sender != my_win.comboBox_region:
        if (len(c)) == 0:
            my_win.textEdit.setText("Нет такого города в базе, выберите регион где находится населенный пункт.")
            my_win.textEdit.setStyleSheet("Color: black")
            my_win.comboBox_region.setCurrentText("")
            return
        else:           
            for pl in c:
                full_stroka = f"{pl.city}"
                if full_stroka not in city_list:
                    city_list.append(full_stroka)
            my_win.listWidget.addItems(city_list) # заполняет лист виджет спортсменами
            return
    else:  # вставляет регион соответсвующий городу
        if city_field != "":
            citys = c.select().where(City.city == city_field)
            if len(citys) == 0:
                ir = my_win.comboBox_region.currentIndex()
                ir = ir + 1
                ct = my_win.lineEdit_city_list.text()
                with db:
                    city = City(city=ct, region_id=ir).save()


def dupl_coach(n_gr):
    """получает список тренеров в группе"""
    coach_list = []
    coach_tmp_list = []
    choices = Choice.select().where((Choice.title_id == title_id()) & (Choice.group == n_gr))
    for k in choices:
        coach = k.coach
        coach_list.append(coach)

    for l in coach_list:
        znak = l.find(",")
        if znak == -1: # один тренер
            coach_tmp_list.append(l)
        else: # у игрока не один тренер и делает из них список тренеров
            coach_1 = l[:znak]
            coach_tmp_list.append(coach_1)
            if l.find(",", znak) == -1:
                znak_1 = l.find(",", znak + 1)
                coach_2 = l[znak: znak_1]
                coach_tmp_list.append(coach_2)
            else:
                coach_2 = l[znak + 2:]
                znak_1 = l.find(",", znak + 1)
                if l.find(",", znak_1) == -1:
                    coach_tmp_list.append(coach_2)
                else:
                    coach_2 = l[znak + 2:znak_1]
                    coach_tmp_list.append(coach_2)
                    coach_3 = l[znak_1 + 2:]
                    coach_tmp_list.append(coach_3)
    return coach_tmp_list


def dupl_regions(n_gr):
    """получает список регионов в группе"""
    region_list = []
    choices = Choice.select().where((Choice.title_id == title_id()) & (Choice.group == n_gr))
    for k in choices:
        region = k.region
        region_list.append(region)
    return region_list


def fill_table(player_list):
    """заполняет таблицу со списком участников QtableView спортсменами из db"""
    data = []
    data_table_tmp = []
    data_table_list = []
    sender = my_win.sender()
    # start = time.time()
    model = MyTableModel(data)
    tb = my_win.tabWidget.currentIndex()

    player_selected = player_list.dicts().execute()
    
    row_count = len(player_selected)  # кол-во строк в таблице
    num_columns = [0, 1, 2, 3, 4, 5, 6]

    # кол-во наваний должно совпадать со списком столбцов
    if tb == 1: # == списки участников
        if my_win.checkBox_6.isChecked():
            num_columns = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
            model.setHorizontalHeaderLabels(['id','Фамилия Имя', 'ДР', 'R', 'Город', 'Регион', 'Разряд', 'Тренер', 'Место', 'id_del'])
        else:
            num_columns = [0, 1, 2, 3, 4, 5, 6, 7, 8]
            model.setHorizontalHeaderLabels(['id','Фамилия Имя', 'ДР', 'R', 'Город', 'Регион', 'Разряд', 'Тренер', 'Место']) 
    elif tb == 2: # Система
        stage = my_win.comboBox_filter_choice_stage.currentText()
        if my_win.comboBox_filter_choice_stage.currentIndex() == 0:
            num_columns = [0, 2, 3, 4, 7, 9, 10, 11, 13, 14, 16]
            model.setHorizontalHeaderLabels(['id','Фамилия Имя', 'Регион', 'Тренер', 'Группа', 'Место гр',
                                              'ПФ', "Группа ПФ", 'Место ПФ', 'Финал', 'Место'])
        elif stage == "Предварительный":
            num_columns = [0, 2, 3, 4, 5, 7, 9]
            model.setHorizontalHeaderLabels(['id','Фамилия Имя', 'Регион', 'Тренер', 'R', 'Группа', 'Место в гр'])
        elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
            num_columns = [0, 2, 3, 4, 5, 10, 11, 13]
            model.setHorizontalHeaderLabels(['id','Фамилия Имя', 'Регион', 'Тренер', 'R', 'ПФ', 'Группа ПФ', 'Место ПФ']) 
        else: 
            num_columns = [0, 2, 3, 4, 5, 14, 16]
            model.setHorizontalHeaderLabels(['id','Фамилия Имя', 'Регион', 'Тренер', 'R', 'Финал', 'Место в финале']) 
    elif tb == 3: # результаты
        num_columns = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
        model.setHorizontalHeaderLabels(['id',' Стадия', 'Группа', 'Встреча', '1-й игрок', '2-й игрок', 'Победитель', 'Очки','Общ. счет', 'Счет в партиях']) 
    elif tb == 5: # рейтинг
        model.setHorizontalHeaderLabels(['id',' Место', 'R', 'Фамилия Имя', 'Дата рождения', 'Город', 'Регион']) 
    elif tb == 6:
        if sender == my_win.lineEdit_find_player_stat:
            num_columns = [0, 1, 2, 3, 4, 5, 6, 7]
            model.setHorizontalHeaderLabels(['id','Фамилия Имя', 'ДР', 'R', 'Город', 'Регион', 'Разряд', 'Тренер']) 
        else:
            num_columns = [0, 1, 4, 5, 6, 7, 8]
            model.setHorizontalHeaderLabels(['id','Этап', 'Игрок-1', 'Игрок-2', 'Победитель', 'Тренер', ''])
    elif tb == 4:
        tb_double = my_win.tabWidget_3.currentIndex()
        if tb_double == 0:
            model.setHorizontalHeaderLabels(['id','Фамилия Имя', 'ДР', 'R', 'Город', 'Регион', 'Разряд', 'Тренер', 'Место']) 
        else:
           model.setHorizontalHeaderLabels(['id',' Стадия', 'Группа', 'Встреча', '1-й игрок', '2-й игрок', 'Победитель', 'Очки','Общ. счет', 'Счет в партиях'])  

    if tb == 1:
        if my_win.checkBox_15.isChecked():
            my_win.tableView.setSelectionMode(QAbstractItemView.MultiSelection) # выделение несколких строк по клику мышью
        else:
            my_win.tableView.setSelectionMode(QAbstractItemView.SingleSelection) # выделение одной строки по клику мышью
        my_win.tableView.setSelectionBehavior(QAbstractItemView.SelectRows) 
    elif tb == 3 or tb == 6:
        my_win.tableView.setSelectionMode(QAbstractItemView.SingleSelection) # выделение одной строки по клику мышью
        my_win.tableView.setSelectionBehavior(QAbstractItemView.SelectRows) # 
    else:
        my_win.tableView.setSelectionMode(QAbstractItemView.NoSelection) # нет выделение строк по клику мышью

    if tb == 5:
        if row_count > 0:
            my_win.label_78.setText(f"Поиск спортсмена в рейтинге: найдено всего {row_count} записей(и).")
        else:
            my_win.label_78.setText(f"Поиск спортсмена в рейтинге: не найдено ни одной записи.")

    if row_count != 0:  # список удаленных игроков пуст если R = 0
       
        for row in range(row_count):  # добавляет данные из базы в TableWidget
            item_1 = str(list(player_selected[row].values())[num_columns[0]])
            item_2 = str(list(player_selected[row].values())[num_columns[1]])
            item_3 = str(list(player_selected[row].values())[num_columns[2]])
            if tb == 1:
                item_3 = format_date_for_view(str_date=item_3) # преобразует дату к виду для экрана
            item_4 = str(list(player_selected[row].values())[num_columns[3]])
            item_5 = str(list(player_selected[row].values())[num_columns[4]])
            item_6 = str(list(player_selected[row].values())[num_columns[5]])
            item_7 = str(list(player_selected[row].values())[num_columns[6]])
            data_table_list = [item_1, item_2, item_3, item_4, item_5, item_6, item_7]
            if tb == 1:
                coach_id = str(list(player_selected[row].values())[num_columns[7]])
                coach = Coach.get(Coach.id == coach_id)
                item_8 = coach.coach
                item_9 = str(list(player_selected[row].values())[num_columns[8]])
                data_table_tmp = [item_8, item_9]
                if my_win.checkBox_6.isChecked():
                    item_10 = str(list(player_selected[row].values())[num_columns[9]])
                    data_table_tmp = [item_8, item_9, item_10]
                data_table_list.extend(data_table_tmp) 
            elif tb == 2:
                if my_win.comboBox_filter_choice_stage.currentIndex() == 0:
                    item_8 = str(list(player_selected[row].values())[num_columns[7]])
                    item_9 = str(list(player_selected[row].values())[num_columns[8]])
                    item_10 = str(list(player_selected[row].values())[num_columns[9]])
                    item_11 = str(list(player_selected[row].values())[num_columns[10]])
                    data_table_tmp = [item_8, item_9, item_10, item_11]
                elif stage == "Предварительный":
                    data_table_tmp = []
                elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
                    item_8 = str(list(player_selected[row].values())[num_columns[7]])
                    data_table_tmp = [item_8]
                else:
                    data_table_tmp = []
                data_table_list.extend(data_table_tmp) 
            elif tb == 3:
                item_8 = str(list(player_selected[row].values())[num_columns[7]])
                item_9 = str(list(player_selected[row].values())[num_columns[8]])
                item_10 = str(list(player_selected[row].values())[num_columns[9]])
                data_table_tmp = [item_8, item_9, item_10]
                data_table_list.extend(data_table_tmp)
            elif tb == 6:
                if sender != my_win.lineEdit_find_player_stat:
                    coach_id = str(list(player_selected[row].values())[num_columns[7]])
                    coach = Coach.get(Coach.id == coach_id)
                    item_8 = coach.coach
                    data_table_tmp = [item_8]
                data_table_list.extend(data_table_tmp)
            data.append(data_table_list.copy()) # данные, которые передаются в tableView (список списков)
        my_win.tableView.setModel(model)
        font = my_win.tableView.font()
        font.setPointSize(11)
        my_win.tableView.setFont(font)
        # my_win.tableView.setSortingEnabled(True)
        my_win.tableView.horizontalHeader().setFont(QFont("Times", 12, QFont.Bold)) # делает заголовки жирный и размер 13
        my_win.tableView.horizontalHeader().setStyleSheet("background-color:yellow;") # делает фон заголовков светлоголубой

        my_win.tableView.verticalHeader().setDefaultSectionSize(16) # высота строки 20 пикселей
        # my_win.tableView.resizeColumnsToContents() # растягивает по содержимому
        my_win.tableView.horizontalHeader().setStretchLastSection(True) # растягивает последнюю колонку до конца
        my_win.tableView.setGridStyle(QtCore.Qt.SolidLine) # вид линии сетки 
    else:
        if tb == 1:
            if my_win.checkBox_15.isChecked() and row_count == 0:
                my_win.statusbar.showMessage(
                "Нет спортсменов из предварительной заявки", 10000)
                my_win.textEdit.setText("Нет спортсменов из предварительной заявки")
            else:
                row = 0
                my_win.statusbar.showMessage(
                    "Нет спортсменов удаленных из списка", 10000)
                my_win.textEdit.setText("Нет спортсменов удаленных из списка")
                my_win.checkBox_6.setChecked(False)
        elif tb == 6:
            row = 0
            my_win.statusbar.showMessage(
                "Такого спортсмена в рейтинг листе нет нет", 10000)
    my_win.tableView.show()
    # finish = time.time()
    # res = finish - start
    # res_msec = res * 1000
    # print('Время работы в миллисекундах: ', res_msec)


def _fill_table(player_list): # ============== вариант эксперемнетальный =============
    """заполняет таблицу со списком участников QtableView спортсменами из db"""
    data = []
    header_list = []
    dict_sample = {}
    sender = my_win.sender()

    item_1_list = []
    item_2_list = []
    item_3_list = []
    item_4_list = []
    item_5_list = []
    item_6_list = []
    item_7_list = []
    item_8_list = []
    item_9_list = []
    item_10_list = []
    item_11_list = []
    item_12_list = []
    item_13_list = []
    item_14_list = []
    item_15_list = []
    item_16_list = []
    list_sample = [item_1_list, item_2_list, item_3_list, item_4_list,
                    item_5_list,  item_6_list, item_7_list, item_8_list, 
                    item_9_list, item_10_list, item_11_list, item_12_list,
                    item_13_list, item_14_list, item_15_list, item_16_list 
                    ]

    tb = my_win.tabWidget.currentIndex()
    player_list_mod = player_list.select() # выборка конкретых столбцов
    # выделение строк
    if tb == 1:
        if my_win.checkBox_15.isChecked():
            my_win.tableView.setSelectionMode(QAbstractItemView.MultiSelection) # выделение несколких строк по клику мышью
        else:
            my_win.tableView.setSelectionMode(QAbstractItemView.SingleSelection) # выделение одной строки по клику мышью
        my_win.tableView.setSelectionBehavior(QAbstractItemView.SelectRows) 
    elif tb == 3 or tb == 7:
        my_win.tableView.setSelectionMode(QAbstractItemView.SingleSelection) # выделение одной строки по клику мышью
        my_win.tableView.setSelectionBehavior(QAbstractItemView.SelectRows) # 
    else:
        my_win.tableView.setSelectionMode(QAbstractItemView.NoSelection) # нет выделение строк по клику мышью

    if tb == 6:
        if row_count > 0:
            my_win.label_78.setText(f"Поиск спортсмена в рейтинге: найдено всего {row_count} записей(и).")
        else:
            my_win.label_78.setText(f"Поиск спортсмена в рейтинге: не найдено ни одной записи.")

    # start = time.time()
    player_selected = player_list_mod.dicts().execute()
    row_count = len(player_selected)  # кол-во строк в таблице
    if tb == 1:
        num_columns = [0, 1, 2, 3, 4, 5, 6, 7, 8]
        header_list = ['id','Фамилия Имя', 'ДР', 'R', 'Город', 'Регион', 'Разряд', 'Тренер', 'Место']
    elif tb == 2:
        stage = my_win.comboBox_filter_choice_stage.currentText()
        if my_win.comboBox_filter_choice_stage.currentIndex() == 0:
            num_columns = [0, 2, 3, 4, 7, 9, 10, 11, 13, 14, 16]
            header_list = ['id','Фамилия Имя', 'Регион', 'Тренер', 'Группа', 'Место гр',
                                              'ПФ', "Группа ПФ", 'Место ПФ', 'Финал', 'Место']
        elif stage == "Предварительный":
            num_columns = [0, 2, 3, 4, 5, 7, 9]
            header_list = ['id','Фамилия Имя', 'Регион', 'Тренер', 'R', 'Группа', 'Место в гр']
        elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
            num_columns = [0, 2, 3, 4, 5, 10, 11, 13]
            header_list = ['id','Фамилия Имя', 'Регион', 'Тренер', 'R', 'ПФ', 'Группа ПФ', 'Место ПФ']
        else: 
            num_columns = [0, 2, 3, 4, 5, 14, 16]
            header_list = ['id','Фамилия Имя', 'Регион', 'Тренер', 'R', 'Финал', 'Место в финале']
    elif tb == 3:
        num_columns = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
        header_list = ['id',' Стадия', 'Группа', 'Встреча', '1-й игрок', '2-й игрок', 'Победитель', 'Очки','Общ. счет', 'Счет в партиях']
    elif tb == 5:
         num_columns = [0, 1, 4, 5, 6, 7, 8]
         header_list = ['id',' Место', 'R', 'Фамилия Имя', 'Дата рождения', 'Город', 'Регион']
    elif tb == 7:
        if sender == my_win.lineEdit_find_player_stat:
            num_columns = [0, 1, 2, 3, 4, 5, 6, 7]
            header_list = ['id','Фамилия Имя', 'ДР', 'R', 'Город', 'Регион', 'Разряд', 'Тренер']
        else:
            num_columns = [0, 1, 4, 5, 6, 7, 8]
            header_list = ['id','Этап', 'Игрок-1', 'Игрок-2', 'Победитель', 'Тренер', '']
    column_count = len(num_columns)

    for n in player_selected:
        val_list = list(n.values())
        p = 0
        for l in num_columns:
            while p < column_count:
                value = val_list[l]
                header = header_list[p]
                if header == "Тренер" and isinstance(value, int):
                    coach = Coach.get(Coach.id == value)
                    value = coach.coach 
                if isinstance(value, date): # узнать тип данных
                    value = format_date_for_view(str_date=value) # преобразует дату к виду для экрана
                elif value is None:
                    value = ''
                list_sample[p].append(value)
                p += 1
                break
    g = 0
    for k in num_columns:
        while g < column_count:
            dict_sample[g] = list_sample[g]
            g += 1

    data = pd.DataFrame(dict_sample) # данные которые передаются в модель
    model = MyTableModel(data)
    model.setHorizontalHeaderLabels(header_list) # список заголовков
    row_count = len(player_selected)  # кол-во строк в таблице
    if row_count != 0:  # список удаленных игроков пуст если R = 0
        font = my_win.tableView.font()
        font.setPointSize(11)
        my_win.tableView.setFont(font)
        my_win.tableView.horizontalHeader().setFont(QFont("Times", 12, QFont.Bold)) # делает заголовки жирный и размер 13
        my_win.tableView.horizontalHeader().setStyleSheet("background-color:yellow;") # делает фон заголовков светлоголубой
        # Установка размеров столбцов
        # header = table_view.horizontalHeader()
        # header.setSectionResizeMode(QHeaderView.Stretch)
        # # Установка размеров строкh
        # eader = table_view.verticalHeader()
        # header.setSectionResizeMode(QHeaderView.ResizeToContents)
        my_win.tableView.verticalHeader().setDefaultSectionSize(16) # высота строки 20 пикселей
        # my_win.tableView.setSortingEnabled(True)
        # my_win.tableView.setSectionResizeMode(QHeaderView.ResizeToContents)
        my_win.tableView.resizeColumnsToContents() # растягивает по содержимому
        my_win.tableView.horizontalHeader().setStretchLastSection(True) # растягивает последнюю колонку до конца
        my_win.tableView.setGridStyle(QtCore.Qt.SolidLine) # вид линии сетки 
    else:
        if tb == 1:
            if my_win.checkBox_15.isChecked() and row_count == 0:
                my_win.statusbar.showMessage(
                "Нет спортсменов из предварительной заявки", 10000)
                my_win.textEdit.setText("Нет спортсменов из предварительной заявки")
            else:
                my_win.statusbar.showMessage(
                    "Нет спортсменов удаленных из списка", 10000)
                my_win.textEdit.setText("Нет спортсменов удаленных из списка")
                my_win.checkBox_6.setChecked(False)
        elif tb == 6:
            my_win.statusbar.showMessage(
                "Такого спортсмена в рейтинг листе нет нет", 10000)
    # font = my_win.tableView.font()
    # font.setPointSize(11)
    # my_win.tableView.setFont(font)
    # my_win.tableView.horizontalHeader().setFont(QFont("Times", 12, QFont.Bold)) # делает заголовки жирный и размер 13
    # my_win.tableView.horizontalHeader().setStyleSheet("background-color:yellow;") # делает фон заголовков светлоголубой
    # my_win.tableView.verticalHeader().setDefaultSectionSize(16) # высота строки 20 пикселей
    # my_win.tableView.resizeColumnsToContents() # растягивает по содержимому
    # my_win.tableView.horizontalHeader().setStretchLastSection(True) # растягивает последнюю колонку до конца
    # my_win.tableView.setGridStyle(QtCore.Qt.SolidLine) # вид линии сетки 
    # my_win.tableView.resizeColumnsToContents() # растягивает по содержимому
   
    my_win.tableView.show()
    my_win.tableView.setModel(model)               
    # finish = time.time()
    # res = finish - start
    # res_msec = res * 1000
    # print('Время работы в миллисекундах: ', res_msec)


def fill_table_R_list():
    """заполняет таблицу списком из текущего рейтинг листа"""
    title = Title.select().where(Title.id == title_id()).get()
    gamer = title.gamer
    if gamer == "Девочки" or gamer == "Девушки" or gamer == "Юниорки" or gamer == "Женщины":
        player_list = R_list_d.select().order_by(R_list_d.r_fname)
    else:
        player_list = R_list_m.select().order_by(R_list_m.r_fname)
    # вставляет в таблицу необходимое кол-во строк
    row_count = len(player_list)
    my_win.label_78.setText(f"Всего {row_count} записей.") 
    fill_table(player_list)


def fill_table_R1_list():
    """заполняет таблицу списком из январского рейтинг листа"""
    title = Title.select().where(Title.id == title_id()).get()
    gamer = title.gamer
    if gamer == "Девочки" or gamer == "Девушки" or gamer == "Юниорки" or gamer == "Женщины":
        player_list = R1_list_d.select().order_by(R1_list_d.r1_fname)
    else:
        player_list = R1_list_m.select().order_by(R1_list_m.r1_fname)
    # вставляет в таблицу необходимое кол-во строк
    row_count = len(player_list)
    my_win.label_78.setText(f"Всего {row_count} записей.")
    fill_table(player_list)


def fill_table_results():
    """заполняет таблицу результатов QtableView из db result"""
    system_id_list = []

    system_stage_list = ["Одна таблица", "Предварительный", "1-й полуфинал", "2-й полуфинал"]
    result = Result.select().where(Result.title_id == title_id())
    system = System.select().where(System.title_id == title_id())
    tb = my_win.tabWidget.currentIndex()
    tab = my_win.tabWidget_stage.currentIndex()
    idx = my_win.tableView.currentIndex() # номер выделенной строки
    row_num = idx.row()
    stage = my_win.tableView.model().index(row_num, 1).data()
    # stage = my_win.tableView.model().index(row_num, 2).data()
    if tb == 3:
        if tab == 0:
            system_id = system.select().where(System.stage == stage).get()
            id_system = system_id.id
            player_list = result.select().where(Result.system_id == id_system)
        elif tab == 1:  # проверка есть ли записи в таблице -result
    # elif tb == 4:
            player_list = result.select().where((Result.system_stage == "1-й полуфинал") | (Result.system_stage == "2-й полуфинал")) # проверка есть ли записи в таблице -result-
        elif tab == 2:
    # elif tb == 5:
            for k in system: # заполняе список ид системы финальных этапов
                id_system = k.id
                if k.stage not in system_stage_list:
                    system_id_list.append(id_system)
            stage = my_win.comboBox_filter_final.currentText()

            if stage == "все финалы":      
                player_list = result.select().where(Result.system_stage == "Финальный")  # проверка есть ли записи в таблице -result 
            else:
                system_id = system.select().where(System.stage == stage).get()
                id_system = system_id.id
                player_list = result.select().where(Result.system_id == id_system)  # проверка есть ли записи в таблице -result 
                    
    fill_table(player_list)


# def fill_table_choice():
#     """заполняет таблицу жеребьевки"""
#     gamer = my_win.lineEdit_title_gamer.text()
#     player_choice = Choice.select().where(Choice.title_id == title_id()).order_by(Choice.rank.desc())
#     choice_list = player_choice.dicts().execute()
#     row_count = len(choice_list)  # кол-во строк в таблице
#     if row_count != 0:
#         column_count = len(choice_list[0])  # кол-во столбцов в таблице
#         # вставляет в таблицу необходимое кол-во строк
#         my_win.tableWidget.setRowCount(row_count)
#         for row in range(row_count):  # добавляет данные из базы в TableWidget
#             for column in range(column_count):
#                 item = str(list(choice_list[row].values())[column])
#                 my_win.tableWidget.setItem(
#                     row, column, QTableWidgetItem(str(item)))
#         # ставит размер столбцов согласно записям
#         my_win.tableWidget.resizeColumnsToContents()
#         for i in range(0, row_count):  # отсортировывает номера строк по порядку
#             my_win.tableWidget.setItem(i, 0, QTableWidgetItem(str(i + 1)))


def fill_table_after_choice():
    """заполняет TableWidget после жеребьевки групп"""
    choice = Choice.select().where(Choice.title_id == title_id())
    # === исправить сортировку по группам там они в текстовом варианте
    pl_choice = choice.select().order_by(Choice.group)
    player_list = pl_choice.select().order_by(Choice.posev_group)
    fill_table(player_list)


def debitor_R():
    """показывает список должников оплаты рейтинга"""
    player_list = Player.select().where(Player.title_id == title_id())
    player_debitor_R = player_list.select().where(Player.pay_rejting == "долг").order_by(Player.player)
    dolg = len(player_debitor_R)
    if dolg == 1:
        end_word = "к"
    elif dolg == 2 or dolg == 3 or dolg == 4:
        end_word = "ка"
    else:
        end_word = "ков"    
    if my_win.checkBox_11.isChecked():       
        if len(player_debitor_R) == 0:
            my_win.label_dolg_R.setText("Нет спортсменов без лицензии.")
            my_win.textEdit.setText("Спортсменов, не оплативших регистрационыый взнос за рейтинг нет.")
            my_win.label_dolg_R.setStyleSheet("color: black")
        else:
            my_win.label_dolg_R.setStyleSheet("color: red")
            my_win.label_dolg_R.setText(f"Без лицензии: {len(player_debitor_R)} участни{end_word}.")
        player_list = player_debitor_R
    else:
        my_win.Button_pay_R.setEnabled(False)
        my_win.textEdit.clear()
        my_win.lineEdit_Family_name.clear()
        my_win.lineEdit_bday.clear()
        my_win.lineEdit_R.clear()
        my_win.lineEdit_city_list.clear()
        my_win.lineEdit_coach.clear()
        my_win.lineEdit_id.clear()
    if len(player_debitor_R) > 0:
        my_win.label_dolg_R.setStyleSheet("color: red")
    else:
        my_win.label_dolg_R.setStyleSheet("color: black")
    my_win.label_dolg_R.setText(f"Без лицензии: {len(player_debitor_R)} участни{end_word}.")
    fill_table(player_list)


def add_player(): 
    """добавляет игрока в список и базу данных"""
    msgBox = QMessageBox()    
    flag = False
    player_list = Player.select().where(Player.title_id == title_id())
    txt = my_win.Button_add_edit_player.text()
    count = len(player_list)
    pl_id = my_win.lineEdit_id.text()
    pl = my_win.lineEdit_Family_name.text()

    otc = my_win.lineEdit_otchestvo.text()
    bd = my_win.lineEdit_bday.text()
    rn = my_win.lineEdit_R.text()
    ct = my_win.lineEdit_city_list.text()
    rg = my_win.comboBox_region.currentText()
    rg = rg.strip() # удаляет лишние пробелы
    rz = my_win.comboBox_razryad.currentText()
    ch = my_win.lineEdit_coach.text()
    player_data_list = [pl, bd, rn, ct, rg, rz, ch, otc]
    for i in player_data_list:
        if i == "":
            result = msgBox.information(my_win, "", "Вы заполнили не все поля данными игрока.",
                                            msgBox.Ok)
            return 
    # ===== проверка на возраст
    znak = bd.find(".")
    check_age_player(znak, bd)
    # =========
    if pl_id == "": # добавляет нового игрока
        flag = check_repeat_player(pl, bd) # проверка повторного ввода игрока
    else:
        if txt == "Редактировать":
            player = Player.select().where(Player.id == pl_id).get()
        else:
            player = Delete_player.select().where(Delete_player.id == pl_id).get()
        pay_R = player.pay_rejting
        comment = player.comment

    num = count + 1
    fn = f"{pl}/{ct}"
    if txt != "Редактировать":
        if flag is True: # если такой игрок присутствует очищает поля 
            my_win.lineEdit_Family_name.clear()
            my_win.lineEdit_bday.clear()
            my_win.lineEdit_R.clear()
            my_win.lineEdit_city_list.clear()
            my_win.lineEdit_coach.clear()
            my_win.lineEdit_otchestvo.clear()
            return
    add_coach(ch, num)
    add_patronymic()
    add_city()
    txt_edit = my_win.textEdit.toPlainText()
    ms = "" # записвыает место в базу как пустое
    idc = Coach.get(Coach.coach == ch) # получает id тренера
    idp = Patronymic.get(Patronymic.patronymic == otc)
    # === вставляет в таблицу player_full =========
    mark = pl.find(' ')
    fam = pl[:mark]
    name = pl[mark + 1:]
    bd =str(bd)
    day = bd[:2]
    month = bd[3:5]
    year = bd[6:]
    bd_new = f"{year}-{month}-{day}"
    flag_player_full = find_player_in_table_players_full(fam, name, ci=ct, bd=bd_new)
    if flag_player_full is None:
        player_full = Players_full(player=pl, bday=bd_new, city=ct, region=rg, razryad=rz, coach_id=idc, patronymic_id=idp).save()
    # ==== определяет завявка предварительная или нет
    title = Title.select().where(Title.id == title_id()).get()
    data_start = title.data_start
    date_current = date.today()
    zayavka = "предварительная" if date_current < data_start else "основная"

    if my_win.checkBox_6.isChecked():  # если отмечен флажок -удаленные-, то восстанавливает игрока и удаляет из
        # таблицы -удаленные-
        with db:
            player_del = Delete_player.get(Delete_player.id == pl_id)         
            pay_R = player_del.pay_rejting # оплачен ли рейтинг
            comment = player_del.comment # коментарий
            player_del.delete_instance()
            year = bd[6:]
            monh = bd[3:5]
            days = bd[:2]
            bd_mod = f"{year}-{monh}-{days}"
            plr = Player(player=pl, bday=bd_mod, rank=rn, city=ct, region=rg,
                         razryad=rz, coach_id=idc, full_name=fn, mesto=ms, title_id=title_id(), pay_rejting=pay_R,
                         comment=comment, coefficient_victories=0, total_game_player=0, total_win_game=0, patronymic_id=idp).save()
                         
        my_win.checkBox_6.setChecked(False)  # сбрасывает флажок -удаленные-
    else:  # просто редактирует игрока
        if txt == "Редактировать":
            # редактирует фамилии тренеров
            bd_new = format_date_for_db(str_date=bd)
            Coach.update(coach = ch).where(Coach.id == idc).execute()
            Player.update(player=pl, bday=bd_new, rank=rn, city = ct, region = rg, razryad = rz,
                            full_name=fn, pay_rejting=pay_R, comment=comment, coach_id = idc, patronymic_id=idp).where(Player.id == pl_id).execute()

            # with db:
            #     plr =  player_list.select().where(Player.id == pl_id).get()
            #     plr.player = pl                
            #     bd_new = format_date_for_db(str_date=bd)
            #     plr.bday = bd_new
            #     plr.rank = rn
            #     plr.city = ct
            #     plr.region = rg
            #     plr.razryad = rz
            #     plr.full_name = fn
            #     plr.pay_rejting = pay_R
            #     plr.comment = comment
            #     plr.save()
        elif txt == "Добавить":
            debt = "долг" if txt_edit == "Спортсмену необходимо оплатить рейтинг!" else ""
            # ==  перевод даты рождения в вид для db
            bd_new = format_date_for_db(str_date=bd)
            # =======
            with db:
                players = Player(player=pl, bday=bd_new, rank=rn, city=ct, region=rg, razryad=rz,
                                coach_id=idc, mesto="", full_name=fn, title_id=title_id(), pay_rejting=debt, comment="", 
                                coefficient_victories=0, total_game_player=0, total_win_game=0, application=zayavka, patronymic_id=idp).save()
            player_predzayavka = Player.select().where((Player.title_id == title_id()) & (Player.application == "предварительная"))
            count_pred = len(player_predzayavka)
            my_win.label_predzayavka.setText(f"По предзаявке: {count_pred} чел.")
            if debt == "долг":
                debitor_R()            
            # =========
            system = System.select().where(System.title_id == title_id())
            system_flag = ready_system() # проверка была создана система
            if system_flag is True:
                result = msgBox.information(my_win, "", "Колличество спортсменов изменилось.\n"
                        "Чтоб изменить число спортсменов в группах nнажмите -ОК-,\nесли обновить систему нажмите -Cancel-",
                                            msgBox.Ok, msgBox.Cancel)
                if result == msgBox.Ok:
                    system_stage = system.select().where(System.stage == "Предварительный").get()
                    system_id = system_stage.id
                    kg = system_stage.total_group
                    player_list = Player.select().where(Player.title_id == title_id())
                    count = len(player_list)  # количество записей в базе
                    # остаток отделения, если 0, то участники равно делится на группы
                    e1 = count % int(kg)
                    # если количество участников равно делится на группы (кол-во групп)
                    p = count // int(kg)
                    g1 = int(kg) - e1  # кол-во групп, где наименьшее кол-во спортсменов
                    g2 = int(p + 1)  # кол-во человек в группе с наибольшим их количеством
                    if e1 == 0:  # то в группах равное количество человек -e1-
                        stroka_kol_group = f"{kg} группы по {str(p)} чел."
                        skg = int((p * (p - 1) / 2) * int(kg))
                        mp = p
                    else:
                        stroka_kol_group = f"{str(g1)} групп(а) по {str(p)} чел. и {str(e1)} групп(а) по {str(g2)} чел."
                        skg = int((((p * (p - 1)) / 2 * g1) + ((g2 * (g2 - 1)) / 2 * e1))) # общее количество игр в группах
                        mp = g2
                    stroka_kol_game = f"{skg} игр"
                    my_win.label_19.setText(stroka_kol_game)
                    System.update(max_player=mp, label_string=stroka_kol_group, kol_game_string=stroka_kol_game).where(System.id == system_id).execute()
                    # ==== если просто добавить игрока последним номером в группу то не очищать -choice-
                    # choice_tbl_made()
                else: # обновление системы
                    pass
                n = 0
                for sys in system:
                    n += 1
                    count = len(system)
                    system_id = sys.id
                    athlet = sys.total_athletes
                    System.update(total_athletes=athlet + 1).where(System.title_id == title_id()).execute()
                    if n == count:
                        System.update(max_player=System.max_player + 1).where((System.id == system_id) & (System.title_id == title_id())).execute()

        pl_id = Player.select().order_by(Player.id.desc()).get() # id нового игрока
        player_id = pl_id.id

        # ======== попробовать вставить одну строку в tableView
    player_list = Player.select().where(Player.title_id == title_id())
    fill_table(player_list)
    count = len(player_list)  # подсчитывает новое кол-во игроков
    my_win.label_46.setText(f"Всего: {count} участников")
    list_player_pdf(player_list)
    my_win.lineEdit_id.clear()
    my_win.lineEdit_Family_name.clear()
    my_win.lineEdit_bday.clear()
    my_win.lineEdit_R.clear()
    my_win.lineEdit_city_list.clear()
    my_win.lineEdit_coach.clear()
    my_win.lineEdit_otchestvo.clear()
    if txt == "Редактировать":
        my_win.Button_add_edit_player.setText("Добавить")
        my_win.Button_del_player.setEnabled(False) 
        my_win.lineEdit_id.clear()       
    my_win.lineEdit_Family_name.setFocus()


def find_otchestvo():
    """ищет отчество в базе данных"""
    txt = my_win.label_63.text()
    my_win.listWidget.clear()
    if txt == "":
        return
    else:
        sex_list = ["Девочки", "Девушки", "Юниорки", "Женщины"]
        my_win.label_63.setText("Отчество")
        titles = Title.select().where(Title.id == title_id()).get()
        pol = titles.gamer
        sex = "w" if pol in sex_list else "m"
        otc = my_win.lineEdit_otchestvo.text()
        otc = otc.capitalize()  # Переводит первую букву в заглавную
        otchestvo_list = Patronymic.select()
        pat = otchestvo_list.where((Patronymic.patronymic ** f'{otc}%') & (Patronymic.sex == sex))  # like
        if (len(pat)) != 0:
            for chp in pat:
                full_stroka = chp.patronymic
                my_win.listWidget.addItem(full_stroka)
 

def format_date_for_db(str_date):
    """первод даты к формату базы данных год-месяц-день"""
    txt =str(str_date)
    day = str_date[:2]
    month = str_date[3:5]
    year = str_date[6:]
    new_str_date = f"{year}-{month}-{day}"
    format_date = datetime.strptime(new_str_date,'%Y-%m-%d')
    return  format_date


def format_date_for_view(str_date):
    """перевод даты к формату для отображения на экране"""
    txt = str(str_date)
    year = txt[:4]
    month = txt[5:7]
    day = txt[8:]  
    format_date = f"{day}.{month}.{year}"
    return format_date


def check_rejting_pay(pl):
    """Проверка игрока на оплату рейтинга и запись в базу данных"""
    txt_edit = my_win.textEdit.toPlainText()
    txt_tmp = my_win.label_63.text()
    if txt_tmp == "Поиск в январском рейтинге." or txt_edit == "":
        b_day = my_win.lineEdit_bday.text()
        year_player = int(b_day[6:])
        date_current = int(datetime.today().strftime("%Y"))
        raznica = date_current - year_player
        if raznica > 11:
            my_win.textEdit.setText("Спортсмену необходимо оплатить рейтинг!")
            my_win.textEdit.setStyleSheet("Color: red")
    elif txt_edit == "Спортсмену необходимо оплатить рейтинг!":
        plr = Player.select().where(Player.title_id == title_id())
        player_id = plr.select().where(Player.player == pl).get()
        with db:
            player_id.pay_rejting = "долг"
            player_id.comment = ""
            player_id.save()


def check_age_player(znak, dr):
    """Проверка возраста участника"""
    msgBox = QMessageBox()
    title = Title.get(Title.id == title_id())
    vozrast_text = title.vozrast
    if vozrast_text != "": # если играют не мужчины или женщины то проверка на соответсвия возраста
        text_1 = vozrast_text.rfind("моложе")
        text_date = vozrast_text[:2]
        if text_1 == -1 and text_date == "до":
            mark = vozrast_text.find(" ")
            total_old = int(vozrast_text[mark + 1:5])
            year_current = int(datetime.today().strftime("%Y")) # текущий год
            year_bday = year_current - total_old + 1
        elif text_1 > -1: # если возраст г.р и моложе
            year_bday = int(vozrast_text[:4])
            year_current = int(datetime.today().strftime("%Y")) # текущий год
        after_date = date(year_bday, 1, 1)
        if znak != -1:
            date_object = datetime.strptime(dr,"%d.%m.%Y")
        else:                    
            date_object = datetime.strptime(dr,"%Y-%m-%d")
        dr_year = int(date_object.strftime('%Y')) # получаем только год рождения в числовом формате
        current_date = date(dr_year, 1, 1)
        if after_date > current_date: # сравниваем две даты
            result = msgBox.information(my_win, "", "Возраст спортсмена не соответсвует\nвозрастной категории соревнования.\n"
                        "Возможно в рейтинге указана\nне правильная дата рождения.\nЕсли дата правильная нажмите -ОК-, или -Cancel-",
                                            msgBox.Ok, msgBox.Cancel)
            if result == msgBox.Ok:
                my_win.lineEdit_Family_name.setText("")               
                return   


def dclick_in_listWidget_double():
    """заносит пары игрок в tableWiev"""
    text = my_win.listWidget_double.currentItem().text()
    ds = len(text)
    sz = text.index(",")
    sz1 = text.index(",", sz + 1)
    fam_name = text[0:sz]
    znak = fam_name.find(" ")
    fam = fam_name[:znak]
    fam = fam.upper()
    name = fam_name[znak + 1:]
    name = name.capitalize()
    city = text[sz + 2:sz1]
    r = text[sz1 + 1:]
    a = my_win.r_pl1.text()
    if a == '':
        my_win.lineEdit_pl1_double.setText(fam_name)
        my_win.lineEdit_city_pl1.setText(city)
        my_win.r_pl1.setText(r)
    else:
        my_win.lineEdit_pl2_double.setText(fam_name)
        my_win.lineEdit_city_pl2.setText(city)
        my_win.r_pl2.setText(r)


def dclick_in_listwidget():
    """Находит фамилию спортсмена в рейтинге или фамилию тренера и заполняет соответсвующие поля списка"""
    msgBox = QMessageBox
    txt_tmp = my_win.label_63.text()
    text = my_win.listWidget.currentItem().text()
    coach_field = my_win.lineEdit_coach.text()
    if txt_tmp == "Список городов.": # если в listwidget список городов которые есть в базе
        my_win.label_63.setText("")
        my_win.lineEdit_city_list.setText(text)    
        cr = City.get(City.city == text)
        rg = Region.get(Region.id == cr.region_id)
        my_win.comboBox_region.setCurrentText(rg.region)
        my_win.listWidget.clear() 
    elif txt_tmp == "Отчество":
        my_win.label_63.setText("") 
        my_win.lineEdit_otchestvo.setText(text)
        my_win.listWidget.clear()
    elif coach_field == "": # если строка "тренер" пустая значит заполняются поля игрока
        ds = len(text)
        sz = text.index(",")
        sz1 = text.index(",", sz + 1)
        sz2 = text.index(",", sz1 + 1)
        fam_name = text[0:sz]
        znak = fam_name.find(" ")
        fam = fam_name[:znak]
        fam = fam.upper()
        name = fam_name[znak + 1:]
        name = name.capitalize()
        r = text[sz + 2:sz1]
        bd = text[sz1 + 2:sz2]
        znak = bd.find(".")       
        check_age_player(znak, bd)
        ci = text[sz2 + 2:ds] # город
        # ==== поиск игрока в базе данных и заполнение полей отчество, разряд и тренер
        player_full = find_player_in_table_players_full(fam, name, ci, bd)
        # ==== переводит строку с датой из базы даннных в строку к обычному виду
        if znak == -1:
            bd = format_date_for_view(str_date=bd)  
        # ==== проверка правильность даты для участия в турнире
        my_win.lineEdit_Family_name.setText(f"{fam} {name}")
        my_win.lineEdit_bday.setText(bd)
        my_win.lineEdit_R.setText(r)
        my_win.lineEdit_city_list.setText(ci)
        # =========== всатвляет данные если они есть в базе =============
        if player_full is not None:
            coaches = Coach.get(Coach.id == player_full[0])
            coach = coaches.coach
            my_win.lineEdit_coach.setText(coach)
            my_win.comboBox_razryad.setCurrentText(player_full[1])
            titles = Title.get(Title.id == title_id())
            flag_otc = titles.otchestvo
            if flag_otc == 1:
                patr = Patronymic.get(Patronymic.id == player_full[2])
                otc = patr.patronymic
                my_win.lineEdit_otchestvo.setText(otc)
         # ======= проверка на рейтинг ====
        if txt_tmp == "Поиск в январском рейтинге.":
            pl = fam_name
            check_rejting_pay(pl) # Если flag is False значит надо оплачивать рейтинг
        c = City.select().where(City.city == ci)  # находит город и соответсвующий ему регион
        if (len(c)) == 0:
            result = msgBox.information(my_win, "", "Выберите регион\nв котором находится город.",
                                            msgBox.Ok)
        else:  # вставляет регион соответсвующий городу
            cr = City.get(City.city == ci)
            rg = Region.get(Region.id == cr.region_id)
            my_win.comboBox_region.setCurrentText(rg.region)
            my_win.listWidget.clear()
        my_win.lineEdit_otchestvo.setFocus()
          # ======= проверка на рейтинг ====
    else:  # идет заполнение поля "тренер" из listWidget
        my_win.lineEdit_coach.setText(text)
        my_win.listWidget.clear()


def find_player_in_table_players_full(fam, name, ci, bd):
    """поиск игрока в базе данных, если он есть извлекает оттуда отчество, разряд и тренеров"""
    old_data = []
    p_full = Players_full.select().order_by(Players_full.player)
    fam_name = f"{fam} {name}"
    for pf in p_full:
        name_full = pf.player
        bd_full = str(pf.bday)
        city_full = pf.city
        if (fam_name == name_full and bd == bd_full):
            if ci == city_full:
                coach_full = pf.coach_id
                raz_full = pf.razryad
                patronymic_full = pf.patronymic_id
                old_data = [coach_full, raz_full, patronymic_full]
                return old_data

  



def load_combobox_filter_final():
    """заполняет комбобокс фильтр финалов для таблицы результаты"""
    my_win.comboBox_filter_final.clear()
    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    fin = []
    for sys in system:
        if sys.stage == "Одна таблица":
            if sys.choice_flag is True:
                fin.append(sys.stage)
                break
        else:
            stage = system.select().where((System.stage ** '%-й финал') | (System.stage == "Суперфинал"))
            fin = [k.stage for k in stage if k.choice_flag is True]
            fin.insert(0, "все финалы")
            break
    my_win.comboBox_filter_final.addItems(fin)
    my_win.comboBox_filter_choice_stage.addItems(fin)


def load_combobox_filter_group():
    """заполняет комбобокс фильтр групп для таблицы результаты"""
    etap = []
    gr_txt = []
    sender = my_win.menuWidget().sender()
    my_win.comboBox_filter_group.clear()
    my_win.comboBox_filter_choice_stage.clear()

    systems = System.select().where(System.title_id == title_id())  # находит system id последнего
    etap = [i.stage for i in systems ] # все этапы системы

    if etap[0] != "":
        fir_e = "Предварительный"
        flag = fir_e in etap # проверка есть ли в списке этап
        if flag == True:
            sf = systems.select().where(System.stage == fir_e).get()
            kg = int(sf.total_group)  # количество групп
        # if sender == my_win.choice_gr_Action or (my_win.tabWidget.currentIndex() == 2 and my_win.radioButton_gr_sort.isChecked()):
        if sender == my_win.choice_gr_Action or (my_win.tabWidget.currentIndex() == 2):
            gr_txt = [f"{i} группа" for i in range(1, kg + 1)]
            gr_txt.insert(0, "все группы")
            my_win.comboBox_filter_choice_stage.addItems(gr_txt)
        elif my_win.tabWidget.currentIndex() == 3:
            my_win.comboBox_filter_group.addItem("все группы")
            gr_txt = [f"{i} группа" for i in range(1, kg + 1)]
            my_win.comboBox_filter_group.addItems(gr_txt)


def load_combobox_filter_group_semifinal():
    """заполняет комбобокс фильтр групп для таблицы результаты"""
    sf_list = ["-все полуфиналы-"]
    gr_txt = []

    my_win.comboBox_filter_semifinal.clear()
    my_win.comboBox_filter_group_sf.clear()

    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    systems_sf = system.select().where(System.stage == "1-й полуфинал").get()
    kg = int(systems_sf.total_group)  # количество групп
    system_sf = system.select().where((System.stage == "1-й полуфинал") | (System.stage == "2-й полуфинал"))
    sf_list = [e.stage for e in system_sf]  # получает список этапов на данных соревнованиях
    my_win.comboBox_filter_semifinal.addItems(sf_list)

    my_win.comboBox_filter_group_sf.addItem("все группы")
    gr_txt = [f"{i} группа" for i in range(1, kg + 1)]
    my_win.comboBox_filter_group_sf.addItems(gr_txt)


def load_comboBox_filter_rejting():
    """Загружает комбобоксы вкладки рейтинг"""
    region_rejting = [""]
    city_rejting = [""]
    b_day = ["", "до 12 лет", "до 13 лет", "до 14 лет", "до 16 лет", "до 20 лет", "до 22 лет"]
    r_data_m = [R_list_m, R1_list_m]
    r_data_w = [R_list_d, R1_list_d]
    id_title = Title.select().where(Title.id == title_id()).get()
    gamer = id_title.gamer
    cur_index = my_win.comboBox_choice_R.currentIndex() # комбобокс выбора рейтинга
    if cur_index == 0: # если выбран текущий рейтинг
        if gamer == "Девочки" or gamer == "Девушки" or gamer == "Юниорки" or gamer == "Женщины":
            r_data = r_data_w[0]
        else:
            r_data = r_data_m[0] 
    elif cur_index == 1: # если рейтинг за январь
        if gamer == "Девочки" or gamer == "Девушки" or gamer == "Юниорки" or gamer == "Женщины":
            r_data = r_data_w[1]
        else:
           r_data = r_data_m[1]
    player_list = r_data.select()

    for k in player_list:
        region = k.r_region
        city = k.r_city
        if region not in region_rejting:
            region_rejting.append(region)
            region_rejting.sort()
        if city not in city_rejting:
            city_rejting.append(city)
            city_rejting.sort()
    my_win.comboBox_filter_region_in_R.addItems(region_rejting)
    my_win.comboBox_filter_city_in_R.addItems(city_rejting)
    my_win.comboBox_filter_date_in_R.addItems(b_day)


def tab():
    """Изменяет вкладку tabWidget в зависимости от вкладки toolBox"""
    tw = my_win.tabWidget.currentIndex()
    my_win.toolBox.setCurrentIndex(tw)


def tab_etap():
    """Включает или выключает вкладки на странице -Результаты-"""
    stage_list = []
    tab_etap = my_win.tabWidget_stage.currentIndex()
    results = Result.select().where(Result.title_id == title_id())    
    systems = System.select().where(System.title_id == title_id())
    for st in systems:
        stage = st.stage
        stage_list.append(stage)
    tab_result()
    if tab_etap == 0:
        player_list = results.select().where(Result.system_stage == "Предварительный")
        etap_text = "квалификации"
        load_combobox_filter_group()
    elif tab_etap == 1:        
        player_list = results.select().where((Result.system_stage == "1-й полуфинал") | (Result.system_stage == "2-й полуфинал"))
        etap_text = "полуфинальном этапе"
        load_combobox_filter_group_semifinal()
    elif tab_etap == 2:
        player_list = results.select().where(Result.system_stage == "Финальный")
        load_combobox_filter_final()
        etap_text = "финальном этапе"
        if "1-й финал" in stage_list:
            system = systems.select().where(System.stage == "1-й финал").get()
            flag_3_place = system.no_game
            if flag_3_place == "3": # отмечает checkBox если не разигрывается 3 место
                my_win.checkBox_no_play_3.setChecked(True)
    count = len(player_list)
    my_win.label_16.setText(f'Всего в {etap_text}\n{count} игры')
    my_win.label_16.show()
    fill_table(player_list)


def tab_result():
    """включает от вкладки надпись этапа"""
    tb = my_win.tabWidget_stage.currentIndex()
    if tb == 0:
        txt = "Предварительный этап"
        txt_tab = "ГР"
        my_win.tabWidget_stage.setTabText(1, "пф")
        my_win.tabWidget_stage.setTabText(2, "фин")
    elif tb == 1:
        txt = "Полуфинальный этап" 
        txt_tab = "ПФ"
        my_win.tabWidget_stage.setTabText(0, "гр")
        my_win.tabWidget_stage.setTabText(2, "фин")
    elif tb == 2:
        txt = "Финальный этап" 
        txt_tab = "ФИН"
        my_win.tabWidget_stage.setTabText(0, "гр")
        my_win.tabWidget_stage.setTabText(1, "пф")   
    my_win.tabWidget_stage.setTabText(tb, txt_tab)
    my_win.label_result.setText(txt)


def tab_double():
    """загружает в зависимости от выбранной вкладке"""
    sender = my_win.sender()
    my_win.listWidget_double.clear()
    tab_double = my_win.tabWidget_3.currentIndex()
    if tab_double == 0:
        player = Player.select().where(Player.title_id == title_id())
        if sender == my_win.lineEdit_pl1_double:
            txt = my_win.lineEdit_pl1_double.text()
            if txt == "":
                my_win.textEdit.clear()
        else:
            txt = my_win.lineEdit_pl2_double.text() 
        txt = txt.upper()
        pl = player.select().where(Player.player ** f'{txt}%')  # like
        if len(pl) > 0:
            for gamer in pl:
                full_stroka = f"{gamer.player}, {str(gamer.city)}, {gamer.rank}"
                my_win.listWidget_double.addItem(full_stroka) # заполняет лист виджет спортсменами
            return
        else:
            my_win.textEdit.setText("Такого спортсмена нет!")


def page_double():
    """Включает вкладку -пары- в зависимости от чекбокса на владке -система-"""
    if my_win.checkBox_double.isChecked():
        my_win.tabWidget.setTabEnabled(4, True)
    else: 
        my_win.tabWidget.setTabEnabled(4, False)


def tool_page():
    """Изменяет вкладку toolWidget в зависимости от вкладки tabWidget"""
    tb = my_win.toolBox.currentIndex()
    my_win.tabWidget.setCurrentIndex(tb)
    page()


def page():
    """Изменяет вкладку toolBox в зависимости от вкладки tabWidget"""
    msgBox = QMessageBox()
    tb = my_win.toolBox.currentIndex()
    sf = System.select().where(System.title_id == title_id())
    if tb == 0: # -титул-    
        my_win.resize(1110, 750)
        my_win.tabWidget_2.setGeometry(QtCore.QRect(260, 290, 841, 411)) # (точка слева, точка сверху, ширина, высота)
        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 841, 285))
        my_win.toolBox.setGeometry(QtCore.QRect(10, 40, 243, 659))
        my_win.comboBox_referee.setPlaceholderText("Введите фамилию судьи")
        my_win.comboBox_referee.setCurrentIndex(-1)
        my_win.comboBox_referee.setEditable(True)
        my_win.comboBox_secretary.setPlaceholderText("Введите фамилию судьи")
        my_win.comboBox_secretary.setCurrentIndex(-1)
        my_win.comboBox_secretary.setEditable(True)
        db_select_title()
        my_win.tableView.hide()
        my_win.tableWidget.hide()
        my_win.widget.hide()
    elif tb == 1:  # -список участников-
        my_win.checkBox_15.setChecked(False)
        my_win.tabWidget_2.setCurrentIndex(0)
        my_win.resize(1110, 750)
        my_win.tabWidget_2.setGeometry(QtCore.QRect(260, 225, 841, 473))
        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 841, 221))
        my_win.toolBox.setGeometry(QtCore.QRect(10, 40, 243, 659))
        load_coach_to_combo()
        load_comboBox_filter()
        region()
        titles = Title.select().where(Title.id == title_id()).get()
        otc = titles.otchestvo
        if otc == 1:
            my_win.lineEdit_otchestvo.setVisible(True)
            my_win.label_28.setVisible(True)
        else:
            my_win.lineEdit_otchestvo.setVisible(False)
            my_win.label_28.setVisible(False)
        my_win.Button_app.setEnabled(False)
        my_win.Button_del_player.setEnabled(False)
        my_win.Button_clear_del.setEnabled(False)
        my_win.Button_pay_R.setEnabled(False)
        my_win.Button_add_edit_player.setText("Добавить")
        my_win.statusbar.showMessage("Список участников соревнований", 5000)
        player_list = Player.select().where((Player.title_id == title_id()) & (Player.bday != "0000-00-00"))

        player_debitor_R = Player.select().where((Player.title_id == title_id()) & (Player.pay_rejting == "долг"))
        player_predzayavka = Player.select().where((Player.title_id == title_id()) & (Player.application == "предварительная"))
        count_debitor_R = len(player_debitor_R)
        count_pred = len(player_predzayavka)
        num_debitor_1 = [1]
        num_debitor_2 = [2, 3, 4]
        if count_debitor_R in num_debitor_1:
            end_word = "участник"
        elif count_debitor_R in num_debitor_2:
            end_word = "участника"
        else:
            end_word = "участников"
        dolg_R()
        fill_table(player_list)  # заполняет TableWidget списком игроков
        count = len(player_list)
        my_win.label_46.setText(f"Всего: {count} участников")
        if count_debitor_R == 0:
            my_win.label_dolg_R.setStyleSheet("color: black")
        else:
            my_win.label_dolg_R.setStyleSheet("color: red")
        my_win.label_dolg_R.setText(f"Без лицензии: {count_debitor_R} {end_word}")
        my_win.label_predzayavka.setText(f"По предзаявке: {count_pred} чел.")
        list_player_pdf(player_list)
        my_win.widget.hide()
        my_win.tableWidget.hide()
    elif tb == 2:  # -система-
        my_win.tabWidget_2.setCurrentIndex(0)
        my_win.resize(1110, 750)
        my_win.tabWidget_2.setGeometry(QtCore.QRect(260, 318, 841, 384))
        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 841, 320))
        my_win.toolBox.setGeometry(QtCore.QRect(10, 40, 243, 659))
        my_win.checkBox_repeat_regions.setChecked(False)
    
        my_win.tableView.setEditTriggers(QAbstractItemView.NoEditTriggers) # запрет редактирования таблицы
        result = Result.select().where(Result.title_id == title_id())
        result_played = result.select().where(Result.winner != "")
        count_result = len(result_played)

        player_list_main = Player.select().where((Player.title_id == title_id()) & (Player.bday != "0000-00-00"))
        count = len(player_list_main)
        my_win.label_8.setText(f"Всего участников: {str(count)} человек")
        my_win.label_52.setText(f"Всего сыграно: {count_result} игр.")
        label_playing_count() # пишет сколько игр сыграно в каждом этапе
        my_win.comboBox_filter_number_group_final.setEnabled(False)
        for k in sf:
            stage = k.stage
            if stage == "Предварительный":
                flag_choice = ready_choice(stage)
                if flag_choice is True:
                    choice_filter_on_system()
                break
        my_win.label_etap_1.hide()
        my_win.label_etap_2.hide()
        my_win.label_etap_3.hide()
        my_win.label_etap_4.hide()
        my_win.label_etap_5.hide()
        my_win.label_etap_6.hide()
        my_win.label_etap_7.hide()
        my_win.label_etap_8.hide()
        my_win.label_etap_9.hide()
        my_win.label_etap_10.hide()
        my_win.label_etap_11.hide()
        my_win.label_etap_12.hide()
        my_win.label_101.hide()
        my_win.label_11.hide()
        my_win.label_12.hide()
        my_win.label_19.hide()
        my_win.label_102.hide()
        my_win.label_27.hide()
        my_win.label_30.hide()
        my_win.label_31.hide()
        my_win.label_103.hide()
        my_win.label_104.hide()
        my_win.label_105.hide()
        my_win.label_53.hide()       
        my_win.label_58.hide()
        my_win.label_106.hide()
        my_win.label_107.hide()
        my_win.label_108.hide()
        my_win.label_109.hide()
        my_win.label_110.hide()
        my_win.label_111.hide()
        my_win.label_112.hide()
        my_win.label_81.hide()
        my_win.label_82.hide()
        my_win.label_83.hide()
        my_win.label_84.hide()
        my_win.label_85.hide()
        my_win.label_86.hide()
        my_win.label_87.hide()

        my_win.comboBox_table_1.hide()
        my_win.comboBox_table_2.hide()
        my_win.comboBox_table_3.hide()
        my_win.comboBox_table_4.hide()
        my_win.comboBox_table_5.hide()
        my_win.comboBox_table_6.hide()
        my_win.comboBox_table_7.hide()
        my_win.comboBox_table_8.hide()
        my_win.comboBox_table_9.hide()
        my_win.comboBox_table_10.hide()
        my_win.comboBox_table_11.hide()
        my_win.comboBox_table_12.hide()

        my_win.spinBox_kol_group.hide()
        stage_list = []
        table = []
        game = []
        sum_game = []

        for i in sf:  # цикл по таблице -system-
            total_player = i.total_athletes
            if total_player == 0: # если система только начала создаваться
                return
            stage_list.append(i.stage)  # добавляет в список этап
            table.append(i.label_string)  # добавляет в список система
            game.append(i.kol_game_string)  # добавляет в список кол-во игр
        count = len(stage_list)
        for i in range(0, count):  # подсчитывает сумму игр
            txt = game[i]
            t = txt.find(" ")
            number_games = int(txt[:t])
            sum_game.append(number_games)
            if i == 0:  # показывает в зависимости от этапов финал, кол-во игр
                my_win.label_101.setText(stage_list[0])
                my_win.label_19.setText(game[0])
                my_win.label_12.setText(table[0])
                my_win.label_101.show()
                my_win.label_12.show()
                my_win.label_19.show()
            elif i == 1:
                my_win.label_102.setText(stage_list[1])
                my_win.label_27.setText(game[1])
                my_win.label_etap_2.setText(table[1])
                my_win.label_102.show()
                my_win.label_27.show()
                my_win.label_etap_2.show()
            elif i == 2:
                my_win.label_103.setText(stage_list[2])
                my_win.label_30.setText(game[2])
                my_win.label_31.setText(table[2])
                my_win.label_30.show()
                my_win.label_31.show()
                my_win.label_103.show()
            elif i == 3:
                my_win.label_104.setText(stage_list[3])
                my_win.label_53.setText(game[3])
                my_win.label_etap_4.setText(table[3])
                my_win.label_104.show()
                my_win.label_53.show()
                my_win.label_etap_4.show()
            elif i == 4:
                my_win.label_105.setText(stage_list[4])
                my_win.label_58.setText(game[4])
                my_win.label_etap_5.setText(table[4])
                my_win.label_105.show()
                my_win.label_58.show()
                my_win.label_etap_5.show()
            elif i == 5:
                my_win.label_106.setText(stage_list[5])
                my_win.label_81.setText(game[5])
                my_win.label_etap_6.setText(table[5])
                my_win.label_106.show()
                my_win.label_81.show()
                my_win.label_etap_6.show()
            elif i == 6:
                my_win.label_107.setText(stage_list[6])
                my_win.label_82.setText(game[6])
                my_win.label_etap_7.setText(table[6])
                my_win.label_107.show()
                my_win.label_82.show()
                my_win.label_etap_7.show()
            elif i == 7:
                my_win.label_108.setText(stage_list[7])
                my_win.label_83.setText(game[7])
                my_win.label_etap_8.setText(table[7])
                my_win.label_108.show()
                my_win.label_83.show()
                my_win.label_etap_8.show()
            elif i == 8:
                my_win.label_109.setText(stage_list[8])
                my_win.label_84.setText(game[8])
                my_win.label_etap_9.setText(table[8])
                my_win.label_109.show()
                my_win.label_84.show()
                my_win.label_etap_9.show()
            elif i == 9:
                my_win.label_110.setText(stage_list[9])
                my_win.label_85.setText(game[9])
                my_win.label_etap_10.setText(table[9])
                my_win.label_110.show()
                my_win.label_85.show()
                my_win.label_etap_10.show()
            elif i == 10:
                my_win.label_111.setText(stage_list[10])
                my_win.label_86.setText(game[10])
                my_win.label_etap_11.setText(table[10])
                my_win.label_111.show()
                my_win.label_86.show()
                my_win.label_etap_11.show()
            elif i == 11:
                my_win.label_112.setText(stage_list[11])
                my_win.label_87.setText(game[11])
                my_win.label_etap_12.setText(table[11])
                my_win.label_112.show()
                my_win.label_87.show()
                my_win.label_etap_12.show()

            total_game = sum(sum_game)
            my_win.comboBox_table_1.hide()
            my_win.comboBox_page_vid.setEnabled(False)
            my_win.Button_etap_made.setEnabled(False)
            my_win.Button_system_made.setEnabled(False)
            my_win.label_33.setText(f"Всего {total_game} игр")
            my_win.label_33.show()
            # сделать правильную сортировку по группам
        player_list = Choice.select().where((Choice.title_id == title_id()) & (Choice.family != "x"))
        fill_table(player_list)
        my_win.widget.hide()
        my_win.tableWidget.hide()
    elif tb == 3:  # вкладка -результаты-       
        for l in range (0, 3): # сначала выключает вкладки этапов
            my_win.tabWidget_stage.setTabEnabled(l, False)
        pf_list = ["1-й полуфинал", "2-й полуфинал"]
        my_win.tabWidget_2.setCurrentIndex(0)
        # выключить вкладки этапы если еще не было жеребьевки
        choice_etap = []
        system_etap = []
        for k in sf:
            ch_flag = k.choice_flag
            sys_etap = k.stage
            system_etap.append(sys_etap)
            if ch_flag == 1:
                choice_etap.append(k.stage)
        for m in choice_etap:
            if m == "Предварительный":
                my_win.tabWidget_stage.setTabEnabled(0, True)
                index = 0
            elif m in pf_list:
                my_win.tabWidget_stage.setTabEnabled(1, True)
                index = 1
            else:
                my_win.tabWidget_stage.setTabEnabled(2, True)
                index = 2
        if len(choice_etap) > 0: # если была жеребьевка этапов, то включает вкладку
            my_win.tabWidget_stage.setCurrentIndex(index)
        # определяет есть соревнования из одной таблице,то включает вкладку -финалы- иначе -группы-
        if "Одна таблица" in system_etap:
            a = 2 # включает вкл финалы
        else:
            a = 0 # включает вкл группы
        my_win.tabWidget_stage.setCurrentIndex(a)
        tb_etap = my_win.tabWidget_stage.currentIndex()

        Button_view = QPushButton(my_win.tabWidget) # (в каком виджете размещена)
        Button_view.resize(100, 80) # размеры кнопки (длина 120, ширина 50)
        Button_view.move(850, 80) # разммещение кнопки (от левого края 850, от верхнего 60) от виджета в котором размещен
        joined_path = os.path.join(pathlib.Path.cwd(), 'icons', 'view_pdf.png')
        Button_view.setIcon(QtGui.QIcon(joined_path))
        Button_view.setIconSize(QtCore.QSize(96, 128))
        Button_view.setFlat(True)
        Button_view.show()
        Button_view.clicked.connect(view)

        # == определяет разигрывается 3-е место или нет и в зависимости от этого включает кнопку и checkBox
        if "1-й финал" in choice_etap:
            system = sf.select().where(System.stage == "1-й финал").get()
            flag_3 = system.no_game
            if flag_3 == '3':
                my_win.checkBox_no_play_3.setChecked(True)
                my_win.Button_3_mesta.setEnabled(True)
            else:
                my_win.checkBox_no_play_3.setChecked(False)
                my_win.Button_3_mesta.setEnabled(False)
        # =============
        Label_view = QLabel(my_win.tabWidget)
        Label_view.resize(150, 80) # размеры кнопки (длина 120, ширина 50)
        Label_view.move(880, 120) # разммещение кнопки (от левого края 850, от верхнего 60) от виджета в котором размещен
        Label_view.setText("Просмотр")
        Label_view.show()
        my_win.widget.hide()
        my_win.tableWidget.hide()
        my_win.resize(1270, 750)

        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 1000, 190))
        my_win.tabWidget_2.setGeometry(QtCore.QRect(260, 195, 1000, 502)) # устанавливает tabWidget_2
        my_win.toolBox.setGeometry(QtCore.QRect(10, 40, 243, 659))
 
        if tb_etap == 0: # подвкладка -Группы-
            stage = "Предварительный"
            load_combobox_filter_group()
        elif tb_etap == 1:  # подвкладка -Полуфиналы-
            stage = my_win.comboBox_filter_semifinal.currentText()
            load_combobox_filter_group_semifinal()
        else:  # подвкладка -Финалы-
            load_combobox_filter_final()
            stage = "1-й финал" if "1-й финал" in choice_etap else "Одна таблица"
            
        # определяет из скольки партий играется этап
        id_system = system_id(stage)
        sys_etap = System.select().where(System.id == id_system).get()
        sc_flag = sys_etap.score_flag # из скольки партий играется встреча
        for i in my_win.groupBox_kolvo_vstrech.findChildren(QRadioButton): # перебирает радиокнопки и включает в зависмости от сделаной жеребьевки
            score_current = int(i.text())
            if score_current == sc_flag:
                i.setChecked(True)
                break

        # my_win.label_result.setText(f"{stage_current} этап")
        game_visible = sys_etap.visible_game # если False, то счет в партиях не писать
        my_win.checkBox_4.setChecked(game_visible)
        my_win.checkBox_7.setEnabled(False)
        my_win.checkBox_8.setEnabled(False)
        my_win.checkBox_7.setChecked(False)
        my_win.checkBox_8.setChecked(False)

        my_win.Button_Ok.setEnabled(False)
        player_list = Result.select().where(Result.system_id == id_system)
        load_combo() # загружает фамилия игроков для поиска
        visible_field(stage)
        fill_table(player_list)
        my_win.label_16.hide()
        my_win.tableView_net.hide() # сетка ручной жеребьевки на 32
        tab_etap()
    elif tb == 4: # парный разряд
        my_win.resize(1110, 750)
        my_win.toolBox.setGeometry(QtCore.QRect(10, 40, 243, 659))
        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 841, 250))
        my_win.tabWidget_2.setGeometry(QtCore.QRect(260, 255, 841, 470)) # устанавливает tabWidget_2
        my_win.groupBox_match_double.setEnabled(True)
        my_win.tabWidget_3.setTabEnabled(0, True)
    elif tb == 5: # вкладка -рейтинг-
        my_win.tabWidget_2.setCurrentIndex(0)
        my_win.resize(1110, 750)
        my_win.tabWidget_2.setGeometry(QtCore.QRect(260, 75, 841, 622))
        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 841, 71))
        my_win.toolBox.setGeometry(QtCore.QRect(10, 40, 243, 659))
        my_win.comboBox_choice_R.clear()
        my_win.comboBox_filter_date_in_R.clear()
        rejting_month = ["За текуший месяц", "За январь месяц"]
        my_win.comboBox_choice_R.addItems(rejting_month)
        load_comboBox_filter_rejting()
    elif tb == 6: # вкладка -дополнительно-
        my_win.tabWidget_2.setCurrentIndex(2)
        my_win.groupBox_4.show()
        my_win.resize(1110, 750)
        my_win.tableWidget.setGeometry(QtCore.QRect(260, 250, 841, 400))
        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 841, 248))
        my_win.tabWidget_2.setGeometry(QtCore.QRect(260, 250, 841, 450))
        my_win.toolBox.setGeometry(QtCore.QRect(10, 40, 243, 659))
        my_win.Button_made_page_pdf.setEnabled(False)
        my_win.Button_up.setEnabled(False)
        my_win.Button_down.setEnabled(False)
        my_win.Button_made_one_file_pdf.setEnabled(False)
        my_win.Button_print_begunki.setEnabled(False)
        my_win.Button_change_player.setEnabled(False)
        my_win.lineEdit_range_tours.hide()
        my_win.comboBox_first_group.setEnabled(False)
        my_win.comboBox_second_group.setEnabled(False)
 
        load_combo_etap_begunki()
   
        # ======
    hide_show_columns(tb)

def otchestvo_input():
    """елси требуется отчество то вклюяает поле и выводит в listView сохраненые данные"""
    pass

def label_playing_count():
    """На вкладке -система- пишет сколько игр сыграно в каждом этапе"""
    result = Result.select().where(Result.title_id == title_id())
    system = System.select().where(System.title_id == title_id())
    n = 0
    my_win.label_playing_etap1.hide()
    my_win.label_playing_etap2.hide()
    my_win.label_playing_etap3.hide()
    my_win.label_playing_etap4.hide()
    my_win.label_playing_etap5.hide()
    my_win.label_playing_etap6.hide()
    my_win.label_playing_etap7.hide()
    my_win.label_playing_etap8.hide()
    my_win.label_playing_etap9.hide()
    my_win.label_playing_etap10.hide()
    my_win.label_playing_etap11.hide()
    my_win.label_playing_etap12.hide()
    for k in system:
        n += 1
        system_id = k.id
        result_playing = result.select().where((Result.system_id == system_id) & (Result.winner != ""))
        count_playing = len(result_playing)
        if n == 1:
            my_win.label_playing_etap1.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap1.show()
        elif n == 2:
            my_win.label_playing_etap2.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap2.show()
        elif n == 3:
            my_win.label_playing_etap3.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap3.show()
        elif n == 4:
            my_win.label_playing_etap4.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap4.show()
        elif n == 5:
            my_win.label_playing_etap5.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap5.show()
        elif n == 6:
            my_win.label_playing_etap6.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap6.show()
        elif n == 7:
            my_win.label_playing_etap7.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap7.show()
        elif n == 8:
            my_win.label_playing_etap8.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap8.show()
        elif n == 9:
            my_win.label_playing_etap9.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap9.show()
        elif n == 10:
            my_win.label_playing_etap10.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap10.show()
        elif n == 11:
            my_win.label_playing_etap11.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap11.show()
        elif n == 12:
            my_win.label_playing_etap12.setText((f"Сыграно: {count_playing} игр."))
            my_win.label_playing_etap12.show()


def add_city():
    """добавляет в таблицу город и соответсвующий ему регион"""
    city_field = my_win.lineEdit_city_list.text()
    city_field = city_field.capitalize()  # Переводит первую букву в заглавную
    cities = City.select().where(City.city == city_field)
    if len(cities) == 0:
        regions = my_win.comboBox_region.currentText()
        reg = Region.select().where(Region.region == regions).get()
        reg_id = reg.id
        City(city=city_field, region_id=reg_id).save()
    index = city_field.find(".")
    if index != -1:
        second_word = city_field[index + 1:]
        second_word = second_word.capitalize()
        city_field = city_field[:index + 1] + second_word
    my_win.lineEdit_city_list.setText(city_field)
    # my_win.textEdit.setText("Выберите регион в котором находится населенный пункт.")


def find_coach():
    """поиск тренера в базе"""
    my_win.label_63.setText("Список тренеров.")
    my_win.listWidget.clear()
    list_coach = []
    cp = my_win.lineEdit_coach.text()
    cp = cp.capitalize()  # Переводит первую букву в заглавную
    if my_win.checkBox_find_player.isChecked():
        player = Player.select().where(Player.title_id == title_id())
        coach_list = Coach.select().where(Coach.coach ** f'%{cp}%')  # создает выборку из базы тренеров фамилии,что начинаются на CP
        for pl in coach_list: #походит циклом и создает список с их ID
            c_id = pl.id
            list_coach.append(c_id)

        player_list = player.select().where(Player.coach_id << list_coach) # окончательная выборка со всеми тренерами (id)
    else:
        c = Coach.select()
        c = c.where(Coach.coach ** f'{cp}%')  # like
        tochka = cp.find(".")
        if tochka == -1:
            if (len(c)) != 0:
                for chp in c:
                    full_stroka = chp.coach
                    my_win.listWidget.addItem(full_stroka)


def add_coach(ch, num):
    """Проверяет наличие тренера в базе и если нет, то добавляет"""
    coach = Coach.select()
    count_coach = len(coach)
    if count_coach == 0:  # если первая запись то добавляет без проверки
        with db:
            cch = Coach(coach=ch, player_id=num).save()
        return
    for c in coach:
        coa = Coach.select().where(Coach.coach == ch)
        if bool(coa):
            return
        else:
            cch = Coach(coach=ch, player_id=num).save()


def add_patronymic():
    """Провкрка отчества если нет, то добавляет в DB"""
    pol_list = ['Мальчики','Юноши', 'Юниоры', 'Мужчины']
    titles = Title.select().where(Title.id == title_id()).get()
    otc = my_win.lineEdit_otchestvo.text()
    otc = otc.capitalize()
    pol = titles.gamer
    sex = "m" if pol in pol_list else "w"
    patronymic = Patronymic.select()
    count = len(patronymic)
    if count == 0:# если 1-я запись
        with db:
            otch = Patronymic(patronymic=otc, sex=sex).save()
    else:
        patron = Patronymic.select().where(Patronymic.patronymic == otc)
        for pat in patronymic:
            patr = pat.patronymic
            if patr == otc:
                return            
        otch = Patronymic(patronymic=otc, sex=sex).save()
        otchestvo = Patronymic.select().where(Patronymic.patronymic == otc).get()
        idp = otchestvo.id
        return idp
 

def find_player():
    """Установка курсора в строку поиска спортсмена в загруженном списке"""
    if my_win.checkBox_find_player.isChecked():
        my_win.lineEdit_Family_name.setFocus()
    else:
        my_win.lineEdit_Family_name.clear()
        player_list = Player.select().where(Player.title_id == title_id())
        fill_table(player_list) 


def find_player_on_tab_system():
    """выделяет строку в tablewidget при поиске фамилии на вкладке -система_"""
    sender = my_win.sender()
    flag_stat = 0
    choice = Choice.select().where(Choice.title_id == title_id())
    if sender == my_win.lineEdit_find_player_in_system:
        txt = my_win.lineEdit_find_player_in_system.text()
        flag_stat = 0
    elif sender == my_win.lineEdit_find_player_stat:
        txt = my_win.lineEdit_find_player_stat.text()
        flag_stat = 1
    txt = txt.upper()
 
    if flag_stat == 0:
        player_list = choice.select().where(Choice.family ** f'{txt}%')  # like поиск в текущем рейтинге
    else:
        player_list = Player.select().where((Player.title_id == title_id()) & (Player.player ** f'{txt}%'))  # like поиск в текущем рейтинге
    count = len(player_list)
    if count == 1:
        pass
    fill_table(player_list)       


def sort():
    """сортировка таблицы QtableView (по рейтингу или по алфавиту)"""
    sender = my_win.sender()  # сигнал от кнопки
    signal_button_list = [my_win.Button_sort_R, my_win.Button_sort_Name, my_win.Button_sort_mesto]
    pl_list = Player.select().where((Player.title_id == title_id()) & (Player.bday != "0000-00-00")) # добавил отделить строки с "X"
    if sender == my_win.Button_sort_R:  # в зависимости от сигала кнопки идет сортировка
        player_list = pl_list.select().where(Player.title_id == title_id()).order_by(Player.rank.desc())  # сортировка по рейтингу
    elif sender == my_win.Button_sort_Name:
        player_list = pl_list.select().where(Player.title_id == title_id()).order_by(Player.player)  # сортировка по алфавиту
    elif sender == my_win.Button_sort_mesto:
        player_list = pl_list.select().where(Player.title_id == title_id()).order_by(Player.mesto)  # сортировка по месту

    fill_table(player_list)
    if sender in signal_button_list:
        list_player_pdf(player_list)


def button_title_made_enable(state):
    """включает кнопку - создание титула - если отмечен чекбокс, защита от случайного нажатия"""
    if state == 2:  # если флажок установлен
        title_str = title_string()
        nm = title_str[0]
        ds = title_str[3]
        de = title_str[4]
        # получение последней записи в таблице
        t = Title.select().order_by(Title.id.desc()).get()
        if t.name == nm and str(t.data_start) == ds and str(t.data_end) == de:
            my_win.Button_title_made.setText("Редактировать")
        else:
            my_win.Button_title_made.setText("Создать")
        my_win.Button_title_made.setEnabled(True)
    else:
        my_win.Button_title_made.setEnabled(False)


def button_system_made_enable(state):
    """включает кнопку - создание системы - если отмечен чекбокс, защита от случайного нажатия"""
    if state == 2:
        my_win.Button_system_made.setEnabled(True)


def date_formated_on_db_or_form(b_day):
    """форматирование даты к виду для формы или базы данных"""
    b_day_formated = b_day.strftime('%d.%m.%Y') # форматирование даты
    return b_day_formated


def proba_perenos_stroki():
    """перенос строки в таблице"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph

    # Создаем документ
    doc = SimpleDocTemplate("table_with_wrap.pdf", pagesize=A4)
    elements = []

    # Подготовка стилей
    styles = getSampleStyleSheet()
    custom_style = styles["Normal"].clone("CustomStyle")
    custom_style.wordWrap = 'LTR' # Перенос слов (LTR - Left-To-Right)
    custom_style.splitLongWords = True # Разделять длинные слова
    custom_style.leading = 14 # Межстрочный интервал

    # Данные таблицы с автоматическим переносом
    data = [[Paragraph("Обычный текст с переносами по словам", custom_style)],]

    # Создаем таблицу с фиксированными ширинами колонок
    table = Table(data, colWidths=[200, 100])

    # Стиль таблицы
    table.setStyle(TableStyle([
    ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ('VALIGN', (0, 0), (-1, -1), 'TOP'), # Выравнивание по верхнему краю
    ('PADDING', (0, 0), (-1, -1), 5), # Отступы в ячейках
    ]))

    elements.append(table)
    doc.build(elements)

    ### Ключевые моменты:
    # 1. **Использование Paragraph**:
    # - Текст нужно оборачивать в `Paragraph`
    # - Обычные строки (`str`) не поддерживают перенос

    # 2. **Настройки стиля**:
    # ```python
    # custom_style = styles["Normal"].clone("CustomStyle")
    # custom_style.wordWrap = 'LTR' # Перенос слов
    # custom_style.splitLongWords = True # Разделять длинные слова
    # custom_style.leading = 14 # Межстрочный интервал
    # ```

    # 3. **Обязательные параметры таблицы**:
    # ```python
    # Table(data, colWidths=[...]) # Фиксированная ширина колонок
    # ```
    # - Без указания `colWidths` перенос работать не будет

    # 4. **Рекомендуемые стили таблицы**:
    # ```python
    # ('VALIGN', (0, 0), (-1, -1), 'TOP') # Выравнивание по верху
    # ('PADDING', (0, 0), (-1, -1), 5) # Отступы для читаемости
    # ```

    ### Для многострочных заголовков
    # ```python
    # header_style = styles["Heading4"].clone("HeaderStyle")
    # header_style.wordWrap = 'LTR'
    # header_style.splitLongWords = True

    # data = [
    # [
    # Paragraph("Колонка с очень длинным заголовком", header_style),
    # Paragraph("Обычный заголовок", header_style)
    # ],
    # # ... остальные строки
    # ]
    # ```

    ### +++++++++++++++++++++++++++++++++++++ Решение проблем
    # 1. **Если текст не переносится**:
    # - Убедитесь, что указаны `colWidths`
    # - Проверьте, что используете `Paragraph`, а не обычную строку
    # - Добавьте `splitLongWords=True`

    # 2. **Если переносы выглядят некрасиво**:
    # ```python
    # custom_style.alignment = 4 # Выравнивание по ширине (0-лево, 1-центр, 2-право, 4-по ширине)
    # custom_style.hyphenation = True # Включить расстановку переносов
    # ```

    # 3. **Для динамической высоты строк**:
    # ```python
    # # ReportLab автоматически рассчитает высоту
    # # Для ручной настройки:
    # table = Table(data, colWidths=[200, 100], rowHeights=None)
    # ```

    # ### Альтернативный вариант (KeepInFrame)
    # Для сложных случаев можно использовать контейнер `KeepInFrame`:
    # ```python
    # from reportlab.platypus import KeepInFrame

    # data = [
    # [
    # KeepInFrame(
    # maxWidth=200,
    # maxHeight=100,
    # content=[Paragraph("Очень длинный текст...", custom_style)]
    # ),
    # "Простой текст"
    # ]
    # ]
# ==============================================================
def _list_player_pdf(player_list):
    """создание списка участников в pdf файл"""
    from reportlab.platypus import Table
    story = []  # Список данных таблицы участников
    elements = []  # Список Заголовки столбцов таблицы
    tit = Title.get(Title.id == title_id())
    short_name = tit.short_name_comp
    gamer = tit.gamer
    count = len(player_list)  # количество записей в базе
    kp = count + 1
    n = 0
    for l in player_list:
        n += 1
        o = l.otchestvo    
        p = l.player
        if o is not None: # если есть отчество то добавляет к фамилии и имени
            p_full = f"{p} {o}"
            dlina = len(p_full)
            if dlina > 27:
                p_full = f"{p}\n{o}"
            p = p_full
        p = f'{p} {o}'
        b = l.bday
        b = format_date_for_view(str_date=b)
        r = l.rank
        c = l.city
        g = l.region
        z = l.razryad
        coach_id = l.coach_id
        t = coach_id.coach
        m = l.mesto
        
        t = chop_line(t) # разбивает строку тренеров не две если строкка длинная
        data = [n, p, b, r, c, g, z, t, m]
        elements.append(data)
    elements.insert(0, ["№", "ФИО", "Дата рожд.", "R", "Город", "Регион", "Разряд", "Тренер(ы)",
                        "Место"])
    t = Table(elements,
            colWidths=(0.8 * cm, 4.4 * cm, 1.6 * cm, 0.8 * cm, 2.5 * cm, 3.2 * cm, 1.1 * cm, 4.6 * cm, 1.0 * cm),
            rowHeights=None, repeatRows=1)  # ширина столбцов, если None-автоматическая
            #   rowHeights=(0.35 * cm), repeatRows=1)  # ширина столбцов, если None-автоматическая
    t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                        #    ('FONTSIZE', (0, 0), (-1, -1), 7),
                            ('FONTSIZE', (0, 0), (-1, -1), 6),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 1),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (-1, kp * -1), 'CENTER'),
                           ('BACKGROUND', (0, 0), (8, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (8, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, kp * -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey, None, (1, 1)),
                        #    ('LINEBELOW', (0, 0), (-1, -1), 0.02, colors.grey, None, (1, 1)),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.5, colors.black)
                           ]))

    h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=150,
            firstLineIndent=-20, textColor="green")  # стиль параграфа
    h3.spaceAfter = 10  # промежуток после заголовка
    story.append(Paragraph(f'Список участников. {gamer}', h3))
    story.append(t)

    doc = SimpleDocTemplate(f"{short_name}_player_list.pdf", pagesize=A4)
    catalog = 1
    change_dir(catalog)
    doc.build(story, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")


def list_player_pdf(player_list):
    """создание списка участников в pdf файл"""
    from reportlab.platypus import Table
     # Подготовка стилей
    styles = getSampleStyleSheet()
    custom_style = styles['Normal'].fontName = 'DejaVuSerif'
    custom_style = styles['Normal'].fontSize = 6
    custom_style = styles["Normal"].clone("CustomStyle")
    custom_style.wordWrap = 'LTR' # Перенос слов (LTR - Left-To-Right)
    custom_style.leading = 6 # Межстрочный интервал
    story = []  # Список данных таблицы участников
    elements = []  # Список Заголовки столбцов таблицы
    tit = Title.get(Title.id == title_id())
    short_name = tit.short_name_comp
    gamer = tit.gamer
    otc = tit.otchestvo # если 1 значит в списках присутсвует отчество
    count = len(player_list)  # количество записей в базе
    kp = count + 1
    n = 0
    for l in player_list:
        n += 1
        if otc == 1:
            pat_id  = l.patronymic_id 
            patronymics = Patronymic.select().where(Patronymic.id == pat_id).get()  
            o = patronymics.patronymic
            p = l.player
            p = f"{p} {o}"
        else:
            p = l.player
        b = l.bday
        b = format_date_for_view(str_date=b)
        r = l.rank
        c = l.city
        g = l.region
        z = l.razryad
        coach_id = l.coach_id
        t = coach_id.coach
        m = l.mesto
        # ========================
        data = [n, [Paragraph(p, custom_style)], b, r, c, g, z, [Paragraph(t, custom_style)], m]
        # =========================

        elements.append(data)
    elements.insert(0, ["№", "ФИО", "Дата рожд.", "R", "Город", "Субъект РФ", "Разряд", "Тренер(ы)",
                        "Место"])
    t = Table(elements,
            colWidths=(0.8 * cm, 5.0 * cm, 1.6 * cm, 0.8 * cm, 2.5 * cm, 3.2 * cm, 1.1 * cm, 4.0 * cm, 1.0 * cm),
            # rowHeights=None, repeatRows=1)  # ширина столбцов, если None-автоматическая
              rowHeights=(0.45 * cm), repeatRows=1)  # ширина столбцов, если None-автоматическая
    t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                        #    ('FONTSIZE', (0, 0), (-1, -1), 7),
                            ('FONTSIZE', (0, 0), (-1, -1), 6),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 1),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (-1, kp * -1), 'CENTER'),
                           ('BACKGROUND', (0, 0), (8, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (8, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, kp * -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey, None, (1, 1)),
                        #    ('LINEBELOW', (0, 0), (-1, -1), 0.02, colors.grey, None, (1, 1)),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.5, colors.black)
                           ]))


    h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=150,
            firstLineIndent=-20, textColor="green")  # стиль параграфа
    h3.spaceAfter = 10  # промежуток после заголовка
    story.append(Paragraph(f'Список участников. {gamer}', h3))
    story.append(t)

    doc = SimpleDocTemplate(f"{short_name}_player_list.pdf", pagesize=A4)
    catalog = 1
    change_dir(catalog)
    doc.build(story, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")


def ReturnCode():
    pass


def exit_comp(flag):   
    """нажата кнопка -выход- и резервное копирование db"""
    msgBox = QMessageBox
    if flag == 0:
        result = msgBox.question(my_win, "Выход из программы", "Вы действительно хотите выйти из программы?",
                                    msgBox.Ok, msgBox.Cancel)
    else:
        result = msgBox.Ok

    if result == msgBox.Ok:
        # my_win.close()
        user = "root"
        password = "db_pass"
        database = "mysql_db"
        current_date = str(datetime.now().strftime('%d_%m_%Y'))
        try:
            dump = f'dump_{database}_{current_date}.sql'
            absolute_path = str(Path(f'backup_db/{dump}').resolve())
            # return dump
            p = subprocess.Popen('mysqldump -u' + user + ' -p' + password + ' --databases ' + database + ' > ' + absolute_path, shell=True)
            # Wait for completion
            p.communicate()
            # Check for errors
            if p.returncode != 0:
                raise ReturnCode
            my_win.statusbar.showMessage("Экспорт базы данных завершен успешно", 5000)            
            my_win.close()
            return dump
        except:
            print('Backup failed for ', db)
    else:
        return


# def exit_comp():
#     """нажата кнопка -выход-"""
#     import subprocess
#     msgBox = QMessageBox
#     result = msgBox.question(my_win, "Выход из программы", "Вы действительно хотите выйти из программы?",
#                              msgBox.Ok, msgBox.Cancel)
#     if result == msgBox.Ok:
#         my_win.close()
#         host = "localhost"
#         user = "root"
#         password = "db_pass"
#         database = "mysql_db"
#         # =====
#         # Open database connection
#         conn = pymysql.connect(host=host, user=user, password=password, database=database)
#         # db = pymysql.connect(host, user, password, database)
#         absolute_path = str(Path('backup_db/mysql_db.sql').resolve())
#         # txt = absolute_path.replace("/", "\\")
#         # SQL query to get all tables in the database
#             # cursor.execute(f'CREATE TABLE {table_name} SELECT * FROM {database}.{table_name}') 

#         tables = cursor.fetchall() # fetching all tables
#         for table in tables:
#             table = table[0]
#             # backup_query = f'CREATE TABLE {table} SELECT * FROM {database}.{table}'
#             backup_query = "SELECT * INTO OUTFILE '/tmp/" + table + ".sql' FIELDS TERMINATED BY ',' LINES TERMINATED BY '\n' FROM " + table
#             cursor.execute(backup_query)

#         # database.close()
#         # # =====
#         # absolute_path = Path('backup_db/mysql_db.sql').resolve()
#         # # backup_file = "mysql_db.sql"
#         command = f"mysqldump -h{host} -u{user} -p{password} {database} > {absolute_path}"
#         process = subprocess.run(command, shell=True)
#         if process.returncode == 0:
#             my_win.statusbar.showMessage("Экспорт базы данных завершен успешно", 5000)
#             print("Database backup completed successfully.")
#         else:
#             print(f"Database backup failed with return code {process.returncode}.")


def add_or_delete_etap_after_choice(stage, flag):
    """добавление этапа после жеребьевки"""
    etap_list = ["Предварительный", "1-й полуфинал", "2-й полуфинал", "1-й финал", "2-й финал", "3-й финал", "4-й финал",
                            "5-й финал", "6-й финал", "7-й финал", "8-й финал", "9-й финал", "10-й финал", "Суперфинал"]
    etap_word = ""
    ind = etap_list.index(stage) # индекс вставляемого этапа 
    system = System.select().where(System.title_id == title_id())
    m = 0
    etap_dict = {}
    id_list = []
    for k in system: # получение словаря текущих этапов и списка их id
        stage_system = k.stage 
        id_s = k.id
        id_list.append(id_s)
        etap_dict[m] = stage_system
        m += 1
    ind_next = etap_list.index(stage_system) # индекс последний в списке всех этапов системы
    if flag == 1: # удаляем этап
        ind = [keys for keys, values in etap_dict.items() if values == stage] # список ключа по значению 
        id_del = id_list[ind[0]] # id удаляемого этапа
        s_d = System.delete().where(System.id == id_del)
        s_d.execute()
        gl_d = Game_list.delete().where(Game_list.system_id == id_del)
        gl_d.execute() 
    else:   
        if ind < ind_next:
            for l in range (len(id_list)):
                if l > ind[0]: # удаляет все что ниже вставляемого этапа
                    s_d = System.delete().where(System.id == id_list[l])
                    s_d.execute()
                    gl_d = Game_list.delete().where(Game_list.system_id == id_list[l])
                    gl_d.execute()


    sb = "Выбор системы проведения соревнования."
    my_win.statusbar.showMessage(sb)
    my_win.spinBox_kol_group.hide()
    my_win.comboBox_etap.clear()
    my_win.comboBox_etap.show()
    my_win.label_10.show()
    my_win.label_10.setText(f"{m + 1}-й этап")

    my_win.Button_etap_made.setEnabled(True)
    my_win.tabWidget.setTabEnabled(2, True)
    my_win.tabWidget.setCurrentIndex(2)
    my_win.comboBox_page_vid.setEnabled(True)

    index = etap_list.index(stage)
    if index == 0:
        etap_word = "Предварительный"
        real_list = ["-выбор этапа-", "Одна таблица", "Предварительный"] # который нужен в комбобокс
    elif index > 0 and index < 3:
        etap_word = "Полуфиналы"
        real_list = ["-выбор этапа-", "Полуфиналы", "Финальный"]
    elif index > 2 and index < 13:
        etap_word = "Финальный"
        real_list = ["-выбор этапа-", "Финальный", "Суперфинал"] 
    else:
        real_list = ["-выбор этапа-", "Суперфинал"] 
        etap_word = "Суперфинал"
    if flag == 0: # уточнить запись в комбобокс этап        
        my_win.comboBox_etap.setCurrentText(etap_word)
    elif flag == 1:
        combobox_etap_compare(real_list)


def system_competition():
    """выбор системы проведения при изменении строки в комбобокс этап или мз меню"""
    msgBox = QMessageBox()
    sender = my_win.sender()
    system_etap_list = []
    semifinal_etap_list = ["1-й полуфинал", "2-й полуфинал"]
    fin_etap_list = ["1-й финал", "2-й финал", "3-й финал", "4-й финал",
                            "5-й финал", "6-й финал", "7-й финал", "8-й финал", "9-й финал", "10-й финал", "Суперфинал"]
    tit = Title.get(Title.id == title_id())
    systems = System.select().where(System.title_id == title_id())
    for p in systems:
        etap = p.stage
        system_etap_list.append(etap)
    gamer = tit.gamer
    id_title = tit.id
    # проверка что все спортсмены подтвердились
    flag_checking = checking_before_the_draw() # TRUE - подтверждены все спортсмены
    if flag_checking is False:
        return
    flag_system = ready_system() # False система еще не создана 
    if sender != my_win.comboBox_etap:
        if sender == my_win.system_edit_Action: # редактирование системы из меню
            sb = "Изменение системы проведения соревнования."
            my_win.statusbar.showMessage(sb)
            # ======
            msgBox.setIcon(QMessageBox.Question)
            msgBox.setText("Вы хотите изменить систему соревнований?")
            # ========
            msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.Cancel)
            msgBox.setDefaultButton(QMessageBox.Cancel)
            ret = msgBox.exec()
            made_list = ["Изменить всю систему", "Отдельные этапы", "Добавить этап", "Удалить этап"]
            if ret == msgBox.Yes:
               my_win.tabWidget.setCurrentIndex(0)
               item_selected, ok = QInputDialog.getItem(
                    my_win, "Системные этапы", "Выберите действия для редактирования", made_list, 0, False) 
            else:
                return
            if item_selected == "Изменить всю систему":  
                system_clear()              
                # clear_db_before_edit() # очищает таблицы перед новой системой соревнования (system, choice)
                tab_enabled(id_title)  # показывает вкладки по новому
                choice_tbl_made()  # заполняет db жеребьевка
                # flag_system = False # ставит флаг, что система еще не создана
                stage = ""
            elif item_selected == "Отдельные этапы":               
                stage, ok = QInputDialog.getItem(
                    my_win, "Системные этапы", "Выберите этап для редактирования", system_etap_list, 0, False)
                id_system = system_id(stage)
                system_exit = systems.select().where(System.stage_exit == stage)
                msgBox.setIcon(QMessageBox.Question)
                msgBox.setText("Изменение системы!")
                msgBox.setInformativeText("Если удалить выбранный этап нажмите -Yes-")
                msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.Cancel)
                msgBox.setDefaultButton(QMessageBox.Cancel)
                ret_1 = msgBox.exec()
                if ret_1 == msgBox.Ok:
                    for m in system_exit:
                        id_sys = m.id
                        System.update(stage_exit="Предварительный", mesta_exit=1).where(System.id == id_sys).execute()
                    sys = System.delete().where(System.id == id_system)
                    sys.execute()
                    return
                elif ret_1 == msgBox.Cancel:
                        return
            elif item_selected == "Добавить этап":
                add_system_etap_list = []
                for k in semifinal_etap_list:
                    if k not in system_etap_list:
                        add_system_etap_list.append(k)
                for k in fin_etap_list:
                    if k not in system_etap_list:
                        add_system_etap_list.append(k)
                        if "Суперфинал" not in system_etap_list:
                            add_system_etap_list.append("Суперфинал")
                            break
                stage, ok = QInputDialog.getItem(
                    my_win, "Системные этапы", "Выберите этап для добавления", add_system_etap_list, 0, False)
                flag = 0 # флаг добавления этапа равен 0
                add_or_delete_etap_after_choice(stage, flag)
                return
            else: # удалить этап
                stage, ok = QInputDialog.getItem(
                    my_win, "Системные этапы", "Выберите этап для удаления", system_etap_list, 0, False)
                id_system = system_id(stage)
                msgBox.setIcon(QMessageBox.Question)
                msgBox.setText("Изменение системы!")
                msgBox.setInformativeText(f"Вы уверны, что хотите удалить\n {stage}?")
                msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.Cancel)
                msgBox.setDefaultButton(QMessageBox.Cancel)
                ret_1 = msgBox.exec()
                if ret_1 == msgBox.Yes:
                    flag = 1 # флаг удаления этапа равен 1
                    add_or_delete_etap_after_choice(stage, flag)
                    return
                else:
                    return

            # =========
        elif sender == my_win.system_made_Action: # создание системы из меню
            sb = "Создание системы проведения соревнования."
            my_win.statusbar.showMessage(sb)
            result = msgBox.question(my_win, "", "Вы хотите создать систему соревнований?",
                                msgBox.Ok, msgBox.Cancel)

            if result == msgBox.Ok:
                choice_tbl_made()  # заполняет db жеребьевка
            else:
                return
        my_win.spinBox_kol_group.hide()
        my_win.comboBox_etap.setEnabled(True)
        my_win.label_102.hide()
        my_win.label_27.hide()
        my_win.label_etap_2.hide()
        my_win.label_etap_3.hide()
        my_win.label_etap_4.hide()
        my_win.label_etap_5.hide() 
        my_win.label_etap_6.hide()
        my_win.label_etap_7.hide()
        my_win.label_etap_8.hide() 
        my_win.label_etap_9.hide()
        my_win.label_etap_10.hide()
        my_win.label_etap_11.hide()
        my_win.label_30.hide()
        my_win.label_31.hide()
        my_win.label_103.hide()
        my_win.label_104.hide()
        my_win.label_105.hide()
        my_win.label_106.hide()
        my_win.label_107.hide()
        my_win.label_108.hide()
        my_win.label_109.hide()
        my_win.label_110.hide()
        my_win.label_111.hide()
        my_win.label_53.hide()
        my_win.label_58.hide() 
        my_win.label_81.hide()
        my_win.label_82.hide()
        my_win.label_83.hide()  
        my_win.label_84.hide()
        my_win.label_85.hide()
        my_win.label_86.hide()
        my_win.label_87.hide()
 
        my_win.tabWidget.setTabEnabled(2, True) # включает вкладку

        if flag_system is True:
            flag_choice = ready_choice(stage)
            if flag_choice is True:
                sb = "Система и жербьевка создана."
            elif flag_choice is False:
                sb = "Система создана, теперь необходимо произвести жеребьевку. " \
                    "Войдите в меню -соревнования- и выберите -жеребьевка-"
            my_win.statusbar.showMessage(sb)
        elif flag_system is False:
            sb = "Выбор системы проведения соревнования."
            my_win.statusbar.showMessage(sb)
            my_win.spinBox_kol_group.hide()
            my_win.comboBox_etap.clear()
            real_list = ["-выбор этапа-", "Одна таблица", "Предварительный"] # который нужен в комбобокс
            combobox_etap_compare(real_list)
            my_win.comboBox_etap.show()
            my_win.comboBox_table_1.hide()
            my_win.label_10.show()
            my_win.label_10.setText("1-й этап")
            my_win.Button_etap_made.setEnabled(False)
            my_win.comboBox_page_vid.setEnabled(True)
            player = Player.select().where((Player.title_id == title_id()) & (Player.bday != '0000-00-00'))
            count = len(player)
            if count != 0:
                my_win.tabWidget.setCurrentIndex(2)
            else:
                reply = QMessageBox.information(my_win, 'Уведомление',
                                                "У Вас нет ни одного спортсмена.\nСначала необходимо создать "
                                                "список участников соревнований.\n Перейти к созданию списка?",
                                                msgBox.Ok,
                                                msgBox.Cancel)
                if reply == msgBox.Ok:
                    my_win.tabWidget.setCurrentIndex(1)
                    my_win.lineEdit_Family_name.setFocus()
                else:
                    return        


def system_clear():
    msgBox = QMessageBox()
    systems = System.select().where(System.title_id == title_id())
    game_lists = Game_list.select().where(Game_list.title_id == title_id())
    results = Result.select().where(Result.title_id == title_id())
    choices = Choice.select().where(Choice.title_id == title_id())
    players = Player.select().where((Player.title_id == title_id()) & (Player.bday == '0000-00-00'))
    result = msgBox.question(my_win, "", "Вы действительно хотите очистить систему соревнований?",
                                msgBox.Ok, msgBox.Cancel)
    if result == msgBox.Ok:
        for i in game_lists:
            gl_d = Game_list.get(Game_list.id == i)
            gl_d.delete_instance()
        for i in choices:
            ch_d = Choice.get(Choice.id == i)
            ch_d.delete_instance()
        for i in results:
            r_d = Result.get(Result.id == i)
            r_d.delete_instance()
        if len(players) > 0: # удаляет запись в -Player- если есть крест сетки
            for i in players:
                pl_d = Player.get(Player.id == i)
                pl_d.delete_instance()
        n = 0
        for i in systems:
            if n == 0:
                System.update(total_athletes=0, total_group=0, max_player=0, stage="", type_table="",
                              page_vid="", label_string="", kol_game_string="", choice_flag=0).where(System.id == i).execute()
            else:
                system_d = System.get(System.id == i)
                system_d.delete_instance()
            n += 1
        Title.update(tab_enabled="Титул Участники").where(Title.id == title_id()).execute()
    else:
        return


def one_table(fin, group):
    """система соревнований из одной таблицы запись в System, Game_list, Result"""
    msgBox = QMessageBox()
    ch = Choice.select().where(Choice.title_id == title_id())
    count = len(Player.select().where(Player.title_id == title_id()))
    visible_game = 1 if my_win.checkBox_visible_game.isChecked() else 0
    # в зависмости сетка или круг
    cur_index = my_win.comboBox_table_1.currentIndex()
    if fin == "Одна таблица":
        if cur_index == 0:
            type_table = "круг"
        elif cur_index == 1:
            vt = "Сетка (-2) на"
            my_win.comboBox_page_vid.setCurrentText("книжная")
            type_table = "сетка"
        elif cur_index == 2:
            vt = "Сетка (с розыгрышем всех мест) на"
            my_win.comboBox_page_vid.setCurrentText("книжная")
            type_table = "сетка"
        elif cur_index == 3:
            vt = "Сетка (с играми за 1-3 места) на"
            my_win.comboBox_page_vid.setCurrentText("книжная")
            type_table = "сетка"
        elif cur_index == 4:
            vt = "Круговая таблица на"
            my_win.comboBox_page_vid.setCurrentText("альбомная")
            type_table = "круг"

        if type_table == "круг":
            total_athletes = count
        else: # на сколько участников таблица
            total_athletes = full_net_player(player_in_final=count)
           

        flag_ready_system = ready_system()
        if flag_ready_system is False:
            sys_m = System.select().where(System.title_id == title_id()).get()
            total_game = numbers_of_games(cur_index, player_in_final=count, kpt=0)
            sys_m.max_player = total_athletes
            # sys_m.total_athletes = total_athletes
            sys_m.total_athletes = count
            sys_m.total_group = group
            sys_m.stage = my_win.comboBox_etap.currentText()
            sys_m.type_table = type_table
            sys_m.page_vid = my_win.comboBox_page_vid.currentText()
            sys_m.label_string = f"{vt} {total_athletes} участников"
            sys_m.kol_game_string =f"{total_game} игр"
            sys_m.visible_game = visible_game
            sys_m.score_flag = my_win.spinBox.text()
            sys_m.save()

            my_win.Button_etap_made.setEnabled(False)
            my_win.comboBox_page_vid.setEnabled(False)

            for k in ch: # записывает в DB после создании системы из одной таблицы basic - Одна таблица
                k.basic = fin
                k.save()
            add_open_tab(tab_page="Система")

            result = msgBox.question(my_win, "", "Система соревнований создана.\n"
                                                 "Теперь необходимо сделать жеребъевку\n"
                                                 "Хотите ее сделать сейчас?",
                                     msgBox.Ok, msgBox.Cancel)
            if result == msgBox.Ok:
                if type_table == "круг":  # функция жеребьевки таблицы по кругу
                    player_in_one_table(fin)
                else:
                    posev_data = player_choice_in_setka(fin)
                    player_in_setka_and_write_Game_list_and_Result(fin, posev_data)
                add_open_tab(tab_page="Результаты")
                flag_choice = True
            else:
                flag_choice = False
                return
            sys_m.stage = fin
            sys_m.choice_flag = flag_choice # запись о том что сделана жеребьевка
            sys_m.save()


def selection_of_the_draw_mode():
    """Выбор режима жеребьевки сетки -автомат- или -ручной-"""
    vid = ["Автоматическая", "Полуавтоматическая", "Ручная"]
    vid, ok = QInputDialog.getItem(
                    my_win, "Жеребьевка", "Выберите режим жеребьевки сетки.", vid, 0, False)
    if vid == "Автоматическая":
        flag = 1
        my_win.tableView_net.hide()
    elif vid == "Полуавтоматическая":
        flag = 2
        my_win.resize(1440, 804)
        my_win.tableView_net.show()
        my_win.tableView_net.setGeometry(QtCore.QRect(1110, 9, 321, 749)) # от лев края, от вверха, ширина и высота)
    elif vid == "Ручная":
        flag = 3
        my_win.resize(1440, 804)
        my_win.tableView_net.show()
        my_win.tableView_net.setGeometry(QtCore.QRect(1110, 9, 321, 749)) # от лев края, от вверха, ширина и высота)
    return flag
    
              
def kol_player_in_group():
    """подсчет кол-во групп и человек в группах"""
    sender = my_win.sender()  # сигнал от кнопки
    flag_visible = my_win.checkBox_visible_game.isChecked()
    kg = my_win.spinBox_kol_group.text()  # количество групп
    score_match = my_win.spinBox.text()
    player_list = Player.select().where((Player.title_id == title_id()) & (Player.bday != '0000-00-00'))
    type_table = "группы"
    count = len(player_list)  # количество записей в базе
    # остаток отделения, если 0, то участники равно делится на группы
    e1 = count % int(kg)
    # если количество участников равно делится на группы (кол-во групп)
    p = count // int(kg)
    g1 = int(kg) - e1  # кол-во групп, где наименьшее кол-во спортсменов
    g2 = int(p + 1)  # кол-во человек в группе с наибольшим их количеством
    if e1 == 0:  # то в группах равное количество человек -e1-
        stroka_kol_group = f"{kg} группы по {str(p)} чел."
        skg = int((p * (p - 1) / 2) * int(kg))
        mp = p
    else:
        stroka_kol_group = f"{str(g1)} групп(а) по {str(p)} чел. и {str(e1)} групп(а) по {str(g2)} чел."
        p = int(p)
        skg = int((((p * (p - 1)) / 2 * g1) + ((g2 * (g2 - 1)) / 2 * e1))) # общее количество игр в группах
        mp = g2
    stroka_kol_game = f"{skg} игр"
    my_win.label_11.hide()
    my_win.label_12.setText(stroka_kol_group)
    my_win.label_12.show()
    my_win.label_19.setText(stroka_kol_game)
    my_win.label_19.show()
    my_win.Button_etap_made.setEnabled(True)
    if int(kg) % 2 != 0: # Если число групп нечетное то вид страницы ставит -книжная-
        my_win.comboBox_page_vid.setCurrentIndex(1)
    else:
        my_win.comboBox_page_vid.setCurrentIndex(0)

    if sender == my_win.Button_etap_made:
        my_win.Button_etap_made.setEnabled(False)
        my_win.comboBox_page_vid.setEnabled(False)
        my_win.spinBox_kol_group.hide()
        # ====== запись в таблицу db -system- первый этап
        s = System.select().where(System.title_id == title_id()).get()
        system = System.get(System.id == s)
        system.max_player = mp
        system.total_athletes = count
        system.total_group = kg
        system.stage = my_win.comboBox_etap.currentText()
        system.type_table = type_table
        system.page_vid = my_win.comboBox_page_vid.currentText()
        system.label_string = stroka_kol_group
        system.kol_game_string = stroka_kol_game
        system.score_flag = score_match
        system.visible_game = flag_visible
        system.save()
    load_combobox_filter_group()


def page_vid():
    """присваивает переменной значение выборат вида страницы"""
    if my_win.comboBox_page_vid.currentText() == "альбомная":
        pv = landscape(A4)
    else:
        pv = A4
    return pv


def view():
    """просмотр PDF файлов средствами OS"""
    from sys import platform
    msgBox = QMessageBox()
    sender = my_win.sender()
    # tab_etap = my_win.tabWidget_stage.currentIndex()
    tab = my_win.tabWidget.currentIndex()
    name_file = ""
    t_id = Title.get(Title.id == title_id())
    short_name = t_id.short_name_comp
    if sender == my_win.view_all_comp_Action: # просмотр полных соревнований в каталоге /competition_pdf
        catalog = 2
        change_dir(catalog)
        view_file = f"{short_name}.pdf"
    else: # просмотр отдельных страниц в каталоге /table_pdf
        if tab == 3: # если просмотр результатов игр
            view_file = made_pdf_table_for_view(sender)
        catalog = 1
        change_dir(catalog)
        if sender == my_win.view_list_Action:
            view_sort = ["По алфавиту", "По рейтингу", "По месту"]
            view_sort, ok = QInputDialog.getItem(
                        my_win, "Сортировка", "Выберите вид сортировки,\n просмотра списка участников.", view_sort, 0, False)
            if view_sort == "По рейтингу":
                player_list_x = Player.select().where(Player.title_id == title_id()).order_by(Player.rank.desc())  # сортировка по рейтингу
            elif view_sort == "По алфавиту": 
                player_list_x = Player.select().where(Player.title_id == title_id()).order_by(Player.player) # сортировка по алфавиту
            elif view_sort == "По месту":
                player_list_x = Player.select().where(Player.title_id == title_id()).order_by(Player.mesto)  # сортировка по месту
            player_list = player_list_x.select().where(Player.player != "x")
            list_player_pdf(player_list)
            change_dir(catalog)
            view_file =  f"{short_name}_player_list.pdf"
        elif sender == my_win.view_referee_list_Action:
            view_file =  f"{short_name}_referee_list.pdf"
            name_file = "Список ГСК"
        elif sender == my_win.view_regions_list_Action:
            view_file =  f"{short_name}_regions_list.pdf"
            name_file = "Список регионов"
        elif sender == my_win.view_winners_list_Action:
            view_file =  f"{short_name}_winners_list.pdf" 
            name_file = "Список победителей"   
        elif sender == my_win.view_title_Action:
            view_file = f"{short_name}_title.pdf"
        elif sender == my_win.view_gr_Action:  # вкладка группы
            view_file = f"{short_name}_table_group.pdf"
        elif sender == my_win.view_fin1_Action:
            view_file = f"{short_name}_1-final.pdf"
        elif sender == my_win.view_fin2_Action:
            view_file = f"{short_name}_2-final.pdf"
        elif sender == my_win.view_fin3_Action:
            view_file = f"{short_name}_3-final.pdf"
        elif sender == my_win.view_fin4_Action:
            view_file = f"{short_name}_4-final.pdf"
        elif sender == my_win.view_fin5_Action:
            view_file = f"{short_name}_5-final.pdf"
        elif sender == my_win.view_fin6_Action:
            view_file = f"{short_name}_6-final.pdf"
        elif sender == my_win.view_fin7_Action:
            view_file = f"{short_name}_7-final.pdf"
        elif sender == my_win.view_fin8_Action:
            view_file = f"{short_name}_8-final.pdf"
        elif sender == my_win.view_fin9_Action:
            view_file = f"{short_name}_9-final.pdf"
        elif sender == my_win.view_fin10_Action:
            view_file = f"{short_name}_10-final.pdf"
        elif sender == my_win.view_superfin_Action:
            view_file = f"{short_name}_superfinal.pdf"
        elif sender == my_win.view_one_table_Action:
            view_file = f"{short_name}_one_table.pdf"
        elif sender == my_win.view_pf1_Action:
            view_file = f"{short_name}_1-semifinal.pdf"
        elif sender == my_win.view_pf2_Action:
            view_file = f"{short_name}_2-semifinal.pdf"
        elif sender == my_win.clear_s32_Action:
            view_file = "clear_32_net.pdf"
        elif sender == my_win.clear_s16_Action:
            view_file = "clear_16_full_net.pdf"
        elif sender == my_win.clear_s32_full_Action:
            view_file = "clear_32_full_net.pdf"
        elif sender == my_win.clear_s32_2_Action:
            view_file = "clear_32_2_net.pdf"
        elif sender == my_win.clear_s16_2_Action:
            view_file = "clear_16_2_net.pdf"   
        elif sender == my_win.clear_s8_2_Action:
            view_file = "clear_8_2_net.pdf"
        elif sender == my_win.clear_s8_full_Action:
            view_file = "clear_8_full_net.pdf"
    flag = open_close_file(view_file)
    if flag is False:
        result = msgBox.information(my_win, "", "Такой файл не существует.\n"
                                                 f"-{name_file}- необходимо создать!",
                                     msgBox.Ok)
        return
    else:
        if platform == "linux" or platform == "linux2":  # linux
            pass
        elif platform == "darwin":  # OS X
            os.system(f"open {view_file}")
        elif platform == "win32":  # Windows...
            os.system(f"{view_file}")
        os.chdir("..")


def player_in_setka_and_write_Game_list_and_Result(fin, posev_data):
    """заполняет таблицу Game_list данными спортсменами из сетки tds - список списков данных из сетки, а затем
    заполняет таблицу -Result-"""
    id_system = system_id(stage=fin)
    system = System.select().where((System.title_id == title_id()) & (System.id == id_system)).get()  # находит system id последнего
    st = "Финальный"
    game = 0
    if fin == "Одна таблица":
        st = "Одна таблица"
    system_table = system.label_string
    mp = system.max_player
    mp = full_net_player(player_in_final=mp)
    if system_table == "Сетка (с розыгрышем всех мест) на 8 участников":
        game = 12
    elif system_table == "Сетка (-2) на 8 участников":
        game = 14
    elif system_table == "Сетка (с розыгрышем всех мест) на 16 участников":
        game = 32
    elif system_table == "Сетка (-2) на 16 участников":
        game = 38
    elif system_table == "Сетка (с розыгрышем всех мест) на 32 участников":
        game = 80
    elif system_table == "Сетка (-2) на 32 участников":
        game = 94
    elif system_table == "Сетка (1-3 место) на 32 участников":
        game = 32
    # создание сетки со спортсменами согласно жеребьевки
    all_list = setka_data(fin, posev_data)
    tds = all_list[0]
    tds_full_name_city = all_list[3]
    k = 0
    for r in tds:
        if r != "X":
            znak = r.find("/")
            family = r[:znak]
            id_pl = all_list[2][family]
            # family_id = f'{family}/{id_pl}'  # фамилия игрока и его id
            player_id = int(id_pl)
        else:
            # === вариант с добавлением игрока вместо фамилии -Х- ====
            pl = Player.select().where(Player.title_id == title_id())
            count = len(pl)
            pl_x = pl.select().where(Player.player == "X")
            if len(pl_x) == 0:
                players = Player.insert(player="X", bday='0000-00-00', city="", region="", razryad="",coach_id=1, 
                                        mesto=0, full_name="X", title_id=title_id(), pay_rejting="", comment="",  coefficient_victories="", 
                                        total_game_player=0, total_win_game=0, application="", patronymic_id=1).execute()
            else:
                player_s = pl.select().where(Player.player == "X").get()
                pl_id = player_s.id
            player_id = players if len(pl_x) == 0 else pl_id
            # ========
        k += 1
    # записывает в Game_List спортсменов участников сетки и присваивает встречи 1-ого тура и записывает в тбл Results

        with db:
            game_list = Game_list(number_group=fin, rank_num_player=k, player_group_id=player_id,
                                  system_id=id_system, title_id=title_id()).save()

    for i in range(1, mp // 2 + 1):  # присваивает встречи 1-ого тура и записывает в тбл Results
        num_game = i
        pl1 = tds_full_name_city[i * 2 - 2]
        pl2 = tds_full_name_city[i * 2 - 1]
        if pl1 is not None and pl2 is not None:
            with db:
                results = Result(number_group=fin, system_stage=st, player1=pl1, player2=pl2,
                                 tours=num_game, title_id=title_id(), system_id=id_system).save()
    for i in range(mp // 2 + 1, game + 1):  # дополняет номера будущих встреч
        pl1 = ""
        pl2 = ""
        with db:
            results = Result(number_group=fin, system_stage=st, player1=pl1, player2=pl2,
                             tours=i, title_id=title_id(),system_id=id_system).save()


def player_in_one_table(fin):
    """Соревнования из одной таблицы, создание и заполнение Game_list, Result (создание жеребьевки в круг)"""
    one_table = []
    id_system = system_id(stage=fin)
    players = Player.select().where(Player.title_id == title_id())
    choice = Choice.select().where(Choice.title_id == title_id())
    system = System.select().where(System.title_id == title_id())
   
    k = 0
    for p in choice:  # цикл заполнения db таблиц -game list-
        k += 1
        player = p.family
        pl_id = p.player_choice_id
        pl_city = players.select().where(Player.id == pl_id).get()
        city = pl_city.city
        player_id = f"{player}/{pl_id}"
        one_table.append(f"{player}/{city}")
        game_list = Game_list(number_group=fin, rank_num_player=k, player_group_id=pl_id, system_id=id_system,
                            title_id=title_id())
        game_list.save()

    tours = tours_list(k - 3)
    round = 0
    for tour in tours: # цикл заполнения db таблиц -Result-
        round += 1
        for match in tour:
            znak = match.find("-")
            first = int(match[:znak])  # игрок под номером в группе
            second = int(match[znak + 1:])  # игрок под номером в группе
            pl1 = one_table[first - 1]
            pl2 = one_table[second - 1]
            results = Result(number_group=fin, system_stage="Одна таблица", player1=pl1, player2=pl2,
                             tours=match, title_id=title_id(), round=round, system_id=id_system)
            results.save()    


def player_fin_on_circle(fin):
    """заполняет таблицу Game_list данными спортсменами из группы, которые будут играть в финале по кругу
     td - список списков данных из групп"""
    msgBox = QMessageBox()
    fin_dict = {}
    sorted_dict = {}
    fin_list = []
    group_dict = {}
    stage = fin
    flag = 0
    id_system = system_id(stage)
    players = Player.select().where(Player.title_id == title_id())
    choice = Choice.select().where(Choice.title_id == title_id())
    system = System.select().where(System.id == id_system).get()  # находит system id последнего

    stage_exit = system.stage_exit # откуда выходят в финал
    # === если stage_exit == "", значит одна таблица в круг
    if stage_exit != "":
        nums = rank_mesto_out_in_group_or_semifinal_to_final(fin) # список мест, выходящих из группы или пф
        count_exit = len(nums) # количество игроков, выходящих в финал
    vid = ["Автоматическая", "Ручная"]
    
    vid, ok = QInputDialog.getItem(my_win, "Жеребьевка", "Выберите режим жеребьевки финала по кругу.", vid, 0, False)
    if vid == "Автоматическая":
        # ++++ вариант с добавлением игрока ++++++=
        result = msgBox.question(my_win, "Жеребьевка", "Будет ли добавлен игрок в финал по кругу.",
                                     msgBox.Ok, msgBox.Cancel)
        if result == msgBox.Ok:
            count_player, ok = QInputDialog.getInt(my_win, "Игроки", "Сколько добавляется в финал")
            player_in_final = system.max_player # количество игроков в финале
            System.update(max_player=(player_in_final + count_player)).where(System.id == id_system).execute()
            player_in_final = player_in_final + count_player # количество игроков в финале

            players = Player.select().where(Player.title_id == title_id())

            pl_list = [k.full_name for k in players]
            pl_list.sort()
            pl, ok = QInputDialog.getItem(my_win, "Жеребьевка", "Выберите спортсмена для добавления.", pl_list, 0, False)
            players_full = players.select().where(Player.full_name == pl).get()
            pl_name = players_full.player
            pl_id = players_full.id
            pl_full = f"{pl_name}/{pl_id}"
            flag = 1
        else:
           player_in_final = system.max_player # количество игроков в финале
        # ==== new variant ===
        # player_in_final = system.max_player # количество игроков в финале

        # == вариант когда осталось 2 человека
        if player_in_final == 2:
            game = 1
            tour = [['1-2']]
            kol_tours = len(tour)  # кол-во туров
        else:
        # ======================
            cp = player_in_final - 3
            tour = tours_list(cp)
            kol_tours = len(tour)  # кол-во туров
            game = len(tour[0])  # кол-во игр в туре
        # ===== получение списка номеров игроков в порядке 1-ого тура
        k = 0
        number_tours = []
        first_tour = tour[0].copy()
        first_tour.sort()

        for n in first_tour:
            z = n.find("-")
            num = int(n[:z])
            number_tours.append(num)
            num = int(n[z + 1:])
            number_tours.append(num)
        # =======
        if stage_exit == "Предварительный":
            choices_fin = choice.select().where(Choice.mesto_group.in_(nums))
            nt = 1
            for b in nums:
                choices_fin = choice.select().where(Choice.mesto_group == b)
                # =====
                for m in choices_fin:
                    num_group_text = m.group
                    znak = num_group_text.find(" ")
                    num_gr_int = int(num_group_text[:znak])
                    group_dict[m] = num_gr_int
                    grouplist = sorted(group_dict.items(), key=lambda x: x[1])
                    sortdict = dict(grouplist)
                    choices_fin_sort_by_group = sortdict.keys()
            # вариант с расстоновкой по 1-му туру
            for n in choices_fin_sort_by_group:
                player = n.family
                pl_id = n.player_choice_id
                player_id = f"{player}/{pl_id}"
                # проверить выход из группы в финал по кругу с неполными групами
                fin_dict[nt] = player_id # словарь (1-й номер наивысшее место в группе, затем место следующее в этой же группе)
                # if count_exit == 1:
                #     fin_dict[nt] = player_id # словарь (1-й номер наивысшее место в группе, затем место следующее в этой же группе)
                # else:
                #     fin_dict[number_tours[nt - 1]] = player_id # словарь (1-й номер наивысшее место в группе, затем место следующее в этой же группе)
                nt += 1
        elif stage_exit in ["1-й полуфинал", "2-й полуфинал"]: # если выход в финал по кругу из ПФ
            nt = 1
            for b in nums:
                choices_fin = choice.select().where((Choice.mesto_semi_final == b) & (Choice.semi_final == stage_exit))
                # ==== вариант перевести текст группы в число а потом отсортировать по группам (выход 1 человек из группы)
                for m in choices_fin:
                    num_group_text = m.sf_group
                    znak = num_group_text.find(" ")
                    num_gr_int = int(num_group_text[:znak])
                    group_dict[m] = num_gr_int
                    grouplist = sorted(group_dict.items(), key=lambda x: x[1])
                    sortdict = dict(grouplist) # словарь id игрока в choice - номер группы по возрастанию
                    choices_fin_sort_by_group = sortdict.keys()
                # ========
            for n in choices_fin_sort_by_group:
                player = n.family
                pl_id = n.player_choice_id
                player_id = f"{player}/{pl_id}"
                if count_exit == 1:
                    fin_dict[nt] = player_id # словарь (1-й номер наивысшее место в группе, затем место следующее в этой же группе)
                else:
                    sorted_dict[number_tours[nt - 1]] = player_id # словарь (1-й номер наивысшее место в группе, затем место следующее в этой же группе)
                nt += 1
            if count_exit > 1:
                fin_dict = dict(sorted(sorted_dict.items()))
            if flag == 1:
                fin_dict[player_in_final] = pl_full
        else:
            nt = 1
            choices_fin = Choice.select().where(Choice.title_id == title_id()).order_by(Choice.rank.desc()) # сортировка по рейтингу
            for n in choices_fin:
                    player = n.family
                    pl_id = n.player_choice_id
                    player_id = f"{player}/{pl_id}"
                    fin_dict[nt] = player_id
                    nt += 1
        #========        
        for nt in range(1, player_in_final + 1):
            fin_list.append(fin_dict[nt]) # список игроков в порядке 1 ого тура
        # == вариант с циклом по словарю
        k = 1
        for l in fin_dict.keys():
            ps_final = k if count_exit == 1 else l # если выход 1 то по порядку, если более то из списка туров
            fam = fin_dict[l]
            id_pl = int(fam[fam.find("/") + 1:])
            if fin == "Одна таблица":
                Choice.update(basic=fin).where((Choice.player_choice_id == id_pl) & (Choice.title_id == title_id())).execute()
            else:
                Choice.update(final=fin, posev_final = ps_final).where((Choice.player_choice_id == id_pl) & (Choice.title_id == title_id())).execute()
            game_list = Game_list(number_group=fin, rank_num_player=ps_final, player_group_id=id_pl, system_id=id_system,
                                title_id=title_id())
            game_list.save()
            k += 1         
        # ==========
        # исправить если из группы выходят больше 2-ух игроков
        for r in range(0, kol_tours):
            round = r + 1
            tours = tour[r]  # игры тура
            for d in range(0, game):  # цикл по играм тура
                match = tours[d]  # матч в туре
                znak = match.find("-")
                first = int(match[:znak])  # игрок под номером в группе
                # игрок под номером в группе
                second = int(match[znak + 1:])
                pl1_fam_id = fin_list[first - 1] # фамилия первого игрока /id
                z = pl1_fam_id.find("/") # находит черту
                pl1_fam = pl1_fam_id[:z] # отделяет фамилия от ид
                pl1_id = int(pl1_fam_id[z + 1:])
                pl1_city = players.select().where(Player.id == pl1_id).get()
                cit1 = pl1_city.city
                pl2_fam_id = fin_list[second - 1] # фамилия второго игрока
                z = pl2_fam_id.find("/")
                pl2_fam = pl2_fam_id[:z]
                pl2_id = int(pl2_fam_id[z + 1:])
                pl2_city = players.select().where(Player.id == pl2_id).get()
                cit2 = pl2_city.city
                full_pl1 = f"{pl1_fam}/{cit1}"
                full_pl2 = f"{pl2_fam}/{cit2}"
                with db:
                    results = Result(number_group=fin, system_stage="Финальный", player1=full_pl1, player2=full_pl2,
                                    tours=match, title_id=title_id(), round=round, system_id=id_system).save()
    else: # ручная расстоновка в финале
        pass
        # vid_input = ["Группа", "Полуфинал"]
        # vid_input, ok = QInputDialog.getItem(my_win, "Жеребьевка", "Выберите откуда выходят в финала по кругу.", vid_input, 0, False)
        # if vid_input == "Полуфинал":
        #     pf, ok = QInputDialog.getInt(my_win, "Полуфинал", "Из кого полуфинал, выходят в финал")

        # mesto, ok = QInputDialog.getInt(my_win, "Места", "Введите место, выходящее в финал")
        # if vid_input == "Группа":
        #     choices_fin = choice.select().where((Choice.mesto_group == mesto))
        # else:
        #     choices_fin = choice.select().where((Choice.mesto_semi_final == mesto) & (Choice.semi_final == pf))
        # # ==== вариант перевести текст группы в число а потом отсортировать по группам (выход 1 человек из группы)
        # for m in choices_fin:
        #     num_group_text = m.sf_group
        #     znak = num_group_text.find(" ")
        #     num_gr_int = int(num_group_text[:znak])
        #     group_dict[m] = num_gr_int
        #     grouplist = sorted(group_dict.items(), key=lambda x: x[1])
        #     sortdict = dict(grouplist) # словарь id игрока в choice - номер группы по возрастанию
        #     choices_fin_sort_by_group = sortdict.keys()
        # # ========
        # for n in choices_fin_sort_by_group:
        #     player = n.family
        #     pl_id = n.player_choice_id
        #     player_id = f"{player}/{pl_id}"
        #     if count_exit == 1:
        #         fin_dict[nt] = player_id # словарь (1-й номер наивысшее место в группе, затем место следующее в этой же группе)
        #     else:
        #         sorted_dict[number_tours[nt - 1]] = player_id # словарь (1-й номер наивысшее место в группе, затем место следующее в этой же группе)
        #     nt += 1    
# =====================
    with db:
        System.update(choice_flag=True).where(System.id == id_system).execute()
    title = Title.select().where(Title.id == title_id()).get()
    page_title = title.tab_enabled
    id_title = title.id
    if "Результаты" not in page_title:
        page_title = f"{page_title} Результаты"
    gamer = title.gamer
    with db:
        title.tab_enabled = page_title
        title.save()
    tab_enabled(id_title)
    pv = system.page_vid
    stage = fin
    table_made(pv, stage)


def _player_fin_on_circle(fin):
    """заполняет таблицу Game_list данными спортсменами из группы, которые будут играть в финале по кругу
     td - список списков данных из групп"""
    msgBox = QMessageBox()
    fin_dict = {}
    sorted_dict = {}
    fin_list = []
    group_dict = {}
    stage = fin
    id_system = system_id(stage)
    players = Player.select().where(Player.title_id == title_id())
    choice = Choice.select().where(Choice.title_id == title_id())
    system = System.select().where(System.id == id_system).get()  # находит system id последнего

    stage_exit = system.stage_exit # откуда выходят в финал
    # === если stage_exit == "", значит одна таблица в круг
    if stage_exit != "":
        nums = rank_mesto_out_in_group_or_semifinal_to_final(fin) # список мест, выходящих из группы или пф
        count_exit = len(nums) # количество игроков, выходящих в финал
    vid = ["Автоматическая", "Ручная"]
    
    vid, ok = QInputDialog.getItem(my_win, "Жеребьевка", "Выберите режим жеребьевки финала по кругу.", vid, 0, False)
    if vid == "Автоматическая":
        # ==== new variant ===
        player_in_final = system.max_player # количество игроков в финале

        # == вариант когда осталось 2 человека
        if player_in_final == 2:
            game = 1
            tour = [['1-2']]
            kol_tours = len(tour)  # кол-во туров
        else:
        # ======================
            cp = player_in_final - 3
            tour = tours_list(cp)
            kol_tours = len(tour)  # кол-во туров
            game = len(tour[0])  # кол-во игр в туре
        # ===== получение списка номеров игроков в порядке 1-ого тура
        k = 0
        number_tours = []
        first_tour = tour[0].copy()
        first_tour.sort()

        for n in first_tour:
            z = n.find("-")
            num = int(n[:z])
            number_tours.append(num)
            num = int(n[z + 1:])
            number_tours.append(num)
        # =======
        if stage_exit == "Предварительный":
            choices_fin = choice.select().where(Choice.mesto_group.in_(nums))
            nt = 1
            for b in nums:
                choices_fin = choice.select().where(Choice.mesto_group == b)
                # =====
                for m in choices_fin:
                    num_group_text = m.group
                    znak = num_group_text.find(" ")
                    num_gr_int = int(num_group_text[:znak])
                    group_dict[m] = num_gr_int
                    grouplist = sorted(group_dict.items(), key=lambda x: x[1])
                    sortdict = dict(grouplist)
                    choices_fin_sort_by_group = sortdict.keys()
            # вариант с расстоновкой по 1-му туру
            for n in choices_fin_sort_by_group:
                player = n.family
                pl_id = n.player_choice_id
                player_id = f"{player}/{pl_id}"
                # проверить выход из группы в финал по кругу с неполными групами
                fin_dict[nt] = player_id # словарь (1-й номер наивысшее место в группе, затем место следующее в этой же группе)
                # if count_exit == 1:
                #     fin_dict[nt] = player_id # словарь (1-й номер наивысшее место в группе, затем место следующее в этой же группе)
                # else:
                #     fin_dict[number_tours[nt - 1]] = player_id # словарь (1-й номер наивысшее место в группе, затем место следующее в этой же группе)
                nt += 1
        elif stage_exit in ["1-й полуфинал", "2-й полуфинал"]: # если выход в финал по кругу из ПФ
            nt = 1
            for b in nums:
                choices_fin = choice.select().where((Choice.mesto_semi_final == b) & (Choice.semi_final == stage_exit))
                # ==== вариант перевести текст группы в число а потом отсортировать по группам (выход 1 человек из группы)
                for m in choices_fin:
                    num_group_text = m.sf_group
                    znak = num_group_text.find(" ")
                    num_gr_int = int(num_group_text[:znak])
                    group_dict[m] = num_gr_int
                    grouplist = sorted(group_dict.items(), key=lambda x: x[1])
                    sortdict = dict(grouplist) # словарь id игрока в choice - номер группы по возрастанию
                    choices_fin_sort_by_group = sortdict.keys()
                # ========
            for n in choices_fin_sort_by_group:
                player = n.family
                pl_id = n.player_choice_id
                player_id = f"{player}/{pl_id}"
                if count_exit == 1:
                    fin_dict[nt] = player_id # словарь (1-й номер наивысшее место в группе, затем место следующее в этой же группе)
                else:
                    sorted_dict[number_tours[nt - 1]] = player_id # словарь (1-й номер наивысшее место в группе, затем место следующее в этой же группе)
                nt += 1
            if count_exit > 1:
                fin_dict = dict(sorted(sorted_dict.items()))
        else:
            nt = 1
            choices_fin = Choice.select().where(Choice.title_id == title_id()).order_by(Choice.rank.desc()) # сортировка по рейтингу
            for n in choices_fin:
                    player = n.family
                    pl_id = n.player_choice_id
                    player_id = f"{player}/{pl_id}"
                    fin_dict[nt] = player_id
                    nt += 1
        #========        
        for nt in range(1, player_in_final + 1):
            fin_list.append(fin_dict[nt]) # список игроков в порядке 1 ого тура
        # == вариант с циклом по словарю
        k = 1
        for l in fin_dict.keys():
            ps_final = k if count_exit == 1 else l # если выход 1 то по порядку, если более то из списка туров
            fam = fin_dict[l]
            id_pl = int(fam[fam.find("/") + 1:])
            if fin == "Одна таблица":
                Choice.update(basic=fin).where((Choice.player_choice_id == id_pl) & (Choice.title_id == title_id())).execute()
            else:
                Choice.update(final=fin, posev_final = ps_final).where((Choice.player_choice_id == id_pl) & (Choice.title_id == title_id())).execute()
            game_list = Game_list(number_group=fin, rank_num_player=ps_final, player_group_id=id_pl, system_id=id_system,
                                title_id=title_id())
            game_list.save()
            k += 1         
        # ==========
        # исправить если из группы выходят больше 2-ух игроков
        for r in range(0, kol_tours):
            round = r + 1
            tours = tour[r]  # игры тура
            for d in range(0, game):  # цикл по играм тура
                match = tours[d]  # матч в туре
                znak = match.find("-")
                first = int(match[:znak])  # игрок под номером в группе
                # игрок под номером в группе
                second = int(match[znak + 1:])
                pl1_fam_id = fin_list[first - 1] # фамилия первого игрока /id
                z = pl1_fam_id.find("/") # находит черту
                pl1_fam = pl1_fam_id[:z] # отделяет фамилия от ид
                pl1_id = int(pl1_fam_id[z + 1:])
                pl1_city = players.select().where(Player.id == pl1_id).get()
                cit1 = pl1_city.city
                pl2_fam_id = fin_list[second - 1] # фамилия второго игрока
                z = pl2_fam_id.find("/")
                pl2_fam = pl2_fam_id[:z]
                pl2_id = int(pl2_fam_id[z + 1:])
                pl2_city = players.select().where(Player.id == pl2_id).get()
                cit2 = pl2_city.city
                full_pl1 = f"{pl1_fam}/{cit1}"
                full_pl2 = f"{pl2_fam}/{cit2}"
                with db:
                    results = Result(number_group=fin, system_stage="Финальный", player1=full_pl1, player2=full_pl2,
                                    tours=match, title_id=title_id(), round=round, system_id=id_system).save()
    else: # ручная расстоновка в финале
        pass
        # vid_input = ["Группа", "Полуфинал"]
        # vid_input, ok = QInputDialog.getItem(my_win, "Жеребьевка", "Выберите откуда выходят в финала по кругу.", vid_input, 0, False)
        # if vid_input == "Полуфинал":
        #     pf, ok = QInputDialog.getInt(my_win, "Полуфинал", "Из кого полуфинал, выходят в финал")

        # mesto, ok = QInputDialog.getInt(my_win, "Места", "Введите место, выходящее в финал")
        # if vid_input == "Группа":
        #     choices_fin = choice.select().where((Choice.mesto_group == mesto))
        # else:
        #     choices_fin = choice.select().where((Choice.mesto_semi_final == mesto) & (Choice.semi_final == pf))
        # # ==== вариант перевести текст группы в число а потом отсортировать по группам (выход 1 человек из группы)
        # for m in choices_fin:
        #     num_group_text = m.sf_group
        #     znak = num_group_text.find(" ")
        #     num_gr_int = int(num_group_text[:znak])
        #     group_dict[m] = num_gr_int
        #     grouplist = sorted(group_dict.items(), key=lambda x: x[1])
        #     sortdict = dict(grouplist) # словарь id игрока в choice - номер группы по возрастанию
        #     choices_fin_sort_by_group = sortdict.keys()
        # # ========
        # for n in choices_fin_sort_by_group:
        #     player = n.family
        #     pl_id = n.player_choice_id
        #     player_id = f"{player}/{pl_id}"
        #     if count_exit == 1:
        #         fin_dict[nt] = player_id # словарь (1-й номер наивысшее место в группе, затем место следующее в этой же группе)
        #     else:
        #         sorted_dict[number_tours[nt - 1]] = player_id # словарь (1-й номер наивысшее место в группе, затем место следующее в этой же группе)
        #     nt += 1    
# =====================
    with db:
        system.choice_flag = True
        system.save()    
    title = Title.select().where(Title.id == title_id()).get()
    page_title = title.tab_enabled
    id_title = title.id
    if "Результаты" not in page_title:
        page_title = f"{page_title} Результаты"
    gamer = title.gamer
    with db:
        title.tab_enabled = page_title
        title.save()
    tab_enabled(id_title)
    pv = system.page_vid
    stage = fin
    table_made(pv, stage)


def player_in_table_group_and_write_Game_list_Result(stage):
    """заполняет таблицу Game_list данными спортсменами из группы td - список списков данных из групп и записывает
    встречи по турам в таблицу -Result- """
    system = System.select().where((System.title_id == title_id()) & (System.stage == stage)).get() # находит system id этапа
    system_id = system.id
    kg = system.total_group
    pv = system.page_vid
    # удаление старых записей в game_list и Result после редактирования жеребьевки групп
    game_list_delete = Game_list.delete().where((Game_list.title_id == title_id()) & (Game_list.system_id == system_id))
    result_delete = Result.delete().where((Result.title_id == title_id()) & (Result.system_id == system_id))
    game_list_delete.execute()
    result_delete.execute()  
    # создание таблиц групп со спортсменами согласно жеребьевки в PDF
    table_made(pv, stage)
    # вызов функции, где получаем список всех участников по группам
    tdt_all = table_data(stage, kg)
    for p in range(0, kg):  # цикл заполнения db таблиц -game list- и  -Results-
        gr = tdt_all[0][p]
        # ======
        gr_id = tdt_all[2][p] # == вариант с id вмест фамилии для Game_list
        # =====
        count_player = len(gr) // 2  # максимальное кол-во участников в группе
        number_group = str(p + 1) + ' группа'
        k = 0  # кол-во спортсменов в группе
        for i in range(0, count_player * 2 - 1, 2):
            family_player = gr[i][1]  # фамилия игрока
            player_id = gr_id[i][1]  # id игрока
            posev = int(gr[i][0]) # посев (номер игрока в группе)
            fp = len(family_player) # кол-во знаков фамилии, если 0 значит игрока нет
                # подсчет кол-во знаков в фамилия, если 0 значит игрока нет
            if fp > 0:  # если строка (фамилия игрока) не пустая идет запись в db
                k += 1
                # записывает в DB id игрока
                player_id = int(gr_id[i][1])  # id игрока
                with db:
                    game_list = Game_list(number_group=number_group, rank_num_player=posev, 
                                            player_group_id=player_id,
                                            system_id=system_id, title_id=title_id()).save()

        # если 1-я строка (фамилия игрока) пустая выход из группы
        if fp == 0 and k != 0 or k == count_player:
            cp = k - 3
            tour = tours_list(cp)
            kol_tours = len(tour)  # кол-во туров
            game = len(tour[0])  # кол-во игр в туре
            for r in range(0, kol_tours):
                round = r + 1
                tours = tour[r]  # игры тура
                for d in range(0, game):  # цикл по играм тура
                    match = tours[d]  # матч в туре
                    znak = match.find("-")
                    first = int(match[:znak])  # игрок под номером в группе
                    # игрок под номером в группе
                    second = int(match[znak + 1:])
                    pl1_id = gr[first * 2 - 2][1]  # фамилия первого игрока
                    # z = pl1_id.find("/") # находит черту
                    # pl1 = pl1_id[:z] # отделяет фамилия от ид
                    pl2_id = gr[second * 2 - 2][1]  # фамилия второго игрока
                    # z = pl2_id.find("/")
                    # pl2 = pl2_id[:z]
                    cit1 = gr[first * 2 - 1][1] # город 1-ого игрока
                    cit2 = gr[second * 2 - 1][1] # город 2-ого игрока
                    full_pl1 = f"{pl1_id}/{cit1}"
                    full_pl2 = f"{pl2_id}/{cit2}"
                    with db:
                        results = Result(number_group=number_group, system_stage=stage, player1=full_pl1, player2=full_pl2,
                                         tours=match, title_id=title_id(), round=round, system_id=system_id).save()


def player_in_setka_and_write_Game_list_Result(stage, posev_list, full_name_list):
    """меняет игроков в сетке на новые места в посеве"""
    gl_id_list = []
    res_id_list = []
    id_system = system_id(stage)
    results = Result.select().where((Result.title_id == title_id()) & (Result.system_id == id_system))
    n = 0
    for g in full_name_list:
        res = results.select().where((Result.player1 == g) | (Result.player2 == g)).get()
        res_id_list.append(res.id)
    b = 0
    for k in res_id_list:
        rt = results.select().where(Result.id == k).get()
        pl1 = rt.player1
        if pl1 == full_name_list[b]:
            Result.update(player1=full_name_list[1 - b]).where(Result.id == k).execute()
        else:
            Result.update(player2=full_name_list[1 - b]).where(Result.id == k).execute()
        b += 1   
    for k in posev_list:
        gl_id = Game_list.select().where((Game_list.rank_num_player == k) & (Game_list.system_id == id_system)).get()
        gl_id_list.append(gl_id)
    for m in gl_id_list:
        Game_list.update(rank_num_player=posev_list[1-n]).where(Game_list.id == m).execute()
        n += 1                  


def chop_line(t, maxline=31):
    """перевод строки если слишком длинный список тренеров"""
    l = len(t)
    if l > maxline:
        s1 = t.find(",", 0, maxline)
        s2 = t.find(",", s1 + 1, maxline)       
        cant = len(t) // maxline
        cant += 1
        strline = ""
        if s2 == -1: # если две фамилии больше 31, перевод после 1-ой фамилии
            for k in range(1, cant):
                index = maxline * k
                strline += "%s\n" % (t[(index - maxline):s1 + 1])
            strline += "%s" % (t[s1 + 1:])
        else:
            for k in range(1, cant):
                index = maxline * k
                strline += "%s\n" % (t[(index - maxline):s2 + 1])
            strline += "%s" % (t[s2 + 1:])
        t = strline
    return t


def chop_line_city(g, maxline=15):
    """перевод строки если слишком длинный список города"""
    l = len(g)
    if l > maxline:
        s1 = g.find(" ", 0, maxline)
        s2 = g.find(" ", s1 + 1, maxline)       
        strline = ""
        if s2 == -1: # если две фамилии больше 31, перевод после 1-ой фамилии
            strline = g[:s1]
        else:
            strline = g[:s2]
        g = strline
    return g


def change_status_visible_and_score_game():
    """изменение статуса колво партий и ввод счета во встречи"""
    sender = my_win.sender()
    system = System.select().where(System.title_id == title_id())
    tab_etap = my_win.tabWidget_stage.currentIndex()
    idx = my_win.tableView.currentIndex() # определиние номера строки
    row_num = idx.row()
    if row_num == -1:
        return
    count = len(system)    
    if row_num == -1: # не выбрана ни одна встреча
        system_stage = True
        match_db = 5
        match_current = 5
        state_visible = True
    else:
        if count == 1:
            stage = "Одна таблица"
        else:
            if tab_etap == 0:
                stage = "Предварительный"
            elif tab_etap == 1:
                stage = my_win.comboBox_filter_semifinal.currentText()
            else:
                stage = my_win.tableView.model().index(row_num, 2).data() #  данные ячейки (из какого финала играют встречу)
        id_system = system_id(stage)
        system_stage = system.select().where(System.id == id_system).get()
        match_db = system_stage.score_flag
        state_visible = system_stage.visible_game  # флаг, показывающий записывать счет в партиях или нет
        match_current = match_db    
        #  ==== изменение состояние =====
    if sender == my_win.checkBox_4:
        for i in my_win.groupBox_kolvo_vstrech.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
            if i.isChecked():
                match_current = int(i.text())
                break
        state_visible = my_win.checkBox_4.isChecked()
    elif (sender == my_win.radioButton_match_3 or 
        sender == my_win.radioButton_match_5 or sender == my_win.radioButton_match_7):
        for i in my_win.groupBox_kolvo_vstrech.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
            if i.isChecked():
                match_current = int(i.text())
                break
    if match_current == 3:
        my_win.radioButton_match_3.setChecked(True)
        my_win.frame_three.setVisible(True)
        my_win.frame_five.setVisible(False)
        my_win.frame_seven.setVisible(False)
    elif match_current == 5:
        my_win.radioButton_match_5.setChecked(True)
        my_win.frame_three.setVisible(True)
        my_win.frame_five.setVisible(True)
        my_win.frame_seven.setVisible(False)
    elif match_current == 7:
        my_win.radioButton_match_7.setChecked(True)
        my_win.frame_three.setVisible(True)
        my_win.frame_five.setVisible(True)
        my_win.frame_seven.setVisible(True)
    my_win.label_22.setVisible(True)
   
    if state_visible is False:
        frame_list = [my_win.frame_three,  my_win.frame_five, my_win.frame_seven]
        # if tab == 3:
        for k in frame_list:
            k.setVisible(False)
        my_win.checkBox_4.setChecked(False)
        my_win.lineEdit_pl1_score_total.setFocus(True)
    my_win.label_22.setVisible(False)
    systems = system.select().where(System.id == id_system).get()
    with db:
        systems.score_flag = match_current
        systems.visible_game = state_visible
        systems.save()
    return state_visible


def visible_field(stage):
    """включает или выключает поля для ввода счета, state - игра со счетом, True если включить поля для счета"""
    sender = my_win.sender()
    system = System.select().where(System.title_id == title_id())
    # ==== текущее состояние радиокнопок и чекбокса кол-во партий и ввод счета =====
    tab = my_win.tabWidget.currentIndex()
    idx = my_win.tableView.currentIndex() # номер выделенной строки
    row_num = idx.row()

    # stage = "Предварительный"
    system_stage = system.select().where(System.stage == stage).get()
    state_visible = system_stage.visible_game
    my_win.checkBox_4.setChecked(state_visible)
        # ======= записывает изменение в базу данных  
    if sender == my_win.checkBox_4: # изменяет состояние чекбокса игра со счетом или нет
        if tab == 3:
            state_visible = my_win.checkBox_4.isChecked()
            if state_visible is True:
                my_win.lineEdit_pl1_s1.setFocus()
            else:
                my_win.lineEdit_pl1_score_total.setFocus()
    change_status_visible_and_score_game() 
    return state_visible


# def change_tab_filter():
#     """Меняет вкладку фильтров страницы результаты ы зависимости от этапа"""
#     # sender = my_win.groupBox_result.sender()
#     stage_current = radiobutton_stage()
#     # for i in my_win.groupBox_result.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
#     #             if i.isChecked():
#     #                 stage_current = i.text()
#     #                 break
#     #             elif (sender == my_win.radioButton_group or 
#     #                 sender == my_win.radioButton_semifinal or sender == my_win.radioButton_final):
#     #                 for i in my_win.groupBox_result.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
#     #                     if i.isChecked():
#     #                         stage_current = i.text()
#     #                         break
#     if stage_current == "Предварительный":
#         my_win.stackedWidget.setCurrentIndex(0)       
#     elif stage_current == "Полуфинальный":
#         my_win.stackedWidget.setCurrentIndex(1)
#     else:
#         my_win.stackedWidget.setCurrentIndex(2)
#         # page()
#     my_win.label_result.setText(f"{stage_current} этап")


def select_player_in_list():
    """выводит данные игрока в поля редактирования или удаления"""
    data_list = []
    titles = Title.select().where(Title.id == title_id()).get()
    for idx in my_win.tableView.selectedIndexes():
        row_num = idx.row()
        col_num = idx.column()
        data = my_win.tableView.model().index(row_num, col_num).data()
        data_list.append(data)

    pl_id = my_win.tableView.model().index(row_num, 0).data() # данные ячейки tableView
    data_list.insert(0, pl_id)

    my_win.lineEdit_id.setText(data_list[0])
    my_win.lineEdit_id.setEnabled(False)
    my_win.lineEdit_Family_name.setText(data_list[1])
    my_win.lineEdit_bday.setText(data_list[2])
    my_win.lineEdit_R.setText(data_list[3])
    my_win.lineEdit_city_list.setText(data_list[4])
    my_win.comboBox_region.setCurrentText(data_list[5])
    my_win.comboBox_razryad.setCurrentText(data_list[6])
    my_win.lineEdit_coach.setText(data_list[7])

    my_win.Button_add_edit_player.setEnabled(True)
    flag_otc = titles.otchestvo # флаг нужно ли отчество в списках или нет (flag-1 да)
    if my_win.checkBox_6.isChecked():  # отмечен флажок -удаленные-
        my_win.Button_del_player.setEnabled(False)
        my_win.Button_add_edit_player.setText("Восстановить")
        # =========== отображение отчества ===
        if flag_otc == 1:
            delete_players = Delete_player.get(Delete_player.id == int(pl_id))
            otc_id = delete_players.patronymic_id        
        # =====
    else:
        my_win.Button_del_player.setEnabled(True)
        my_win.Button_add_edit_player.setEnabled(True)
        my_win.Button_add_edit_player.setText("Редактировать")
        # =========== отображение отчества ===
        if flag_otc == 1:
            players = Player.get(Player.id == int(pl_id))
            otc_id = players.patronymic_id        
        # =====
    if flag_otc == 1:
        otchestvo = Patronymic.get(Patronymic.id == otc_id)
        otc = otchestvo.patronymic
        my_win.lineEdit_otchestvo.setText(otc)
    if my_win.checkBox_11.isChecked():  # отмечен флажок -оплата R-
        my_win.Button_pay_R.setEnabled(True)
    else:
        my_win.Button_pay_R.setEnabled(False)


def save_in_db_pay_R():
    """запись в базу данных оплату рейтинга"""
    idx = my_win.tableView.currentIndex() # номер выделенной строки
    row_num = idx.row()
    pl_id = my_win.tableView.model().index(row_num, 0).data() # данные ячейки tableView
    comment, ok = QInputDialog.getText(my_win, "Коментарий", "Введите коментарий о месте нахождении квитанции.")
    if ok:
        query = Player.update(pay_rejting="оплачен", comment=comment).where(Player.id == pl_id)
        query.execute()
    else:
        return
    debitor_R()


def check_repeat_player(pl, bd):
    """фукция проверки повтора ввода одно и того же игрока"""
    dr = []
    player_list = Player.select().where(Player.title_id == title_id())
    repeat = player_list.select().where(Player.player == pl) 
    count_family = len(repeat)
    if count_family != 0:
        for l in repeat:
            b_day = l.bday
            b_day_formated = b_day.strftime('%d.%m.%Y') # форматирование даты
            dr.append(b_day_formated)
        if bd in dr:
            my_win.textEdit.setText("Такой игрок уже присутствует в списках!")   
            flag = True
        else:
            flag = False
    else:
        flag = False
    return flag


def select_player_in_game():
    """выводит фамилии игроков встречи"""
    tab = my_win.tabWidget.currentIndex()
    # tab_etap = my_win.tabWidget_stage.currentIndex()
    row_num= my_win.tableView.currentIndex().row() # определиние номера строки
    numer_game = my_win.tableView.model().index(row_num, 3).data()
    if tab == 1:
        select_player_in_list()
    elif tab ==2:
        change_choice_group()
    elif tab == 3:
        my_win.groupBox_kolvo_vstrech.setEnabled(True)
        state_visible = change_status_visible_and_score_game()
        pl1 = my_win.tableView.model().index(row_num, 4).data()
        pl2 = my_win.tableView.model().index(row_num, 5).data()
        pl_win = my_win.tableView.model().index(row_num, 6).data()
        win_pole = my_win.tableView.model().index(row_num, 7).data()
        sc = my_win.tableView.model().index(row_num, 8).data()

        if win_pole != "None" and win_pole != "":  # встреча сыграна, то заполняет поля общий счет
            if pl1 == pl_win:
                # если в сетке недостающие игроки (bye), то нет счета
                if sc != "":
                    sc1 = sc[0]
                    sc2 = sc[4]
                else:  # оставляет поля общий счет пустыми
                    sc1 = ""
                    sc2 = ""
            else:
                # если в сетке недостающие игроки (bye), то нет счета
                if sc != "":
                    sc1 = sc[4]
                    sc2 = sc[0]
                else:
                    sc1 = ""
                    sc2 = ""
            my_win.lineEdit_pl1_score_total.setText(sc1)
            my_win.lineEdit_pl2_score_total.setText(sc2)
            my_win.lineEdit_player1.setText(pl1)
            my_win.lineEdit_player2.setText(pl2)
            my_win.lineEdit_pl1_s1.setFocus()
        else: # новая встреча
                my_win.checkBox_7.setEnabled(True)
                my_win.checkBox_8.setEnabled(True)
                my_win.checkBox_7.setChecked(False)
                my_win.checkBox_8.setChecked(False)
                my_win.groupBox_match.setTitle(f"Встреча №{numer_game}")
                my_win.lineEdit_player1.setText(pl1)
                my_win.lineEdit_player2.setText(pl2)
 
                if pl1 == "X" or pl2 == "X":
                    my_win.Button_Ok.setEnabled(True)
                    my_win.Button_Ok.setFocus()                   
                else:
                    if state_visible is True:
                        my_win.lineEdit_pl1_s1.setFocus()
                    else:
                        my_win.lineEdit_pl1_score_total.setFocus()
        my_win.tableView.selectRow(row_num)
    elif tab == 4:
        tb_double = my_win.tabWidget_3.currentIndex()
        fam = my_win.tableView.model().index(row_num, 1).data()
        city = my_win.tableView.model().index(row_num, 4).data()
        r =  my_win.tableView.model().index(row_num, 3).data()
        if tb_double == 0:
            my_win.lineEdit_pl1_double.clear()
            my_win.lineEdit_pl1_double.setText(fam)
            my_win.lineEdit_city_pl1.setText(city)
            my_win.r_pl1.setText(r)
    elif tab == 7:
        player_id = my_win.tableView.model().index(row_num, 0).data()
        players = Player.select().where(Player.id == player_id).get()
        player = players.full_name

        player_list = Result.select().where(((Result.player1 == player) | (Result.player2 == player)) & (Result.title_id == title_id()))
        fill_table(player_list)
    


def delete_player():
    """удаляет игрока из списка и заносит его в архив"""
    msgBox = QMessageBox
    game_list = Game_list.select().where(Game_list.title_id == title_id())
    system = System.select().where(System.title_id == title_id())
    result = Result.select().where(Result.title_id == title_id())
    titles = Title.select().where(Title.id == title_id()).get()
    flag_otc = titles.otchestvo
    idx = my_win.tableView.currentIndex() # определиние номера строки
    row_num = idx.row()

    player_id  = my_win.tableView.model().index(row_num, 0).data() # данные ячейки tableView
    player_del  = my_win.tableView.model().index(row_num, 1).data() # данные ячейки tableView
    birthday  = my_win.tableView.model().index(row_num, 2).data() # данные ячейки tableView
    rank   = my_win.tableView.model().index(row_num, 3).data() # данные ячейки tableView
    player_city_del  = my_win.tableView.model().index(row_num, 4).data() # данные ячейки tableView
    region = my_win.tableView.model().index(row_num, 5).data() # данные ячейки tableView
    razryad  = my_win.tableView.model().index(row_num,6).data() # данные ячейки tableView
    coach  = my_win.tableView.model().index(row_num, 7).data() # данные ячейки tableView
    full_name = f"{player_del}/{player_city_del}"
    coach_id = Coach.get(Coach.coach == coach)
    player = Player.select().where(Player.id == player_id).get()
    pay_R = player.pay_rejting
    comment = player.comment
    if flag_otc == 1:
        patronymic_id = player.patronymic_id
    else:
        # patronymic_id = ""
        patronymic_id = 0
    question = msgBox.question(my_win, "", f"Вы действительно хотите удалить\n"
                                         f" {player_del} город {player_city_del}?",
                             msgBox.Ok, msgBox.Cancel)
    if question == msgBox.Ok:
        system_flag = ready_system() # проверка была создана система True - система создана
        if system_flag is True:
            count = len(system)
        # ============ корректрует запись в таблице -system- после удаления игрока
            if count == 1: # значит система состоит из одной таблицы
                sys = system.id
                athlet = sys.total_athletes
                kg = sys.total_group
            else:
                for sys in system:
                    stage = sys.stage
                    if stage == "Предварительный":
                        id_system = system_id(stage)
                        athlet = sys.total_athletes # кол-во участников
                        athlet -= 1
                        kg = sys.total_group # кол-во групп
                        e1 = athlet % int(kg)
                        # если количество участников равно делится на группы (кол-во групп)
                        p = athlet // int(kg)
                        g1 = int(kg) - e1  # кол-во групп, где наименьшее кол-во спортсменов
                        g2 = int(p + 1)  # кол-во человек в группе с наибольшим их количеством
                        if e1 == 0:  # то в группах равное количество человек -e1-
                            stroka_kol_group = f"{kg} группы по {str(p)} чел."
                        else:
                            stroka_kol_group = f"{str(g1)} групп(а) по {str(p)} чел. и {str(e1)} групп(а) по {str(g2)} чел."
                        System.update(total_athletes=athlet, label_string=stroka_kol_group).where(System.id == id_system).execute()
                    check_flag = check_choice(stage)
                    if check_flag is True:
                        question = msgBox.information(my_win, "", f"Уже была произведена жеребьевка!\n"
                                                        f" {player_del} город {player_city_del}\n"
                                                        "будет удален(а) из посева.",
                                            msgBox.Ok)
                        
                        choices = Choice.delete().where(Choice.player_choice_id == player_id)
                        choices.execute()
                        game_lists = game_list.select().where(Game_list.player_group_id == player_id).get()
                        posev = game_lists.rank_num_player
                        number_group = game_lists.number_group
                        # === изменяет номера посева, если удаляемый игрок не в последний посев ==
                        g_list = game_list.select().where((Game_list.system_id == id_system) & 
                                                        (Game_list.number_group == number_group))
                        for k in g_list:
                            gl_id = k.id
                            ps = k.rank_num_player # посев игрока
                            if posev < ps:
                                rank_in_group = ps - 1
                                gl = Game_list.update(rank_num_player=rank_in_group).where(Game_list.id == gl_id)
                                gl.execute()
                            elif posev == ps:
                        # === удаляет игрока из Game_list ===
                                gl = Game_list.delete().where(Game_list.id == gl_id)
                                gl.execute()
                        # ==== заменяет туры (удаляет встречи с удаленным игроком)
                        result_game = result.select().where((Result.system_id == id_system) & 
                                                                (Result.number_group == number_group))
                        fam_city_del = f"{player_del}/{player_city_del}"
                        for k in result_game:
                            pl1 = k.player1
                            pl2 = k.player2
                            if pl1 == fam_city_del or pl2 == fam_city_del:
                                res = Result.delete().where(Result.id == k)
                                res.execute()
                        for k in result_game:
                            tour = k.tours
                            znak = tour.find("-")
                            p1 = int(tour[:znak])  # игрок под номером в группе
                            p2 = int(tour[znak + 1:])  # игрок под номером в группе
                            if p1 > posev and p2 > posev:
                                p1 -= 1
                                p2 -= 1
                            elif p1 > posev:
                                p1 -= 1
                            elif p2 > posev:
                                p2 -= 1
                            new_tour = f"{p1}-{p2}"
                            res = Result.update(tours=new_tour).where(Result.id == k)
                            res.execute()                
        # записывает в таблицу -Delete player-
        birthday_mod = format_date_for_db(str_date=birthday)
        with db: 
            del_player   = Delete_player(bday=birthday_mod, rank=rank, city=player_city_del,
                                        region=region, razryad=razryad, coach_id=coach_id, full_name=full_name,
                                        player=player_del, title_id=title_id(), pay_rejting=pay_R, comment=comment, patronymic_id=patronymic_id).save()
            
        if system_flag is True:  # удаляет игрока из таблицы -Choice-, если была создана система 
            choices = Choice.select().where((Choice.title_id == title_id()) & (Choice.player_choice_id == player_id))
            count = len(choices)
            if count == 1: # если -0-, значит удалили после жеребьевки
                id_choice = choices.id
                pl_choice_del = Choice.get(Choice.id == id_choice)
                pl_choice_del.delete_instance()     
        pl_del = Player.get(Player.id == player_id)
        pl_del.delete_instance() # удаляет игрока из таблицы -PLayer-

        my_win.lineEdit_id.clear()
        my_win.lineEdit_Family_name.clear()
        my_win.lineEdit_bday.clear()
        my_win.lineEdit_R.clear()
        my_win.lineEdit_city_list.clear()
        my_win.lineEdit_coach.clear()
        my_win.lineEdit_otchestvo.clear()
        player_list_pred = Player.select().where((Player.title_id == title_id()) & (Player.application == "предварительная"))
        count = len(player_list_pred)    
        my_win.label_predzayavka.setText(f"По предзаявке {count} чел.")
        my_win.label_predzayavka.setStyleSheet("color: black")
        player_list = Player.select().where(Player.title_id == title_id())
        count = len(player_list)
        my_win.label_46.setText(f"Всего: {count} участников")
        fill_table(player_list)
    else:
        return


def sortByAlphabet(inputStr):
    inputStr = inputStr.lower()
    return inputStr[0]
  

def load_comboBox_filter():
    """загрузка комбобокса регионами для фильтрации списка"""
    my_win.comboBox_fltr_region.clear()
    my_win.comboBox_fltr_city.clear()
    reg = []
    gorod = []
    player = Player.select().where(Player.title_id == title_id())
    # player = Player.select().where((Player.title_id == title_id()) & (Player.bday != "0000-00-00"))
    if my_win.comboBox_fltr_region.count() > 0:  # проверка на заполненность комбобокса данными
        return
    else:
        for r in player:
            reg_n = r.region
            reg_n = reg_n.strip() # удаляет лишние пробелы
            if reg_n not in reg:
                reg.append(reg_n)
        if "" in reg:
            reg.remove("")
        reg.sort(key=sortByAlphabet)
        reg.insert(0, "")
        my_win.comboBox_fltr_region.addItems(reg)
    
    if my_win.comboBox_fltr_city.count() < 0:  # проверка на заполненность комбобокса данными
        for c in player:
            cityes = c.city
            if cityes not in gorod:
                gorod.append(cityes)
        gorod.sort(key=sortByAlphabet)
        gorod.insert(0, "")
        my_win.comboBox_fltr_city.addItems(gorod)


def change_city_from_region_in_R():
    """изменяет список городов в комбобоксе фильтра списка в зависимости от региона на вкладке рейтинг"""
    gorod = []
    my_win.comboBox_filter_city_in_R.clear()
    region = my_win.comboBox_filter_region_in_R.currentText()
    r_data_m = [R_list_m, R1_list_m]
    r_data_w = [R_list_d, R1_list_d]
    gamer_w = ["Девочки", "Девушки", "Юниорки", "Женщины"]
    id_title = Title.select().where(Title.id == title_id()).get()
    gamer = id_title.gamer
    cur_index = my_win.comboBox_choice_R.currentIndex()
    flag = 0
    if cur_index == 0: # если выбран текущий рейтинг
        if gamer in gamer_w:
            r_data = r_data_w[0]
        else:
            r_data = r_data_m[0]
        r_region = r_data.select().where(r_data.r_region == region)
        flag = 0        
    elif cur_index == 1: # если рейтинг за январь
        if gamer in gamer_w:
            r_data = r_data_w[1]
        else:
           r_data = r_data_m[1]
        r_region = r_data.select().where(r_data.r1_region == region) 
        flag = 1
    for pl_reg in r_region:
        r_gorod = pl_reg.r_city if flag == 0 else pl_reg.r1_city 
        if r_gorod not in gorod:
                gorod.append(r_gorod)
    gorod.sort(key=sortByAlphabet)
    gorod.insert(0, "")
    my_win.comboBox_filter_city_in_R.addItems(gorod)


def change_city_from_region():
    """изменяет список городов в комбобоксе фильтра списка в зависимости от региона"""  
    gorod = []
    my_win.comboBox_fltr_city.clear()
    player = Player.select().where(Player.title_id == title_id())
    region = my_win.comboBox_fltr_region.currentText()
    if region == "":
        player_region = player.select()
    else:
        player_region = player.select().where(Player.region == region)
    for pl_reg in player_region:
        if pl_reg.city not in gorod:
            gorod.append(pl_reg.city)
    if "" in gorod:
        gorod.remove("")
    gorod.sort(key=sortByAlphabet)
    gorod.insert(0, "")
    my_win.comboBox_fltr_city.addItems(gorod)


def filter_player_list(sender):
    """фильтрация списка участников по областям, тренерам, городам"""
    sender = my_win.sender()
    player = Player.select().where(Player.title_id == title_id())
    if sender == my_win.Button_fltr_list: # кнопка применить 
        region = my_win.comboBox_fltr_region.currentText()
        city = my_win.comboBox_fltr_city.currentText()
        coach = my_win.comboBox_fltr_coach.currentText()
        if region != "" and city != "":

            player_list = player.select().where((Player.region == region)  & (Player.city == city))
        elif region == "" and city != "":
            player_list = player.select().where(Player.city == city)
        elif region != "" and city == "":
            player_list = player.select().where(Player.region == region)
        else:
            coach_id_list = []
            ch = Coach.select().where(Coach.coach.contains(f'{coach}')) # поиск по части предложения
            for j in ch:
                id_ch = j.id
                if id_ch not in coach_id_list:
                    coach_id_list.append(id_ch)
            player_list = Player.select().where((Player.title_id == title_id()) & (Player.coach_id.in_(coach_id_list)))
    elif sender == my_win.checkBox_15: # отмечен чекбокс предзаявка
        if my_win.checkBox_15.isChecked():
            region = my_win.comboBox_fltr_region.currentText()
            my_win.Button_app.setEnabled(True)
            if region != "":
                player_list = player.select().where((Player.application == "предварительная") & (Player.region == region))
            else:
                player_list = player.select().where(Player.application == "предварительная")
            count = len(player_list)
        else:
            my_win.Button_app.setEnabled(False)
            my_win.textEdit.clear()
            player_list_pred = player.select().where(Player.application == "предварительная")
            count = len(player_list_pred)
            player_list = Player.select().where(Player.title_id == title_id())
    elif sender == my_win.Button_reset_fltr_list:
        player_list = Player.select().where(Player.title_id == title_id())
        my_win.comboBox_fltr_region.setCurrentIndex(0)
        my_win.comboBox_fltr_city.setCurrentIndex(0)
        my_win.comboBox_fltr_coach.setCurrentIndex(0) 
        my_win.checkBox_15.setChecked(False)      
        load_comboBox_filter()
    player_list_pred = player.select().where(Player.application == "предварительная")
    count = len(player_list_pred)    
    my_win.label_predzayavka.setText(f"По предзаявке {count} чел.")
    my_win.label_predzayavka.setStyleSheet("color: black")
    fill_table(player_list)


def find_in_player_list():
    """поиск спортсмена или тренера"""
    player = Player.select().where(Player.title_id == title_id())
    txt = my_win.lineEdit_Family_name.text()
    if txt == "":
        my_win.textEdit.clear()
    txt = txt.upper()
    player_list = player.select().where(Player.player ** f'{txt}%')  # like
    if len(player_list) > 0:
        fill_table(player_list)
    else:
        my_win.textEdit.setText("Такого спортсмена нет!")


def find_in_player_rejting_list():
    """поиск спортсмена в рейтинг листе"""
    r_data_m = [R_list_m, R1_list_m]
    r_data_w = [R_list_d, R1_list_d]
    gamer_w = ["Девочки", "Девушки", "Юниорки", "Женщины"]
    id_title = Title.select().where(Title.id == title_id()).get()

    gamer = id_title.gamer
    txt_r = ""
    cur_index = my_win.comboBox_choice_R.currentIndex()
    txt_r = my_win.lineEdit_find_player_in_R.text()
    txt_r = txt_r.capitalize()
    if cur_index == 0: # если выбран текущий рейтинг
        if gamer in gamer_w:
            r_data = r_data_w[0]
        else:
            r_data = r_data_m[0]
        player_list = r_data.select().where(r_data.r_fname ** f'{txt_r}%')   
    elif cur_index == 1: # если рейтинг за январь
        if gamer in gamer_w:
            r_data = r_data_w[1]
        else:
           r_data = r_data_m[1]
        player_list = r_data.select().where(r_data.r1_fname ** f'{txt_r}%')
 
    fill_table(player_list) # заполняет таблицу -tablewidget- списком спортсменов


def filter_rejting_list():
    """Фильтрует вкладку -рейтинг-"""
    sender = my_win.sender()
    r_data_m = [R_list_m, R1_list_m]
    r_data_w = [R_list_d, R1_list_d]

    gamer_w = ["Девочки", "Девушки", "Юниорки","Женщины"]
    id_title = Title.select().where(Title.id == title_id()).get()
    gamer = id_title.gamer
    cur_index = my_win.comboBox_choice_R.currentIndex()
    region_txt = my_win.comboBox_filter_region_in_R.currentText()
    city_txt = my_win.comboBox_filter_city_in_R.currentText()
    date_txt = my_win.comboBox_filter_date_in_R.currentText()
 
    if cur_index == 0:
        r_data = r_data_w[0] if gamer in gamer_w else r_data_m[0] # текущий рейтинг
        rejting_name = r_data.r_fname
        rejting_list = r_data.r_list
        rejting_region = r_data.r_region
        rejting_city = r_data.r_city
        rejting_date = r_data.r_bithday
    else:
        r_data = r_data_w[1] if gamer in gamer_w else r_data_m[1] # январский рейтинг 
        rejting_name = r_data.r1_fname
        rejting_list = r_data.r1_list
        rejting_region = r_data.r1_region
        rejting_city = r_data.r1_city
        rejting_date = r_data.r1_bithday

    if date_txt != "":
        znak = date_txt.find(" ")
        year_fltr = int(date_txt[znak: znak + 3])
        year_current = int(datetime.today().strftime("%Y")) # текущий год
        year_bday = year_current - year_fltr + 1
        after_date = date(year_bday, 1, 1)
        player_list = r_data.select().where(rejting_date > after_date).order_by(rejting_list.desc())

    if region_txt == "" and city_txt == "" and date_txt == "":
        player_list = r_data.select()
    elif region_txt != "" and city_txt != "" and date_txt == "":
        player_list = r_data.select().where((rejting_region == region_txt) & (rejting_city == city_txt))
    elif region_txt != "" and city_txt == "" and date_txt != "":
        player_list = r_data.select().where((rejting_date > after_date) & (rejting_region == region_txt))
    elif region_txt == "" and city_txt != "" and date_txt != "":
        player_list = r_data.select().where((rejting_date > after_date) & (rejting_city == city_txt))
    elif region_txt == "" and city_txt != "" and date_txt == "":
        player_list = r_data.select().where(rejting_city == city_txt)
    elif region_txt != "" and city_txt == "" and date_txt == "":
        player_list = r_data.select().where(rejting_region == region_txt)
    elif region_txt != "" and city_txt != "" and date_txt != "":
        player_list = r_data.select().where((rejting_date > after_date) & (rejting_region == region_txt) & (rejting_city == city_txt))

    if sender == my_win.Button_sort_rejting_in_R: 
        if date_txt != "" and region_txt == "" and city_txt == "":
            player_list = r_data.select().where(rejting_date > after_date).order_by(rejting_list.desc())       
        elif region_txt == "" and city_txt == "" and date_txt == "":
            player_list = r_data.select().order_by(rejting_list.desc()) 
        elif region_txt != "" and city_txt == "" and date_txt == "":
            player_list = r_data.select().where(rejting_region == region_txt).order_by(rejting_list.desc())
        elif region_txt != "" and city_txt != "" and date_txt == "":
            player_list = r_data.select().where((rejting_region == region_txt) & (rejting_city == city_txt)).order_by(rejting_list.desc())  
        elif region_txt != "" and city_txt == "" and date_txt != "":
            player_list = r_data.select().where((rejting_date > after_date) & (rejting_region == region_txt)).order_by(rejting_list.desc())  
        elif region_txt != "" and city_txt != "" and date_txt != "":
            player_list = r_data.select().where((rejting_date > after_date) & (rejting_region == region_txt) & (rejting_city == city_txt)).order_by(rejting_list.desc())    
    elif sender == my_win.Button_sort_alf_R: 
        if region_txt == "" and city_txt == "" and date_txt == "":
            player_list = r_data.select().order_by(rejting_name)
        elif date_txt != "" and region_txt == "" and city_txt == "":
            player_list = r_data.select().where(rejting_date > after_date).order_by(rejting_name)
        elif region_txt != "" and city_txt == "" and date_txt == "":
            player_list = r_data.select().where(rejting_region == region_txt).order_by(rejting_name)
        elif region_txt != "" and city_txt == "" and date_txt != "":
            player_list = r_data.select().where((rejting_date > after_date) & (rejting_region == region_txt)).order_by(rejting_name)
        elif region_txt != "" and city_txt != "" and date_txt == "":
            player_list = r_data.select().where((rejting_region == region_txt) & (rejting_city == city_txt)).order_by(rejting_name)
        elif region_txt != "" and city_txt != "" and date_txt != "":
            player_list = r_data.select().where((rejting_date > after_date) & (rejting_region == region_txt) & (rejting_city == city_txt)).order_by(rejting_name)
    fill_table(player_list) # заполняет таблицу -tablewidget- списком спортсменов


def enter_total_score():
    """ввод счета во встречи без счета в партиях"""
    msgBox = QMessageBox
    sender = my_win.sender()
    tab = my_win.tabWidget.currentIndex()
    mark = 0
    flag = 0
    mistake = 0
    if sender == my_win.lineEdit_pl1_score_total:
        mark = my_win.lineEdit_pl1_score_total.text()
        flag = 0
    elif sender == my_win.lineEdit_pl2_score_total:
        mark = my_win.lineEdit_pl2_score_total.text()
        flag = 1 
    if mark != "":  
        mark = int(mark)
        mistake = check_input_total_score(mark, flag)
        if tab == 3 and flag == 0:
            my_win.lineEdit_pl2_score_total.setFocus() if mistake == 0 else my_win.lineEdit_pl1_score_total.setFocus()
        elif tab == 3 and flag == 1:
            enter_score(none_player=0) if mistake == 0 else my_win.lineEdit_pl2_score_total.setFocus()
    else:
        reply = msgBox.information(my_win, 'Уведомление',
                                                "Проверьте правильность ввода счета!",
                                                msgBox.Ok)
        return
    

def check_input_total_score(mark, flag):
    """проверка ввода счета встречи и его правильность"""
    msgBox = QMessageBox
    score_list = []
    tab = my_win.tabWidget.currentIndex() 
    mark_int = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
    if tab == 3:
        for i in my_win.groupBox_kolvo_vstrech.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
            if i.isChecked():
                match_current = int(i.text())
                break
        s1 = my_win.lineEdit_pl1_score_total.text()
        s2 = my_win.lineEdit_pl2_score_total.text()

    if mark in mark_int:
        if flag == 1:
            score_list.append(int(s1))
            score_list.append(int(s2))
            if match_current // 2 + 1 not in score_list:
                reply = msgBox.information(my_win, 'Уведомление',
                                                "Проверьте правильность ввода счета!\nСчет меньше необходимого.",
                                               msgBox.Ok)
                return
        if match_current // 2 + 1 < mark:
            reply = msgBox.information(my_win, 'Уведомление',
                                                "Проверьте правильность ввода счета!\nЧисло не соответсвует из скольки партий матч.",
                                               msgBox.Ok)
            mistake = 1
        else:
            mistake = 0
            return mistake
    else:
        reply = msgBox.information(my_win, 'Уведомление',
                                                "Вы ввели не правильно символ!",
                                                msgBox.Ok)
        mistake = 1
        return mistake
                 

def focus():
    """переводит фокус на следующую позицию
    sum_total_game список (1-й колво очков которые надо набрать, 2-й сколько уже набрали)"""
    sender = my_win.sender()  # в зависимости от сигала кнопки идет сортировка
    system = System.select().where(System.title_id == title_id())
    tab_etap = my_win.tabWidget_stage.currentIndex()
    # stage = my_win.comboBox_filter_final.currentText()
    idx = my_win.tableView.currentIndex() # определиние номера строки
    row_num = idx.row()
    mark_list = [my_win.lineEdit_pl1_s1, my_win.lineEdit_pl2_s1, my_win.lineEdit_pl1_s2, my_win.lineEdit_pl2_s2,
            my_win.lineEdit_pl1_s3, my_win.lineEdit_pl2_s3, my_win.lineEdit_pl1_s4, my_win.lineEdit_pl2_s4,
            my_win.lineEdit_pl1_s5, my_win.lineEdit_pl2_s5, my_win.lineEdit_pl1_s6, my_win.lineEdit_pl2_s6,
            my_win.lineEdit_pl1_s7, my_win.lineEdit_pl2_s7]
    if tab_etap == 0:
        stage = "Предварительный"       
    elif tab_etap == 1:
        # stage = my_win.tableView.model().index(row_num, 2).data() # данные ячейки tableView
        stage = my_win.comboBox_filter_semifinal.currentText()
    else:
        stage = my_win.tableView.model().index(row_num, 2).data() # данные ячейки tableView
    id_system = system_id(stage)
    sys = system.select().where(System.id == id_system).get()
    sf = sys.score_flag  # флаг из скольки партий играется матч
    mark_index = mark_list.index(sender)
    mark = mark_list[mark_index].text()
    flag_mistake = control_mark_in_score(mark, sf)
    if flag_mistake is True:
        return
    if mark_index % 2 == 1:
        if mark_index >= sf:
            sum_total_game = score_in_game()  # подсчет очков в партии
            if len(sum_total_game) == 0: # значит была ошибка в счете и поэтому он вернул пустой список
                return
            if sum_total_game[0] != sum_total_game[1]:
                mark_list[mark_index + 1].setFocus()
            else:
                my_win.Button_Ok.setFocus()
                return
        mark_list[mark_index + 1].setFocus()    
    else:
        mark_list[mark_index + 1].setFocus()
 

def control_mark_in_score(mark, sf):
    """проверка ввода счета в ячейку """
    msgBox = QMessageBox
    tab = my_win.tabWidget.currentIndex()
    if tab == 3:
        score_list = [my_win.lineEdit_pl1_score_total.text(), my_win.lineEdit_pl2_score_total.text()] # список общий счет в партии

    flag = True if str((sf + 1) // 2) in [score_list[0], score_list[1]] else False

    if flag is False:
        mark_number = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "0"]
        if mark == "":
            msgBox.critical(my_win, "", "Ошибка при вводе счета!\nвведите счет")
            flag_mistake = True
            return flag_mistake
        else:
            count = len(mark)
            if count > 2:
                msgBox.critical(my_win, "", "Ошибка при вводе счета!\nпроверьте правильность ввода")
                flag_mistake = True
                return flag_mistake
            else:
                for k in range(0, count):
                    mark_zn = mark[k]
                    if mark_zn not in mark_number:
                        msgBox.critical(my_win, "", "Ошибка при вводе счета!\nпроверьте правильность ввода")
                        flag_mistake = True
                        return flag_mistake
                flag_mistake = False
                return flag_mistake 


def score_in_game():
    """считает общий счет в партиях"""
    msgBox = QMessageBox

    system = System.select().where(System.title_id == title_id())
    t = my_win.tabWidget_stage.currentIndex()
    if t == 0:
        stage = "Предварительный"
    elif t == 1:
        stage = my_win.comboBox_filter_semifinal.currentText()
    elif t == 2:
        if my_win.comboBox_filter_final.currentIndex() == 0:
            row_num = my_win.tableView.currentIndex().row() # определиние номера строки
            stage = my_win.tableView.model().index(row_num, 2).data() # данные ячейки tableView
        else:
            stage = my_win.comboBox_filter_final.currentText()
    total_score = []
    ts1 = []
    ts2 = []
    total_game = []
    sum_total_game = []
    # row_num = my_win.tableView.currentIndex().row() # определиние номера строки

    tab = my_win.tabWidget.currentIndex()
    s11 = s21 = s12 = s22 = s13 = s23 = s14 = s24 = s15 = s25 = s16 = s26 = s17 = s27 = 0
    # поля ввода счета в партии
    if tab == 3:
        sys = system.select().where(System.stage == stage).get()
        sf = sys.score_flag  # флаг из скольки партий играется матч
        # ==========
        s11 = my_win.lineEdit_pl1_s1.text()
        s21 = my_win.lineEdit_pl2_s1.text()
        s12 = my_win.lineEdit_pl1_s2.text()
        s22 = my_win.lineEdit_pl2_s2.text()
        s13 = my_win.lineEdit_pl1_s3.text()
        s23 = my_win.lineEdit_pl2_s3.text()
        s14 = my_win.lineEdit_pl1_s4.text()
        s24 = my_win.lineEdit_pl2_s4.text()
        s15 = my_win.lineEdit_pl1_s5.text()
        s25 = my_win.lineEdit_pl2_s5.text()
        s16 = my_win.lineEdit_pl1_s6.text()
        s26 = my_win.lineEdit_pl2_s6.text()
        s17 = my_win.lineEdit_pl1_s7.text()
        s27 = my_win.lineEdit_pl2_s7.text()
# определяет из скольки партий играется встреча
    if sf == 3: 
        total_score = [s11, s21, s12, s22, s13, s23]
        max_game = 2
    elif sf == 5:
        total_score = [s11, s21, s12, s22, s13, s23, s14, s24, s15, s25]
        max_game = 3
    elif sf == 7:
        total_score = [s11, s21, s12, s22, s13, s23, s14, s24, s15, s25, s16, s26, s17, s27]
        max_game = 4
    point = 0

    n = len(total_score)
# ==========================================
    for i in range(0, n, 2):
        if total_score[i] != "":
            sc1 = int(total_score[i])
            sc2 = int(total_score[i + 1])
            
            flag = control_score(sc1, sc2)

            if flag is True:
                if sc1 > sc2:
                    point = 1
                    ts1.append(point)
                else:
                    point = 1
                    ts2.append(point)
                st1 = sum(ts1)
                st2 = sum(ts2)
                # ==============
                if tab == 3:
                    my_win.lineEdit_pl1_score_total.setText(str(st1))
                    my_win.lineEdit_pl2_score_total.setText(str(st2))
                    if st1 == max_game or st2 == max_game:  # сравнивает максимальное число очков и набранные очки одним из игроков
                        # если игрок набрал макс очки активиоует кнопку ОК и переводит на нее фокус
                        my_win.Button_Ok.setEnabled(True)
                        my_win.Button_Ok.setFocus()
                total_game.append(st1)
                total_game.append(st2)
                # находит максимальное число очков из сыгранных партий
                max_score = max(total_game)
                if i == 0:
                    # добавляет в список макс число очков которые надо набрать
                    sum_total_game.append(max_game)
                    # добавляет в список макс число очков которые уже набрал игрок
                    sum_total_game.append(max_score)
                else:
                    sum_total_game[0] = max_game
                    sum_total_game[1] = max_score
            elif flag is False:
                sum_total_game = []
                return sum_total_game
                # желательно сюда ввести чтобы фокус ставился на туже ячейку
    return sum_total_game


def control_score(sc1, sc2):
    """проверка на правильность ввода счета"""
    msgBox = QMessageBox
    if sc1 == 11:
        flag = True if sc2 in [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 13] else False
    elif sc1 > 11:
        flag = True if sc2 == sc1 - 2 or sc2 == sc1 + 2 else False
    elif 0 <= sc1 <= 9:
        flag = True if sc2 == 11 else False
    elif sc1 == 10:
        flag = True if sc2 == 12 else False

    if flag == False:
        result = msgBox.information(my_win, "", "Проверьте правильность ввода\n счета в партии!",
                                    msgBox.Ok)
        flag = False
    return flag


def control_winner_player(winner, loser):
    """ Проверка условия победителя матча, различие рейтинга"""
    msgBox = QMessageBox
    if loser == "X":
        return
    flag = True
    player = Player.select().where(Player.title_id == title_id())
    result = Result.select().where(Result.title_id == title_id())
    if winner.rfind("/") == -1: # если спортсмены имеют только имя и фамилию без города (старый вариант)
        f_name_win_id = player.select().where(Player.player == winner).get()
        f_name_los_id = player.select().where(Player.player == loser).get()
        winner = f_name_win_id.full_name
        loser = f_name_los_id.full_name
    
    player_win_id = player.select().where(Player.full_name == winner).get()
    player_los_id = player.select().where(Player.full_name == loser).get()
    r_win = player_win_id.rank
    coefficient_win = player_win_id.coefficient_victories
    total_game_win = player_win_id.total_game_player
    r_los = player_los_id.rank
    coefficient_los = player_los_id.coefficient_victories
    total_game_los = player_los_id.total_game_player
    if total_game_win < 5 or total_game_los < 5: # если еще сыграно мало игр и определяем по разности рейтинга
        if r_win - r_los > 0:
            flag = True
        elif abs(r_los - r_win) < 15:
            flag = True 
        else:
            flag = False      
    else:
        if coefficient_win - coefficient_los > 0:
            flag = True
        elif abs(coefficient_los - coefficient_win) <= 0.3:
            flag = True
        else:
            if (r_win - r_los) < 0 and abs(r_win - r_los) > 30:
                flag = False
    if flag is False:            
        result = msgBox.information(my_win, "", f"Вы уверенны в победе\n {winner} рейтинг{r_win}!\n над {loser} рейтинг {r_los}",
                                    msgBox.Yes, msgBox.No)
        if  result == msgBox.No:
            flag = False
            return flag
    sum_win = player_win_id.total_win_game + 1
    all_game_win = total_game_win + 1
    sum_los = player_los_id.total_win_game
    all_game_los = total_game_los + 1  
    koef_win = sum_win / all_game_win
    koef_los = sum_los / all_game_los
    koef_win = float('{:.3f}'.format(koef_win)) # соотношение выйгранных партий ко всем играм
    koef_los = float('{:.3f}'.format(koef_los)) # соотношение выйгранных партий ко всем играм
 
    with db:
        player_win_id.total_game_player = all_game_win # все игр сыгранных игроком
        player_win_id.total_win_game = sum_win # сколько побед он одержал
        player_win_id.coefficient_victories = koef_win
        player_win_id.save()
        player_los_id.total_game_player = all_game_los
        player_los_id.total_win_game = sum_los
        player_los_id.coefficient_victories = koef_los
        player_los_id.save()


def check_real_player():
    """Изменяет спортсменов по предварительной заявке на реальных"""
    my_win.tabWidget.setCurrentIndex(1)

    indices = my_win.tableView.selectionModel().selectedRows()
    for index in sorted(indices):
        rows = index.row()
        id_pl = my_win.tableView.model().index(rows, 0).data() # данные ячейки tableView
        app = Player.update(application="основная").where(Player.id == id_pl)
        app.execute()
    player_list = Player.select().where((Player.title_id == title_id()) & (Player.application == "предварительная"))
    fill_table(player_list)


def enter_score(none_player=0):
    """заносит в таблицу -результаты- победителя, счет и т.п. sc_total [партии выигранные, проигранные, очки победителя
     очки проигравшего]"""
    sender = my_win.sender()
    tab_etap = my_win.tabWidget_stage.currentIndex()
    row_num = my_win.tableView.currentIndex().row()
    id = my_win.tableView.model().index(row_num, 0).data() # данные ячейки tableView
    fin = my_win.tableView.model().index(row_num, 2).data() # данные ячейки tableView
    num_game = my_win.tableView.model().index(row_num, 3).data() # данные ячейки tableView
    if sender != my_win.Button_3_mesta:
        if row_num == -1: # значит не выбрана ни одна строка и нажат ентер
            return
    else:
        fin = "1-й финал"
    sys = System.select().where(System.title_id == title_id()) 
    if tab_etap == 0: # группы
        stage = "Предварительный"
    elif tab_etap == 1: # полуфиналы
        if row_num == -1: # не выбрана строка и идет ПФ по умочанию
            stage = "1-й полуфинал"
        else:
            id_res = my_win.tableView.model().index(row_num, 0).data() #  данные ячейки (из какого финала играют встречу)
            result = Result.select().where(Result.id == id_res).get()
            stage = result.system_stage
    else:  # финальный
        if fin == "1 группа":
            stage = "Одна таблица"
        else:
            stage = fin
    id_system = system_id(stage)
    system = sys.select().where(System.id == id_system).get()
    type = system.type_table
    flag = 0
    if stage in ["Предварительный", "1-й полуфинал", "2-й полуфинал"]:
        sc_total = circle_type(none_player, stage)
    elif stage == "Одна таблица":
        if type == "сетка":
            sc_total = setka_type(none_player)
            flag = 1
        else:
            sc_total = circle_type(none_player, stage)
    else:  # финалы
        if type == "сетка":
            sc_total = setka_type(none_player) # список 1-й счет победителя, 2-й счет проигравшего, 3-й очки победителя и очки проигравшего
            flag = 1
        else:  # по кругу
            sc_total = circle_type(none_player, stage)
    st1 = sc_total[0]  # партия выигранные
    st2 = sc_total[1]  # партии проигранные
    w = sc_total[2]  # очки победителя
    l = sc_total[3]  # очки проигравшего
# ===================
    if my_win.lineEdit_player1.text() != "X" and my_win.lineEdit_player2.text() != "X":
        pl1 = my_win.lineEdit_player1.text()
        pl2 = my_win.lineEdit_player2.text()
        if none_player == 0: # встреча состоялась
            winner = pl1 if st1 > st2 else pl2
            loser = pl2 if st1 > st2 else pl1
            ts_winner = f"{st1} : {st2}" if st1 > st2 else f"{st2} : {st1}"
            ts_loser = f"{st2} : {st1}" if st1 > st2 else f"{st1} : {st2}"
            winner_string = string_score_game()  # пишет счет в партии
        elif none_player == 1: # не явился 1-й игрок
            winner = pl2
            loser = pl1
            ts_winner = f"{st2} : {st1}"
            ts_loser = f"{st1} : {st2}"
        else:
            winner = pl1
            loser = pl2
            ts_winner = f"{st1} : {st2}"
            ts_loser = f"{st2} : {st1}"
        if none_player != 0: # если победа по неявке
            if type == "сетка":
                winner_string = ""
            elif type == "круг" or type == "группы":
                winner_string = "В : П"
    else: # если нет одного игрока -X-
        if my_win.lineEdit_player1.text() == "X":
            winner = my_win.lineEdit_player2.text()
            loser = my_win.lineEdit_player1.text()
        else:
            winner = my_win.lineEdit_player1.text()
            loser = my_win.lineEdit_player2.text()
            # loser_fam_name = loser # оставляет -X-
        winner_string = ""
        ts_winner = ""
        ts_loser = ""
    loser_fam_name = loser # 
    # if none_player == 0: # если победа по неявке, то не проверяет победу
    #     flag = control_winner_player(winner, loser) # проверка реальности победы игрока (маленький рейтинг над большим)
    #     if flag is False:
    #         return
    res = Result.get(Result.id == id)
    flag_edit_match = res.points_win # если он не None, значит будет редактирование строки
    with db:  # записывает в таблицу -Result- сыгранный матч в группах или финал по кругу
        result = Result.get(Result.id == id)
        result.winner = winner
        result.points_win = w
        result.score_win = winner_string
        result.score_in_game = ts_winner
        result.loser = loser
        result.points_loser = l
        result.score_loser = ts_loser
        result.save()
    #  == попытка удалить встречи с игроками задействованных в редактировани счета по сетке
    if flag_edit_match is not None and type == "сетка":
        player_match = []
        player_match = [winner, loser] # список игроков редактируемом матче
        id_system = system_id(stage)
        res_edit = Result.select().where(Result.system_id == id_system)
        tour_edit = int(result.tours) # номера тура в сетке, котороый редактируется
        for k in res_edit:
            tour = int(k.tours)
            if tour > tour_edit:
                pl1 = k.player1
                pl2 = k.player2
                if pl1 in player_match:
                    app = Result.update(player1="", winner="", points_win=0, score_in_game="", score_win="", loser="", points_loser="", score_loser="").where(Result.id == k)
                    app.execute()
                elif pl2 in player_match:
                    app = Result.update(player2="", winner="", points_win=0, score_in_game="", score_win="", loser="", points_loser="", score_loser="").where(Result.id == k)
                    app.execute()
    # ==============================================
    if tab_etap == 2:  # записывает в -Result- сыгранный матч со сносками на соответствующие строки победителя и проигравшего
        if type == "сетка":
            vid_setki = system.label_string  # вид сетки и кол-во участников
            # список 1-й номер победителя 2-й проигравшего
            snoska = numer_game(num_game, vid_setki) # snoska список [номер встречи победителя, номер встречи приогравшего, номер в сетке куда сносится проигравший]
            res = Result.select().where(Result.title_id == title_id())
            sys_id = sys.select().where(System.stage == fin).get()
            s_id = sys_id.id
            results = res.select().where(Result.system_id == s_id)
            if snoska[0] != 0:
                with db:  # записывает в db таблицу Result победителя и проигравшего
                    player = winner
                    match_num = result.tours  # номер встречи, в строке
                    game = snoska[2] * -1 # номер встречи число
                    for k in range(0, 2):
                        if int(match_num) == game:
                            res_id = results.select().where(Result.tours == snoska[k]).get() # id встречи, куда попадает победитель и проигравший
                            # =========                            
                            if res_id.player1 == "":
                                res_id.player1 = player
                            else:
                                res_id.player2 = player
                            res_id.save()
                            player = loser_fam_name
        elif type == "круг":
            pass
    fill_table_results()
    line_edit_list = [my_win.lineEdit_pl1_s1, my_win.lineEdit_pl2_s1, my_win.lineEdit_pl1_s2, my_win.lineEdit_pl2_s2,
                        my_win.lineEdit_pl1_s3, my_win.lineEdit_pl2_s3, my_win.lineEdit_pl1_s4, my_win.lineEdit_pl2_s4,
                        my_win.lineEdit_pl1_s5, my_win.lineEdit_pl2_s5, my_win.lineEdit_pl1_s6, my_win.lineEdit_pl2_s6,
                        my_win.lineEdit_pl1_s7, my_win.lineEdit_pl2_s7, my_win.lineEdit_player1,  my_win.lineEdit_player2,
                        my_win.lineEdit_pl1_score_total, my_win.lineEdit_pl2_score_total]
    my_win.checkBox_7.setChecked(False)
    my_win.checkBox_8.setChecked(False)
    if tab_etap == 0:
        filter_gr()
    elif tab_etap == 1:
        filter_sf()
    elif tab_etap == 2:
        filter_fin()

    for line in line_edit_list:
            line.clear()


def made_pdf_table_for_view(sender):
    """вызов функции заполнения таблицы pdf группы сыгранными играми"""
    # group_list = ["Предварительный", "1-й полуфинал", "2-й полуфинал"]
    tab_etap = my_win.tabWidget_stage.currentIndex()
    t_id = Title.get(Title.id == title_id())
    short_name = t_id.short_name_comp
    
    if tab_etap == 0:
        view_file = f"{short_name}_table_group.pdf" 
    elif tab_etap == 1:
        stage = my_win.comboBox_filter_semifinal.currentText()
        n_fin = stage[:1]
        view_file = f"{short_name}_{n_fin}-semifinal.pdf"
    elif tab_etap == 2:
        stage = my_win.tableView.model().index(0, 2).data() # данные ячейки tableView номер финала для просмотра в пдф пот нажатию кнопки
        if stage == "Одна таблица":
            view_file = f"{short_name}_one_table.pdf"
        elif stage == "Суперфинал":
            view_file = f"{short_name}_superfinal.pdf"
        else:
            n_fin = stage[:1]
            view_file = f"{short_name}_{n_fin}-final.pdf"
        fin = stage

    if sender == my_win.view_gr_Action or tab_etap == 0:  # вкладка группы
        stage = "Предварительный"
    elif sender == my_win.view_fin1_Action:
        stage = "1-й финал"
        fin = stage
    elif sender == my_win.view_fin2_Action:
        stage = "2-й финал"
        fin = stage
    elif sender == my_win.view_fin3_Action:
        stage = "3-й финал"
        fin = stage
    elif sender == my_win.view_fin4_Action:
        stage = "4-й финал"
        fin = stage
    elif sender == my_win.view_fin5_Action:
        stage = "5-й финал"
        fin = stage
    elif sender == my_win.view_fin6_Action:
        stage = "6-й финал"
        fin = stage
    elif sender == my_win.view_fin7_Action:
        stage = "7-й финал"
        fin = stage
    elif sender == my_win.view_fin8_Action:
        stage = "8-й финал"
        fin = stage
    elif sender == my_win.view_fin9_Action:
        stage = "9-й финал"
        fin = stage
    elif sender == my_win.view_fin10_Action:
        stage = "10-й финал"
        fin = stage
    elif sender == my_win.view_superfin_Action:
        stage = "Суперфинал"
        fin = stage
    elif sender == my_win.view_one_table_Action:
        stage = "Одна таблица"
    elif sender == my_win.view_pf1_Action:
        stage = "1-й полуфинал"
    elif sender == my_win.view_pf2_Action:
        stage = "2-й полуфинал"
    my_win.tabWidget.setCurrentIndex(3)
    # ==== новый вариант с использованием system id
    id_system = system_id(stage)
    systems = System.select().where(System.id == id_system).get()
    pv = systems.page_vid
    type_table = systems.type_table
    system_table = systems.label_string
    # ========
    if type_table == "круг" or type_table == "группы":
        table_made(pv, stage)
    else:
        if system_table == "Сетка (с розыгрышем всех мест) на 8 участников":
            setka_8_full_made(fin)
        elif system_table == "Сетка (-2) на 8 участников":
            setka_8_2_made(fin)
        elif system_table == "Сетка (с розыгрышем всех мест) на 16 участников":
            setka_16_full_made(fin)
        elif system_table == "Сетка (-2) на 16 участников":
            setka_16_2_made(fin)
        elif system_table == "Сетка (с розыгрышем всех мест) на 32 участников":
            setka_32_full_made(fin)
        elif system_table == "Сетка (-2) на 32 участников":
            setka_32_2_made(fin)
        elif system_table == "Сетка (с розыгрышем всех мест) на 32 участников":
            setka_32_made(fin)  
    return view_file


def setka_type(none_player):
    """сетка ставит очки в зависимости от неявки игрока, встреча состоялась ли пропуск встречи -X-"""
    sc_total = []
    if my_win.lineEdit_player1.text() == "X" or my_win.lineEdit_player2.text() == "X":
        w = ""
        l = ""
        st1 = ""
        st2 = ""
    else:
        if none_player == 0:
            st1 = int(my_win.lineEdit_pl1_score_total.text())
            st2 = int(my_win.lineEdit_pl2_score_total.text())
            w = 2
            l = 1
        else:
            if none_player == 1: # не явился 1-й игрок
                st1 = "П"
                st2 = "В"
            elif none_player == 2:
                st1 = "В"
                st2 = "П"
            w = 2
            l = 0
            my_win.lineEdit_pl1_score_total.setText(st1)
            my_win.lineEdit_pl2_score_total.setText(st2)
    sc_total.append(st1)
    sc_total.append(st2)
    sc_total.append(w)
    sc_total.append(l)
    return sc_total


def circle_type(none_player, stage):
    """круговая таблица"""
    sc_total = []
    st1 = ""
    st2 = ""
    w = ""
    l = ""
    # if stage == "Предварительный":
    if none_player == 0:
        st1 = int(my_win.lineEdit_pl1_score_total.text())
        st2 = int(my_win.lineEdit_pl2_score_total.text())
        w = 2
        l = 1
    else:
        if none_player == 1:  # не явился 1-й игрок
            st1 = "П"
            st2 = "В"
        elif none_player == 2:  # не явился 2-й игрок
            st1 = "В"
            st2 = "П"
        w = 2
        l = 0
        my_win.lineEdit_pl1_score_total.setText(st1)
        my_win.lineEdit_pl2_score_total.setText(st2)
    # elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
    #     if none_player == 0:
    #         st1 = int(my_win.lineEdit_pl1_score_total_pf.text())
    #         st2 = int(my_win.lineEdit_pl2_score_total_pf.text())
    #         w = 2
    #         l = 1
    #     else:
    #         if none_player == 1:  # не явился 1-й игрок
    #             st1 = "П"
    #             st2 = "В"
    #         elif none_player == 2:  # не явился 2-й игрок
    #             st1 = "В"
    #             st2 = "П"
    #         w = 2
    #         l = 0
    #         my_win.lineEdit_pl1_score_total_pf.setText(st1)
    #         my_win.lineEdit_pl2_score_total_pf.setText(st2)    
    # else:
    #     if none_player == 0:
    #         st1 = int(my_win.lineEdit_pl1_score_total_fin.text())
    #         st2 = int(my_win.lineEdit_pl2_score_total_fin.text())
    #         w = 2
    #         l = 1
    #     else:
    #         if none_player == 1:  # не явился 1-й игрок
    #             st1 = "П"
    #             st2 = "В"
    #         elif none_player == 2:  # не явился 2-й игрок
    #             st1 = "В"
    #             st2 = "П"
    #         w = 2
    #         l = 0
    #         my_win.lineEdit_pl1_score_total_fin.setText(st1)
    #         my_win.lineEdit_pl2_score_total_fin.setText(st2)
    sc_total.append(st1)
    sc_total.append(st2)
    sc_total.append(w)
    sc_total.append(l)
    return sc_total


def string_score_game():
    """создает строку со счетом победителя"""
    tab = my_win.tabWidget.currentIndex()
    visible_flag = True
    if tab == 3:
        visible_flag = my_win.checkBox_4.isChecked()
        for i in my_win.groupBox_kolvo_vstrech.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
            if i.isChecked():
                g = (int(i.text()) + 1) // 2 # число, максимальное кол-во партий для победы
                break
    # поля ввода счета в партии
    st1 = int(my_win.lineEdit_pl1_score_total.text())
    st2 = int(my_win.lineEdit_pl2_score_total.text())
    s11 = my_win.lineEdit_pl1_s1.text()
    s21 = my_win.lineEdit_pl2_s1.text()
    s12 = my_win.lineEdit_pl1_s2.text()
    s22 = my_win.lineEdit_pl2_s2.text()
    s13 = my_win.lineEdit_pl1_s3.text()
    s23 = my_win.lineEdit_pl2_s3.text()
    s14 = my_win.lineEdit_pl1_s4.text()
    s24 = my_win.lineEdit_pl2_s4.text()
    s15 = my_win.lineEdit_pl1_s5.text()
    s25 = my_win.lineEdit_pl2_s5.text()
    s16 = my_win.lineEdit_pl1_s6.text()
    s26 = my_win.lineEdit_pl2_s6.text()
    s17 = my_win.lineEdit_pl1_s7.text()
    s27 = my_win.lineEdit_pl2_s7.text()

    # создание строки счета победителя
    if st1 > st2:
        if visible_flag is True:
            if int(s11) > int(s21):  # 1-й сет
                n1 = s21
            else:
                n1 = str(f"-{s11}")

            if int(s12) > int(s22):  # 2-й сет
                n2 = s22
            else:
                n2 = str(f"-{s12}")
            if (g == 2 and st1 == 2 and st2 == 0):  # из 3-х партий 2-0
                winner_string = f"({n1},{n2})"
                return winner_string
            
            if int(s13) > int(s23):  # 3-й сет
                n3 = s23
            else:
                n3 = str(f"-{s13}")
            if (g == 2 and st1 == 2 and st2 == 1) or (g == 3 and st1 == 3 and st2 == 0):  # из 3-х  2-1 или из 5-и 3-0
                winner_string = f"({n1},{n2},{n3})"
                return winner_string
            
            if int(s14) > int(s24):  # 4-й сет
                n4 = s24
            else:
                n4 = str(f"-{s14}")
            if (g == 4 and st1 == 4 and st2 == 0) or (g == 3 and st1 == 3 and st2 == 1):  # из 5-и 3-1 или из 7-и 4-0
                winner_string = f"({n1},{n2},{n3},{n4})"
                return winner_string

            if int(s15) > int(s25):  # 5-й сет
                n5 = s25
            else:
                n5 = str(f"-{s15}")
            if (g == 4 and st1 == 4 and st2 == 1) or (g == 3 and st1 == 3 and st2 == 2):  # из 5-и 3-2 или из 7-и 4-1
                winner_string = f"({n1},{n2},{n3},{n4},{n5})"

            if s16 == "" and s26 == "":
                return winner_string
             
            if int(s16) > int(s26):  # 6-й сет
                n6 = s26
            else:
                n6 = str(f"-{s16}")
            if (g == 4 and st1 == 4 and st2 == 2):  # из 7-и 4-2
                winner_string = f"({n1},{n2},{n3},{n4},{n5},{n6})"

            if int(s17) > int(s27):  # 7-й сет
                n7= s27
            else:
                n7 = str(f"-{s17}")
            if (g == 4 and st1 == 4 and st2 == 3):  # из 7-и 4-3
                winner_string = f"({n1},{n2},{n3},{n4},{n5},{n6},{n7})"
        else:
            # if visible_flag is True:
            winner_string = f"{st1} : {st2}" 
            # else:
            #     winner_string = f"{st1} : {st2}"      
        return winner_string
    else:
        if visible_flag is True:
            if int(s11) < int(s21):  # 1-й сет
                n1 = s11
            else:
                n1 = str(f"-{s21}")
            if int(s12) < int(s22):  # 2-й сет
                n2 = s12
            else:
                n2 = str(f"-{s22}")
            if (g == 2 and st1 == 0 and st2 == 2):  # из 3-х партий 2-0
                winner_string = f"({n1},{n2})"
                return winner_string
            
            if int(s13) < int(s23):  # 3-й сет
                n3 = s13
            else:
                n3 = str(f"-{s23}")
            if (g == 2 and st1 == 1 and st2 == 2) or (g == 3 and st1 == 0 and st2 == 3):  # из 3-х  2-1 или из 5-и 3-0
                winner_string = f"({n1},{n2},{n3})"
                return winner_string
            
            if int(s14) < int(s24):  # 4-й сет
                n4 = s14
            else:
                n4 = str(f"-{s24}")
            if (g == 4 and st1 == 0 and st2 == 4) or (g == 3 and st1 == 1 and st2 == 3):  # из 5-и 3-1 или из 7-и 4-0
                winner_string = f"({n1},{n2},{n3},{n4})"
                return winner_string
            
            if int(s15) < int(s25):  # 5-й сет
                n5 = s15
            else:
                n5 = str(f"-{s25}")
            if  (g == 4 and st1 == 1 and st2 == 4) or (g == 3 and st1 == 2 and st2 == 3):  # из 5-и 3-2 или из 7-и 4-1
                winner_string = f"({n1},{n2},{n3},{n4},{n5})"

            if s16 == "" and s26 == "":
                return winner_string

            if int(s16) < int(s26):  # 6-й сет
                n6 = s16
            else:
                n6 = str(f"-{s26}")
            if (g == 4 and st1 == 2 and st2 == 4):  # из 7-и 4-2
                winner_string = f"({n1},{n2},{n3},{n4},{n5},{n6})"

            if int(s17) < int(s27):  # 7-й сет
                n7 = s17
            else:
                n7 = str(f"-{s27}")
            if (g == 4 and st1 == 3 and st2 == 4):  # из 7-и 4-3
                winner_string = f"({n1},{n2},{n3},{n4},{n5},{n6},{n7})"
            
        else:
            # if visible_flag is True:
            winner_string = f"{st2} : {st1}"

        return winner_string


def result_filter_name():
    """отсортировывает встречи с участием игрока"""
    cp = my_win.comboBox_find_name.currentText()
    cp = cp.title()  # Переводит первую букву в заглавную
    c = Result.select().where(Result.title_id == title_id())
    c = c.where(Result.player1 ** f'{cp}%')  # like
    result_list = c.dicts().execute()
    row_count = len(result_list)  # кол-во строк в таблице
    column_count = len(result_list[0])  # кол-во столбцов в таблице
    # вставляет в таблицу необходимое кол-во строк
    my_win.tableWidget.setRowCount(row_count)

    for row in range(row_count):  # добавляет данные из базы в TableWidget
        for column in range(column_count):
            item = str(list(result_list[row].values())[column])
            my_win.tableWidget.setItem(
                row, column, QTableWidgetItem(str(item)))


def filter_fin(pl=False):
    """фильтрует таблицу -Result- на вкладке финалы"""
    sender = my_win.sender()
    num_game_fin = my_win.lineEdit_num_game_fin.text()
    final = my_win.comboBox_filter_final.currentText()

    name = my_win.comboBox_find_name_fin.currentText()
    round = my_win.lineEdit_tour.text()
    played = my_win.comboBox_filter_played_fin.currentText()
    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    filter = Result.select().where(Result.title_id == title_id())
    count = 0                             
    if sender == my_win.lineEdit_num_game_fin:
        fltr = filter.select().where((Result.system_stage == "Финальный") & (Result.tours == num_game_fin))
        count = len(fltr)
        my_win.label_16.setText(f'Всего в {final} {count} игры')
    if final == "":
        return
    elif final != "все финалы":
        id_system = system_id(stage=final)
        fltr = filter.select().where(Result.system_id == id_system)
    fin = []
    if final == "Одна таблица":
        if my_win.comboBox_find_name_fin.currentText() != "":
            if pl == False:
                fltr = filter.select().where(Result.player1 == name)
            else:
                fltr = filter.select().where(Result.player2 == name)
            c = len(fltr)
        else:
            if final == "Одна таблица" and played == "все игры" and num_game_fin == "" and round == "":
                fltr = filter.select().where(Result.system_stage == "Одна таблица")
                count = len(fltr)
                my_win.label_16.setText(f'Всего {count} игры')
            elif final == "Одна таблица" and played == "завершенные":
                fl = filter.select().where(Result.system_stage == "Одна таблица")
                fltr = fl.select().where(Result.winner != "")
                count = len(fltr)
                my_win.label_16.setText(f'Сыграно {count} игры')
            elif final == "Одна таблица" and played == "не сыгранные":
                fltr = filter.select().where(Result.system_stage ==
                                             "Одна таблица" and Result.points_win == None)
                count = len(fltr)
                my_win.label_16.setText(f'Не сыграно {count} игры')
            elif final == "Одна таблица" and played == "все игры" and num_game_fin == "" and round != "":
                fl = filter.select().where(Result.system_stage == "Одна таблица")
                fltr = fl.select().where(Result.round == round)
                count = len(fltr)
                my_win.label_16.setText(f'Всего {count} игры')
            elif final == "Одна таблица" and played == "все игры" and num_game_fin != "" and round == "":
                fl = filter.select().where(Result.system_stage == "Одна таблица")
                fltr = fl.select().where(Result.tours == num_game_fin)
    else:
        if final == "все финалы" and played == "все игры" and num_game_fin == "" and round == "":
            fltr = filter.select().where(Result.system_stage == "Финальный")
            if name == "":
                count = len(fltr)
                my_win.label_16.setText(f'Всего в финалах {count} игры')
            else:  # выбор по фамилии спортсмена
                row = 0
                fl = filter.select().where(Result.system_stage == "Финальный")
                fltr = fl.select().where((Result.player1 == name)| (Result.player2 == name)) # объединение запросов (отбор по 2-ум столбцам)
                count = len(fltr)
        # один из финалов встречи которые не сыгранные
        elif final != "все финалы" and played == "не сыгранные" and num_game_fin == "" and round == "":
            fl = filter.select().where(Result.system_id == id_system)
            fltr = fl.select().where(Result.points_win == None)
            count = len(fltr)
            my_win.label_16.setText(
                f'Всего в {final}\nне сыгранно {count} игры')
        elif final != "все финалы" and played == "завершенные" and num_game_fin == "" and round == "":
            fltr = filter.select().where((Result.system_id == id_system) & (Result.points_win == 2))
            count_pl = len(fltr)
            my_win.label_16.setText(f'Завершено в {final} {count_pl} игры')
        elif final != "все финалы" and played == "все игры" and num_game_fin == "" and round == "":
            fltr = filter.select().where(Result.system_id == id_system)
            count = len(fltr)
            my_win.label_16.setText(f'Всего в {final} {count} игры')
        elif final != "все финалы" and played == "все игры" and num_game_fin != "":
            fltr = filter.select().where((Result.system_id == id_system) & (Result.tours == num_game_fin))
            count = len(fltr)
            my_win.label_16.setText(f'Всего в {final} {count} игры')
        elif final == "все финалы" and played == "все игры" and num_game_fin != "":
            fltr = filter.select().where((Result.system_stage == "Финальный") & (Result.tours == num_game_fin))
            count = len(fltr)
            my_win.label_16.setText(f'Всего в {final} {count} игры')
        elif final == "все финалы" and played == "все игры" and num_game_fin == "" and round != "":
            fltr = filter.select().where((Result.system_stage == "Финальный") & (Result.round == int(round)))
            count = len(fltr)
            my_win.label_16.setText(f'Всего в {final} {count} игры')
        elif final != "все финалы" and played == "все игры" and num_game_fin != "":
            fltr = filter.select().where((Result.system_id == id_system) & (Result.tours == num_game_fin))
        elif final != "все финалы" and played == "все игры" and num_game_fin == "" and round != "":
            fltr = filter.select().where((Result.system_id == id_system) & (Result.round == int(round)))
            count = len(fltr)
            my_win.label_16.setText(f'Всего в {final} {count} игры')  
        elif final != "все финалы" and played == "завершенные" and num_game_fin == "" and round != "":
            fltr_fin = filter.select().where(Result.system_id == id_system)
            fltr = fltr_fin.select().where((Result.round == int(round)) & (Result.points_win == 2))
            count = len(fltr)
            my_win.label_16.setText(f'Всего в {final} {count} игры')   
        elif final == "все финалы" and played == "завершенные" and num_game_fin == "":
            fltr = filter.select().where((Result.system_stage == "Финальный") & (Result.points_win == 2))
            count_pl = len(fltr)
            my_win.label_16.setText(f' Всего сыграно во всех финалах {count_pl} игры')
        elif final == "все финалы" and played == "не сыгранные" and num_game_fin == "":
            # fltr = filter.select().where((Result.system_stage == "Финальный") & (Result.points_win == None))
            fltr = filter.select().where((Result.system_stage == "Финальный") & (Result.winner == None))
            count = len(fltr)
            my_win.label_16.setText(f'Всего в {final} не сыгранно {count} игры')
        else:
            if final != "все финалы" and num_game_fin != "" and round != "":
                fltr = filter.select().where((Result.system_id == id_system) & (Result.round == round))
            else:
                for sys in system:  # отбирает финалы с сеткой
                    if sys.stage != "Предварительный" and sys.stage != "Полуфиналы":
                        txt = sys.label_string
                        txt = txt[:5]
                        if txt == "Сетка":
                            fin.append(sys.stage)
                fin, ok = QInputDialog.getItem(
                    my_win, "Финалы", "Выберите финал, где искать номер встречи.", fin, 0, False)
                fltr = filter.select().where(Result.number_group == fin)
            row = 0
            for result_list in fltr:
                row += 1
                if result_list.tours == num_game_fin:
                    num_game_fin = int(num_game_fin)
                    row_num = num_game_fin - 1
                    my_win.tableView.selectRow(row_num)
                    break

    player_list = fltr
    fill_table(player_list)
    if count == 0: # если в финал по сетке ввели номер тура
        my_win.lineEdit_tour.clear()
        my_win.statusbar.showMessage("Финалы по сетке", 5000)
        return 
    my_win.label_16.show()


def filter_sf():
    """фильтрует таблицу -результаты- на вкладке полуфиналы"""
    data = []
    data_table_tmp = []
    find_player = []
    model = MyTableModel(data)
    sf = ['1-й полуфинал', '2-й полуфинал']
    semifinal = my_win.comboBox_filter_semifinal.currentText()
    id_system = system_id(stage=semifinal)
    group = my_win.comboBox_filter_group_sf.currentText()
    name = my_win.comboBox_find_name_sf.currentText()
    played = my_win.comboBox_filter_played_sf.currentText()
    find_player.append(name)
    fltr_id = Result.select().where(Result.title_id == title_id())
    if group == "все группы" and my_win.comboBox_find_name_sf.currentText() != "":
        if semifinal == "-все полуфиналы-":
            pl1_query = fltr_id.select().where(Result.system_stage.in_(sf) & (Result.player1 == name))
            pl2_query = fltr_id.select().where(Result.system_stage.in_(sf) & (Result.player2 == name))            
        else:
            pl1_query = fltr_id.select().where((Result.system_stage == semifinal) & (Result.player1 == name))
            pl2_query = fltr_id.select().where((Result.system_stage == semifinal) & (Result.player2 == name)) 
        fltr = pl1_query | pl2_query # объдиняет два запроса в один
    elif group == "все группы" and played == "все игры":
        # filter_sf = fltr_id.select().where((Result.system_id == id_system) & (Result.title_id == title_id()))
        # filter_sf = fltr_id.select().where((Result.system_stage == semifinal) & (Result.title_id == title_id()))
        # fltr = filter_sf.select().where(Result.system_stage.in_(sf))
        fltr = Result.select().where(Result.system_id == id_system)
    elif group == "все группы" and played == "завершенные":
        if semifinal == "-все полуфиналы-":
            fltr = fltr_id.select().where(Result.system_stage.in_(sf) & (Result.points_win == 2))
        else:
            fltr = fltr_id.select().where((Result.system_stage == semifinal) & (Result.points_win == 2))
    elif group != "все группы" and played == "завершенные":
        if semifinal == "-все полуфиналы-":
            fl = fltr_id.select().where((Result.system_stage.in_(sf)) & (Result.number_group == group))
        else:
            fl = fltr_id.select().where((Result.system_stage == semifinal) & (Result.number_group == group))
        fltr = fl.select().where(Result.points_win == 2)
    elif group != "все группы" and played == "не сыгранные":
        if semifinal == "-все полуфиналы-":
            fl = fltr_id.select().where(Result.system_stage.in_(sf)  & (Result.number_group == group))
        else:
            fl = fltr_id.select().where((Result.system_stage == semifinal) & (Result.number_group == group))
        fltr = fl.select().where(Result.points_win == None)
    elif group == "все группы" and played == "не сыгранные":
        filter_sf = fltr_id.select().where(Result.system_id == id_system)
        fltr = filter_sf.select().where(Result.points_win != 2 and Result.points_win == None)
    elif group != "все группы" and played == "все игры":
        if semifinal == "-все полуфиналы-":
            fltr = fltr_id.select().where(Result.system_stage.in_(sf) & (Result.number_group == group))
        else:
            fltr = fltr_id.select().where((Result.system_stage == semifinal) & (Result.number_group == group))
    count = len(fltr)
    result_list = fltr.dicts().execute()
    row_count = len(result_list)  # кол-во строк в таблице
    if row_count != 0:
        column_count = len(result_list[0])  # кол-во столбцов в таблице
    if played == "завершенные":
        my_win.label_16.setText(f"сыграно {row_count} встреч")
    elif played == "не сыгранные":
        my_win.label_16.setText(f"не сыграно {row_count} встреч(а)")
    else:
        my_win.label_16.setText(f"всего {row_count} встреч(а)")
    my_win.label_16.show()
 
    if row_count != 0:
        for row in range(row_count):  # добавляет данные из базы в TableWidget
            for column in range(column_count):
                item = str(list(result_list[row].values())[column])
                data_table_tmp.append(item)
            data.append(data_table_tmp.copy())
            data_table_tmp.clear()
        my_win.tableView.setModel(model)


def filter_gr():
    """фильтрует таблицу -результаты- на вкладке группы"""
    find_player = []

    group = my_win.comboBox_filter_group.currentText()
    name = my_win.comboBox_find_name.currentText()
    played = my_win.comboBox_filter_played.currentText()
    find_player.append(name)
 
    if group == "":
        return
    fltr_id = Result.select().where((Result.title_id == title_id()) & (Result.system_stage == "Предварительный"))
    if group != "все группы":
        player_list = fltr_id.select().where(Result.number_group == group)

    if group == "все группы" and my_win.comboBox_find_name.currentText() != "":
        pl1_query = fltr_id.select().where(Result.player1 == name)
        pl2_query = fltr_id.select().where(Result.player2 == name)
        player_list = pl1_query | pl2_query # объдиняет два запроса в один
    elif group == "все группы" and played == "все игры":
        player_list = fltr_id.select()
    elif group == "все группы" and played == "завершенные":
        player_list = fltr_id.select().where(Result.points_win == 2)
    elif group != "все группы" and played == "завершенные":
        fl = fltr_id.select().where(Result.number_group == group)
        player_list = fl.select().where(Result.points_win == 2)
    elif group != "все группы" and played == "не сыгранные":
        fl = fltr_id.select().where(Result.number_group == group)
        player_list = fl.select().where(Result.points_win != 2 & Result.points_win == None)
    elif group == "все группы" and played == "не сыгранные":
        player_list = fltr_id.select().where((Result.points_win != 2 & Result.points_win == None))

    row_count = len(player_list)  # кол-во строк в таблице

    if played == "завершенные":
        my_win.label_16.setText(f"сыграно {row_count} встреч")
    elif played == "не сыгранные":
        my_win.label_16.setText(f"не сыграно еще {row_count} встреч(а)")
    else:
        my_win.label_16.setText(f"всего {row_count} встреч(а)")
    my_win.label_16.show()

    fill_table(player_list)

 
def load_comboBox_referee():
    """Загружает комбобокс списком судей"""
    msgBox = QMessageBox()
    my_win.comboBox_referee.clear()
    my_win.comboBox_secretary.clear()
    referee_list = []
    referee = Referee.select()
    if len(referee) == 0:
       result = msgBox.information(my_win, "Уведомление", "База данных судей еще пуста!", msgBox.Ok)
       if result == msgBox.Ok:
            return
    else:
        for k in referee:
            family = k.family
            city = k.city
            fam_city = f"{family}/ {city}"
            if fam_city not in referee_list:
                referee_list.append(fam_city)
        my_win.comboBox_referee.addItems(referee_list)
        my_win.comboBox_secretary.addItems(referee_list)      
    return referee_list


def load_combo():
    """загружает комбобокс поиска спортсмена на вкладке группы, пф и финалы фамилиями спортсменов"""
    text = []
    my_win.comboBox_find_name.clear()
    my_win.comboBox_find_name_sf.clear()
    my_win.comboBox_find_name_fin.clear()
    players = Player.select().where(Player.title_id == title_id())
    for i in players:  # цикл по таблице базы данных (I это id строк)
        family = i.player
        city = i.city
        text.append(f"{family}/{city}")
    my_win.comboBox_find_name.addItems(text)
    my_win.comboBox_find_name_sf.addItems(text)
    my_win.comboBox_find_name_fin.addItems(text)
    my_win.comboBox_find_name.setCurrentText("")
    my_win.comboBox_find_name_sf.setCurrentText("")
    my_win.comboBox_find_name_fin.setCurrentText("")


def load_combo_etap_begunki():
    """загружает комбобокс выбора этапов системы на вкладке дополнительно"""
    my_win.comboBox_select_stage_begunki.clear()
    my_win.comboBox_edit_etap1.clear()
    my_win.comboBox_edit_etap2.clear()
    stage_system = ["-Выбор спортсменов-", "Списки участников"]
    results = Result.select().where(Result.title_id == title_id())
    for i in results:
        stage = i.system_stage
        if stage not in stage_system:
            stage_system.append(stage)
    my_win.comboBox_select_stage_begunki.addItems(stage_system)
    my_win.comboBox_edit_etap1.addItems(stage_system)
    my_win.comboBox_edit_etap2.addItems(stage_system)


def reset_filter():
    """сбрасывает критерии фильтрации"""
    sender = my_win.sender()
    if sender == my_win.Button_reset_filter_gr:
        my_win.comboBox_find_name.setCurrentText("")
        my_win.comboBox_filter_played.setCurrentText("все игры")
        my_win.comboBox_filter_group.setCurrentText("все группы")
        filter_gr()
    elif sender == my_win.Button_reset_filter_sf:
        my_win.comboBox_find_name_sf.setCurrentText("")
        my_win.comboBox_filter_played_sf.setCurrentText("все игры")
        my_win.comboBox_filter_group_sf.setCurrentText("все группы")
        filter_sf()
    elif sender == my_win.Button_reset_filter_fin:
        my_win.comboBox_find_name_fin.setCurrentText("")
        my_win.comboBox_filter_played_fin.setCurrentText("все игры")
        my_win.lineEdit_tour.setText("")
        my_win.lineEdit_num_game_fin.setText("")
        if my_win.comboBox_filter_final.currentText() == "Одна таблица":
            my_win.comboBox_filter_final.setCurrentText("Одна таблица")
        else:
            my_win.comboBox_filter_final.setCurrentText("все финалы")
        filter_fin()
    load_combo()


def choice_semifinal_automat(stage):
    """жеребьевка полуфиналов"""
    mesto_first = 0
    system = System.select().where(System.title_id == title_id())
    systems = system.select().where(System.stage == "Предварительный").get()
    total_group = systems.total_group
    # ===== new
    id_system = system_id(stage)
    system_stage = system.select().where(System.id == id_system).get()
    mesta_exit = system_stage.mesta_exit
    # определение мест в ПФ, выходящих из группы
    if stage == "1-й полуфинал": 
        mesto_first = 1
    else:
        system_stage = system.select().where(System.stage == "1-й полуфинал").get()
        mesta_exit = system_stage.mesta_exit
        mesto_first = mesta_exit + 1

    for k in range(1, total_group + 1):
        choices = Choice.select().where((Choice.title_id == title_id()) & (Choice.group == f"{k} группа"))
        p = 0 if k <= total_group // 2 else mesta_exit
        n = k if k <= total_group // 2 else total_group - k + 1
        for i in range(mesto_first, mesta_exit + mesto_first):
            p += 1
            choice_mesta = choices.select().where(Choice.mesto_group == i)
            count = len(choice_mesta)
            if count == 0:
                break
            with db:
                choice_mesta = choices.select().where(Choice.mesto_group == i).get() # записывает в db номер полуфинала
                choice_mesta.semi_final = stage
                choice_mesta.sf_group = f"{n} группа" # номера группы полуфинала
                choice_mesta.posev_sf = p # номер посева
                choice_mesta.save()
    with db:  # записывает в систему, что произведена жеребъевка
        system = System.get(System.id == id_system)
        system.choice_flag = True
        system.save()
    player_in_table_group_and_write_Game_list_Result(stage)


def choice_gr_automat():
    "новая система жеребьевки групп"
    " current_region_group - словарь (регион - список номеров групп куда можно сеять)"
    " reg_player - словарь регион ид игрока, player_current - список сеящихся игроков, posev - словарь всего посева"
    import ast
    msgBox = QMessageBox()
    posev_tmp = {}
    reg_player = {}
    gr_region = {}
    posev_group = {}
    player_current = []
    pgt = []
    posev = {}
    group_list = []
    start = 0
    end = 1
    step = 0
    vid = ["Автоматическая", "Полуавтоматическая", "Ручная"]
    vid, ok = QInputDialog.getItem(my_win, "Жеребьевка", "Выберите режим жеребьевки групп.", vid, 0, False)
    my_win.tabWidget.setCurrentIndex(3)
    my_win.tabWidget_2.setCurrentIndex(3)
    my_win.tableView_choice_group.setGeometry(QtCore.QRect(0, 0, 1000, 550)) # (точка слева, точка сверху, ширина, высота)
    my_win.tableView_choice_group.show()
    txt_tmp = []
    # получаем список списков чистый для начала жеребьевки
    id_fam_region_list = []
    id_fam_region_list_tmp = []
    id_family_region_list = []
    stage = "Предварительный"
    sys = System.select().where(System.title_id == title_id())
    sys_id = sys.select().where(System.stage == stage).get()
    group = sys_id.total_group
    total_player = sys_id.total_athletes
    max_player = sys_id.max_player  # максимальное число игроков в группе, оно же число посевов
    for b in range(1,max_player + 1):  # цикл создания словарей (номер посева, списки списков(номер группы и 0 вместо номера регионов))
        for x in range(1, group + 1):
            posev_group[x] = 0
        gr_region = posev_group.copy()
        posev[f"{b}_посев"] = gr_region
        posev_group.clear()
    
    pl_choice = Choice.select().where((Choice.title_id == title_id()) & (Choice.family != "X")).order_by(Choice.rank.desc())
    m = 1  # начальное число посева
    p = 0
    number_poseva = 0  # общий счетчик посева игроков
    reg_list = []
    player_list = [] # список всех игроков в порядке посева
    if vid == "Автоматическая":
        for np in pl_choice:
            choice = np.get(Choice.id == np)
            regio_n = choice.region
            region = regio_n.rstrip()
            pl_id = choice.player_choice_id
            reg = Region.select().where(Region.region == region).get()
            region_id = reg.id 
            reg_list.append(region_id)
            player_list.append(pl_id)
        while number_poseva < total_player:
            p += 1
            if number_poseva == 0 or number_poseva % group == 0 :
                group_list = list(range(1, group + 1))  # получение списка групп с помощью функции range
            #  +++ вариант с упорядовычинем списка групп в реверси
                if m % 2 == 0:
                    group_list.sort(reverse = True)
                else:
                    group_list.sort(reverse = False)
            # +++++++++++++
            region_id = reg_list[number_poseva]
            pl_id = player_list[number_poseva]
            posev_tmp = posev[f"{m}_посев"]

            if m == 1:  # 1-й посев       
                posev_tmp[p] = region_id  # создает словарь группа - номер региона
                number_poseva += 1
                player_current.append(pl_id)
                reg_player[pl_id] = number_poseva  # словарь ид игрока его группа при посеве
                if number_poseva == group:  # если доходит окончания данного посева идет запись в db
                    choice_save(m, player_current, reg_player)
            else:  # 2-й посев и т.д.
                current_region_group = {}  # словарь регион - список номеров групп куда можно сеять
                key_reg_previous = []
                current = region_player_current(number_poseva, reg_list, group, player_list)  # должен быть получен список текущих регионов посева
                key_reg_current = current[0]  # номера регионов текущего посева
                player_current = current[1]  # номера игроков (id)

                for o in previous_region_group.keys():  # цикл получения списка регионов предыдущих посевов уникальный
                    key_reg_previous.append(o)
                pgt.clear()
                remains = total_player - number_poseva  # остаток посева
                finish = 0
                if remains > group: 
                    finish = group  # если остаток больше кол-во групп
                else:
                    finish = remains            
                for y in range(0, finish):
                    group_list_tmp = []  
                    z = key_reg_current[y] # список регионов которые уже были посеяны
                    pgt.append(y + 1)  # номера групп которые уже посеяны будут удалены из списка

                    if z not in key_reg_previous:  # если нет в списке, то добавляет полный список групп
                        current_region_group[z] = group_list
                    else:
                        gr_del = previous_region_group[z]  # список групп где уже есть этот регион
                        group_list_tmp = list((Counter(group_list) - Counter(gr_del)).elements()) # удаляет из списка номера групп где уже есть регионы
                        if m % 2 == 0:
                            group_list_tmp.sort(reverse = True)
                        else:
                            group_list_tmp.sort(reverse = False)
                        r = len(group_list_tmp)
                        if r == 0:  # если во всех группах уже есть, то начинает опять полный список групп
                            current_region_group[z] = group_list  # получает словарь со списком групп куда сеять
                        else:
                            current_region_group[z] = group_list_tmp  # получает словарь со списком групп куда сеять
                    # система распределения по группам (посев), где m - номер посева начина со 2-ого посева
                sv = add_delete_region_group(key_reg_current, current_region_group, posev_tmp, m, posev, start, end, step, player_current)
                current.clear()
                number_poseva = number_poseva + sv
            if number_poseva != total_player:  # выход из системы жеребьевки при достижении оканчания
                if number_poseva == group * m:  # смена направления посева
                    if m % 2 != 0:
                        start = group
                        end = 0
                        step = -1
                    else:
                        start = 0
                        end = group
                        step = 1
                    m += 1
                    previous_region_group = posev_test(posev, group, m)  # возвращает словарь регион  - список номера групп, где он есть
            else:
                fill_table_after_choice()
                System.update(choice_flag=1).where(System.id == sys_id).execute() # записывает, что жеребьевка сделана
                player_in_table_group_and_write_Game_list_Result(stage)
            group_list.clear()
    elif vid == "Полуавтоматическая":
        pass
        # gr_region_dict = {}
        # # gr_region_list = []
        # gr_region_temp = []
        # psv = 0
        # for n_posev in range(0, (max_player) * 2):
        #     psv = n_posev // 2 + 1
        #     for player_in_group in range(0, group + 1): # внутренний посев
        #         if player_in_group == 0:
        #            id_fam_region_list_tmp.append(psv) 
        #         else:
        #             id_fam_region_list_tmp.append("-")
                
        #     id_fam_region_list.append(id_fam_region_list_tmp.copy()) # список списков в который помещаются игроки и регионы согласно жеребьевки
        #     id_fam_region_list_tmp.clear()   
        # # ==================================
        # for np in pl_choice:
        #     choice = np.get(Choice.id == np)
        #     regio_n = choice.region
        #     region = regio_n.rstrip()
        #     family_player = np.family
        #     # coach_player = np.coach
        #     pl_id = choice.player_choice_id 
        #     # full_player_str = f"{pl_id}/{family_player}/{region}/{coach_player}" 
        #     full_player_str = f"{pl_id}/{family_player}/{region}" # полные данные спортсмены          
        #     choice_list = [full_player_str]                                                         
        #     player_list.append(choice_list)
        # k = 1
        # posev_list = []
        # for posev in range(0, group * max_player):
        #     if posev < total_player:
        #         one_player = player_list[posev]
        #         txt_tmp.append(one_player)
        #         if posev == group * k - 1:
        #             posev_tmp = txt_tmp.copy()
        #             posev_list.append(posev_tmp)
        #             txt_tmp.clear()
        #             k += 1 
        #     else:
        #         posev_tmp = txt_tmp.copy()
        #         posev_list.append(posev_tmp)
        #         break
        # all_player = 0      
        # for number_posev in range(0, max_player): # полный посев
        # # ============== вариант ручной жеребьевки ========   
        #     if number_posev % 2 == 0: # меняет направления групп в зависимости от посева
        #         nums = [i for i in range(1, group + 1)] # генератор списка
        #     else:
        #         nums = [i for i in range(group, 0, -1)] # генератор списка    
        #     txt_tmp.clear()
        #     id_family_region_list.clear()
        #     a = 0
        #     count = len(posev_list[number_posev])
        #     count_gr = group if number_posev < max_player - 1 else count
        #     while a < count_gr: # создает список отдельного посева
        #         ps = posev_list[number_posev] # список игроков одного посева
        #         txt_temp = ps[a] # один игрок в посеве
        #         # отделяет регион
        #         txt_region = txt_temp[0]
        #         mark = txt_region.rfind("/")
        #         region_pl = txt_region[mark + 1:] # регион игрока
        #         gr_region_temp.append(region_pl)
        #         gr_region_dict[player_in_group + 1] = gr_region_temp.copy()
        #         gr_region_temp.clear()
        #         # ===============================
        #         txt_id_str = f"{txt_temp[0]}" # ролучение id_фамилию и регион в строковой форме
        #         id_family_region_list.append(txt_id_str)
        #         text_str = (',\n'.join(id_family_region_list)) # список игроков посева для формы выбора номера группы
        #         a += 1
        # # ===============================================
            
        #     if number_posev == 0: # 1-й посев сразу записывает в таблицу, а остальные группы заполняет пробелами
        #         number_group = 0
        #         for player_in_group in range(0, group): # внутренний посев 
        #             id_fam_region_str = id_family_region_list[player_in_group]
        #             mark = id_fam_region_str.rfind("/")
        #             id_family = id_fam_region_str[:mark] # id и фамилия игрока
        #             region_pl = id_fam_region_str[mark + 1:] # регион игрока                    
        #             id_fam_region_list[number_posev * 2][player_in_group + 1] =  id_family
        #             id_fam_region_list[number_posev * 2 + 1][player_in_group + 1] =  region_pl
        #             all_player += 1 # число игроков, посеянных
        #             # создание списка регионов по группам
        #             gr_region_temp.append(region_pl)
        #             gr_region_dict[player_in_group + 1] = gr_region_temp.copy()
        #             gr_region_temp.clear()
        #         view_table_group_choice(id_fam_region_list, max_player, group) # функция реального просмотра жеребьевки 
        #     else: # 2-й посев и следующие
        #         # if number_posev % 2 == 0: # меняет направления групп в зависимости от посева
        #         #     nums = [i for i in range(1, group + 1)] # генератор списка
        #         # else:
        #         #     nums = [i for i in range(group, 0, -1)] # генератор списка 
        #         for player_in_group in range(0, group):  # внутренний посев
        #             if all_player == total_player: # если все спортсмены прожеребились  
        #                 msgBox.information(my_win, "Уведомление", "Все спортсмены, распределены по группам.")
        #                 choice_save_manual_group(id_fam_region_list, group)
        #                 System.update(choice_flag=1).where(System.id == sys_id).execute() # Отмечает, что ручная жеребьевка выполнена
        #                 fill_table_after_choice()
        #                 player_in_table_group_and_write_Game_list_Result(stage)
        #                 break
        #             else:
        #                 tx = f"Список спортсменов в порядке посева:\n\n{text_str}\n\n" + "Выберите номер группы и нажмите -ОК-"
        #                 txt = (','.join(list(map(str, nums)))) # номера групп
        #                 number_group, ok = QInputDialog.getText(my_win, f'Номера групп: {txt}', tx)
        #                 number_group = int(number_group)     
    elif vid == "Ручная":
        # my_win.tabWidget.setCurrentIndex(3)
        # my_win.tabWidget_2.setCurrentIndex(3)
        # my_win.tableView_choice_group.setGeometry(QtCore.QRect(0, 0, 1000, 550)) # (точка слева, точка сверху, ширина, высота)
        # my_win.tableView_choice_group.show()
        # txt_tmp = []
        # # получаем список списков чистый для начала жеребьевки
        # id_fam_region_list = []
        # id_fam_region_list_tmp = []
        # id_family_region_list = []
        psv = 0
        for n_posev in range(0, (max_player) * 2):
            psv = n_posev // 2 + 1
            for player_in_group in range(0, group + 1): # внутренний посев
                if player_in_group == 0:
                   id_fam_region_list_tmp.append(psv) 
                else:
                    id_fam_region_list_tmp.append("-")
                
            id_fam_region_list.append(id_fam_region_list_tmp.copy()) # список списков в который помещаются игроки и регионы согласно жеребьевки
            id_fam_region_list_tmp.clear()   
        # ==================================
        for np in pl_choice:
            choice = np.get(Choice.id == np)
            regio_n = choice.region
            region = regio_n.rstrip()
            family_player = np.family
            # coach_player = np.coach
            pl_id = choice.player_choice_id 
            # full_player_str = f"{pl_id}/{family_player}/{region}/{coach_player}" 
            full_player_str = f"{pl_id}/{family_player}/{region}" # полные данные спортсмены          
            choice_list = [full_player_str]                                                         
            player_list.append(choice_list)
        k = 1
        posev_list = []
        for posev in range(0, group * max_player):
            if posev < total_player:
                one_player = player_list[posev]
                txt_tmp.append(one_player)
                if posev == group * k - 1:
                    posev_tmp = txt_tmp.copy()
                    posev_list.append(posev_tmp)
                    txt_tmp.clear()
                    k += 1 
            else:
                posev_tmp = txt_tmp.copy()
                posev_list.append(posev_tmp)
                break
        all_player = 0      
        for number_posev in range(0, max_player): # полный посев
        # ============== вариант ручной жеребьевки ======== 
            family_region_list = []       
            txt_tmp.clear()
            id_family_region_list.clear()
            a = 0
            count = len(posev_list[number_posev])
            count_gr = group if number_posev < max_player - 1 else count
            while a < count_gr: # создает список отдельного посева
                ps = posev_list[number_posev] # список игроков одного посева
                txt_temp = ps[a] # один игрок в посеве
                txt_id_str = f"{txt_temp[0]}" # ролучение id_фамилию и регион в строковой форме
                znak = txt_id_str.find("/")
                family_region_list.append(txt_id_str[znak + 1:])
                id_family_region_list.append(txt_id_str)
                text_str = (',\n'.join(family_region_list)) # список игроков посева для формы выбора номера группы
                # text_str = (',\n'.join(id_family_region_list)) # список игроков посева для формы выбора номера группы
                a += 1
        # ===============================================
            if number_posev == 0: # 1-й посев сразу записывает в таблицу, а остальные группы заполняет пробелами
                number_group = 0
                for player_in_group in range(0, group): # внутренний посев 
                    id_fam_region_str = id_family_region_list[player_in_group]
                    mark = id_fam_region_str.rfind("/")
                    id_family = id_fam_region_str[:mark] # id и фамилия игрока
                    region_pl = id_fam_region_str[mark + 1:] # регион игрока
                    id_fam_region_list[number_posev * 2][player_in_group + 1] =  id_family
                    id_fam_region_list[number_posev * 2 + 1][player_in_group + 1] =  region_pl
                    all_player += 1 # число игроков, посеянных
                view_table_group_choice(id_fam_region_list, max_player, group) # функция реального просмотра жеребьевки
            else:
                if number_posev % 2 == 0: # меняет направления групп в зависимости от посева
                    nums = [i for i in range(1, group + 1)] # генератор списка
                else:
                    nums = [i for i in range(group, 0, -1)] # генератор списка 
                for player_in_group in range(0, group):  # внутренний посев
                    if all_player == total_player: # если все спортсмены прожеребились  
                        msgBox.information(my_win, "Уведомление", "Все спортсмены, распределены по группам.")
                        choice_save_manual_group(id_fam_region_list, group)
                        System.update(choice_flag=1).where(System.id == sys_id).execute() # Отмечает, что ручная жеребьевка выполнена
                        fill_table_after_choice()
                        player_in_table_group_and_write_Game_list_Result(stage)
                        break
                    else:
                        tx = f"Список спортсменов в порядке посева:\n\n{text_str}\n\n" + "Выберите номер группы и нажмите -ОК-"
                        txt = (','.join(list(map(str, nums))))
                        number_group, ok = QInputDialog.getText(my_win, f'Номера групп: {txt}', tx)
                        number_group = int(number_group)
                        number_correct = False # группа введена не правильно
                        if number_group in nums:
                            number_correct = True # группа введена правильно
                        while not number_correct: # проверка на правильность ввода
                            if int(number_group) not in nums:
                                msgBox.information(my_win, "Уведомление", "Вы не правильно ввели номер, повторите снова.")
                            else:
                                number_correct = True
                                continue
                            number_group, ok = QInputDialog.getText(my_win, f'Номера групп: {txt}', tx)
                        number_group = int(number_group)
                        znak = text_str.find(",")
                        fam_city = text_str[:znak]
                        msgBox.information(my_win, "Жеребьевка участников", f"{fam_city} идет на номер группы: {number_group}") 

                        id_fam_region_str = id_family_region_list[player_in_group]
                        mark = id_fam_region_str.rfind("/")
                        id_family = id_fam_region_str[:mark] # id и фамилия игрока
                        region_pl = id_fam_region_str[mark + 1:] # регион игрока
                        id_fam_region_list[number_posev * 2][number_group] =  id_family
                        id_fam_region_list[number_posev * 2 + 1][number_group] =  region_pl
                        view_table_group_choice(id_fam_region_list, max_player, group) # функция реального просмотра жеребьевки
                        nums.remove(number_group) # удаляет посеянную группу
                        text_str = text_str.replace(f'{fam_city},', '')
                        all_player += 1 
        if all_player == total_player:   
            msgBox.information(my_win, "Уведомление", "Все спортсмены, распределены по группам.")
            choice_save_manual_group(id_fam_region_list, group)
            System.update(choice_flag=1).where(System.id == sys_id).execute() # Отмечает, что ручная жеребьевка выполнена
            fill_table_after_choice()
            player_in_table_group_and_write_Game_list_Result(stage)



def choice_save_manual_group(id_fam_region_list, group):
    """записывает в таблицу -Choice- результаты ручной жеребьевки"""
    posev = 0
    row = 0
    for l in id_fam_region_list:       
        if row % 2 == 0:
            posev += 1
            for m in range(1, group + 1):
                txt_str = l[m]
                if txt_str == "-":
                    continue
                else:
                    mark = txt_str.find("/")
                    id_pl = int(txt_str[:mark]) # id игрока
                    choice = Choice.select().where((Choice.title_id == title_id()) & (Choice.player_choice_id == id_pl)).get()
                    with db:  # запись в таблицу Choice результата жеребъевки                        
                        choice.group = f"{m} группа"
                        choice.posev_group = posev
                        choice.save()
        row += 1


def out_red(text):
    "\033[34m{}".format(text)
    return(text)      


# def progress_bar(step_bar):
#     """прогресс бар""" 
#     msgBox = QMessageBox 
#     my_win.progressBar.setValue(step_bar)
#     if step_bar >= 99:
#        result = msgBox.information(my_win, "Уведомление", "Жеребьевка завершена, проверьте ее результаты!", msgBox.Ok)
#        if result == msgBox.Ok:
#             my_win.progressBar.setValue(0)
#     return step_bar


def check_one_region_in_choice(fin):
    """Проверка на спортсменов одного регионоа в жеребьевке"""
    system = System.select().where(System.title_id == title_id())
    stage_exit = system.stage_exit
    mesta_exit = system.mesta_exit
    choice = Choice.select().where(Choice.stage)


def rank_mesto_out_in_group_or_semifinal_to_final(fin):
    """определение мест, выходящих из группы или полуфинала в финал"""
    stage = fin
    id_system = system_id(stage)
    systems_stage = System.select().where(System.title_id == title_id())
    # == словарь этап число игроков в группе или полуфинале
    player_in_stage = {} # словарь -этап- количество в нем участников
    etap_out_and_player = {}
    num_pl = []
    num_pl_sf1 = []
    num_pl_sf2 = []
    player_out_sf1 = 0
    player_out_sf2 = 0
    for l in systems_stage:
        etap = l.stage
        if etap == "Предварительный":
            num_pl.clear()
            max_pl = l.max_player
            for k in range(1, max_pl +1):
                num_pl.append(k)
            etap_out_and_player[etap] = num_pl
        elif etap == "1-й полуфинал":
            num_pl_sf1.clear()
            max_pl = l.max_player // l.total_group
            for k in range(1, max_pl +1):
                num_pl_sf1.append(k)
            etap_out_and_player[etap] = num_pl_sf1
            player_out_sf1 = l.mesta_exit
            list_mest = etap_out_and_player["Предварительный"]
            del list_mest[:player_out_sf1]
        elif etap == "2-й полуфинал":
            num_pl_sf2.clear()
            max_pl = l.max_player // l.total_group
            for k in range(1, max_pl +1):
                num_pl_sf2.append(k)
            etap_out_and_player[etap] = num_pl_sf2
            player_out_sf2 = l.mesta_exit
            list_mest = etap_out_and_player["Предварительный"]
            del list_mest[:player_out_sf2]
        else:
            system_fin = System.select().where((System.title_id == title_id()) & (System.stage == etap)).get()
            if etap == "Одна таблица":
                end = system_fin.total_athletes
                nums = [i for i in range(1, end + 1)] # генератор списка
            else:
                etap_out_fin = system_fin.stage_exit # из какого этапа выходят в финал
                pl_out = system_fin.mesta_exit # сколько мест
                if etap == "Суперфинал":
                    pl_out_list = [i for i in range(1, pl_out + 1)] # список мест из 1 финала, играющих в суперфинале
                    etap_out_and_player[etap_out_fin] = pl_out_list
                list_mest = etap_out_and_player[etap_out_fin]

                if fin != etap:
                    del list_mest[:pl_out]
                    player_in_stage[etap] = etap_out_and_player[etap_out_fin]
                else:
                    del list_mest[pl_out:]
                    nums = list_mest
                    break
    return nums


def choice_setka_automat(fin, flag, count_exit):
    """автоматическая жеребьевка сетки, fin - финал, count_exit - сколько выходят в финал
    flag - флаг вида жеребьевки ручная или автомат""" 
    msgBox = QMessageBox 
    full_posev = []  # список полного списка участников 1-ого посева
    group_last = []
    number_last = [] # посеянные номера в сетке
    reg_last = []  # посеянные регионы в сетке
    number_posev = []  # список по порядку для посева
    current_region_posev = {} # в текущем посеве список регионов по порядку
    posev_data = {} # окончательные посев номер в сетке - игрок/ город
    num_id_player = {} # словарь номер сетки - id игрока
    possible_number = {}
    flag_stop_manual_choice = 0 # флаг окончания ручной жеребьевки
    #===================================
    id_system = system_id(stage=fin)
    system = System.select().where((System.title_id == title_id()) & (System.id == id_system)).get()
    choice = Choice.select().where(Choice.title_id == title_id())

    max_player = system.max_player
    stage_exit = system.stage_exit
  
    posevs = setka_choice_number(fin, count_exit) # выбор номеров сетки для посева
    player_net = posevs[0] # максимальное число игроков в сетке
 
    if fin == "Суперфинал" or fin == "Одна таблица":
        count_exit = 1

    if count_exit == 1:
        posev_1 = posevs[1]
    elif count_exit == 2:
        posev_1 = posevs[1]
        posev_2 = posevs[2]
    elif count_exit == 3:
        posev_1 = posevs[1]
        posev_2 = posevs[2]
        posev_3 = posevs[3]
    elif count_exit == 4:
        posev_1 = posevs[1]
        posev_2 = posevs[2]
        posev_3 = posevs[3]
        posev_4 = posevs[4]

    free_seats = 0 # кол-во свободных мест в сетке
    step = 0
    del_num = 0
    free_num = []
    real_all_player_in_final = []

    nums = rank_mesto_out_in_group_or_semifinal_to_final(fin) # получение списка номеров мест, выходящих в финал, суперфинал

    n = 0  
    # end_posev = count_exit if flag != 3 else 1
    end_posev = count_exit
    while n < end_posev:  #  ======   НАЧАЛО ПОСЕВА   =========   добавил n=0 и n+=1 стр 7098
        if system.stage == "Одна таблица":
            real_all_player_in_final = len(choice.select().where(Choice.basic == fin))
            choice_posev = choice.select().order_by(Choice.rank)
        elif fin == "1-й финал":
            if stage_exit == "Предварительный":
                # == реальное число игроков в финале
                real_all_player_in_final = len(choice.select().where(Choice.mesto_group.in_(nums)))
                # == число игроков в конкретном посеве финала
                if flag == 3:
                    choice_posev = choice.select().where(Choice.mesto_group.in_(nums))
                else:
                    # if n == 0:
                    #     choice_posev = choice.select().where(Choice.mesto_group == nums[n]) # если 1-й финал и 1-й посев то сортирует по группам 
                    # else:
                    choice_posev = choice.select().where(Choice.mesto_group == nums[n])          
            elif stage_exit == "1-й полуфинал" or stage_exit == "2-й полуфинал": # выходят из полуфинала
                real_all_player_in_final = len(choice.select().where((Choice.semi_final == stage_exit) & (Choice.mesto_semi_final.in_(nums))))
                # == число игроков в конкретном посеве финала
                choice_posev = choice.select().where((Choice.semi_final == stage_exit) & (Choice.mesto_semi_final == nums[n])) 
        elif fin == "Суперфинал":
            choice_posev = choice.select().where((Choice.final == stage_exit) & (Choice.mesto_final.in_(nums)))
            real_all_player_in_final = len(choice.select().where((Choice.final == stage_exit) & (Choice.mesto_final.in_(nums))))
        else: # финалы по сетке начиная со 2-ого и т.д.
            if stage_exit == "Предварительный": # откуда выход в финал
                if flag == 3:
                #    choice_posev = choice.select().where(Choice.mesto_group == nums[n]) 
                    choice_posev = choice.select().where(Choice.mesto_group.in_(nums)) 
                else:
                    if count_exit > 1:
                        choice_posev = choice.select().where(Choice.mesto_group == nums[n])
                    else:
                        choice_posev = choice.select().where(Choice.mesto_group.in_(nums))
                real_all_player_in_final = len(choice.select().where(Choice.mesto_group.in_(nums))) # реальное число игроков в сетке
            elif stage_exit == "1-й полуфинал" or stage_exit == "2-й полуфинал": # выходят из полуфинала
                choice_posev = choice.select().where((Choice.semi_final == stage_exit) & (Choice.mesto_semi_final == nums[n]))
                real_all_player_in_final = len(choice.select().where((Choice.semi_final == stage_exit) & (Choice.mesto_semi_final.in_(nums))))
        count_player_in_final = len(choice_posev) # количество игроков в отдельном посева
        # ищет свободные номера только в последнем посеве        
        # if real_all_player_in_final != max_player:
        max_player = full_net_player(player_in_final=max_player)
        if real_all_player_in_final != max_player and n == end_posev - 1:
            free_num = free_place_in_setka(max_player, real_all_player_in_final)
            del_num = 1 # флаг, что есть свободные номера
        full_posev.clear()
        for posevs in choice_posev: # отбор из базы данных согласно местам в группе для жеребьевки сетки
            psv = []
        
            family = posevs.family
            if fin == "Суперфинал":
                count_exit = 1
                group = ""
                group_number = 1
                mesto_group = posevs.mesto_final
            elif fin != "Одна таблица":
                if stage_exit == "Предварительный":
                    group = posevs.group
                    mesto_group = posevs.mesto_group
                elif stage_exit == "1-й полуфинал" or stage_exit == "2-й полуфинал":
                    group = posevs.sf_group # номер группы ПФ
                    mesto_group = posevs.mesto_semi_final # место в группе (ПФ)
                ind = group.find(' ')
                group_number = int(group[:ind]) # номер группы
            else:
                group = ""
                group_number = 1
                mesto_group = ""
            pl_id = posevs.player_choice_id # id игрока
            region = posevs.region
            region = region.strip()
            player = Player.get(Player.id == pl_id)
            city = player.city
            rank = player.rank

            psv = [pl_id, family, region, group_number, group, city, rank, mesto_group]
            full_posev.append(psv) 

        if fin == "Суперфинал":
            full_posev.sort(key=lambda k: k[7]) # сортировка списка участников по месту в 1-ом финале
        elif count_exit == 1 or fin == "Одна таблица":
            full_posev.sort(key=lambda k: k[6], reverse=True) # сортировка списка участников по рейтингу
        elif fin == "1-й финал":
            if count_exit > 1:
                full_posev.sort(key=lambda k: (k[7], k[3])) # сортировка списка участников сначала по месту в группе а потом по группач
            else:
                full_posev.sort(key=lambda k: k[3]) # сортировка списка участников по группам
        elif count_exit != 1 or fin != "1-й финал":
            full_posev.sort(key=lambda k: k[6], reverse=True) # сортировка списка участников по рейтингу
   

        for k in full_posev: # цикл по списку игроков
            k.pop(3)
            k.pop(6)
        # ======== начало жеребьевки =========
        end = player_net // count_exit if flag == 1 else count_player_in_final
        number_posev = [i for i in range(0, end)] # генератор списка номеров посева

        # присваивает posev список списков, разбитый на отдельные подпосевы
        if n == 0:
            posev = posev_1
        elif n == 1:
            posev = posev_2
        elif n == 2:
            posev = posev_3
        elif n == 3:
            posev = posev_4
          # если ручная жеребьевка то обьединяет все посевы
        if flag == 3:
            if count_exit == 1:
                posev = posev_1
            elif count_exit == 2:
                posev = posev_1 + posev_2
            elif count_exit == 3:
                posev = posev_1 + posev_2 + posev_3
            elif count_exit == 4:
                posev = posev_1 + posev_2 + posev_3 + posev_4

        count_posev = len(posev) # количество подпосевов в посеве

        for i in range(0, count_posev):  # список посева, разделеный на отдельные посевы
            if flag_stop_manual_choice == 1: # выход из цикла если окончена ручная жеребьевка
                break
            current_region_posev.clear()
            sev_tmp = posev[i].copy()
            sev = sev_tmp.copy() # список в отдельном подпосеве
            sev_tmp.clear()
            count = len(posev[i]) # всего количество номеров в посеве
            if del_num == 1 and i == count_posev - 1:                    
                for h in free_num:
                    sev.remove(h)
                free_seats = len(free_num) # сколько свободных мест в сетке
                count = len(posev[i]) - free_seats
            # if del_num == 1 and n == count_exit - 1: 
            # if del_num == 1 and i == end_posev - 1:                    
            #     for h in free_num:
            #         sev.remove(h)
            #     free_seats = len(free_num) # сколько свободных мест в сетке
            #     count = len(posev[i]) - free_seats
            for w in range(0, count): # внутренний цикл посева
                l = number_posev[0] # общий список всего посева (порядковый номер посева)
                if i == 0 and n == 0: #  ===== 1-й посев
                    sev = posev[i]  # список номеров в сетке для посева
                    num_set = sev[w] # номер в сетке на который идет сев
                    count_sev = len(sev) # количество номеров в посеве
                else:
                    num_set = sev[0] # проверить
                    if len(posev[i]) > count_player_in_final and count_exit > 1:
                        count_sev = count_player_in_final
                    else:
                        count_sev = len(sev) # конкретное число оставшихся в посеве минус свободных мест(если они есть)
                    if count_sev > 1: # если сеющихся номеров больше одного
                        if w == 0: # 1-й основной посев
                            gr_region_tmp = []
                            for k in range(l, l + count_sev):
                                region = full_posev[k][2]
                                gr = full_posev[k][3]
                                gr_region_tmp.append(region)
                                gr_region_tmp.append(gr)
                                gr_region = gr_region_tmp.copy()
                                current_region_posev[k] = gr_region # словарь регионы, в текущем посеве по порядку
                                gr_region_tmp.clear()
                        number_last.clear()
                        number_last = list(num_id_player.keys()) # список уже посеянных номеров в сетке
                        # ==== вариант  с сорт по кол ву регионов начиная с 5 по 8 посев
                        if l > 3 and l < 15:
                            current_region_posev = sort_region(current_region_posev)
                        # =====
                        reg_last.clear()
                        group_last.clear()
                        for v in num_id_player.values():
                            reg_last.append(v[1]) # список уже посеянных регионов
                            group_last.append(v[2]) # список номеров групп уже посеянных
                        if n != 0 or (n == 0 and l > 1):
                        # =========== определения кол-во возможны вариантов посева у каждого региона 
                            possible_number = possible_draw_numbers(current_region_posev, reg_last, number_last, group_last, n, sev, num_id_player, player_net, count_exit)                        

                            if i != 0 or n != 0: # отсортирововаем список по увеличению кол-ва возможных вариантов
                                possible_number = {k:v for k,v in sorted(possible_number.items(), key=lambda x:len(x[1]))}
                                num_posev = list(possible_number.keys())   
                            l = list(possible_number.keys())[0]
                            num_set = possible_number[l] # номера куда можно сеять
                            # === выбор ручная или автомат ====
                            if flag == 1: # автоматичекая
                                if len(num_set) == 0:
                                    result = msgBox.information(my_win,"Уведомление", "Автоматическая жеребьевка не получилась.\n"
                                    "Если хотите повторите снова.\nНажмите -RETRY-\n"
                                    "Если хотите изменить значение мультирегиональность\nНажмите -OK-\n"
                                    "Если отменить жеребьевку\nНажмите -Cancel-", msgBox.Retry| msgBox.Ok| msgBox.Cancel)
                                    if result == msgBox.Retry:
                                        flag = selection_of_the_draw_mode() # выбор ручная или автоматическая жеребьевка
                                        choice_setka_automat(fin, flag, count_exit)
                                    elif result == msgBox.No:
                                        Title.update(multiregion=0).where(Title.id == title_id()).execute()
                                    elif result == msgBox.Cancel:
                                        return
                                    sorted_tuple = sorted(num_id_player.items(), key=lambda x: x[0])
                                    dict(sorted_tuple)                                    
                                    player_choice_in_setka(fin)
                                    step = 0
                                elif len(num_set) != 1: # есть выбор из номеров случайно
                                    num_set = random_generator(num_set)
                                elif len(num_set) == 1: # остался только один номер
                                    num_set = num_set[0]
                            elif flag == 2: # полуавтомат
                                my_win.tableView.setGeometry(QtCore.QRect(260, 241, 841, 540))
                                player_list = []
                                player_list_tmp = []

                                for j in possible_number.keys():
                                    posev_list = full_posev[j]
                                    pl = posev_list[1] # фамилия
                                    reg = posev_list[2] # регион
                                    pn = possible_number[j] # возможные номера посева
                                    player_list_tmp = [pl, reg, pn]
                                    player_list.append(player_list_tmp.copy())
                                    player_list_tmp.clear()
                                txt_tmp = []
    
                                for g in player_list:
                                    if len(num_id_player) == 2:
                                        fam_city = ""
                                        number_net = ""
                                        view_table_choice(fam_city, number_net, num_id_player) # функция реального просмотра жеребьевки
                                    t_str = str(g[2])
                                    txt_str = f"{g[0]} - {g[1]} номера: {t_str}" 
                                    txt_tmp.append(txt_str)
                                text_str = (',\n'.join(txt_tmp))
                                tx = f"Список спортсменов в порядке посева:\n\n{text_str}\n\n" + "Выберите один из номеров и нажмите\n - ОК - если выбрали сами или - Cancel - если хотите выбор случайный"
                                txt = (','.join(list(map(str, num_set))))
                                while True:
                                    try:
                                        number_net, ok = QInputDialog.getText(my_win, f'Возможные номера посева: {txt}', tx)
                                        znak = text_str.find(":")
                                        fam_city = text_str[:znak - 7]
                                        if not ok:
                                            number_net = random.choice(num_set)
                                        msgBox.information(my_win, "Жеребьевка участников", f"{fam_city} идет на номер: {number_net}")
                                        number_net = int(number_net)
                                        view_table_choice(fam_city, number_net, num_id_player) # функция реального просмотра жеребьевки
                                    except ValueError:
                                        msgBox.information(my_win, "Уведомление", "Вы не правильно ввели номер, повторите снова.")
                                        continue
                                    else:
                                        if number_net in num_set:
                                            num_set = number_net
                                            break
                                        else:
                                            msgBox.information(my_win, "Уведомление", "Вы не правильно ввели номер, повторите снова.") 
                            elif flag == 3: # ручная жеребьевка
                                q = 1
                                num_list = "-"
                                for r in range(1, max_player + 1 ):
                                    num_id_player[r] = num_list
                                my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 841, 274))
                                my_win.tableView.setGeometry(QtCore.QRect(260, 318, 841, 430))
                                txt_tmp = []
                                player_list = []
                                pl_id_list = []
                                player_list_tmp = []
                                for k in range(len(full_posev)):
                                    posev_list = full_posev[k]
                                    id_pl = posev_list[0] # id игрока
                                    pl = posev_list[1] # фамилия
                                    reg = posev_list[2] # регион
                                    gr = posev_list[3]
                                    r = posev_list[5] # рейтинг
                                    player_list_tmp = [pl, reg, r]
                                    player_list.append(player_list_tmp.copy())
                                    pl_id_list_tmp = [id_pl, reg, gr] 
                                    pl_id_list.append(pl_id_list_tmp.copy())
                                    player_list_tmp.clear()
                                    pl_id_list_tmp.clear()
                                    q += 1
                                # if fin != "Суперфинал":
                                #     pl_id_list.sort(key=lambda x: x[2], reverse=True) # отсортировывает списки списков по 3-му элементу
                                #     player_list.sort(key=lambda x: x[2], reverse=True) # отсортировывает списки списков по 3-му элементу
                                for i in range(0, count_posev):
                                    n_poseva = posev[i] 
                                    count_sev = len(n_poseva)
                                    if i == 0: 
                                        for l in player_list: # цикл создания списка для ручной жеребьевки
                                            pl = l[0]
                                            region = l[1]
                                            pl_reg = f"{pl}/ {region}"
                                            txt_tmp.append(pl_reg)
                                        text_str = (',\n'.join(txt_tmp))
                                    m = 0
                                    for k in list(full_posev):
                                        if m == count_sev:
                                            break
                                        else:
                                            id_player = k[0]
                                            region = k[2]
                                            gr = k[3]  
                                            id_region = [id_player, region, gr]
                                            n_sev = n_poseva
                                            f_text = txt_tmp[0]                                  
                                            tx = f"Сеятся игрок:\n{f_text}\n\nСписок спортсменов в порядке посева:\n\n{text_str}\n\n"\
                                                "Выберите один из номеров и нажмите - ОК -" 
                                            number_net, ok = QInputDialog.getText(my_win, f'Возможные номера посева: {n_sev}', tx, QLineEdit.Normal) 
                                            if number_net == '' or int(number_net) not in [j for j in range(1, player_net + 1)]:
                                                msgBox.information(my_win, "Уведомление", "Вы не правильно ввели номер, повторите снова.")
                                                break
                                            fam_city = ""
                                            old = num_id_player[int(number_net)]

                                            num_id_player[int(number_net)] = id_region
                                            view_table_choice(fam_city, number_net, num_id_player)
                                            txt_tmp.remove(f_text)
                                            text_str = (',\n'.join(txt_tmp))
                                            full_posev.pop(0)

                                            if old != "-":
                                                id_old = old[0]
                                                players = Player.select().where(Player.id == id_old).get()
                                                fio = players.player
                                                city = players.city
                                                old_player = f"{fio}/{city}"
                                                number_net, ok = QInputDialog.getText(my_win,'Замена игрока', f'На какой номер в сетке\nпереместить игрока:\n {old_player}', QLineEdit.Normal) 
                                                num_id_player[int(number_net)] = old
                                                fam_city = ""
                                                view_table_choice(fam_city, number_net, num_id_player)
                                            m += 1
                                        end -= 1 
                                        if end == 0:
                                        # if end == 0 or real_all_player_in_final == (len(num_id_player) - len(free_num)):
                                            flag_stop_manual_choice = 1
                                            step = 100
                if flag_stop_manual_choice == 0:                      
                    id_player = full_posev[l][0]
                    region = full_posev[l][2]
                    gr = full_posev[l][3]  
                    id_region = [id_player, region, gr]
                    num_id_player[num_set] = id_region
                    # ======== модуль удаления посеянных номеров =========
                    if count_sev > 1:
                        c = len(current_region_posev)
                        if c != 0:
                            del possible_number[l] # удаляет из словаря возможных номеров посеянный порядковый номер
                            del current_region_posev[l] # удаляет из словаря текущий посеянный регион
                            if num_set in sev: # проверяет посеянный номер в посеве
                                sev.remove(num_set)  # удаляет посеянный номер из всех номеров этого посева
                            for z in possible_number.keys():
                                possible_tmp = possible_number[z]
                                #=====
                                if flag is False and len(possible_number) == 1:
                                    number_net = sev[0]
                                    fam_city = f"{pl}/{reg}"
                                    view_table_choice(fam_city, number_net, num_id_player) # функция реального просмотра жеребьевки 
                                #======
                                if num_set in possible_tmp: # проверяет посеянный номер в возможных номерах
                                    possible_tmp.remove(num_set) # удаляет посеянный номер из возможных номеров
                    elif count_sev == 1: # удаляет последний номер в посеве
                        sev.clear()
                        possible_number.clear()
                    number_posev.remove(l)
                    if i != 0:
                        num_posev.remove(l)

                    sp = 100 / (real_all_player_in_final)
                    step += sp
                else:
                    break
        if step > 99:
            del_num_list = []  
            for i in num_id_player.keys():
                tmp_list = list(num_id_player[i])
                if len(tmp_list) == 1:
                    del_num_list.append(i)                    
                else:
                    id = tmp_list[0]
                    pl_id = Player.get(Player.id == id)
                    family_city = pl_id.full_name
                    posev_data[i] = family_city
                    with db:
                        choice_final = choice.select().where(Choice.player_choice_id == pl_id).get()
                        if fin == "Суперфинал":
                            choice_final.super_final = i
                        else:
                            choice_final.final = fin
                            choice_final.posev_final = i
                        choice_final.save()
            if len(del_num_list) > 0:
                for e in del_num_list:
                    del num_id_player[e]
            key_set = set(num_id_player.keys()) # получаем сет всех ключей (номеров сетки)
            all_num_set = [j for j in range(1, player_net + 1)]
            free_number = set((all_num_set))
            free_number.difference_update(key_set) # вычитаем из всех номеров те которые посеяны и остается номера -X-
            for h in free_number:
                posev_data[h] = "X"
        n += 1 # добавил в связи со сменой цикла
    return posev_data


def _choice_setka_automat(fin, flag, count_exit):
    """автоматическая жеребьевка сетки, fin - финал, count_exit - сколько выходят в финал
    flag - флаг вида жеребьевки ручная или автомат""" 
    msgBox = QMessageBox 
    full_posev = []  # список полного списка участников 1-ого посева
    group_last = []
    number_last = [] # посеянные номера в сетке
    reg_last = []  # посеянные регионы в сетке
    number_posev = []  # список по порядку для посева
    current_region_posev = {} # в текущем посеве список регионов по порядку
    posev_data = {} # окончательные посев номер в сетке - игрок/ город
    num_id_player = {} # словарь номер сетки - id игрока
    possible_number = {}
    flag_stop_manual_choice = 0 # флаг окончания ручной жеребьевки
    #===================================
    id_system = system_id(stage=fin)
    system = System.select().where((System.title_id == title_id()) & (System.id == id_system)).get()
    choice = Choice.select().where(Choice.title_id == title_id())

    max_player = system.max_player
    stage_exit = system.stage_exit
  
    posevs = setka_choice_number(fin, count_exit) # выбор номеров сетки для посева
    player_net = posevs[0] # максимальное число игроков в сетке
 
    if fin == "Суперфинал" or fin == "Одна таблица":
        count_exit = 1

    if count_exit == 1:
        posev_1 = posevs[1]
    elif count_exit == 2:
        posev_1 = posevs[1]
        posev_2 = posevs[2]
    elif count_exit == 3:
        posev_1 = posevs[1]
        posev_2 = posevs[2]
        posev_3 = posevs[3]
    elif count_exit == 4:
        posev_1 = posevs[1]
        posev_2 = posevs[2]
        posev_3 = posevs[3]
        posev_4 = posevs[4]

    free_seats = 0 # кол-во свободных мест в сетке
    step = 0
    del_num = 0
    free_num = []
    real_all_player_in_final = []

    flag_new = 0

    nums = rank_mesto_out_in_group_or_semifinal_to_final(fin) # получение списка номеров мест, выходящих в финал, суперфинал

    n = 0  
    # end_posev = count_exit if flag != 3 else 1
    end_posev = count_exit
    while n < end_posev:  #  ======   НАЧАЛО ПОСЕВА   =========   добавил n=0 и n+=1 стр 7098
        if system.stage == "Одна таблица":
            real_all_player_in_final = len(choice.select().where(Choice.basic == fin))
            choice_posev = choice.select().order_by(Choice.rank)
        elif fin == "1-й финал":
            if stage_exit == "Предварительный":
                # == реальное число игроков в финале
                real_all_player_in_final = len(choice.select().where(Choice.mesto_group.in_(nums)))
                # == число игроков в конкретном посеве финала
                if flag == 3:
                    choice_posev = choice.select().where(Choice.mesto_group.in_(nums))
                else:
                    # if n == 0:
                    #     choice_posev = choice.select().where(Choice.mesto_group == nums[n]) # если 1-й финал и 1-й посев то сортирует по группам 
                    # else:
                    choice_posev = choice.select().where(Choice.mesto_group == nums[n])          
            elif stage_exit == "1-й полуфинал" or stage_exit == "2-й полуфинал": # выходят из полуфинала
                real_all_player_in_final = len(choice.select().where((Choice.semi_final == stage_exit) & (Choice.mesto_semi_final.in_(nums))))
                # == число игроков в конкретном посеве финала
                choice_posev = choice.select().where((Choice.semi_final == stage_exit) & (Choice.mesto_semi_final == nums[n])) 
        elif fin == "Суперфинал":
            choice_posev = choice.select().where((Choice.final == stage_exit) & (Choice.mesto_final.in_(nums)))
            real_all_player_in_final = len(choice.select().where((Choice.final == stage_exit) & (Choice.mesto_final.in_(nums))))
        else: # финалы по сетке начиная со 2-ого и т.д.
            if stage_exit == "Предварительный": # откуда выход в финал
                if flag == 3:
                #    choice_posev = choice.select().where(Choice.mesto_group == nums[n]) 
                    choice_posev = choice.select().where(Choice.mesto_group.in_(nums)) 
                else:
                    if count_exit > 1:
                        choice_posev = choice.select().where(Choice.mesto_group == nums[n])
                    else:
                        choice_posev = choice.select().where(Choice.mesto_group.in_(nums))
                real_all_player_in_final = len(choice.select().where(Choice.mesto_group.in_(nums))) # реальное число игроков в сетке
            elif stage_exit == "1-й полуфинал" or stage_exit == "2-й полуфинал": # выходят из полуфинала
                choice_posev = choice.select().where((Choice.semi_final == stage_exit) & (Choice.mesto_semi_final == nums[n]))
                real_all_player_in_final = len(choice.select().where((Choice.semi_final == stage_exit) & (Choice.mesto_semi_final.in_(nums))))
        count_player_in_final = len(choice_posev) # количество игроков в отдельном посева
        # ищет свободные номера только в последнем посеве        
        # if real_all_player_in_final != max_player:
        max_player = full_net_player(player_in_final=max_player)
        if real_all_player_in_final != max_player and n == end_posev - 1:
            free_num = free_place_in_setka(max_player, real_all_player_in_final)
            del_num = 1 # флаг, что есть свободные номера
        full_posev.clear()
        for posevs in choice_posev: # отбор из базы данных согласно местам в группе для жеребьевки сетки
            psv = []
        
            family = posevs.family
            if fin == "Суперфинал":
                count_exit = 1
                group = ""
                group_number = 1
                mesto_group = posevs.mesto_final
            elif fin != "Одна таблица":
                if stage_exit == "Предварительный":
                    group = posevs.group
                    mesto_group = posevs.mesto_group
                elif stage_exit == "1-й полуфинал" or stage_exit == "2-й полуфинал":
                    group = posevs.sf_group
                    mesto_group = posevs.mesto_semi_final
                ind = group.find(' ')
                group_number = int(group[:ind]) # номер группы
            else:
                group = ""
                group_number = 1
                mesto_group = ""
            pl_id = posevs.player_choice_id # id игрока
            region = posevs.region
            region = region.strip()
            player = Player.get(Player.id == pl_id)
            city = player.city
            rank = player.rank

            psv = [pl_id, family, region, group_number, group, city, rank, mesto_group]
            full_posev.append(psv) 

        if fin == "Суперфинал":
            full_posev.sort(key=lambda k: k[7]) # сортировка списка участников по месту в 1-ом финале
        elif count_exit == 1 or fin == "Одна таблица":
            full_posev.sort(key=lambda k: k[6], reverse=True) # сортировка списка участников по рейтингу
        elif fin == "1-й финал":
            if count_exit > 1:
                full_posev.sort(key=lambda k: (k[7], k[3])) # сортировка списка участников сначала по месту в группе а потом по группач
            else:
                full_posev.sort(key=lambda k: k[3]) # сортировка списка участников по группам
        elif count_exit != 1 or fin != "1-й финал":
            full_posev.sort(key=lambda k: k[6], reverse=True) # сортировка списка участников по рейтингу
   

        for k in full_posev:
            k.pop(3)
            k.pop(6)
        # ======== начало жеребьевки =========
        end = player_net // count_exit if flag == 1 else count_player_in_final
        number_posev = [i for i in range(0, end)] # генератор списка

        if n == 0:
            posev = posev_1
            # list_of_lists = posev_1
            # # === вариант обьединить все псоевы в один список
            # posev = list(itertools.chain(*list_of_lists)) # соединяет отдельный посевы в один список    
        elif n == 1:
            posev = posev_2
        elif n == 2:
            posev = posev_3
        elif n == 3:
            posev = posev_4
          
        if flag == 3: # если ручная жеребьевка то обьединяет все посевы
            if count_exit == 1:
                posev = posev_1
            elif count_exit == 2:
                posev = posev_1 + posev_2
            elif count_exit == 3:
                posev = posev_1 + posev_2 + posev_3
            elif count_exit == 4:
                posev = posev_1 + posev_2 + posev_3 + posev_4

        count_posev = len(posev) # количество подпосевов в посеве

        for i in range(0, count_posev):  # список посева, разделеный на отдельные посевы
            if flag_stop_manual_choice == 1: # выход из цикла если окончена ручная жеребьевка
                break

            current_region_posev.clear()
            sev_tmp = posev[i].copy()
            sev = sev_tmp.copy()
            sev_tmp.clear()
            count = len(posev[i]) # всего количество номеров в посеве
            if del_num == 1 and i == count_posev - 1:                    
                for h in free_num:
                    sev.remove(h)
                free_seats = len(free_num) # сколько свободных мест в сетке
                count = len(posev[i]) - free_seats
            for w in range(0, count): # внутренний цикл посева
                l = number_posev[0] # общий список всего посева (порядковый номер посева)
                if i == 0 and n == 0: #  ===== 1-й посев (сеятся игроки на 1-й и последний номер в сетке)
                    sev = posev[i]  # список номеров в сетке для посева
                    num_set = sev[w] # номер в сетке на который идет сев
                    count_sev = len(sev) # количество номеров в посеве
                else: # посев остальных номеров
                    # num_set = sev[0] # проверить
                    # if len(posev[i]) > count_player_in_final and count_exit > 1:
                    if count > count_player_in_final and count_exit > 1:
                        count_sev = count_player_in_final
                    else:
                        count_sev = len(sev) # конкретное число оставшихся в посеве минус свободных мест(если они есть)
                    if count_sev > 1: # если сеющихся номеров больше одного
                        if w == 0: # 1-й основной посев
                            gr_region_tmp = []
                            for k in range(l, l + count_sev): # цикл получения словаря (регионы в предстоящем посеве)
                                region = full_posev[k][2]
                                gr = full_posev[k][3]
                                gr_region_tmp.append(region)
                                gr_region_tmp.append(gr)
                                gr_region = gr_region_tmp.copy()
                                current_region_posev[k] = gr_region # словарь регионы, в текущем посеве по порядку
                                gr_region_tmp.clear()
                        number_last.clear()
                        number_last = list(num_id_player.keys()) # список уже посеянных номеров в сетке
                        # ==== вариант  с сорт по кол ву регионов начиная с 5 по 8 посев
                        if l > 3 and l < 15:
                            current_region_posev = sort_region(current_region_posev)
                        # =====
                        reg_last.clear()
                        group_last.clear()
                        for v in num_id_player.values():
                            reg_last.append(v[1]) # список уже посеянных регионов
                            group_last.append(v[2]) # список номеров групп уже посеянных
                        if n != 0 or (n == 0 and l > 1):
                        # =========== определения кол-во возможны вариантов посева у каждого региона 
                            possible_number = possible_draw_numbers(current_region_posev, reg_last, number_last, group_last, n, sev, num_id_player, player_net, count_exit)                        

                            if i != 0 or n != 0: # отсортирововаем список по увеличению кол-ва возможных вариантов
                                possible_number = {k:v for k,v in sorted(possible_number.items(), key=lambda x:len(x[1]))}
                                num_posev = list(possible_number.keys())
                            # ============== проба получить возможные варианты посева
                            # flag_new = 0
                            variant_poseva = posev_variant(possible_number, sev)
                            choice_dict = variant_poseva[0]
                            # === выбор ручная или автомат ====
                            if flag == 1: # автоматичекая
                                count = len(choice_dict) # сколько номеров сеятся
                                for j in choice_dict.keys():
                                    num_set = choice_dict[j] # номера куда можно сеять
                                    id_player = full_posev[l][0]
                                    region = full_posev[l][2]
                                    gr = full_posev[l][3]  
                                    id_region = [id_player, region, gr]
                                    num_id_player[num_set] = id_region # словарь номера в сетке - регион

                                    del possible_number[j] # удаляет из словаря возможных номеров посеянный порядковый номер
                                    del current_region_posev[j] # удаляет из словаря текущий посеянный регион
                                    if num_set in sev: # проверяет посеянный номер в посеве
                                        sev.remove(num_set)  # удаляет посеянный номер из всех номеров этого посева
                                    num_posev.remove(l)
                                    number_posev.remove(l)
                                    l += 1
                                flag_new = 1
                            elif flag == 2: # полуавтомат
                                my_win.tableView.setGeometry(QtCore.QRect(260, 241, 841, 540))
                                player_list = []
                                player_list_tmp = []

                                for j in possible_number.keys():
                                    posev_list = full_posev[j]
                                    pl = posev_list[1] # фамилия
                                    reg = posev_list[2] # регион
                                    pn = possible_number[j] # возможные номера посева
                                    player_list_tmp = [pl, reg, pn]
                                    player_list.append(player_list_tmp.copy())
                                    player_list_tmp.clear()
                                txt_tmp = []
    
                                for g in player_list:
                                    if len(num_id_player) == 2:
                                        fam_city = ""
                                        number_net = ""
                                        view_table_choice(fam_city, number_net, num_id_player) # функция реального просмотра жеребьевки
                                    t_str = str(g[2])
                                    txt_str = f"{g[0]} - {g[1]} номера: {t_str}" 
                                    txt_tmp.append(txt_str)
                                text_str = (',\n'.join(txt_tmp))
                                tx = f"Список спортсменов в порядке посева:\n\n{text_str}\n\n" + "Выберите один из номеров и нажмите\n - ОК - если выбрали сами или - Cancel - если хотите выбор случайный"
                                txt = (','.join(list(map(str, num_set))))
                                while True:
                                    try:
                                        number_net, ok = QInputDialog.getText(my_win, f'Возможные номера посева: {txt}', tx)
                                        znak = text_str.find(":")
                                        fam_city = text_str[:znak - 7]
                                        if not ok:
                                            number_net = random.choice(num_set)
                                        msgBox.information(my_win, "Жеребьевка участников", f"{fam_city} идет на номер: {number_net}")
                                        number_net = int(number_net)
                                        view_table_choice(fam_city, number_net, num_id_player) # функция реального просмотра жеребьевки
                                    except ValueError:
                                        msgBox.information(my_win, "Уведомление", "Вы не правильно ввели номер, повторите снова.")
                                        continue
                                    else:
                                        if number_net in num_set:
                                            num_set = number_net
                                            break
                                        else:
                                            msgBox.information(my_win, "Уведомление", "Вы не правильно ввели номер, повторите снова.") 
                            elif flag == 3: # ручная жеребьевка
                                q = 1
                                num_list = "-"
                                for r in range(1, max_player + 1 ):
                                    num_id_player[r] = num_list
                                my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 841, 274))
                                my_win.tableView.setGeometry(QtCore.QRect(260, 318, 841, 430))
                                txt_tmp = []
                                player_list = []
                                pl_id_list = []
                                player_list_tmp = []
                                for k in range(len(full_posev)):
                                    posev_list = full_posev[k]
                                    id_pl = posev_list[0] # id игрока
                                    pl = posev_list[1] # фамилия
                                    reg = posev_list[2] # регион
                                    gr = posev_list[3]
                                    r = posev_list[5] # рейтинг
                                    player_list_tmp = [pl, reg, r]
                                    player_list.append(player_list_tmp.copy())
                                    pl_id_list_tmp = [id_pl, reg, gr] 
                                    pl_id_list.append(pl_id_list_tmp.copy())
                                    player_list_tmp.clear()
                                    pl_id_list_tmp.clear()
                                    q += 1
                                # if fin != "Суперфинал":
                                #     pl_id_list.sort(key=lambda x: x[2], reverse=True) # отсортировывает списки списков по 3-му элементу
                                #     player_list.sort(key=lambda x: x[2], reverse=True) # отсортировывает списки списков по 3-му элементу
                                for i in range(0, count_posev):
                                    n_poseva = posev[i] 
                                    count_sev = len(n_poseva)
                                    if i == 0: 
                                        for l in player_list: # цикл создания списка для ручной жеребьевки
                                            pl = l[0]
                                            region = l[1]
                                            pl_reg = f"{pl}/ {region}"
                                            txt_tmp.append(pl_reg)
                                        text_str = (',\n'.join(txt_tmp))
                                    m = 0
                                    for k in list(full_posev):
                                        if m == count_sev:
                                            break
                                        else:
                                            id_player = k[0]
                                            region = k[2]
                                            gr = k[3]  
                                            id_region = [id_player, region, gr]
                                            n_sev = n_poseva
                                            f_text = txt_tmp[0]                                  
                                            tx = f"Сеятся игрок:\n{f_text}\n\nСписок спортсменов в порядке посева:\n\n{text_str}\n\n"\
                                                "Выберите один из номеров и нажмите - ОК -" 
                                            number_net, ok = QInputDialog.getText(my_win, f'Возможные номера посева: {n_sev}', tx, QLineEdit.Normal) 
                                            if number_net == '' or int(number_net) not in [j for j in range(1, player_net + 1)]:
                                                msgBox.information(my_win, "Уведомление", "Вы не правильно ввели номер, повторите снова.")
                                                break
                                            fam_city = ""
                                            old = num_id_player[int(number_net)]

                                            num_id_player[int(number_net)] = id_region
                                            view_table_choice(fam_city, number_net, num_id_player)
                                            txt_tmp.remove(f_text)
                                            text_str = (',\n'.join(txt_tmp))
                                            full_posev.pop(0)

                                            if old != "-":
                                                id_old = old[0]
                                                players = Player.select().where(Player.id == id_old).get()
                                                fio = players.player
                                                city = players.city
                                                old_player = f"{fio}/{city}"
                                                number_net, ok = QInputDialog.getText(my_win,'Замена игрока', f'На какой номер в сетке\nпереместить игрока:\n {old_player}', QLineEdit.Normal) 
                                                num_id_player[int(number_net)] = old
                                                fam_city = ""
                                                view_table_choice(fam_city, number_net, num_id_player)
                                            m += 1
                                        end -= 1 
                                        if end == 0:
                                        # if end == 0 or real_all_player_in_final == (len(num_id_player) - len(free_num)):
                                            flag_stop_manual_choice = 1
                                            step = 100
                if flag_stop_manual_choice == 0:
                    if flag_new == 0:                      
                        id_player = full_posev[l][0]
                        region = full_posev[l][2]
                        gr = full_posev[l][3]  
                        id_region = [id_player, region, gr]
                        num_id_player[num_set] = id_region # словарь номера в сетке - регион
                    # ======== модуль удаления посеянных номеров =========
                    if count_sev > 1:
                        c = len(current_region_posev)
                        if c != 0:
                            del possible_number[l] # удаляет из словаря возможных номеров посеянный порядковый номер
                            del current_region_posev[l] # удаляет из словаря текущий посеянный регион
                            if num_set in sev: # проверяет посеянный номер в посеве
                                sev.remove(num_set)  # удаляет посеянный номер из всех номеров этого посева
                            for z in possible_number.keys():
                                possible_tmp = possible_number[z]
                                #=====
                                if flag is False and len(possible_number) == 1:
                                    number_net = sev[0]
                                    fam_city = f"{pl}/{reg}"
                                    view_table_choice(fam_city, number_net, num_id_player) # функция реального просмотра жеребьевки 
                                #======
                                if num_set in possible_tmp: # проверяет посеянный номер в возможных номерах
                                    possible_tmp.remove(num_set) # удаляет посеянный номер из возможных номеров
                    elif count_sev == 1: # удаляет последний номер в посеве
                        sev.clear()
                        possible_number.clear()
                    number_posev.remove(l)
                    if i != 0 and flag == 0:
                        num_posev.remove(l)

                    sp = 100 / (real_all_player_in_final)
                    step += sp
                else:
                    break
        if step > 99:
            del_num_list = []  
            for i in num_id_player.keys():
                tmp_list = list(num_id_player[i])
                if len(tmp_list) == 1:
                    del_num_list.append(i)                    
                else:
                    id = tmp_list[0]
                    pl_id = Player.get(Player.id == id)
                    family_city = pl_id.full_name
                    posev_data[i] = family_city
                    with db:
                        choice_final = choice.select().where(Choice.player_choice_id == pl_id).get()
                        if fin == "Суперфинал":
                            choice_final.super_final = i
                        else:
                            choice_final.final = fin
                            choice_final.posev_final = i
                        choice_final.save()
            if len(del_num_list) > 0:
                for e in del_num_list:
                    del num_id_player[e]
            key_set = set(num_id_player.keys()) # получаем сет всех ключей (номеров сетки)
            all_num_set = [j for j in range(1, player_net + 1)]
            free_number = set((all_num_set))
            free_number.difference_update(key_set) # вычитаем из всех номеров те которые посеяны и остается номера -X-
            for h in free_number:
                posev_data[h] = "X"
        n += 1 # добавил в связи со сменой цикла
    return posev_data


def posev_variant(possible_number, sev):
    """определяет возможные варианты посева"""
    all_var = {}
    variant = {}
    variant_tmp_list = []
    sev_all = []
    sev_num_list = []
    count = len(possible_number) # количество номеров в посеве
    # # ============
    for i in permutations(sev, count):  # получает список всех вариантов жеребьевки
        i = list(i)
        sev_all.append(i)
    count_sev = len(sev_all)

    for j in range(0, count_sev):
        v = 0
        one_variant = sev_all[j]
        for p in possible_number.keys():
            num = one_variant[v]
            possible_num = possible_number[p] # возможные номера посева у региона
            if num in possible_num:
                variant[p] = one_variant[v]
                v += 1
        variant_tmp_list.append(variant.copy())
        sev_num_list = variant_tmp_list.copy() 
        variant_tmp_list.clear()
        r = len(sev_num_list[0])
        if count == r: # оставляет только рабочие варианты
            all_var[j] = sev_num_list
        # выбирает случайнйю компбинация посева
    key_list = list(all_var.keys())
    numbers_poseva = random.choice(key_list)
    choice_dict = all_var[numbers_poseva]
    return choice_dict      


def sort_region(current_region_posev):
    """сортировка регионов по их количеству в посеве"""
    m = 0
    sum_reg = {}
    for r in current_region_posev.values():
        reg_list = r[0]
        x = sum_reg.setdefault(reg_list, m) 
        sum_reg[reg_list] = x + 1
    for y in sum_reg.keys():
        for d in current_region_posev.values():
            if y == d[0]:
                zn = sum_reg[y]
                d.append(zn)
    marklist = sorted(current_region_posev.items(), key=lambda item: item[1][2], reverse=True)
    current_region_posev = dict(marklist)
    for r in current_region_posev.values():
        reg_list = r[0]
        r.pop()
        sum_reg[reg_list] = r
    return current_region_posev


def sortkey(e):
    return e[1]



def _view_table_group_choice(id_fam, number_group, number_posev):
    """показ таблицы жеребьевки"""
    stage = "Предварительный" # менять при ручной жеребьевка пф или игр в круг 
    sys = System.select().where(System.title_id == title_id())
    sys_id = sys.select().where(System.stage == stage).get()
    group = sys_id.total_group
    max_player = sys_id.max_player
    # создает таблицу посева групп
    my_win.tableWidget_chioce_group.setColumnCount(group) # устанавливает колво столбцов
    my_win.tableWidget_chioce_group.setRowCount(max_player) # кол-во строк (максимальное число игроков в группе)
    column_label = [f'{i} группа' for i in range(1, group + 1)] # генератор списка
    my_win.tableWidget_chioce_group.resizeColumnsToContents()
    for i in range(0, group):  # закрашивает заголовки таблиц  рейтинга зеленым цветом
        my_win.tableWidget_chioce_group.showColumn(i)
        item = QtWidgets.QTableWidgetItem()
        brush = QtGui.QBrush(QtGui.QColor(76, 100, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        my_win.tableWidget_chioce_group.setHorizontalHeaderItem(i, item)
    my_win.tableWidget_chioce_group.setHorizontalHeaderLabels(column_label) # заголовки столбцов в tableWidget
    znak = id_fam.find("\n")
    znak_1 = id_fam.find("/")
    id_pl = id_fam[:znak_1] # получение id игрока
    if znak == 0:
        fam_region = id_fam[1:]
    else:
        fam_region = id_fam
    fam_region = fam_region.replace(f'{id_pl}/', '') # удаляет из строки id игрока
    fam_region = fam_region.replace('/', '\n') 
    num_group = f'{number_group} группа'
    Choice.update(posev_group=number_posev + 1, group=num_group).where(Choice.player_choice == id_pl).execute() # записывает в Choice
    my_win.tableWidget_chioce_group.setItem(number_posev, number_group - 1, QTableWidgetItem(fam_region)) # (номер строки, номер столбца, значения)
    # my_win.tableWidget_chioce_group.resizeColumnsToContents() # растягивает ячейку по ширине
    # my_win.tableWidget_chioce_group.resizeRowsToContents()  # растягивает ячейку по высоте
    my_win.tableWidget_chioce_group.show()


def view_table_group_choice(id_fam_region_list, max_player, group):
    """показ таблицы жеребьевки с отдельной строкой регионы"""
    header_list = []
    data = id_fam_region_list
    for b in range(1, group + 1):
        header_list.append(f"{b} группа")
    header_list.insert(0, "посев")
    model = MyTableModel(data)
    model.setHorizontalHeaderLabels(header_list) # список заголовков
    my_win.tableView_choice_group.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
    font = my_win.tableView_choice_group.font()
    font.setPointSize(11)
    my_win.tableView_choice_group.setFont(font)
    my_win.tableView_choice_group.horizontalHeader().setFont(QFont("Times", 12, QFont.Bold)) # делает заголовки жирный и размер 13
    my_win.tableView_choice_group.horizontalHeader().setStyleSheet("background-color:yellow;") # делает фон заголовков светлоголубой
    my_win.tableView_choice_group.verticalHeader().setDefaultSectionSize(15)
    my_win.tableView_choice_group.setGridStyle(QtCore.Qt.DashDotLine) # вид линии сетки 
    for k in range(0, max_player * 2, 2):
        my_win.tableView_choice_group.setSpan(k, 0, 2, 1) # нач строка, нач столбец, кол-во строк, кол-во столбцов
    my_win.tableView_choice_group.setModel(model)
    my_win.tableView_choice_group.show()


def view_table_choice(fam_city, number_net, num_id_player):
    """показ таблицы жеребьевки"""
    data = []
    num_fam = []
    manual_choice_dict = {}
    player = Player.select().where(Player.title_id == title_id())
    count_player = max(num_id_player.keys()) # наибольшой ключ в словаре (на сколько сетка)

    manual_choice_dict = num_id_player.copy()

    for r in range(1, count_player + 1):
        manual_choice_dict.setdefault(r, "-")
        list_net = manual_choice_dict[r]
        if r == number_net:
            num_fam_tmp = [r, fam_city, ""]
        elif list_net == "-":
            num_fam_tmp = [r, list_net, ""]
        else:
            id_player = list_net[0]
            group = list_net[2]
            group.replace("группа", "гр.")
            pl_full = player.select().where(Player.id == id_player).get()
            player_full = pl_full.full_name
            num_fam_tmp = [r, player_full, group]
        num_fam = num_fam_tmp.copy()
        num_fam_tmp.clear()
        data.append(num_fam) # список списков
        model = MyTableModel(data)
        my_win.tableView_net.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        my_win.tableView_net.verticalHeader().setDefaultSectionSize(15)
        my_win.tableView_net.setGridStyle(QtCore.Qt.DashDotLine) # вид линии сетки 
        my_win.tableView_net.setModel(model)
        my_win.tableView_net.show()


def setka_choice_number(fin, count_exit):
    """номера сетки при посеве"""
    posevs = []
    posev_1 = []
    posev_2 = []
    posev_3 = []
    posev_4 = []
    id_system = system_id(stage=fin)
    system = System.select().where((System.title_id == title_id()) & (System.id == id_system)).get()
    type_setka = system.label_string
    if fin == "Суперфинал":
        count_exit = 1
        posev_1 = [[1, 8], [4, 5], [2, 3, 6, 7]]
        player_net = 8
    else:
        if count_exit == 1 or fin == "Одна таблица":
            if type_setka == "Сетка (с розыгрышем всех мест) на 8 участников" or type_setka == "Сетка (-2) на 8 участников":
                posev_1 = [[1, 8], [4, 5], [2, 3, 6, 7]]
                player_net = 8
            elif type_setka == "Сетка (с розыгрышем всех мест) на 16 участников" or type_setka == "Сетка (-2) на 16 участников":
                posev_1 = [[1, 16], [8, 9], [4, 5, 12, 13], [2, 3, 6, 7, 10, 11, 14, 15]]
                player_net = 16
            elif type_setka == "Сетка (с розыгрышем всех мест) на 32 участников" or type_setka == "Сетка (-2) на 32 участников":
                posev_1 = [[1, 32], [16, 17], [8, 9, 24, 25], [4, 5, 12, 13, 20, 21, 28, 29], [2, 3, 6, 7, 10, 11, 14, 15, 18, 19, 22, 23, 26, 27, 30, 31]]
                player_net = 32
        elif count_exit == 2:
            if type_setka == "Сетка (с розыгрышем всех мест) на 8 участников" or type_setka == "Сетка (-2) на 8 участников":
                posev_1 = [[1, 8], [4, 5]]
                posev_2 = [[2, 3, 6, 7]]
                player_net = 8
            elif type_setka == "Сетка (с розыгрышем всех мест) на 16 участников" or type_setka == "Сетка (-2) на 16 участников":
                posev_1 = [[1, 16], [8, 9], [4, 5, 12, 13]]
                posev_2 = [[2, 3, 6, 7, 10, 11, 14, 15]]
                player_net = 16
            elif type_setka == "Сетка (с розыгрышем всех мест) на 32 участников" or type_setka == "Сетка (-2) на 32 участников":
                posev_1 = [[1, 32], [16, 17], [8, 9, 24, 25], [4, 5, 12, 13, 20, 21, 28, 29]]
                posev_2 = [[2, 3, 6, 7, 10, 11, 14, 15, 18, 19, 22, 23, 26, 27, 30, 31]]
                player_net = 32
        elif count_exit == 3:
            if type_setka == "Сетка (с розыгрышем всех мест) на 8 участников" or type_setka == "Сетка (-2) на 8 участников":
                posev_1 = [[1, 8]]
                posev_2 = [[4, 5]]
                posev_3 = [[3, 6]]
                # posev_4 = [[2, 7]]
                player_net = 8
            elif type_setka == "Сетка (с розыгрышем всех мест) на 16 участников" or type_setka == "Сетка (-2) на 16 участников":
                posev_1 = [[1, 16], [8, 9]]
                posev_2 = [[4, 5, 12, 13]]
                posev_3 = [[3, 6, 11, 14]]
                # posev_4 = [[2, 7, 10, 15]]
                player_net = 16
            elif type_setka == "Сетка (с розыгрышем всех мест) на 32 участников" or type_setka == "Сетка (-2) на 32 участников":
                posev_1 = [[1, 32], [16, 17], [8, 9, 24, 25]]
                posev_2 = [[4, 5, 12, 13, 20, 21, 28, 29]]
                posev_3 = [[3, 6, 11, 14, 19, 22, 27, 30]]
                # posev_4 = [[2, 7, 10, 15, 18, 23, 26, 31]] 
                player_net = 32
        elif count_exit == 4:
            if type_setka == "Сетка (с розыгрышем всех мест) на 8 участников" or type_setka == "Сетка (-2) на 8 участников":
                posev_1 = [[1, 8]]
                posev_2 = [[4, 5]]
                posev_3 = [[3, 6]]
                posev_4 = [[2, 7]]
                player_net = 8
            elif type_setka == "Сетка (с розыгрышем всех мест) на 16 участников" or type_setka == "Сетка (-2) на 16 участников":
                posev_1 = [[1, 16], [8, 9]]
                posev_2 = [[4, 5, 12, 13]]
                posev_3 = [[3, 6, 11, 14]]
                posev_4 = [[2, 7, 10, 15]]
                player_net = 16
            elif type_setka == "Сетка (с розыгрышем всех мест) на 32 участников" or type_setka == "Сетка (-2) на 32 участников":
                posev_1 = [[1, 32], [16, 17], [8, 9, 24, 25]]
                posev_2 = [[4, 5, 12, 13, 20, 21, 28, 29]]
                posev_3 = [[3, 6, 11, 14, 19, 22, 27, 30]]
                posev_4 = [[2, 7, 10, 15, 18, 23, 26, 31]]
                player_net = 32
    posevs.append(player_net)
    if len(posev_1) != 0:
        posevs.append(posev_1)
        if len(posev_2) != 0:
            posevs.append(posev_2)
            if len(posev_3) != 0:
                posevs.append(posev_3)
                if len(posev_4) != 0:
                    posevs.append(posev_4)
    return posevs


def free_place_in_setka(max_player, real_all_player_in_final):
    """вычеркиваем свободные номера в сетке"""
    free_num = []
    free_number_8 = [2, 7, 6, 3]
    free_number_16 = [2, 15, 7, 10, 6, 11, 3, 14]
    free_number_24 = [5, 20, 8, 17, 11, 14, 2, 23]
    free_number_32 = [2, 31, 15, 18, 10, 23, 7, 26, 6, 27, 11, 22, 14, 19, 3, 30]
    count = max_player - real_all_player_in_final # кол-во свободных мест


    if max_player == 8:
        free_number = free_number_8
    if max_player == 16:
        free_number = free_number_16
    elif max_player == 24:
        free_number = free_number_24
    elif max_player == 32:
        free_number = free_number_32

    for i in range (0, count):
        k = free_number[i]
        free_num.append(k)
    return free_num
    

def possible_draw_numbers(current_region_posev, reg_last, number_last, group_last, n, sev, num_id_player, player_net, count_exit):
    """возможные номера посева new"""
    possible_number = {}
    proba_possible = {} 
    num_tmp = []
    reg_tmp = []
    # =========
    titles = Title.get(Title.id == title_id())
    multi_reg = titles.multiregion
    #============
    current_region = list(current_region_posev.values()) # список регионов и группы в предстоящем посеве
    y = 0
    for reg in current_region_posev.keys():
        cur_reg = current_region[y][0] # текущий регион посева
        cur_gr = current_region[y][1] # номер группы, которая сеятся
        #=======
        # if multi_reg == 0 or (len(num_id_player) >= player_net // 2 and count_exit == 1): # если спортсмены одного региона нет рассеивания
        if multi_reg == 0: # если спортсмены одного региона нет рассеивания
            possible_number[reg] = sev
        else:
            if n == 0:
                if cur_reg in reg_last: # если регион который сеятся есть в уже посеянных областях
                    reg_tuple = tuple(reg_last)
                    count = reg_tuple.count(cur_reg) # количество регионов уже посеянных 
                    if count == 1: # значит только один регион в посеве
                        cur_gr = current_region[y][1]
                        number_posev = number_setka_posev(cur_gr, group_last, reg_last, number_last, n, cur_reg, sev, player_net)
                        possible_number[reg] = number_posev
                    else: # если есть уже областей более двух
                        number_tmp = []
                        num_tmp.clear()
                        start = 0
                        for k in reg_last: # получаем список номеров сетки областей в той половине куда идет сев
                            if k == cur_reg:
                                index = reg_last.index(k, start)
                                set_number = number_last[index] # номер где уже посеянна такая же область
                                num_tmp.append(set_number)
                            start += 1
                        # if count % 2 == 0: # если число четное
                        if count == 2: # посеяны 2 области разводит по четвертям
                            for h in num_tmp:
                                if h <= player_net // 4: # если номер в сетке вверху, то наде сеять вниз
                                    f = [i for i in sev if i >= player_net // 4 + 1 and i <= player_net // 2] # отсеивает в списке номера 9-16
                                elif h > player_net // 4 and h <= player_net // 2: 
                                    f = [i for i in sev if i <= player_net // 4] # отсеивает в списке номера 1-8
                                elif h >= player_net // 2 + 1 and h <= int(player_net * 3 / 4): 
                                    f = [i for i in sev if i > player_net * 3 / 4] # отсеивает в списке номера 25-32
                                elif h > player_net * 3 / 4: 
                                    f = [i for i in sev if i >= player_net // 2 + 1 and i <= int(player_net * 3 / 4)] # отсеивает в списке номера 17-24
                                number_tmp += f
                        elif count == 3:
                            number_posev = sev
                            number_tmp = alignment_in_half(player_net, num_tmp, sev, count, number_posev)
                        elif count >= 4: # посеяны 4 области разводит по восьмушкам
                            number_tmp = sev
                                # if player_net == 16:
                                #     for h in num_tmp:
                                #         if h <= 2: # если номер в сетке 1-2
                                #             f = [i for i in sev if i >= 3 and i <= 4] # отсеивает в списке номера 3-4 ()
                                #         elif h >= 3 and h <= 4: # если номер в сетке 3-4
                                #             f = [i for i in sev if i < 3] # отсеивает в списке номера 1-2 ()
                                #         elif h >= 5 and h <= 6: # если номер в сетке 5-6
                                #             f = [i for i in sev if i >= 7 and i <= 8] # отсеивает в списке номера 25-32
                                #         elif h >= 7 and h <= 8: # если номер в сетке 7-8
                                #             f = [i for i in sev if i >= 5 and i <= 6] # отсеивает в списке номера 17-24
                                #         elif h >= 9 and h <= 10: # если номер в сетке вверху, то наде сеять вниз
                                #             f = [i for i in sev if i >= 11 and i <= 12] # отсеивает в списке номера 9-16
                                #         elif h >= 11 and h <= 12: 
                                #             f = [i for i in sev if i <= 9 and i <= 10] # отсеивает в списке номера 1-8
                                #         elif h >= 13 and h <= 14: 
                                #             f = [i for i in sev if i > 14] # отсеивает в списке номера 25-32
                                #         elif h > 14: 
                                #             f = [i for i in sev if i >= 12 and i <= 13] # отсеивает в списке номера 17-24    
                                #         number_tmp += f
                                # elif player_net == 32:
                                    # ==== новый вариант где уже 4 области разведены
                                    # number_tmp = sev
                                    # =======
                                    # for h in num_tmp:
                                    #     if h <= player_net // 8: # если номер в сетке вверху, то наде сеять вниз
                                    #         f = [i for i in sev if i >= 5 and i <= 8] # отсеивает в списке номера 3-4 ()
                                    #     elif h >= 5 and h <= 8: 
                                    #         f = [i for i in sev if i < 5] # отсеивает в списке номера 1-2 ()
                                    #     elif h >= 9 and h <= 12: 
                                    #         f = [i for i in sev if i >= 13 and i <= 16] # отсеивает в списке номера 25-32
                                    #     elif h >= 13 and h <= 16: 
                                    #         f = [i for i in sev if i >= 9 and i <= 12] # отсеивает в списке номера 17-24
                                    #     elif h >= 17 and h <= 20: # если номер в сетке вверху, то наде сеять вниз
                                    #         f = [i for i in sev if i >= 21 and i <= 24] # отсеивает в списке номера 9-16
                                    #     elif h >= 21 and h <= 24: 
                                    #         f = [i for i in sev if i >= 17 and i <= 20] # отсеивает в списке номера 1-8
                                    #     elif h >= 25 and h <= 28: 
                                    #         f = [i for i in sev if i >= 29] # отсеивает в списке номера 25-32
                                    #     elif h > 28: 
                                    #         f = [i for i in sev if i >= 25 and i <= 28] # отсеивает в списке номера 17-24    
                                    #     number_tmp += f
                        # elif count > 2:
                        #     # number_posev = number_setka_posev(cur_gr, group_last, reg_last, number_last, n, cur_reg, sev, player_net, count_exit)
                        #     # if 
                        #     number_posev = sev
                        #     number_tmp = alignment_in_half(player_net, num_tmp, sev, count, number_posev)
                        
                        number_posev = number_tmp.copy()
                        possible_number[reg] = number_posev
                else: # все номера в той части куда можно сеять
                    possible_number[reg] = sev
            else: # 2-й посев и последующие 
                number_posev = number_setka_posev(cur_gr, group_last, reg_last, number_last, n, cur_reg, sev, player_net) # возможные номера после ухода от своей группы без учета регионов
                number_posev_old = number_setka_posev_last(cur_gr, group_last, number_last, n, player_net)
                reg_tmp.clear()
                # ======
                # if n > 1:# при выоде в сетку двоих не разводил регионы
                if n >= 1:
                    for k in number_posev_old: # получаем список прошлых посеянных областей в той половине куда идет сев
                        d = number_last.index(k)
                        reg_tmp.append(reg_last[d]) # список регионов     
                    if cur_reg in reg_tmp: # если сеянная область есть в прошлом посеве конкретной половины
                        num_tmp = [] # список номеров сетки где есть такой же регион (в той половине или четверти с номером который сеятся)
                        for d in number_posev_old: # номер в сетке в предыдущем посеве
                            posev_tmp = num_id_player[d]
                            if cur_reg in posev_tmp:
                                num_tmp.append(d) # список номеров в сетке, где уже есть такой же регион
                        count = len(num_tmp) # количество областей в той части сетки, куде сеятся регион
                    # ======== отбирает номера из -number_posev- , где учитывается регион ======
                        if count == 1 and n == 1: # есть только одна область в той же половине другой четверти (1 место и 2-е место в группе)
                            if num_tmp[0] <= player_net // 4: # в первой четверти (1-8)
                                number_posev = [i for i in number_posev if i > player_net // 4 and i <= player_net // 2] # номера 8-16
                            elif num_tmp[0] >= (player_net // 4 + 1) and num_tmp[0] <= player_net // 2: # в первой четверти (9-16)
                                number_posev = [i for i in number_posev if i < 9] # номера 1-8
                            elif num_tmp[0] >= (player_net // 2 + 1) and num_tmp[0] <= player_net // 4 * 3: # в первой четверти (16-24)
                                number_posev = [i for i in number_posev if i > player_net // 4 * 3] # номера 25-32
                            elif num_tmp[0] >= (player_net // 4 * 3 + 1) and num_tmp[0] <= player_net: # в первой четверти (25-32)
                                number_posev = [i for i in number_posev if i > player_net // 2 and i < (player_net // 4 * 3 + 1)] # номера 17-24
                        elif (count == 1 and n == 2):
                            if num_tmp[0] <= player_net // 8: # в первой четверти (1-4)
                                number_posev = [i for i in number_posev if i > player_net // 8 and i <= player_net // 4] # номера 5-8
                            elif num_tmp[0] >= player_net // 8 + 1 and num_tmp[0] <= player_net // 4: # в первой четверти (5-8)
                                number_posev = [i for i in number_posev if i < player_net // 8 + 1] # номера 1-4
                            elif num_tmp[0] >= player_net // 4 + 1 and num_tmp[0] <= player_net // 8 * 3: # в первой четверти (9-12)
                                number_posev = [i for i in number_posev if i > player_net // 8 * 3 and i <= player_net // 2] # номера 13-16
                            elif num_tmp[0] >= (player_net // 8 * 3 + 1) and num_tmp[0] <= player_net // 2: # в первой четверти (13-16)
                                number_posev = [i for i in number_posev if i >= player_net // 4 + 1 and i <= player_net // 8 * 3] # номера 9-12
                            elif num_tmp[0] >= player_net // 2 + 1 and num_tmp[0] <= player_net // 8 * 5: # в первой четверти (17-20)
                                number_posev = [i for i in number_posev if i > player_net // 8 * 5 and i <= player_net // 4 * 3] # номера 21-24
                            elif num_tmp[0] >= player_net // 8 * 5 and num_tmp[0] <= (player_net // 4 * 3): # в первой четверти (21-24)
                                number_posev = [i for i in number_posev if i >(player_net // 2 + 1) and i <= player_net // 8 * 5] # номера 17-20
                            elif num_tmp[0] >= (player_net // 4 * 3 + 1) and num_tmp[0] <= player_net // 8 * 7: # в первой четверти (25-28)
                                number_posev = [i for i in number_posev if i > player_net  // 8 * 7 + 1] # номера 29-32
                            elif num_tmp[0] >= player_net // 8 * 7 + 1: # в первой четверти (29-32)
                                number_posev = [i for i in number_posev if i >= player_net // 4 * 3 + 1 and i <= player_net  // 8 * 7] # номера 25-28
                        elif n == 3:
                            if count == 1 and len(number_posev) != 1:
                                if num_tmp[0] <= player_net // 8: # в первой четверти (1-4)
                                    number_posev = [i for i in number_posev if i > player_net // 8 and i <= player_net // 4] # номера 5-8
                                elif num_tmp[0] >= player_net // 8 + 1 and num_tmp[0] <= player_net // 4: # в первой четверти (5-8)
                                    number_posev = [i for i in number_posev if i < player_net // 8 + 1] # номера 1-4
                                elif num_tmp[0] >= player_net // 4 + 1 and num_tmp[0] <= player_net // 8 * 3: # в первой четверти (9-12)
                                    number_posev = [i for i in number_posev if i > player_net // 8 * 3 and i <= player_net // 2] # номера 13-16
                                elif num_tmp[0] >= (player_net // 8 * 3 + 1) and num_tmp[0] <= player_net // 2: # в первой четверти (13-16)
                                    number_posev = [i for i in number_posev if i >= player_net // 4 + 1 and i <= player_net // 8 * 3] # номера 9-12
                                elif num_tmp[0] >= player_net // 2 + 1 and num_tmp[0] <= player_net // 8 * 5: # в первой четверти (17-20)
                                    number_posev = [i for i in number_posev if i > player_net // 8 * 5 and i <= player_net // 4 * 3] # номера 21-24
                                elif num_tmp[0] >= player_net // 8 * 5 and num_tmp[0] <= (player_net // 4 * 3): # в первой четверти (21-24)
                                    number_posev = [i for i in number_posev if i >= (player_net // 2 + 1) and i <= player_net // 8 * 5] # номера 17-20
                                elif num_tmp[0] >= (player_net // 4 * 3 + 1) and num_tmp[0] <= player_net // 8 * 7: # в первой четверти (25-28)
                                    number_posev = [i for i in number_posev if i > player_net  // 8 * 7 + 1] # номера 29-32
                                elif num_tmp[0] >= player_net // 8 * 7 + 1: # в первой четверти (29-32)
                                    number_posev = [i for i in number_posev if i >= player_net // 4 * 3 + 1 and i <= player_net  // 8 * 7] # номера 25-28
                        else:  
                            number_tmp = alignment_in_half(player_net, num_tmp, sev, count, number_posev) # номер (а)куда можно сеять
                            number_posev.clear()
                            number_posev = number_tmp.copy()         

                possible_number[reg] = number_posev
                proba_possible[cur_gr] = number_posev
        y += 1
    return possible_number


def alignment_in_half(player_net, num_tmp, sev, count, number_posev):
    """выравнивание количество областей по половинам 
    -num_tmp- номера где уже есть эта область
    -number_tmp- номера куда можно сеять,
    -number_posev- возможные номера посева
    -count- число регионов посеянных """
    number_tmp = [] 
    upper_half = 0
    quarter_num = -1
    su = 0
    sd = 0
    max_num = max(num_tmp)
    min_num = min(num_tmp)
    if count % 2 != 0: # нечетное число регионов
        upper_half = len([i for i in num_tmp if i <= player_net // 2]) # количество областей в верхней половине сетки 1-16
        if upper_half == count: # все области в верху сетки  1-16
            quarter_num = len([i for i in num_tmp if i <= player_net // 4]) # количество областей в верхней четверти сетки 1-8
            sev_tmp = [i for i in sev if i <= player_net // 2] # оставляет номера нижней половины
        elif upper_half == 0: # все области в низу сетки  17-32:
            quarter_num = len([i for i in num_tmp if i <= player_net * 3 / 4]) # количество областей в верхней четверти сетки 17-24
            sev_tmp = [i for i in sev if i > player_net // 2] # оставляет номера нижней половины
        else: # посеянные области в разных половинах
            for t in num_tmp:
                if t > player_net / 2:
                    sd += 1 # в нижней половине
                else:
                    su += 1 # в вверхней половине
            if sd > su: # больше областей в низу
                sev_tmp = [i for i in sev if i <= player_net // 2] # оставляет номера вверхней половины
                num_tmp = [i for i in num_tmp if i <= player_net // 2] # получает номер, который один в половине сетки 
            else: # больше областей в вверху
                sev_tmp = [i for i in sev if i > player_net // 2] # оставляет номера нижней половины
                num_tmp = [i for i in num_tmp if i > player_net // 2] # получает номер, который один в половине сетки 

            for k in num_tmp:  
                if k <= player_net // 4: # если номер в сетке вверху, то наде сеять вниз 1-8 (1-4)
                    np = [i for i in sev_tmp if i > player_net // 4] # 1-я четверть 32(9-16) 16(5-8)
                elif k > player_net // 4 and k <= player_net // 2: # 9-16 (5-8)
                    np = [i for i in sev_tmp if i <= player_net // 4] # 2-я четверть 32(1-8) 16(1-4)
                elif k > player_net // 2  and k <= player_net * 3 / 4: # 17-24 (9-12)
                    np = [i for i in sev_tmp if i > player_net * 3 / 4] # 3-я четверть 32(17-24) 16(9-12)
                elif k > player_net * 3 / 4: 
                    np = [i for i in sev_tmp if i <= player_net * 3 / 4] # 4-я четверть 32(25-32) 16(13-16)
                number_tmp += np

        if quarter_num == -1:
            return number_tmp
        else:
            if quarter_num == 0:
                pass
            elif quarter_num == 1:
                num_tmp = [min_num]
            elif quarter_num == 2:
                num_tmp = [max_num]
            elif quarter_num == count: # все области с 1 по 8
                pass
        number_posev = sev_tmp

    for k in num_tmp:
        if k <= 4: # в первой четверти (1-4)
            np = [i for i in number_posev if i >= 5 and i <= 8]
        elif k >= 5 and k <= 8: # в первой четверти (5-8)
            np = [i for i in number_posev if i >= 1 and i <= 4]
        elif k >= 9 and k <= 12: # в первой четверти (9-12)
            np = [i for i in number_posev if i >= 13 and i <= 16]
        elif k >= 13 and k <= 16: # в первой четверти (13-16)
            np = [i for i in number_posev if i >= 9 and i <= 12]
        elif k >= 17 and k <= 20: # в первой четверти (17-20)
            np = [i for i in number_posev if i >= 21 and i <= 24]
        elif k >= 21 and k <= 24: # в первой четверти (21-24)
            np = [i for i in number_posev if i >= 17 and i <= 20]
        elif k >= 25 and k <= 28: # в первой четверти (25-28)
            np = [i for i in number_posev if i >= 29 and i <= 32]
        elif k >= 29: # в первой четверти (29-32)
            np = [i for i in number_posev if i >= 25 and i <= 28]
        number_tmp += np
    return number_tmp


def number_setka_posev(cur_gr, group_last, reg_last, number_last, n, cur_reg, sev, player_net):
    """промежуточные номера для посева в сетке после ухода от своей группы при выоде из группы больше двух"""
    
    if n == 0:
        if cur_reg in reg_last:
            index = reg_last.index(cur_reg)
            set_number = number_last[index] # номер где уже посеянна такая же область 
            if set_number <= player_net // 2: # если номер в сетке вверху, то наде сеять вниз
                number_posev = [i for i in sev if i > player_net // 2] # отсеивает в списке номера больше 16
            else: 
                number_posev = [i for i in sev if i <= player_net // 2] # отсеивает в списке номера больше 16 
    elif n == 1: # уводит 2-е место от 1-ого в другую половину
        index = group_last.index(cur_gr)
        set_number = number_last[index] # номер где посеянна группа, от которой надо увести 
        if set_number <= player_net // 2: # если номер в сетке вверху, то наде сеять вниз
            number_posev = [i for i in sev if i > player_net // 2] # номера от 17 до 32
        else: 
            number_posev = [i for i in sev if i <= player_net // 2] # номера от 1 до 16 
    elif n > 1: 
        quatro = player_net // 4 # количество номеров в четверти
        if n == 2: # уводит 3-е место от 2-ого в другую четверть
            group_last = group_last[quatro:]
            number_last = number_last[quatro:] # список номеров 2-ого посева
        elif n == 3: # уводит 4-е место от 1-ого в другую четверть
            group_last = group_last[:quatro] 
            number_last = number_last[:quatro] # номера 1 мест в группах
        index = group_last.index(cur_gr)
        set_number = number_last[index] # номер где посеянна группа, во 4-ом посеве от которой надо увести

        if set_number <= player_net // 4: # если номер в сетке вверху, то наде сеять вниз
            number_posev = [i for i in sev if i >= (player_net // 4 + 1) and i < (player_net // 2 + 1)] # номера от 9 до 17
        elif set_number > player_net // 4 and set_number < (player_net // 2 + 1): # если номер в сетке вверху, то наде сеять вниз: 
            number_posev = [i for i in sev if i <= player_net // 4] # номера от 1 до 8 
        elif set_number > player_net // 2 and set_number < (player_net // 4 * 3 + 1): # если номер в сетке вверху, то наде сеять вниз: 
            number_posev = [i for i in sev if i >= (player_net // 4 * 3 + 1)] # номера от 25 до 32   
        elif set_number >= (player_net // 4 * 3 + 1): # если номер в сетке вверху, то наде сеять вниз: 
            number_posev = [i for i in sev if i >= (player_net // 2 + 1) and i < (player_net // 4 * 3 + 1)] # номера от 17 до 24

    return number_posev


def number_setka_posev_last(cur_gr, group_last, number_last, n, player_net):
    """промежуточные номера для посева в сетке
     -number_last- посеянные номера""" 
    if n == 0:
        if cur_gr in group_last:
            index = group_last.index(cur_gr)
            set_number = number_last[index] # номер где уже посеянна такая же область 
            if set_number <= player_net // 2: # если номер в сетке вверху, то надо сеять вниз
                number_posev_old  = [i for i in number_last if i > player_net // 2] # отсеивает в списке номера больше 16
            else: 
                number_posev_old  = [i for i in number_last if i <= player_net // 2] # отсеивает в списке номера больше 16 
    elif n == 1: # уводит 2-е место от 1-ого в другую половину
        index = group_last.index(cur_gr)
        set_number = number_last[index] # номер где посеянна группа, от которой надо увести 
        if set_number <= player_net // 2: # если номер в сетке вверху, то наде сеять вниз
            number_posev_old  = [i for i in number_last if i > player_net // 2] # номера от 17 до 32
        else: 
            number_posev_old  = [i for i in number_last if i <= player_net // 2] # номера от 1 до 16 
    elif n > 1: 
        quatro = player_net // 4 # количество номеров в четверти
        if n == 2: # уводит 3-е место от 2-ого в другую четверть
            group_last = group_last[quatro:] 
            number_last = number_last[quatro:]  
        elif n == 3: # уводит 4-е место от 1-ого в другую четверть
            group_last = group_last[:quatro] 
            number_last = number_last[:quatro]

        index = group_last.index(cur_gr)
        set_number = number_last[index] # номер где посеянна группа, от которой надо увести 
        if set_number <= 8: # если номер в сетке вверху, то наде сеять вниз
            number_posev_old  = [i for i in number_last if i >= 9 and i < 17] # номера от 9 до 17
        elif set_number > 8 and set_number < 17: # если номер в сетке вверху, то наде сеять вниз: 
            number_posev_old  = [i for i in number_last if i <= 8] # номера от 1 до 8 
        elif set_number > 16 and set_number < 25: # если номер в сетке вверху, то наде сеять вниз: 
            number_posev_old  = [i for i in number_last if i >= 25] # номера от 25 до 32   
        elif set_number >= 25: # если номер в сетке вверху, то наде сеять вниз: 
            number_posev_old  = [i for i in number_last if i >= 17 and i < 25] # номера от 17 до 24

    return number_posev_old


def random_generator(posev_tmp):
    """выдает случайное число из предложенного списка"""
    num_set = random.choice(posev_tmp)
    return num_set


def add_delete_region_group(key_reg_current, current_region_group, posev_tmp, m, posev, start, end, step, player_current):
    """при добавлении в группу региона удалении номера группы из списка сеянных -b- номер группы
    -m- номер посева, kol_group_free - словарь регион и кол-во свободных групп"""
    free_list = []
    reg_list = []
    kol_group_free = {}
    reg_player = dict.fromkeys(player_current, 0)
    player_list = player_current.copy()
    sv = 0
    if start == 0:
        end = len(key_reg_current)
    else:
        start = len(key_reg_current)

    for s in range(start, end, step):
        sv += 1
        # === новый вариант посева ===
        for i in key_reg_current:  # получение словаря (регион и кол-во мест (групп) куда можно сеять)
            kol_reg = len(current_region_group[i]) # колво регионов (посевов)
            kol_group_free[i] = kol_reg
 
        free_list = list(kol_group_free.values())  # список кол-во свободных групп, куда можно сеять
        reg_list = list(kol_group_free.keys())  # список ключей (регионов)

        last = len(reg_list)  # кол-во остатка посева
        region = key_reg_current[0]
        free_gr = kol_group_free[region]  # кол-во групп куда можно сеять
        # ==== сделать последний посев по наименшему количеству вариантов посева
 
        if 1 in free_list and last > 1 or last == 1 and free_gr == 1 :  # проверка есть ли группа где осталось только одно места для посева
            # сделать посев 1 регион но много групп
            ind = free_list.index(1) # получаем индекс в списке ключа -1-
            region = reg_list[ind]
            u = current_region_group[region][0]  # номер группы 
            values = posev_tmp[u] 
            if values == 0:
                posev_tmp[u] = region  # запись региона в группу (посев)
        else:
            f = current_region_group[region]  # список номеров групп для посева текущего региона
            temp_list = []
            if free_gr != 1:
                if len(f) == 0: # значит во всех группах есть уже данный регион
                    posev[f"{m}_посев"] = posev_tmp
                    for p in posev_tmp.keys():
                        if 0 == posev_tmp[p]:
                            temp_list.append(p)
                    # # =====
                    # finish = end if end > start else start
                    # # =====
                    # for i in range(1, finish + 1):
                    #     temp_list.append(i)
                    if m % 2 != 0:  # в зависимости от четности посева меняет направления посева групп в списке
                        temp_list.sort()
                    else:
                        temp_list.sort(reverse = True)   
                    current_region_group[region] = temp_list
                    f = current_region_group[region]  # список номеров групп для посева текущего региона
                    # ===== удалить посеянные группы ====
                    # posev[f"{m}_посев"] = posev_tmp
                    # num_gr_posev = []
                    # for p in posev_tmp.keys():
                    #     if 0 != posev_tmp[p]:
                    #         num_gr_posev.append(p)
                    # for d in num_gr_posev:  # цикл удаления посеянных групп                        
                    #     list_group = current_region_group[region]
                    #     if d in list_group:  # находит сеяную группу и удаляет ее из списка групп
                    #         list_group.remove(d)

                if s in f: #  присваивает переменной u - номер группы, если она идет по порядку
                    posev_tmp[s] = region
                    u = s 
                else:  # присваивает переменной u - номер группы, если она идет не по порядку
                    g = f[0] # номер группы
                    posev_tmp[g] = region
                    u = g   
            elif free_gr == 0:                
                for i in range (1, len(posev) + 1):
                    gr_dict = posev[f"{m}_посев"]
                    gr = gr_dict[i]
                    if gr == 0:
                        temp_list.append(i)
                current_region_group[region] = temp_list
                f = current_region_group[region]                      
            if s in f:  #  присваивает переменной u - номер группы, если она идет по порядку
                posev_tmp[s] = region
                u = s
            else:  # присваивает переменной u - номер группы, если она идет не по порядку
                g = f[0]
                posev_tmp[g] = region
                u = g    
        # ====не правильное соответствие номера региона и номера группы
        index = key_reg_current.index(region)
        p = player_list[index]
        reg_player[p] = u
        #=====================
        posev[f"{m}_посев"] = posev_tmp
        for d in key_reg_current:  # цикл удаления посеянных групп
            list_group = current_region_group[d]
            if u in list_group:  # находит посеянную группу и удаляет ее из списка групп
                list_group.remove(u)
        player_list.remove(p)  # список id игроков посева (и удаляет игрока -P-, как посеянный)     
        key_reg_current.remove(region)  # удаляет регион из списка как посеянный
        count_in_list = key_reg_current.count(region)
        if count_in_list == 0:  # если в посеве больше одного региона, то пропускает удаление из словаря
            del current_region_group[region] 
            del kol_group_free[region]

        if start > end:
            start -= 1
        else:
            start += 1 
    choice_save(m, player_current, reg_player)        
    return sv


def choice_save(m, player_current, reg_player):
    """запись в db результаты жеребьевки конкретного посева"""
    for i in player_current:
        num_group = reg_player[i]
        with db:  # запись в таблицу Choice результата жеребъевки
            choice = Choice.get(Choice.player_choice_id == i)
            choice.group = f"{num_group} группа"
            choice.posev_group = m
            choice.save()


def region_player_current(number_poseva, reg_list, group, player_list):
    """ создание списка номеров регионов в порядке посева для текущего номера посева """
    key_reg_current = []
    key_tmp = []
    player_current = []
    pl_tmp = []
    current = []
    r = 0
    p = 0
    start = number_poseva
    count = len(player_list)  # кол-во игрок
    remains = count - number_poseva  # остаток посева
    if remains > group: 
        end = start + group  # если остаток больше кол-во групп
    else:
        end = start + remains
    for k in range(start, end):
        r = reg_list[k]
        key_tmp.append(r)
        p = player_list[k]
        pl_tmp.append(p)

    key_reg_current = key_tmp.copy()
    player_current = pl_tmp.copy()
    key_tmp.clear()
    pl_tmp.clear()
    current.append(key_reg_current)
    current.append(player_current)
    return current


def posev_test(posev, group, m):
    """возвращает словарь предыдущих посевов регион - группы, где они есть, m - номер посева"""
    uniq_region = []  # уникальный список регионов которые уже посеяны
    tmp_posev = {}
    previous_region_group = {} 
    gr = [] 
    gr_tmp = []
    # список регионов данного посева
    for p in range(1, m):
        tmp_posev = posev[f"{p}_посев"]
        for a in range(1, group + 1):
            v = tmp_posev.setdefault(a)
            if v not in uniq_region:
                uniq_region.append(v)
    # уникальный список регионов
    for val in uniq_region:  # цикл получения словаря (номер региона - список групп где они уже есть)
        for d in range(1, m):
            for key, value in posev[f"{d}_посев"].items():
                if val == value:
                    gr_tmp.append(key) # добавляет в список номер групп для региона val
        gr = gr_tmp.copy()
        gr_set = set(gr)
        gr = list(gr_set)
        previous_region_group[val] = gr
        gr_tmp.clear()
    return previous_region_group


def choice_setka(fin):
    """жеребьевки сетки"""
    system = System.select().where((System.title_id == title_id()) & (System.stage == fin)).get()# находит system id последнего

    flag = system.choice_flag
    if flag is True:  # перед повторной жеребьевкой
        del_choice = Game_list.select().where((Game_list.title_id == title_id()) & (Game_list.number_group == fin))
        for i in del_choice:
            i.delete_instance()  # удаляет строки финала (fin) из таблицы -Game_list
        
        del_result = Result.select().where((Result.title_id == title_id()) & (Result.number_group == fin))
        for i in del_result:
            i.delete_instance()  # удаляет строки финала (fin) из таблицы -Result-

        with db:  # запись в таблицу Choice результата жеребъевки
            system.choice_flag = False
            system.save()


def select_stage_for_edit():
    """выбор финалов или номеров групп для редактирования игроков """
    group_list = []
    stage = ""
    title = ""
    sender = my_win.sender()
    etap_list = ["Предварительный", "1-й полуфинал", "2-й полуфинал"]
    if sender == my_win.comboBox_edit_etap1:
        my_win.comboBox_first_group.clear()
        my_win.lineEdit_change_pl1.clear()
        my_win.lineEdit_change_pl1_2.clear()
    else:
        my_win.comboBox_second_group.clear()
        my_win.lineEdit_change_pl2
        my_win.lineEdit_change_pl2_2.clear()

    systems = System.select().where(System.title_id == title_id())
    players = Player.select().where(Player.title_id == title_id())
    if sender == my_win.comboBox_edit_etap1:
        index = my_win.comboBox_edit_etap1.currentIndex()
        stage = my_win.comboBox_edit_etap1.currentText()
    elif sender ==  my_win.comboBox_edit_etap2:
        index = my_win.comboBox_edit_etap2.currentIndex()
        stage = my_win.comboBox_edit_etap2.currentText()
    else:
        stage = "-Выбор этапа-"

    if index == 0:
       return
    elif index == 1:
        for k in players:
            player_full = k.full_name
            # f_name = k.player
            # region = k.region
            # id_coach = k.coach_id
            # coaches = Coach.select().where(Coach.id == id_coach).get()
            # coach_fam = coaches.coach
            # player_full = f"{f_name}/{region}/{coach_fam}"
            group_list.append(player_full)
        group_list.sort()
        title = "-Выбор спортсменов-"
    elif stage == "Одна таблица":
        pass
    elif stage in etap_list:
        sys_id = systems.select().where(System.stage == stage).get()
        group = sys_id.total_group
        group_list = [f"{i} группа" for i in range(1, group + 1)] # генератор списка
        title = "-Выбор группы-"
    elif stage not in etap_list:
        etap_fin_list = []
        for s in systems:
            etap = s.stage
            if etap not in etap_list:
                system_choice = systems.select().where(System.stage == etap).get()
                flag = system_choice.choice_flag
                if flag is True:
                    etap_fin_list.append(etap)

        group_list = [i for i in etap_fin_list] # генератор списка
        title = "-Выбор финала-"
    group_list.insert(0, title)
    if sender == my_win.comboBox_edit_etap1:
        my_win.comboBox_first_group.addItems(group_list)
        my_win.comboBox_first_group.setEnabled(True)
    else:
        my_win.comboBox_second_group.addItems(group_list)
        my_win.comboBox_second_group.setEnabled(True)
    if my_win.comboBox_edit_etap1.currentIndex() != 0 and my_win.comboBox_edit_etap2.currentIndex() != 0:
        my_win.Button_change_player.setEnabled(True)


def edit_group_after_draw():
    """редактирование групп после жеребьевки"""
    if my_win.comboBox_edit_etap1.currentText() == "-Выбор этапа-":
        return
    else:
        stage = my_win.comboBox_edit_etap1.currentText()
    if my_win.comboBox_edit_etap2.currentText() == "-Выбор этапа-":
        return
    else:
        stage = my_win.comboBox_edit_etap2.currentText()
    my_win.tableView.setVisible(False)
    my_win.comboBox_first_group.clear()
    my_win.comboBox_second_group.clear()
    system = System.select().where(System.title_id == title_id())
    system_group = system.select().where(System.stage == stage).get()

    players = Player.select().where(Player.title_id == title_id())
    total_gr = system_group.total_group
    group = [f"{i} группа" for i in range(1, total_gr + 1)] # генератор списка
    group.insert(0, "-выберите группу-")   
    my_win.comboBox_first_group.addItems(group)
    my_win.comboBox_second_group.addItems(group)
    player = [k.full_name for k in players]
    player.sort()
    my_win.comboBox_player_group_edit.addItems(player)


def add_item_listwidget():
    """добавление элементов в листвиджет"""
    flag_combo = 0
    flag_fin = 0
    sender = my_win.sender()
    choices = Choice.select().where(Choice.title_id == title_id())
    my_win.tableView.setVisible(False)
    coach_list = []
    coach = ""
    if sender == my_win.comboBox_first_group:
        my_win.listWidget_first_group.clear()
        gr = my_win.comboBox_first_group.currentText()
    else:
        my_win.listWidget_second_group.clear()
        gr = my_win.comboBox_second_group.currentText()

    if gr == "":
        return
    else:
        if sender == my_win.comboBox_first_group:
            if my_win.comboBox_edit_etap1.currentText() == "Предварительный":
                group = choices.select().where(Choice.group == gr).order_by(Choice.posev_group)
            elif my_win.comboBox_edit_etap1.currentText() == "1-й полуфинал":
                group = choices.select().where(Choice.sf_group == gr).order_by(Choice.posev_sf)
            else: # финалы
                flag_fin = 1
                final = my_win.comboBox_first_group.currentText()
                group = choices.select().where(Choice.final == final).order_by(Choice.posev_final)
               
        else:
            if my_win.comboBox_edit_etap2.currentText() == "Предварительный":
                group = choices.select().where(Choice.group == gr).order_by(Choice.posev_group)
            elif my_win.comboBox_edit_etap2.currentText() == "1-й полуфинал":
                group = choices.select().where(Choice.sf_group == gr).order_by(Choice.posev_sf) 
            else: # финалы
                flag_fin = 1
                final = my_win.comboBox_second_group.currentText()
                group = choices.select().where(Choice.final == final).order_by(Choice.posev_final)

        n = 0
        for k in group:
            item = QListWidgetItem()
            if flag_fin == 1:
                n = k.posev_final # нумерация посева в сетке
            else:
                n += 1
            family = k.family
            region = k.region
            coach = k.coach
            text = f"{n}:{family}/{region}/{coach}"
            item.setText(text) 
            if sender == my_win.comboBox_first_group:
                my_win.listWidget_first_group.addItem(item)
            else:
                my_win.listWidget_second_group.addItem(item)
            coach_list.append(coach)
   
        # duplicat = duplicat_coach_in_group(coach_list)
        # if duplicat is not None:
        #     color_coach_in_listwidget(duplicat, flag_combo)
        # color_coach_in_tablewidget(duplicat, coach_list)


def color_coach_in_listwidget(duplicat, flag_combo):
    """отмечает строки с повторяющимися тренерами"""
    if flag_combo == 1:
        item = my_win.listWidget_first_group.item
        count = my_win.listWidget_first_group.count()
    else:
        item = my_win.listWidget_second_group.item
        count = my_win.listWidget_second_group.count()
    for row in range(count):
        find_coach = []
        data_lw = item(row).text()
        mark = data_lw.rfind("/")
        coach_in_row = data_lw[mark + 1:]
        find_mark_1 = coach_in_row.find(",")
        if find_mark_1 != -1:
            coach_first = coach_in_row[:find_mark_1]
            find_mark_2 = coach_in_row.find(",", find_mark_1 + 1)
            find_coach.append(coach_first)
            if find_mark_2 == -1:
                coach_second = coach_in_row[find_mark_1 + 2:]
                find_coach.append(coach_second)
            else:
                coach_second = coach_in_row[find_mark_1 + 2:find_mark_2]
                coach_third = coach_in_row[find_mark_2 + 2:]
                find_coach.append(coach_second)
                find_coach.append(coach_third)
        else:
            find_coach.append(coach_in_row)
        for k in duplicat:
            if k in find_coach:
                item(row).setForeground(QColor(0, 0, 255)) # изменяет весь текст на синий


def list_player_in_group_after_draw():
    """Смена игроков в группах после жеребьевки при отметки в listwidget при редакитровании"""
    sender = my_win.sender()

    if sender == my_win.Button_add_pl1:
        index_1 = my_win.comboBox_edit_etap2.currentIndex()
        if index_1 == 1:
            pl1_fam = my_win.comboBox_first_group.currentText()
            my_win.lineEdit_change_pl1.setText(pl1_fam)
        else:
            for row in range(my_win.listWidget_first_group.count()):
                select_item = my_win.listWidget_first_group.selectedItems()
            for i in select_item:
                player_first = i.text()
                if my_win.lineEdit_change_pl1.text() == "":
                    my_win.lineEdit_change_pl1.setText(player_first)
                else:
                    my_win.lineEdit_change_pl1_2.setText(player_first)
    else:
        index_2 = my_win.comboBox_edit_etap2.currentIndex()
        if index_2 == 1:
            pl2_fam = my_win.comboBox_second_group.currentText()
            my_win.lineEdit_change_pl2.setText(pl2_fam)
        else:
            for row in range(my_win.listWidget_second_group.count()):
                select_item = my_win.listWidget_second_group.selectedItems()
            for i in select_item:
                player_second = i.text()
                if my_win.lineEdit_change_pl2.text() == "":
                    my_win.lineEdit_change_pl2.setText(player_second)
                else:
                    my_win.lineEdit_change_pl2_2.setText(player_second)


def change_player_between_group_after_draw():
    """Смена игроков в группах после жеребьевки при отметки в listwidget при редакитровании"""
    msgBox = QMessageBox
    flag_change = 0
    player_dict = {}
    player_new = 0 # если игрок добавляется после жеребьевке то 1
    game_list = Game_list.select().where(Game_list.title_id == title_id())
    choices = Choice.select().where(Choice.title_id == title_id())
    systems = System.select().where(System.title_id == title_id())
    players = Player.select().where(Player.title_id == title_id())

    etap_1 = my_win.comboBox_edit_etap1.currentText()
    etap_2 = my_win.comboBox_edit_etap2.currentText()
    player1 = my_win.lineEdit_change_pl1.text()
    player2 = my_win.lineEdit_change_pl2.text()
    player1_2 = my_win.lineEdit_change_pl1_2.text() # 2-й игрок из группы для смены в ПФ
    player2_2 = my_win.lineEdit_change_pl2_2.text() # 2-й игрок из группы для смены в ПФ
    gr_pl1 = my_win.comboBox_first_group.currentText() # номер группы (финал)
    gr_pl2 = my_win.comboBox_second_group.currentText() # номер группы (финал)
    group_list = [gr_pl1, gr_pl2]
    # etap_list = [etap_1, etap_2]
    posev_list = []
    player_list = [player1, player2, player1_2, player2_2]
    #  === получаем full_name для определения id игроков
    full_name_list = []
    for pl in player_list:
        if pl != "":
            znak = pl.find(":")
            znak1 = pl.find("/") 
            znak2 = pl.rfind("/")
            family_name = pl[znak + 1:znak1]
            # posev_number = pl[:znak]
            if znak == -1:
                player_id = players.select().where(Player.full_name == pl).get()               
            else:
                region = pl[znak1 + 1:znak2]
                player_id = players.select().where((Player.player == family_name) & (Player.region == region)).get()
            pl_id = player_id.id
            full_name = player_id.full_name
            full_name_list.append(full_name)
            player_dict[full_name] = pl_id
    # подсчитывает колличество не пустых значений (кол-во участников)
    element_count = len([item for item in player_list if item != ""]) # подсчитывает колличество не пустых значений (кол-во участников) участвующих в редактировании
    # колличество игроков в группе если добавляется игрок из списка (который не прожеребьен)
    if element_count  == 0:
        result = msgBox.information(my_win, "Уведомление", "Вы не выбрали игроков группы!", msgBox.Ok)
        return
    elif element_count == 1: 
        if etap_1 == "Списки участников" or etap_2 == "Списки участников": # добавляет игрока из списка участников в группу
            stage = etap_1 if etap_1 == "Предварительный" else etap_2
            gr = gr_pl1 if etap_1 != "Списки участников" else gr_pl2
            # колличество игроков в группе если добавляется игрок из списка (который не прожеребьен)
            count_in_group = my_win.listWidget_first_group.count() if my_win.listWidget_first_group.count() != 0 else my_win.listWidget_second_group.count()
            # все данные игрока, которого не было в жерербьевке
            for family_pl in full_name_list:
                if family_pl != "":
                    id_pl = player_dict[family_pl]
                    players_data = players.select().where(Player.id == id_pl).get()
                    region = players_data.region
                    coach_id = players_data.coach_id
                    coachs = Coach.select().where(Coach.id == coach_id).get()
                    coach_family = coachs.coach
                    rank = players_data.rank
            player_new = 1
        else: # перемещает игрока в другую группу
            stage = etap_1 if etap_1 == "Предварительный" else etap_2
            
            if my_win.listWidget_first_group.count() < my_win.listWidget_second_group.count():
                count_in_group = my_win.listWidget_first_group.count() 
                gr = gr_pl1
            else:
                count_in_group = my_win.listWidget_second_group.count()
                gr = gr_pl2
        id_system = system_id(stage)
        # system = systems.select().where(System.stage == stage).get()
        # system_etap_id = system.id # id этапа

        posev, ok = QInputDialog.getInt(my_win, "Номер посева", "Введите номер посева", min=1, max=(count_in_group + 1))
        if not ok:
            return
        else:
            if posev <= count_in_group: # если пытаются заменить игрока в группе
                result = msgBox.question(my_win, "Уведомление", "Вы хотите заменить игрока группы\n"
                f"{posev} посева?", msgBox.No, msgBox.Ok) 
                if result == msgBox.No:
                    return
            elif posev > count_in_group and player_new == 0:  #=== заменяем спортсмена в группу на последний посев и  обновляет Choice (хотя он потом встает в группе по R)  
                query = Choice.update(group=gr, posev_group=posev).where(Choice.player_choice_id == pl_id) # обновляет запись в Choice                 
                query.execute()
            else: # == если добавляют игрока в конец группы                   
                with db:
                    game_list = Game_list(number_group=gr, 
                                        rank_num_player=posev, 
                                        player_group_id=id_pl,
                                        system_id=id_system, 
                                        title_id=title_id()
                                        ).save()
                    # если новый игрок, которого не было в жеребьевке
                    choice = Choice(player_choice_id=id_pl,
                                    family=family_name,
                                    region=region,
                                    coach=coach_family,
                                    rank=rank,
                                    group=gr,
                                    posev_group=posev,
                                    title_id=title_id()
                                    ).save()  
    elif element_count == 2: # меняют местами игроков (в группе, ПФ, финале)
        if etap_1 == etap_2: # оба игрока из одного этапа соревнования
            stage = etap_1
            for pl in player_list:
                if pl != "":
                    znak = pl.find(":")
                    posev = int(pl[:znak]) 
                    posev_list.append(posev)
                    # break
            n = 0      
            for pl in full_name_list:
                pl_id = player_dict[pl]
                gr = group_list[1 - n]
                posev = posev_list[1 - n]
                if stage == "Предварительный":
                    query = Choice.update(group=gr, posev_group=posev).where(Choice.player_choice_id == pl_id) # обновляет запись в Choice  
                elif stage == "1-й полуфинал" or  stage == "2-й полуфинал":
                    query = Choice.update(semi_final=stage, sf_group=gr, posev_sf=posev).where(Choice.player_choice_id == pl_id) 
                elif stage == "Финальный":
                    flag_change = 1
                    query = Choice.update(final=gr, posev_final=posev).where(Choice.player_choice_id == pl_id) # обновляет запись в Choice                    
                query.execute()
                n += 1
        else: # игроки из разных этапов соревнования для исправления ошибок
            for pl in full_name_list:
                pl_id = player_dict[pl]
                choice = choices.select().where(Choice.player_choice_id == pl_id).get()
                if etap_1 == "Финальный":
                    stage = my_win.comboBox_first_group.currentText()
                    systems_1 = systems.select().where(System.stage == stage).get()
                    system_id_1 = systems_1.id 
                elif etap_2 == "Финальный":
                    stage = my_win.comboBox_second_group.currentText()
                    systems_2 = systems.select().where(System.stage == stage).get()
                    system_id_2 = systems_2.id       
    else:
        # ====== если меняет в полуфинале группы (менять результат) ======
        if element_count == 4: # если присутствуют 2-е игроки для обмена (ПФ смена регионов)
            stage = etap_1
            results = Result.select().where((Result.title_id == title_id()) & (Result.system_stage == stage))
            players = Player.select().where(Player.title_id == title_id())

            for k in range(0, 2): # перезаписывает таблицу Result
                result1 = results.select().where(Result.player1 == full_name_list[k])
                result2 = result1.select().where(Result.player2 == full_name_list[k + 2]).get()
                with db:
                    result2.number_group = group_list[1 - k]
                    result2.save()
            n = 0      
            for pl in full_name_list:
                pl_id = player_dict[pl]
                gr = group_list[1 - n]
                if stage == "1-й полуфинал" or  stage == "2-й полуфинал":
                    query = Choice.update(sf_group=gr).where(Choice.player_choice_id == pl_id) 
                query.execute()
                n += 1
        # =====================
    if flag_change == 1:
        player_in_setka_and_write_Game_list_Result(gr, posev_list, full_name_list)
    else:
        player_in_table_group_and_write_Game_list_Result(stage) 

    if element_count == 4:   
        load_playing_game_in_table_for_semifinal(stage)  
        # ========
    my_win.lineEdit_change_pl1.clear()
    my_win.lineEdit_change_pl2.clear()
    my_win.lineEdit_change_pl1_2.clear()
    my_win.lineEdit_change_pl2_2.clear()
    my_win.comboBox_first_group.setCurrentText("-выберите группу-")
    my_win.comboBox_second_group.setCurrentText("-выберите группу-")
    my_win.listWidget_first_group.clear()
    my_win.listWidget_second_group.clear()
    my_win.comboBox_edit_etap1.setCurrentIndex(0)
    my_win.comboBox_edit_etap2.setCurrentIndex(0)
    if stage == "Предварительный":
        my_win.tabWidget.setCurrentIndex(3)
    elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
         my_win.tabWidget.setCurrentIndex(4)            
    my_win.tableView.setVisible(True)

# def add_player_to_group():
#     """добавление игрока в группу при редактировании"""
#     player_choice_tmp = []
#     n_group = my_win.comboBox_number_group_edit.currentText()
#     player_gr = my_win.comboBox_player_group_edit.currentText()
#     edit_group_after_draw()


def choice_tbl_made():
    """создание таблицы жеребьевка, заполняет db списком участников для жеребьевки"""
    players = Player.select().order_by(Player.rank.desc()).where((Player.title_id == title_id()) & (Player.bday != '0000-00-00'))
    choice = Choice.select().where(Choice.title_id == title_id())
    if len(choice) != 0:
        for i in choice:
            ch_d = Choice.get(Choice.id == i)
            ch_d.delete_instance()
    for i in players:
        family = i.player
        region = i.region
        rank = i.rank
        coach_id = i.coach_id
        coachs =Coach.select().where(Coach.id == coach_id).get()
        coach = coachs.coach
        chc = Choice(player_choice=i, family=family, region=region, coach=coach, rank=rank,
                    title_id=title_id()).save()


def filter_player_on_system():
    """Фильтрует игроков на вкладке системы по группам ПФ или финалам"""
    choice = Choice.select().where(Choice.title_id == title_id())
    number_group = my_win.comboBox_filter_number_group_final.currentText()
    stage = my_win.comboBox_filter_choice_stage.currentText()
    if number_group == "все группы" or number_group == "все финалы":
        return
    if  stage == "Предварительный":
        player_list = choice.select().where(Choice.group == number_group)
    elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
        player_list = choice.select().where((Choice.semi_final == stage) & (Choice.sf_group == number_group)).order_by(Choice.sf_group)
    else:
        player_list = choice.select().where(Choice.final == number_group)
    fill_table(player_list)


def choice_filter_on_system():
    """фильтрует таблицу жеребьевка по группам, полуфиналам или финалам"""
    systems = System.select().where(System.title_id == title_id())
    group_list = ["Предварительный", "1-й полуфинал", "2-й полуфинал"]
    stage = my_win.comboBox_filter_choice_stage.currentText()
    number_group = my_win.comboBox_filter_number_group_final.currentText()
    if stage != "-выберите этап-" and stage != "":
        my_win.comboBox_filter_number_group_final.setEnabled(True)
    else:
        my_win.comboBox_filter_number_group_final.setEnabled(False)
    if stage == "-выберите этап-" and number_group == "все группы" and number_group == "все финалы":
        return
    else:
        if my_win.checkBox_repeat_regions.isChecked():
            return
        if stage == "":
            my_win.comboBox_filter_choice_stage.clear()
            etaps_list = [i.stage for i in systems if i.choice_flag is True] # все этапы системы
            etaps_set = set(etaps_list)
            groups_set = set(group_list)
            etaps_set.intersection_update(groups_set)
            etaps = list(etaps_set)
            etaps.sort(reverse = True)
            etaps.append("Финальный")
            etaps.insert(0, "-выберите этап-")
            my_win.comboBox_filter_choice_stage.addItems(etaps)
        elif stage in group_list:
            id_system = system_id(stage)
            my_win.comboBox_filter_number_group_final.clear()
            systems_sf = systems.select().where(System.id == id_system).get()
            kg = int(systems_sf.total_group)  # количество групп
            gr_txt = [f"{i} группа" for i in range(1, kg + 1)]
            gr_txt.insert(0, "все группы")
            my_win.comboBox_filter_number_group_final.addItems(gr_txt)
        else:
            my_win.comboBox_filter_number_group_final.clear()
            systems_sf = systems.select().where(System.choice_flag == 1)
            gr_txt = [i.stage for i in systems_sf if i.stage not in group_list]
            gr_txt.insert(0, "все финалы")
            my_win.comboBox_filter_number_group_final.addItems(gr_txt)
           
        player_list = Choice.select().where(Choice.title_id == title_id())
        fill_table(player_list)
 
 
    # duplicat = duplicat_coach_in_group(coach_list)
    # color_coach_in_tablewidget(duplicat, coach_list)


def duplicat_coach_in_group(coach_list):
    """поиск совпадения тренеров в одной группе"""
    tmp_list = []
    count = len(coach_list)
    for i in coach_list:
        znak = i.find(",")
        if znak == -1: # один тренер
            tmp_list.append(i)
        else:
            coach_1 = i[:znak]
            tmp_list.append(coach_1)
            if i.find(",", znak) == -1:
                znak_1 = i.find(",", znak + 1)
                coach_2 = i[znak: znak_1]
                tmp_list.append(coach_2)
            else:
                coach_2 = i[znak + 2:]
                znak_1 = i.find(",", znak + 1)
                if i.find(",", znak_1) == -1:
                    tmp_list.append(coach_2)
                else:
                    coach_2 = i[znak + 2:znak_1]
                    tmp_list.append(coach_2)
                    coach_3 = i[znak_1 + 2:]
                    tmp_list.append(coach_3)
    count_list = len(tmp_list)
    count_uniq = len(set(tmp_list)) 
    if count_list > count_uniq:
        duplicat = [x for i, x in enumerate(tmp_list) if i != tmp_list.index(x)]
        return duplicat


def load_coach_to_combo():
    """загружает список тренеров в комбобокс для фильтра на странце участники"""
    coach_list = []
    my_win.comboBox_fltr_coach.clear()
    players = Player.select().where(Player.title_id == title_id())
    for k in players:
        coachs_id = k.coach_id
        coachs = Coach.select().where(Coach.id == coachs_id).get()
        coach_family = coachs.coach
        if coach_family not in coach_list:
            coach_list.append(coach_family)
    # =====
    tmp_list = []
    for i in coach_list:
        znak = i.find(",")
        if znak == -1: # один тренер
            if i not in tmp_list:
                tmp_list.append(i)
        else:
            coach_1 = i[:znak]
            if coach_1 not in tmp_list:
                tmp_list.append(coach_1)
            znak_1 = i.find(",", znak + 1)
            if znak_1 == -1:
                coach_2 = i[znak + 2:]
                if coach_2 not in tmp_list:
                    coach_2.rstrip()
                    tmp_list.append(coach_2)
            else: # три тренера
                coach_2 = i[znak + 2: znak_1]
                if coach_2 not in tmp_list:
                    tmp_list.append(coach_2)
                coach_3 = i[znak_1 + 2:]   
                if coach_3 not in tmp_list:
                    coach_3.rstrip()
                    tmp_list.append(coach_3)
    tmp_list.sort() 
    tmp_list.insert(0, "")
    my_win.comboBox_fltr_coach.addItems(tmp_list)            
    # count_list = len(tmp_list)
    # count_uniq = len(set(tmp_list)) 
    # if count_list > count_uniq:
    #     duplicat = [x for i, x in enumerate(tmp_list) if i != tmp_list.index(x)]
    #     return duplicat


def color_region_in_tableWidget(fg):
    """смена цвета шрифта в QtableWidget -fg- номер группы"""
    reg = []
    rid = []

    if fg != "все группы" and fg != "":
        choice = Choice.select().where(Choice.title_id == title_id())
        line = choice.select().where(Choice.group == fg)
        for i in line:
            region = i.region
            region = str(region.rstrip())  # удаляет пробел в конце строки
            reg.append(region)
        if len(reg) != 0:
            for x in reg:
                count_region = reg.count(x)
                if count_region > 1:  # если повторяющихся регионов больше одного
                    p = 0
                    for m in range(len(reg)):
                        ind = reg.index(x, p)
                        p = ind + 1
                        rid.append(ind)                       
                        if m == count_region - 1:
                            break
            rid = list(set(rid))  # получает список индексов повторяющихся регионов
            rows = my_win.tableWidget.rowCount()  # кол-во строк в отсортированной таблице
            if rows != 0:
                for k in rid:
                    my_win.tableWidget.item(k, 3).setForeground(QBrush(QColor(255, 0, 0)))  # окрашивает текст в красный цвет


def hide_show_columns(tb):
    """скрывает или показывает столбцы TableView"""
    for k in range(0, 19):
        my_win.tableView.hideColumn(k)
    if tb == 0: # титул
        pass
    elif tb == 1: # участники
        my_win.tableView.showColumn(1) # фамилия имя
        my_win.tableView.showColumn(2) # др
        my_win.tableView.showColumn(3) # рейтинг
        my_win.tableView.showColumn(4) # город
        my_win.tableView.showColumn(5) # регион
        my_win.tableView.showColumn(6) # разряд
        my_win.tableView.showColumn(7) # тренеры
        my_win.tableView.showColumn(8) # место
        # my_win.tableView.showColumn(16) # заявка
    elif tb == 2: # система
        my_win.tableView.showColumn(1) # фамилия имя
        my_win.tableView.showColumn(2) # регион
        my_win.tableView.showColumn(3) # предварительный
        my_win.tableView.showColumn(4) # место группы
        my_win.tableView.showColumn(5) # пф
        my_win.tableView.showColumn(6) # тренеры
        my_win.tableView.showColumn(7) # место
        my_win.tableView.showColumn(8) # место в группе

        my_win.tableView.showColumn(9) # тренеры
        my_win.tableView.showColumn(10) # место
        my_win.tableView.showColumn(11) # место в группе
    elif tb == 3:
        my_win.tableView.showColumn(2)
        my_win.tableView.showColumn(3) # регион
        my_win.tableView.showColumn(4) # игрок 1
        my_win.tableView.showColumn(5) # игрок 2
        my_win.tableView.showColumn(6) # победитель        
        my_win.tableView.showColumn(8) # общий счет
        my_win.tableView.showColumn(9) # счет в партиях
    elif tb == 6:
        # my_win.tableView.showColumn(0)
        my_win.tableView.showColumn(1)
        my_win.tableView.showColumn(2)
        my_win.tableView.showColumn(3)
        my_win.tableView.showColumn(4)
        my_win.tableView.showColumn(5)
        my_win.tableView.showColumn(6)
    # elif tb == 7:
        # my_win.tableWidget.showColumn(0)
        # my_win.tableWidget.showColumn(1)


def etap_made(stage):
    """создание этапов соревнований"""
    titles = Title.select().where(Title.id == title_id()).get()
    id_title = titles.id
    system = System.select().where(System.title_id == title_id())
    for l in system:
        total_athletes = l.total_athletes
        break
    sum_game = []
    etap = my_win.comboBox_etap.currentText() if stage == "" or stage is False else stage
    if etap == "Одна таблица":
        fin = my_win.comboBox_etap.currentText()
        one_table(fin, group=1)
        gamer = my_win.lineEdit_title_gamer.text()
        tab_enabled(id_title)
        return
    elif etap == "Предварительный":    
        kol_player_in_group() # кол-во участников в группах
    elif etap == "Финальный":
        systems = system.select().order_by(System.id.desc()).get()
        stage = systems.stage
        total_game_table(exit_stage="", kpt=0, fin=stage, pv="") # сколько игр в финале или пф 
        kol_game = my_win.spinBox.text()
        state_visible = my_win.checkBox_visible_game.isChecked() # записывает в DB измененный статус видимости
        with db:
            systems.score_flag = kol_game
            systems.visible_game = state_visible
            systems.save()
        # суммирует все игры этапов  
    group_list = ["Предварительный", "1-й полуфинал", "2-й полуфинал"]  
    player_in_final = []
    # ====
    systems = System.select().where(System.title_id == id_title)
    for k in systems:
        stage = k.stage
        pl_final = k.max_player
        if stage not in group_list:
            player_in_final.append(pl_final)
        kol_game_str = k.kol_game_string
        zn = kol_game_str.find(" ")
        number = int(kol_game_str[:zn]) # переводит строку кол-во игр в числа и записывает в списки
        sum_game.append(number)
    all_sum_game = sum(sum_game) # всего игр в турнире
    # встаивить число игроков в последнем финале (он может быть не полным) и заменить число в списке all_sum_player_final
    # if stage == "Суперфинал":
    #     player_in_final.pop()
    # player_in_final.pop()
    # sum_pl_whithout_last_final = sum(player_in_final)
    # player_last_final = total_athletes - sum_pl_whithout_last_final
    # player_in_final.append(player_last_final)
    # =====================
    all_sum_player_final = sum(player_in_final) # кол-во игроков в финалах
    my_win.label_33.setText(f"Всего:{all_sum_game} игр.")
    my_win.label_52.setText(f"Посеяно {all_sum_player_final} чел.")
    my_win.checkBox_visible_game.setChecked(True)
    flag = control_all_player_in_final(etap, all_sum_player_final) # проверяет все ли игроки распределены по финалам
    if flag is True: # продолжает выбор этапа
        made_system_load_combobox_etap()
    my_win.Button_etap_made.setEnabled(False)


def total_game_table(exit_stage, kpt, fin, pv):
    """количество участников и кол-во игр"""
    msgBox = QMessageBox()
    sender = my_win.sender()
    sum_player = [0]
    no_game3 = ""
    etap_text = my_win.comboBox_etap.currentText()
    flag_visible = my_win.checkBox_visible_game.isChecked()
    system = System.select().where(System.title_id == title_id()) # находит system id последнего
    systems = system.select().where(System.stage == "Предварительный").get()
    total_athletes = systems.total_athletes
    total_gr = systems.total_group
    score_match = my_win.spinBox.text()
    for sys in system:
        fin_type = sys.type_table
        if fin_type == "круг" or fin_type == "сетка":
            fin_player = sys.max_player
            sum_player.append(fin_player)

        sum_pl = sum(sum_player)
    if kpt != 0:  # подсчет кол-во игр из выбора кол-ва игроков вышедших из группы и системы финала
        if etap_text == "Полуфиналы":
            vt = "группы"
            type_table = "группы"
            # +++ вариант с полуфиналом с выходом из 1-ого пф
            # if exit_stage == "Предварительный":
            #     gr_pf = total_gr // 2
            #     # player_in_final = gr_pf * kpt * 2 # колво участников в полуфинале
            # elif exit_stage == "1-й полуфинал":
            #     gr_pf = total_gr // 4
            gr_pf = total_gr // 2
            player_in_final = gr_pf * kpt * 2 # колво участников в полуфинале    
            cur_index = 0
        elif etap_text == "Финальный" or etap_text == "Суперфинал":
            cur_index = current_index_combobox_table(sender)
            if cur_index == 1:
                vt = "Сетка (-2) на"
                my_win.comboBox_page_vid.setCurrentText("книжная")
                type_table = "сетка"
            elif cur_index == 2:
                vt = "Сетка (с розыгрышем всех мест) на"
                my_win.comboBox_page_vid.setCurrentText("книжная")
                type_table = "сетка"
            elif cur_index == 3:
                vt = "Сетка (с играми за 1-3 места) на"
                my_win.comboBox_page_vid.setCurrentText("книжная")
                type_table = "сетка"
            elif cur_index == 4:
                vt = "Круговая таблица на"
                type_table = "круг"
        # ======
            if fin == "1-й финал" and type_table == "сетка":
                result = msgBox.question(my_win, "Уведомление", "Будет ли разигрываться 3-е место\n в 1-ом финале?"
                , msgBox.No, msgBox.Yes) 
                if result == msgBox.Yes:
                    no_game3 = ""
                else:
                    no_game3 = 3 # два 3-х места

            if exit_stage == "1-й полуфинал" or exit_stage == "2-й полуфинал":
                system_exit = system.select().where(System.stage == exit_stage).get()
                total_gr = system_exit.total_group 

            if etap_text == "Суперфинал":
                player_in_final = kpt
                player_in_final_full = full_net_player(player_in_final)
            else:
                player_in_final_full = total_gr * kpt # колво участников в конкретном финале, если в группах полный состав
                player_in_final_current = total_athletes - sum_pl # кол-во участников в последнем финале (разница всех игроков минус уже разведенных по финалам)
                
                if player_in_final_current <  player_in_final_full:
                   player_in_final = player_in_final_current
                   player_in_final_full = full_net_player(player_in_final)
                else: 
                    player_in_final = player_in_final_full

        total_games = numbers_of_games(cur_index, player_in_final, kpt) # подсчет кол-во игр

        if etap_text == "Полуфиналы":
            gr_pf = total_gr // 2
            str_setka = f"{gr_pf} {vt} по {kpt * 2} участника"
            total_gr = gr_pf
        else:
            player_in_final_full = full_net_player(player_in_final)
            str_setka = f"{vt} {player_in_final_full} участников" # пишет в базе данных полное кол-во игроков сетке
            total_gr = 0
 
        stroka_kol_game = f"{total_games} игр"
        # заполняет max_player в зависиости от кол игроков
        m_pl = player_in_final
        # if type_table == "круг": # если финал по кругу
        #     m_pl = player_in_final
        # elif type_table == "группы": # если ПФ
        #     m_pl = player_in_final
        # else: # если финал сетка
        #     m_pl = player_in_final
        #     # m_pl = full_net_player(player_in_final)

        # ======
        system = System(title_id=title_id(), total_athletes=total_athletes, total_group=total_gr, kol_game_string=stroka_kol_game,
                        max_player=m_pl, stage=fin, type_table=type_table, page_vid=pv, label_string=str_setka,
                        choice_flag=0, score_flag=score_match, visible_game=flag_visible, stage_exit=exit_stage, mesta_exit=kpt, no_game=no_game3).save()    
        
        return [str_setka, player_in_final, total_athletes, stroka_kol_game]


def full_net_player(player_in_final):
    """максимальное количество игроков в сетке при не полном составе"""
    if player_in_final == 4:
       player_in_final_full = 8
    else: 
        for m in range(1, 6):
            game = 2** m
            if game >= player_in_final:
                break
        player_in_final_full = 2 ** m
    return player_in_final_full


def current_index_combobox_table(sender):
    """определяет индекс значения комбобокса"""
    if sender == my_win.comboBox_table_2:
        cur_index = my_win.comboBox_table_2.currentIndex()
    elif sender == my_win.comboBox_table_3:
        cur_index = my_win.comboBox_table_3.currentIndex()
    elif sender == my_win.comboBox_table_4:
        cur_index = my_win.comboBox_table_4.currentIndex()
    elif sender == my_win.comboBox_table_5:
        cur_index = my_win.comboBox_table_5.currentIndex()
    elif sender == my_win.comboBox_table_6:
        cur_index = my_win.comboBox_table_6.currentIndex()
    elif sender == my_win.comboBox_table_7:
        cur_index = my_win.comboBox_table_7.currentIndex() 
    elif sender == my_win.comboBox_table_8:
        cur_index = my_win.comboBox_table_8.currentIndex()  
    elif sender == my_win.comboBox_table_9:
        cur_index = my_win.comboBox_table_9.currentIndex()
    elif sender == my_win.comboBox_table_10:
        cur_index = my_win.comboBox_table_10.currentIndex() 
    elif sender == my_win.comboBox_table_11:
        cur_index = my_win.comboBox_table_11.currentIndex()  
    elif sender == my_win.comboBox_table_12:
        cur_index = my_win.comboBox_table_12.currentIndex()    
    return cur_index


def control_all_player_in_final(etap, all_sum_player_final):
    """проверка все ли игроки распределены по финалам и дает сигнал об окончании создании системы"""
     # титул id и стадия содержит слово финал (1 и 2 заменяет %)
    msgBox = QMessageBox
    titles = Title.select().where(Title.id == title_id()).get()
    id_title = titles.id
    system = System.select().order_by(System.id).where(System.title_id == title_id())
    system_stage = system.select().where(System.stage == "Предварительный").get()
    total_player = system_stage.total_athletes
    t = total_player - all_sum_player_final # оставшиеся не распределенные участники по финалам
    txt = ""
    if all_sum_player_final == total_player or t <= 2: # все игроки посеяны по финалам или остался 1 или 2 игрока окончание создание системы
        if t == 1 and etap != "Суперфинал":     
            txt = "Остался 1 участник, не вошедший в финальную часть"
            msgBox.information(my_win, "Уведомление", txt)
        elif t == 2 and etap != "Суперфинал":
            txt = "Остались 2 игрока, они могут сыграть за место между собой"
            msgBox.information(my_win, "Уведомление", txt)
                    # ====== вставить вопрос о суперфинале и игры за 3 место если финал сетка
        if etap != "Суперфинал":           
            add_open_tab(tab_page="Система")
            result = msgBox.question(my_win, "", "Система соревнований создана.\n"
                                                        "Теперь необходимо сделать жеребъевку\n"
                                                        "предварительного этапа.\n"
                                                        "Хотите ее сделать сейчас?",
                                        msgBox.Ok, msgBox.Cancel)
            if result == msgBox.Ok:                
                flag_checking = checking_before_the_draw() # проверка что все спортсмены подтвердились
                if flag_checking is False:
                    return
                    
                choice_gr_automat()
                add_open_tab(tab_page="Результаты")
                tab_enabled(id_title)
                with db:
                    system_stage.choice_flag = True
                    system_stage.save()
                    flag = True
            else:
                my_win.choice_gr_Action.setEnabled(True)
                flag_ch = system_stage.choice_flag
                if flag_ch is True: # была уже сделана жеребьевка групп, и идет процесс добавления или удаления этапа при редактировании системы
                    return
                else:
                    with db:
                        system_stage.choice_flag = False
                        system_stage.save()
                        flag = True
                return
        else:
            return
        flag = False    
    elif t >= 3: # продолжает создание системы
        flag = True
    elif t == 0:
        flag = False
    return flag


def checking_before_the_draw():
    """Проверка перед жеребьевкой групп, что все игроки подтверждены"""
    checking_flag = False
    msgBox = QMessageBox()
    players = Player.select().where((Player.title_id == title_id()) & (Player.application == "предварительная"))
    count = len(players)
    if count > 0:
        msgBox.information(my_win, 'Уведомление',
                                                "В списке присутствуют спортсмены,\nиз предварительной заявке"
                                                "\nне подтвержденые о своем участии!",
                                        msgBox.Ok)
        player_list = players.select().where(Player.application == "Предварительная")
        my_win.tabWidget.setCurrentIndex(1)
        fill_table(player_list)
        checking_flag = False
    else:
        checking_flag = True
    return checking_flag


def combobox_etap_compare(real_list):
    """сравнение и изменение значение комбокса в зависиости от выбора этапа
    -real_list- список значений для будущего этапа"""
    count_items = my_win.comboBox_etap.count()
    item_list = [] # текущий набор комбобокса
    for i in range(0, count_items):
        cur_item = my_win.comboBox_etap.itemText(i)
        item_list.append(cur_item) # значение которые есть в комбобокс после добавления отсутствующих
    res = [x for x in real_list + item_list if x not in real_list or x not in item_list] # список, который надо убрать из комбобокса

    if len(res) != 0:
        my_win.comboBox_etap.clear()
        my_win.comboBox_etap.addItems(real_list)


def made_system_load_combobox_etap():
    """подготовка комбобокса для дальнейшего создания системы"""
    sender = my_win.sender()
    ct = my_win.comboBox_etap.currentText()
    label_etap_list = [my_win.label_101, my_win.label_102, my_win.label_103, my_win.label_104, my_win.label_105, my_win.label_106,
                       my_win.label_107, my_win.label_108, my_win.label_109, my_win.label_110,  my_win.label_111, my_win.label_112]
    combobox_etap_list = [my_win.comboBox_table_1, my_win.comboBox_table_2, my_win.comboBox_table_3, my_win.comboBox_table_4, 
                          my_win.comboBox_table_5, my_win.comboBox_table_6, my_win.comboBox_table_7, my_win.comboBox_table_8,
                          my_win.comboBox_table_9, my_win.comboBox_table_10, my_win.comboBox_table_11, my_win.comboBox_table_12]
    label_text = my_win.label_10.text() # номер текущего этапа
    real_list = []
 # после выбора из комбобокса соответственно этапу включает label   
    if ct == "" or ct == "-выбор этапа-" or sender == my_win.Button_etap_made: # после нажатия кнопки создать этап готовит следующий этап
        if ct == "-выбор этапа-" or ct == "":
            return   
        if label_text == "1-й этап": # пишет следующий этап
            if ct == "Одна таблица":
                my_win.comboBox_table_1.show()
                my_win.spinBox_kol_group.hide()
                my_win.label_11.hide()
                my_win.label_101.hide()
            elif ct == "Предварительный":
                my_win.spinBox_kol_group.hide()
                my_win.comboBox_table_1.hide()
                my_win.label_11.hide()
                my_win.label_12.show()
                real_list = ["-выбор этапа-", "Полуфиналы", "Финальный"] # который нужен в комбобокс
                combobox_etap_compare(real_list)
                my_win.label_10.setText("2-й этап")
        elif  label_text == "2-й этап":  
            # ==== поиск всех занчений комбобокса
            if ct == "Полуфиналы":
                real_list = ["-выбор этапа-", "Полуфиналы", "Финальный"] # который нужен в комбобокс
            elif ct == "Финальный": 
                real_list = ["-выбор этапа-", "Финальный", "Суперфинал"] 
            combobox_etap_compare(real_list)
            my_win.label_10.setText("3-й этап")
        elif  label_text == "3-й этап": # текущий этап
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_104.show()
            my_win.label_10.setText("4-й этап")
        elif  label_text == "4-й этап": 
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_105.show()
            my_win.label_10.setText("5-й этап")
        elif  label_text == "5-й этап": 
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_106.show()
            my_win.label_10.setText("6-й этап")
        elif  label_text == "6-й этап": 
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_107.show()
            my_win.label_10.setText("7-й этап")
        elif  label_text == "7-й этап": 
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_108.show()
            my_win.label_10.setText("8-й этап")
        elif  label_text == "8-й этап": 
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_109.show()
            my_win.label_10.setText("9-й этап")
        elif  label_text == "9-й этап": 
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_110.show()
            my_win.label_10.setText("10-й этап")
        elif  label_text == "10-й этап": 
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_111.show()
            my_win.label_10.setText("11-й этап")
        elif  label_text == "11-й этап": 
            real_list = ["-выбор этапа-", "Финальный", "Суперфинал"]
            combobox_etap_compare(real_list)
            my_win.label_112.show()
            my_win.label_10.setText("12-й этап")
        my_win.comboBox_etap.setCurrentText("-выбор этапа-")   
    else:   # выбор значения из комбобокса создания этапов
        if ct == "Одна таблица":
            my_win.comboBox_table_1.show()
            my_win.spinBox_kol_group.hide()
            my_win.label_101.show()
            my_win.label_101.setText("Одна таблица")
            # my_win.label_11.show()
            my_win.label_11.hide()
        elif ct == "Предварительный":
            my_win.spinBox_kol_group.show()
            my_win.comboBox_table_1.hide()
            my_win.label_101.show()
            my_win.label_101.setText("Предварительный этап")
            my_win.label_11.show()
            my_win.label_12.hide()
        elif ct == "Полуфиналы":
            mark = label_text.find("-")
            number_etap = int(label_text[:mark])
            label_etap_list[number_etap - 1].show()
            label_etap_list[number_etap - 1].setText(f"{number_etap - 1}-й полуфинал")
            kol_player_in_final()
        elif ct == "Финальный":
            if label_text == "2-й этап":
                my_win.label_102.show()
                my_win.label_102.setText("1-й финал")
                my_win.comboBox_table_2.show()
            elif label_text == "3-й этап":
                last_etap = my_win.label_102.text() 
                my_win.label_103.show()               
                if last_etap == "1-й полуфинал":
                    my_win.label_103.setText("1-й финал")
                else:
                    my_win.label_103.setText("2-й финал")
                my_win.comboBox_table_3.show()               
            elif label_text == "4-й этап":
                last_etap = my_win.label_103.text() 
                if last_etap == "2-й полуфинал":
                    my_win.label_104.setText("1-й финал")
                else: 
                    txt = my_win.label_103.text()
                    znak = txt.find("-") 
                    fin = int(txt[:znak])
                    final = f"{fin + 1}-й финал"    
                    my_win.label_104.setText(final)
                my_win.comboBox_table_4.show()
            elif label_text == "5-й этап":
                txt = my_win.label_104.text()
                znak = txt.find("-") 
                fin = int(txt[:znak])
                final = f"{fin + 1}-й финал"    
                my_win.label_105.setText(final)
                my_win.comboBox_table_5.show()
            elif label_text == "6-й этап":
                txt = my_win.label_105.text()
                znak = txt.find("-") 
                fin = int(txt[:znak])
                final = f"{fin + 1}-й финал"    
                my_win.label_106.setText(final)
                my_win.comboBox_table_6.show()
            elif label_text == "7-й этап":
                txt = my_win.label_106.text()
                znak = txt.find("-") 
                fin = int(txt[:znak])
                final = f"{fin + 1}-й финал"    
                my_win.label_107.setText(final)
                my_win.comboBox_table_7.show()
            elif label_text == "8-й этап":
                txt = my_win.label_107.text()
                znak = txt.find("-") 
                fin = int(txt[:znak])
                final = f"{fin + 1}-й финал"    
                my_win.label_108.setText(final)
                my_win.comboBox_table_8.show()
            elif label_text == "9-й этап":
                txt = my_win.label_108.text()
                znak = txt.find("-") 
                fin = int(txt[:znak])
                final = f"{fin + 1}-й финал"    
                my_win.label_109.setText(final)
                my_win.comboBox_table_9.show()
            elif label_text == "10-й этап":
                txt = my_win.label_109.text()
                znak = txt.find("-") 
                fin = int(txt[:znak])
                final = f"{fin + 1}-й финал"    
                my_win.label_110.setText(final)
                my_win.comboBox_table_10.show()
            elif label_text == "11-й этап":
                txt = my_win.label_110.text()
                znak = txt.find("-") 
                fin = int(txt[:znak])
                final = f"{fin + 1}-й финал"    
                my_win.label_111.setText(final)
                my_win.comboBox_table_11.show()
            elif label_text == "12-й этап":
                txt = my_win.label_111.text()
                znak = txt.find("-") 
                fin = int(txt[:znak])
                final = f"{fin + 1}-й финал"    
                my_win.label_112.setText(final)
                my_win.comboBox_table_12.show()
            else:
                mark = label_text.find("-")
                fin = int(txt[:mark])
                final = f"{fin + 1}-й финал"  
                label_etap_list[number_etap - 1].setText(final) 
                combobox_etap_list[number_etap - 1].show() 
        elif ct == "Суперфинал":
            mark = label_text.find("-")
            number_etap = int(label_text[:mark])
            label_etap_list[number_etap - 1].show()
            label_etap_list[number_etap - 1].setText("Суперфинал")
            combobox_etap_list[number_etap - 1].show()


def total_games_in_final_without_group_games(player_in_final, total_gr, kpt):
    """всего игр в финале без учета сыгранных игр в предварительном этапе"""
    # остаток от деления, если 0, то участники равно делится на группы
    remains = player_in_final % int(total_gr)
    if remains == 0:  # если в группах равное количество человек
        playing_game = (kpt * (kpt - 1)) // 2 * total_gr
    else:
        full_group = player_in_final // kpt # кол-во групп с полным количеством участников
        no_full_group = total_gr - remains
        playing_game_in_full_group = (kpt * (kpt - 1)) // 2 * full_group
        kpt_min = kpt - 1
        playing_game_in_no_full_group = (kpt_min * (kpt_min - 1)) // 2 * no_full_group
        playing_game = playing_game_in_full_group + playing_game_in_no_full_group
    total_games = (player_in_final * (player_in_final - 1)) // 2 - playing_game
    return total_games


def total_games_in_final_with_group_games(player_in_final, gr_pf, kpt):
    """всего игр в полуфинале с учетом сыгранных игр в предварительном этапе"""
    # остаток от деления, если 0, то участники равно делится на группы
    remains = player_in_final % int(gr_pf)
    if remains == 0:  # если в группах равное количество человек
        playing_game_in_group = (kpt * (kpt - 1)) # кол-во игр, сыгранных в группе
        total_games = (((kpt * 2 * (kpt * 2 - 1)) // 2) - playing_game_in_group) * gr_pf # всего игр в пф
    else:
        full_group = player_in_final // kpt # кол-во групп с полным количеством участников
        no_full_group = gr_pf - remains
        playing_game_in_group = (kpt * (kpt - 1)) # кол-во игр, сыгранных в группе
        playing_game_in_full_group = (((kpt * (kpt - 1)) // 2) - playing_game_in_group) * full_group
        kpt_min = kpt - 1
        playing_game_in_no_full_group = ((kpt_min * (kpt_min - 1)) // 2 - playing_game_in_group) * no_full_group
        total_games = playing_game_in_full_group + playing_game_in_no_full_group
    return total_games


def numbers_of_games(cur_index, player_in_final, kpt):
    """подсчет количество игр в зависимости от системы"""
    systems = System.select().where(System.title_id == title_id())
    system_etap = my_win.comboBox_etap.currentText() 
    if system_etap == "Предварительный":
        system = systems.select().where(System.stage == "Предварительный").get()
        gr = system.total_group
    elif system_etap == "Полуфиналы":
        system = systems.select().where(System.stage == "Предварительный").get()
        gr = system.total_group
        gr_pf = gr // 2
        total_games = total_games_in_final_with_group_games(player_in_final, gr_pf, kpt)
    else:
        if cur_index == 1:  # сетка - 2
            if player_in_final == 8:
                total_games = 14
            elif player_in_final > 4 and player_in_final < 8: # если игроков не полная сетка
                tours = 3
                free = 8 - player_in_final
                if free == 1:
                    total_games = 14 - free * tours
                elif free > 1:
                    total_games = 14 - (free * tours - 1)
            elif player_in_final == 16:
                total_games = 38
            elif player_in_final > 8 and player_in_final < 16: # если игроков не полная сетка
                tours = 4
                free = 16 - player_in_final
                if free == 1:
                    total_games = 38 - free * tours
                elif free > 1:
                    total_games = 38 - (free * tours - 1)
            elif player_in_final == 32:
                total_games = 94
        elif cur_index == 2:  # прогрессивная сетка
            full_net = full_net_player(player_in_final)
            tours = int(math.log2(full_net))
            all_game_net = full_net // 2 * tours # количество игр в сетке при полном составе
            free = full_net - player_in_final
            if free == 0:
                total_games = all_game_net
            elif free == 1 or free == 2 or free == 3:
                # total_games = all_game_net - free * tours
                total_games = all_game_net - (free * tours - (free - 1))
            elif free == 4:
                 total_games = all_game_net - (free * (tours - 1))
            else:
                total_games = all_game_net - (free * tours - (free - 1))
            #    total_games = all_game_net - (free * tours - (free - 1))
            # ========
            # if player_in_final <= 8:
            #     free = 8 - player_in_final
            #     total_games = 12
            # elif player_in_final == 16:
            #     total_games = 32
            # elif player_in_final > 8 and player_in_final < 16:
            #     tours = 4
            #     free = 16 - player_in_final
            #     if free == 1:
            #         total_games = 32 - free * tours
            #     elif free == 2:
            #         total_games = 32 - (free * tours - 1)
            #     else:
            #         total_games = 32 - (free * 2 + 4)
            # elif player_in_final == 32: 
            #      total_games = 80  
            # elif player_in_final > 16 and player_in_final < 32:
            #     tours = 5
            #     free = 32 - player_in_final
            #     if free == 1:
            #         total_games = 80 - free * tours
            #     elif free > 1:
            #         total_games = 80 - (free * tours - 1)
            # else:
            #     total_games= 80
        elif cur_index == 3:  # сетка с розыгрышем призовых мест
            if player_in_final == 32:
                total_games = 32    
            elif player_in_final < 32:                
                total_games = player_in_final
        elif cur_index == 4: # игры в круг
            if system_etap == "Одна таблица":
                gr = 1 
            else:
                system = systems.select().where(System.stage == "Предварительный").get()
                gr = system.total_group
                gr = gr // 2
            total_games = total_games_in_final_without_group_games(player_in_final, gr, kpt)

    return total_games


# def clear_db_before_edit():
#     """очищает таблицы при повторном создании системы"""
#     system = System.select().where(System.title_id == title_id())
#     title = Title.select().where(Title.id == title_id()).get()
#     gl = Game_list.select().where(Game_list.title_id == title_id())
#     pl = Player.select().where((Player.title_id == title_id()) & (Player.bday == '0000-00-00'))
#     for i in gl:
#         gl_d = Game_list.get(Game_list.id == i)
#         gl_d.delete_instance()
#     chc = Choice.select().where(Choice.title_id == title_id())
#     for i in chc:
#         ch_d = Choice.get(Choice.id == i)
#         ch_d.delete_instance()
#     rs = Result.select().where(Result.title_id == title_id())
#     for i in rs:
#         r_d = Result.get(Result.id == i)
#         r_d.delete_instance()
#     if len(pl) > 0: # удаляет запись в -Player- если есть крест сетки
#         for i in pl:
#             pl_d = Player.get(Player.id == i)
#             pl_d.delete_instance()
#     for i in system:  # удаляет все записи
#         i.delete_instance()
#     sys = System(title_id=title_id(), total_athletes=0, total_group=0, max_player=0, stage="", type_table="", page_vid="",
#                  label_string="", kol_game_string="", choice_flag=False, score_flag=5, visible_game=True,
#                  stage_exit="", mesta_exit="", no_game="").save()

#     with db:
#         # записывает в таблицу -Title- новые открытые вкладки
#         title.tab_enabled = "Титул Участники"
#         title.save() 


def clear_db_before_choice(stage):
    """очищает систему перед повторной жеребьевкой и изменяет кол-во участников если они изменились"""
    msgBox = QMessageBox
    sys = System.select().where(System.title_id == title_id())
    player = Player.select().where((Player.title_id == title_id()) & (Player.bday != '0000-00-00'))
    id_system = system_id(stage)
    system = sys.select().where(System.id == id_system).get()
    tg = system.total_group
    total_player = system.total_athletes
    max_pl = system.max_player
    new_total_player = len(player)
    if total_player != new_total_player: #  если изменилось число участников
        result = msgBox.question(my_win, "Список участников", "Был изменено число участников.\n"
        "вы хотите изменить систему соревнований?",
                                    msgBox.Ok, msgBox.Cancel)
        if result == msgBox.Ok:
             # очищает таблицы перед новой системой соревнования (system, choice)
            # clear_db_before_edit()
            system_clear()
            choice_tbl_made()  # заполняет db жеребьевка
        else:
            e1 = new_total_player % tg  # остаток до полного посева групп, где спортсменов на одного больше
                # если количество участников равно делится на группы (кол-во групп)
            p_min = new_total_player // tg  # минимальное число спортсменов в группах
            g1 = int(tg) - e1  # кол-во групп, где наименьшее кол-во спортсменов
            p_max= p_min + 1  # кол-во человек в группе с наибольшим их количеством
            if e1 == 0:  # то в группах равное количество человек -e1-
                stroka_kol_group = f"{tg} группы по {str(p_min)} чел."
                skg = int((p_min * (p_min - 1) / 2) * int(tg))
                max_pl = p_min
            else:
                stroka_kol_group = f"{str(g1)} групп(а) по {str(p_min)} чел. и {str(e1)} групп(а) по {str(p_max)} чел."
                skg = int((((p_min * (p_min - 1)) / 2 * g1) + ((p_max * (p_max - 1)) / 2 * e1)))
                max_pl = p_max
            kgs = f"{skg} игр"
            sys_t = System.select().where(System.id == id_system).get()
            sys_t.max_player = max_pl
            sys_t.label_string = stroka_kol_group
            sys_t.kol_game_string = kgs
            sys_t.save()

            for x in sys:
                x.total_athletes = new_total_player
                x.save()
    # else:  # если число спортсменов не изменилось (просто смена участников)
    gl = Game_list.select().where(Game_list.title_id == title_id())
    # system_id = sys.select().where(System.stage == stage).get()
    gamelists = gl.select().where(Game_list.system_id == id_system)
    for i in gamelists:
        gl_d = Game_list.get(Game_list.id == i)
        gl_d.delete_instance()
    choices = Choice.select().where(Choice.title_id == title_id())
    if stage == "Предварительный":
        for i in choices:
            ch_d = Choice.get(Choice.id == i)
            ch_d.delete_instance()
    results = Result.select().where((Result.title_id == title_id()) & (Result.system_stage == stage))
    for i in results:
        r_d = Result.get(Result.id == i)
        r_d.delete_instance()
    if stage == "Предварительный":
        choice_tbl_made()


def clear_db_before_choice_final(fin):
    """очищает базу данных -Game_list- и -Result- перед повторной жеребьевкой финалов"""
    stage = fin
    id_system = system_id(stage)
    gamelist = Game_list.select().where((Game_list.title_id == title_id()) & (Game_list.system_id == id_system))
    for i in gamelist:
        gl_d = Game_list.get(Game_list.id == i)
        gl_d.delete_instance()
    results = Result.select().where((Result.title_id == title_id()) & (Result.system_id == id_system))
    for i in results:
        r_d = Result.get(Result.id == i)
        r_d.delete_instance()
    choice = Choice.select().where((Choice.title_id == title_id()) & (Choice.final == stage))
    for i in choice:
        Choice.update(posev_final="").where(Choice.id == i).execute()
    # System.update(choice_flag=0).where(System.id == id_system).execute() # обновляет запись, что жеребьевка не сделана


def clear_db_before_choice_semifinal(stage):
    """очищает базу данных -Game_list- и -Result- перед повторной жеребьевкой полуфиналов"""
    id_system = system_id(stage)
    gamelist = Game_list.select().where(Game_list.title_id == title_id())
    gl = gamelist.select().where(Game_list.system_id == id_system)
    for i in gl:
        gl_d = Game_list.get(Game_list.id == i)
        gl_d.delete_instance()
    results = Result.select().where(Result.title_id == title_id())
    rs = results.select().where(Result.system_id == id_system)
    for i in rs:
        r_d = Result.get(Result.id == i)
        r_d.delete_instance()


def ready_system():
    """проверка на готовность системы"""
    all_player_in_final = []
    system = System.select().where(System.title_id == title_id())  # находит system id первого
    count = len(system)
    flag = False
    if count == 1:
        for k in system:
            stage = k.stage
        if stage == "Одна таблица":
            my_win.statusbar.showMessage("Система соревнований создана", 5000)
            flag = True
        else:
            my_win.statusbar.showMessage("Необходимо создать систему соревнований", 500)
            flag = False
    elif count > 1:
        sys = system.select().order_by(System.id.desc()).where(System.title_id == title_id()).get()
        stage_system = sys.stage
        if stage_system == "Предварительный" or stage_system == "1-й полуфинал" or stage_system == "2-й полуфинал":
            my_win.statusbar.showMessage("Необходимо создать систему соревнований", 500)
            flag = False
        else:
            sys_min = system.select().order_by(System.id).where(System.title_id == title_id()).get()
            total_player = sys_min.total_athletes
            system_id = system.select().where(System.stage ** '%финал')
            for k in system_id:
                tot_player = k.max_player
                all_player_in_final.append(tot_player)
            all_player = sum(all_player_in_final)
            if all_player >= total_player:
                my_win.statusbar.showMessage("Система соревнований создана", 5000)
                flag = True
            else:
                my_win.statusbar.showMessage("Необходимо создать систему соревнований", 500)
                flag = False
    return flag


def ready_choice(stage):
    """проверка на готовность жеребьевки групп"""
    sys = System.select().where(System.title_id == title_id())
    id_system = system_id(stage)
    greb_flag = False
    if stage != "":
        system = sys.select().where(System.id == id_system).get()
        greb_flag = system.choice_flag   
    if greb_flag is True:
        my_win.statusbar.showMessage("Жеребьевка сделана", 5000)
        flag = True
    else:
        my_win.statusbar.showMessage("Жеребьевка групп еще не выполнена", 5000)
        flag = False
    return flag


def select_choice_final():
    """выбор жеребьевки финала"""
    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    fin = []
    for sys in system:
        if sys.stage != "Предварительный" and sys.stage != "1-й полуфинал" and sys.stage != "2-й полуфинал":
            fin.append(sys.stage)
    fin, ok = QInputDialog.getItem(my_win, "Выбор финала", "Выберите финал для жеребъевки", fin, 0, False)
    if ok:
        return fin
    else:
        fin = None
        return fin


def select_choice_semifinal():
    """выбор жеребьевки полуфинала"""
    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    semifinal = []
    for sys in system:
        if sys.stage == "1-й полуфинал" or sys.stage == "2-й полуфинал":
            semifinal.append(sys.stage)
    semifinal, ok = QInputDialog.getItem(my_win, "Выбор полуфинала", "Выберите полуфинал для жеребъевки", semifinal, 0, False)
    if ok:
        return semifinal
    else:
        semifinal = None
        return semifinal


def manual_choice_setka(fin, count_exit, mesto_first_poseva):
    """Ручная жеребьевка сетки"""
    choice = Choice.select().where(Choice.title_id == title_id())
    posevs = setka_choice_number(fin, count_exit)
    player_net = posevs[0]
    posev_1 = posevs[1]
    z = len(posevs)

    if z == 3:
        posev_2 = posevs[2]
    elif z == 4:
        posev_2 = posevs[2]
        posev_3 = posevs[3]
    elif z == 5:
        posev_2 = posevs[2]
        posev_3 = posevs[3]
        posev_4 = posevs[4]
    for n in range (0, count_exit): # начало основного посева
        if fin == "1-й финал":
            choice_posev = choice.select().where(Choice.mesto_group == mesto_first_poseva + n)
        else:
            choice_posev = choice.select().where(Choice.mesto_group == mesto_first_poseva + n).order_by(Choice.rank)


def check_choice(stage):
    """Проверяет перед жеребьевкой финалов, была ли она произведена ранее или еще нет"""
    # stage_list = []
    id_system = system_id(stage)
    system = System.select().where((System.title_id == title_id()) & (System.id == id_system)).get()  # находит system id последнего
    check_flag = system.choice_flag
    return check_flag


def checking_possibility_choice(stage):
    """Проверяет перед жеребьевкой финалов, сыграны ли все партиии в группах или полуфиналах"""
    msg = QMessageBox
    id_system = system_id(stage)
    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    system_final = system.select().where(System.id == id_system).get() # получаем запись конкретного финала
    check_flag = False
    
    if stage == "Одна таблица":
        check_flag = system_final.choice_flag
    else:    
        exit = system_final.stage_exit  # запись откуда идет выход в финал
        id_system = system_id(stage=exit) # id системы откуда выход в суперфинал

        if exit == "1-й полуфинал" or exit == "2-й полуфинал":
            exit_str = f"{exit}е"
        elif exit == "предварительном этапе":            
            exit_str = "предварительном этапе"
        else:
            exit_str = "1-ом финале"
        # gr = Result.select().where((Result.title_id == title_id()) & (Result.system_stage == exit)) # отбираем записи с выходом в финал
        gr = Result.select().where((Result.title_id == title_id()) & (Result.system_id == id_system)) # отбираем записи с выходом в финал

        for i in gr:
            game = i.points_win 
            check_flag = True      
            if game is None:
                result = msg.information(my_win, "Предварительный этап", "Еще не все встречи сыграны в" f"{exit_str}.",
                                        msg.Ok)
                check_flag = False
                break                        
    return check_flag


def del_player_table():
    """таблица удаленных игроков на данных соревнованиях"""
    if my_win.checkBox_6.isChecked():
        my_win.Button_clear_del.setEnabled(True)
        player_list = Delete_player.select().where(Delete_player.title_id == title_id())
        count = len(player_list)
        if count == 0:
            my_win.statusbar.showMessage(
                "Удаленных участников соревнований нет", 10000)
            fill_table(player_list)
        else:
            my_win.tableView.hideColumn(8)
            my_win.tableView.hideColumn(10)
            my_win.tableView.hideColumn(11)
            my_win.tableView.hideColumn(12)
            my_win.tableView.hideColumn(13)
            fill_table(player_list)
            my_win.statusbar.showMessage(
                "Список удаленных участников соревнований", 10000)
            if my_win.lineEdit_Family_name.text() != "":
                my_win.Button_add_edit_player.setText("Восстановить")
                my_win.Button_add_edit_player.setEnabled(True)
            else:
                my_win.Button_add_edit_player.setEnabled(False)
    else:
        player_list = Player.select().where(Player.title_id == title_id())
        fill_table(player_list)
        my_win.tableView.showColumn(8)
        my_win.Button_add_edit_player.setText("Добавить")
        my_win.Button_add_edit_player.setEnabled(True)
        my_win.Button_clear_del.setEnabled(False)
        my_win.statusbar.showMessage("Список участников соревнований", 10000)


def clear_del_player():
    """Очистка базы данных удаленных игроков"""
    msgBox = QMessageBox
    del_player = Delete_player.select().where(Delete_player.title_id == title_id())
    result = msgBox.question(my_win, "Участники", "Вы действительно хотите очистить список\n"
                                "удаленных игроков?",
                                        msgBox.Ok, msgBox.Cancel)
    if result == msgBox.Ok:
        for i in del_player:
            i.delete_instance()
        my_win.Button_clear_del.setEnabled(False)  
        my_win.checkBox_6.setChecked(False)      
    else:
        return


def remains_in_group(etap_system, etap_system_dict):
    """подсчет игроков в группе и полуфиналов после создания финалов"""
    stage_dict = {} # словарь (этап: кол0во игроков)
    number_player_gr = 0
    number_player_pf1 = 0
    number_player_pf2 = 0
    out_pf1 = 0
    out_pf2 = 0
    out_f = 0
    system = System.select().where(System.title_id == title_id())
    for m in range(0, 2):
        for k  in system:
            etap_system = k.stage
            if etap_system == "Предварительный":
                number_player_gr = k.max_player
                stage_dict[etap_system] = number_player_gr if m == 0 else number_player_gr - out_pf1 - out_pf2
            elif etap_system == "1-й полуфинал":
                number_player_pf1 = k.max_player // k.total_group # кол-во игрок в 1-ом пф
                out_pf1 = etap_system_dict[etap_system] # сколько вышло из группы 1-й пф
                stage_dict[etap_system] = number_player_pf1
            elif etap_system == "2-й полуфинал":
                number_player_pf2 = k.max_player // k.total_group
                out_pf2 = etap_system_dict[etap_system] # сколько вышло из группы 1-й пф
                stage_dict[etap_system] = number_player_pf2
            elif (etap_system == "1-й финал" or etap_system == "2-й финал" or etap_system == "3-й финал" or
                etap_system == "4-й финал" or etap_system == "5-й финал" or etap_system == "6-й финал" or
                etap_system == "7-й финал" or etap_system == "8-й финал" or etap_system == "9-й финал" or
                etap_system == "10-й финал"):
                if m == 1:
                    systems = system.select().where(System.stage == etap_system).get()
                    exit_stage = systems.stage_exit # откуда выходят в финал
                    out_f = etap_system_dict[etap_system] # сколько вышло из группы 1-й пф
                    stage_dict[exit_stage] = stage_dict[exit_stage ] - out_f  # сколько вышло из 1-й пф в 1-й финал
    return stage_dict


def max_player_and_exit_stage(etap):
    """определяет максимальное число спортсменов в комбобоксе и стадию откуда выход в финал
    etap - текущий этап, stage - предыдущий этап, label_text - номер этапа, mx_pl - максимальное число в комбобоксе
    # etap_list список [этап, кол-во игроков, из какого этапа вышли"""
    exit_player_stage = []
    etap_list = []
    etap_list_tmp = []
    total_stage = []
    etap_dict = {}
    etap_system_dict = {}
    system = System.select().where(System.title_id == title_id())
    i = 0
    for k in system: # получение словаря этапов
        i += 1
        etap_system = k.stage
        mesta_exit = k.mesta_exit
        stage_exit = k.stage_exit
        etap_list_tmp.append(etap_system)
        etap_list_tmp.append(mesta_exit)
        etap_list_tmp.append(stage_exit)
        total_stage.append(etap_system)
        etap_list = etap_list_tmp.copy()
        etap_list_tmp.clear()
        etap_dict[i] = etap_list
        etap_system_dict[etap_system] = mesta_exit
    number_etap = i + 1
    dict_etap = remains_in_group(etap_system, etap_system_dict)

    listing_etap = etap_dict[i] # список этапа (название, выход)
    last_etap = listing_etap[0] 
    system_last = system.select().where(System.stage == last_etap).get()
    mesta_exit = listing_etap[1]
    stage_exit = listing_etap[2]

    if number_etap == 2:
        if etap == "Полуфиналы":
            fin = "1-й полуфинал"
        elif etap == "Финальный":
            fin = "1-й финал"
        exit_stage = "Предварительный" # откуда попадают в полуфинал игроки
        max_pl = system_last.max_player # максимальное допустимое число игроков при выборе в комбобоксе
    elif number_etap == 3:
        if etap == "Полуфиналы":
            fin = "2-й полуфинал"
            # ++++ вариант выбора выхода в 2-й полуфинал (из группы или 1-щ=ого пф)
            # group_list = ["-выбор этапа", "Предварительный", "1-й полуфинал"]
            # exit_stage, ok = QInputDialog.getItem(my_win, "Полуфиналы", "Выберите этап откуда\n"
            #                             "выходят во 2-й полуфинал", group_list)
            # +++++++                                
            exit_stage = "Предварительный"
        elif etap == "Финальный":
            fin = "1-й финал" if "1-й полуфинал" in total_stage else "2-й финал"
            exit_stage = "1-й полуфинал" if "1-й полуфинал" in total_stage else "Предварительный"
        max_pl = dict_etap[exit_stage]
    elif (number_etap == 4 or number_etap == 5 or number_etap == 6 or number_etap == 7
        or number_etap == 8 or number_etap == 9 or number_etap == 10 or number_etap == 11 or number_etap == 12):
        fin = number_final(last_etap) # текущий этап
        if "2-й полуфинал" in total_stage:
            if dict_etap["1-й полуфинал"] == 0 and dict_etap["2-й полуфинал"] != 0:
                exit_stage = "2-й полуфинал"
            elif dict_etap["1-й полуфинал"] != 0 and dict_etap["2-й полуфинал"] != 0:
                exit_stage = "1-й полуфинал"
            elif dict_etap["1-й полуфинал"] == 0 and dict_etap["2-й полуфинал"] == 0:
                exit_stage = "Предварительный"
        elif "1-й полуфинал" in total_stage:
            if dict_etap["1-й полуфинал"] == 0:
                exit_stage = "Предварительный"
            elif dict_etap["1-й полуфинал"] != 0:
                exit_stage = "1-й полуфинал"
        else:
            exit_stage = "Предварительный"
        max_pl = dict_etap[exit_stage]

    exit_player_stage.append(max_pl)
    exit_player_stage.append(exit_stage)
    exit_player_stage.append(fin)

    return exit_player_stage


def number_final(last_etap):
    """определяет номер финала исходя из предыдущего"""
    if last_etap == "2-й полуфинал":
        fin = "1-й финал"
    else:
        znak = last_etap.find("-") 
        fin_num = int(last_etap[:znak])
        fin = f"{fin_num + 1}-й финал"
    return fin


def kol_player_in_final():
    """после выбора из комбобокс системы финала подсчитывает сколько игр в финале"""
    sender = my_win.sender()
    pv = my_win.comboBox_page_vid.currentText()
    etap = my_win.comboBox_etap.currentText()
    player = Player.select().where(Player.title_id == title_id())
    count = len(player)
    fin = ""
    exit_stage = ""
    label_text = my_win.label_10.text()
    flag_one_table = False
    if etap != "Суперфинал":
        if my_win.comboBox_etap.currentText() == "Одна таблица":
            if my_win.comboBox_table_1.currentText() == "Круговая система":
                kol_game = count * (count - 1) // 2
                my_win.label_etap_1.show()
                my_win.label_19.show()
                my_win.label_101.show()
                my_win.label_11.hide()
                my_win.label_101.setText(my_win.comboBox_etap.currentText())
                my_win.label_19.setText(f"{kol_game} игр.")
                my_win.label_33.setText(f"Всего: {kol_game} игр.")
                my_win.label_etap_1.setText(f"{count} человек по круговой системе.")
                my_win.comboBox_etap.hide()
                my_win.comboBox_table_1.hide()
                my_win.comboBox_page_vid.setCurrentText("альбомная")
            else: # система из одной таблицы по олимпийской системе
                my_win.comboBox_page_vid.setCurrentText("книжная")
                cur_index = my_win.comboBox_table_1.currentIndex()
                total_game = 0
                if cur_index != 0:
                    player_in_final = count
                    kpt = 1
                    total_game = numbers_of_games(cur_index, player_in_final, kpt)
                my_win.label_etap_1.show()
                my_win.label_19.show()
                my_win.label_19.setText(f"{total_game} игр.")
                my_win.label_33.setText(f"Всего: {total_game} игр.")
                my_win.label_etap_1.setText(f"{count} человек в сетке.")
                my_win.comboBox_table_1.hide()
            flag_one_table = True
        else:
            exit_player_stage = max_player_and_exit_stage(etap)
            max_exit_group = exit_player_stage[0]
            exit_stage = exit_player_stage[1]
            fin = exit_player_stage[2]
        # изменение падежа этапов в комбобоксе
        if exit_stage == "Предварительный":
            exit_stroka = "Предварительного этапа"
        elif exit_stage == "1-й полуфинал":
            exit_stroka = "1-ого полуфинала"
        elif exit_stage == "2-й полуфинал":
            exit_stroka = "2-ого полуфинала" 
    else:
        fin = "Суперфинал"
        exit_stage = "1-й финал"
        exit_stroka = "1-ого финала"
        max_exit_group = 12
    if flag_one_table is False:
        kpt, ok = QInputDialog.getInt(my_win, "Число участников", "Введите число участников, выходящих\n "
                                                                    f"из {exit_stroka} в {fin}", min=1, max=max_exit_group)
                                                                            # возвращает из функции несколько значения в списке
        list_pl_final = total_game_table(exit_stage, kpt, fin, pv)
    else:
        list_pl_final = ['', count, '', pv]
        my_win.Button_etap_made.setEnabled(True)
        return
                    
  
    if ok is True: # заполняет этапы значениями (label)
        if label_text == "1-й этап":
            my_win.label_19.show()
            my_win.label_19.setText(list_pl_final[3])
            my_win.label_etap_1.show()
            my_win.label_etap_1.setText(list_pl_final[0])
        elif label_text == "2-й этап":
            my_win.label_27.setText(list_pl_final[3])
            my_win.label_27.show()
            my_win.label_etap_2.setText(list_pl_final[0])
            my_win.label_etap_2.show()
            my_win.comboBox_table_2.hide()
        elif label_text == "3-й этап":
            my_win.label_30.setText(list_pl_final[3])
            my_win.label_30.show()
            my_win.label_etap_3.setText(list_pl_final[0])
            my_win.label_etap_3.show()
            my_win.comboBox_table_3.hide()
        elif label_text == "4-й этап":
            my_win.label_53.setText(list_pl_final[3])
            my_win.label_53.show()
            my_win.label_etap_4.setText(list_pl_final[0])
            my_win.label_etap_4.show()
            my_win.comboBox_table_4.hide()
        elif label_text == "5-й этап":
            my_win.label_58.setText(list_pl_final[3])
            my_win.label_58.show()
            my_win.label_etap_5.setText(list_pl_final[0])
            my_win.label_etap_5.show()
            my_win.comboBox_table_5.hide()
        elif label_text == "6-й этап":
            my_win.label_81.setText(list_pl_final[3])
            my_win.label_81.show()
            my_win.label_etap_6.setText(list_pl_final[0])
            my_win.label_etap_6.show()
            my_win.comboBox_table_6.hide()
        elif label_text == "7-й этап":
            my_win.label_82.setText(list_pl_final[3])
            my_win.label_82.show()
            my_win.label_etap_7.setText(list_pl_final[0])
            my_win.label_etap_7.show()
            my_win.comboBox_table_7.hide()
        elif label_text == "8-й этап":
            my_win.label_83.setText(list_pl_final[3])
            my_win.label_83.show()
            my_win.label_etap_8.setText(list_pl_final[0])
            my_win.label_etap_8.show()
            my_win.comboBox_table_8.hide()
        elif label_text == "9-й этап":
            my_win.label_84.setText(list_pl_final[3])
            my_win.label_84.show()
            my_win.label_etap_9.setText(list_pl_final[0])
            my_win.label_etap_9.show()
            my_win.comboBox_table_9.hide()
        elif label_text == "10-й этап":
            my_win.label_85.setText(list_pl_final[3])
            my_win.label_85.show()
            my_win.label_etap_10.setText(list_pl_final[0])
            my_win.label_etap_10.show()
            my_win.comboBox_table_10.hide()
        elif label_text == "11-й этап":
            my_win.label_86.setText(list_pl_final[3])
            my_win.label_86.show()
            my_win.label_etap_11.setText(list_pl_final[0])
            my_win.label_etap_11.show()
            my_win.comboBox_table_11.hide()
        elif label_text == "12-й этап":
            my_win.label_87.setText(list_pl_final[3])
            my_win.label_87.show()
            my_win.label_etap_12.setText(list_pl_final[0])
            my_win.label_etap_12.show()
            my_win.comboBox_table_12.hide()
        my_win.Button_etap_made.setEnabled(True)
        my_win.comboBox_page_vid.setEnabled(True)
        my_win.Button_etap_made.setFocus(True)


def max_exit_player_out_in_group(exit_stage):
    """максимальное число игроков для комбобокса"""
    system = System.select().where(System.title_id == title_id())
    systems = system.select().where(System.stage == exit_stage).get()
    stroka = systems.label_string
    ind = stroka.find("по")
    max_exit_group = int(stroka[ind + 3:ind + 5]) # получаем число игроков в полуфинале
    return max_exit_group


def no_play():
    """победа по неявке соперника"""
    tb = my_win.tabWidget.currentIndex() 
    check_pl1 = my_win.checkBox_7.isChecked() 
    check_pl2 = my_win.checkBox_8.isChecked()
   
    if tb == 3:
        if check_pl1 is False and check_pl2 is False:
            return

    none_player = 1 if check_pl1 is True else 2
    enter_score(none_player)


def backup_mysql_database(host, port, username, password, database, backup_path):

    # Check if the backup directory exists
    if not os.path.exists(backup_path):
        print(f"Error: Backup directory '{backup_path}' does not exist.")
        sys.exit(1)
    # Create a filename for the backup with the current date and time
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d")
    backup_file = f"{backup_path}/{database}_{timestamp}.sql"

    # Command to create a database backup using mysqldump
    dump_command = f"mysqldump --no-tablespaces --host={host} --port={port} --user={username} --password={password} {database} > {backup_file} 2>/dev/null"
    # Execute the mysqldump command
    subprocess.run(dump_command, shell=True)

    return backup_file


def backup():
    """резервное копирование базы данных"""
    # === вариант с Mysql ====
    # Define database connection details
    host = "localhost"
    user = "root"
    password = "db_pass"
    database = "mysql_db"
    backup_file = "backup.sql"

    # Establish connection with the MySQL database
    connection = pymysql.connect(host=host, user=user, password=password, database=database)
    # Execute the mysqldump command
    command = f"mysqldump -h{host} -u{user} -p{password} {database} > {backup_file}"
    subprocess.run(command, shell=True)
    # Close the database connection
    connection.close()
    # ======
    # try:
    #     db = sqlite3.connect('comp_db.db')
    #     db_backup = sqlite3.connect(f'comp_db_backup.db')
    #     with db_backup:
    #         db.backup(db_backup, pages=3, progress=None)
    #     # показывает статус бар на 5 секунд
    #     my_win.statusbar.showMessage(
    #         "Резервное копирование базы данных завершено успешно", 5000)
    # except sqlite3.Error as error:
    #     # показывает статус бар на 5 секунд
    #     my_win.statusbar.showMessage(
    #         "Ошибка при копировании базы данных", 5000)
    # finally:
    #     if (db_backup):
    #         db_backup.close()
    #         db.close()
    #         my_win.close()


def title_id():
    """возвращает title id в зависимости от соревнования"""
    name = my_win.lineEdit_title_nazvanie.text()  # определяет название соревнований из титула
    if name != "":       
        data = my_win.dateEdit_start.text()
        gamer = my_win.lineEdit_title_gamer.text()
        age = my_win.lineEdit_title_vozrast.text()
        titles_data = Title.select().where((Title.name == name) & (Title.gamer == gamer)) # получает эту строку в db
        titles = titles_data.select().where((Title.data_start == data) & (Title.vozrast == age)).get()
        title_id = titles.id
    else:
        # получение последней записи в таблице
        t_id = Title.select().order_by(Title.id.desc()).get()
        title_id = t_id.id
    return title_id


def system_id(stage):
    """получения id системы данного этапа"""
    systems = System.select().where((System.stage == stage) & (System.title_id == title_id())).get()
    id_system = systems.id
    return id_system


def func_zagolovok(canvas, doc):
    """создание заголовка страниц"""
    pagesizeW = doc.width
    pagesizeH = doc.height
    current_button = "0"
    if pagesizeH > pagesizeW:
        pv = A4
    else:
        pv = landscape(A4)
    (width, height) = pv

    title = Title.get(Title.id == title_id())

    nz = title.name
    ind = title.perenos
    ms = title.mesto
    sr = f"среди {title.sredi} {title.vozrast}"
    data_comp = data_title_string()
    # === определяет если список судей то подпись ставит выше
    tb = my_win.tabWidget.currentIndex()
    if tb == 7:
        for i in my_win.tabWidget.findChildren(QRadioButton): # перебирает радиокнопки и определяет какая отмечена
                if i.isChecked():
                    current_button = i.text()
                    break
    # =========   
    canvas.saveState()
    canvas.setFont("DejaVuSerif-Italic", 14)
    # центральный текст титула
    if ind > 0 and pv == A4:
        nz_list = nz.split()
        word = nz_list[ind]
        s1 = nz.find(word) + len(word)
        strline1 = nz[:s1]
        strline2 = nz[s1 + 1:]
        # nz = f"{strline1}\n{strline2}"
        canvas.drawCentredString(width / 2.0, height - 1.1 * cm, strline1)
        canvas.drawCentredString(width / 2.0, height - 1.5 * cm, strline2)
        canvas.setFont("DejaVuSerif-Italic", 11)
        canvas.drawCentredString(width / 2.0, height - 1.9 * cm, sr)
        canvas.setFont("DejaVuSerif-Italic", 10)
    else:
        canvas.drawCentredString(width / 2.0, height - 1.1 * cm, nz)
        canvas.drawCentredString(width / 2.0, height - 1.5 * cm, sr)
        canvas.setFont("DejaVuSerif-Italic", 11)
    # текста титула по основным
    # canvas.drawCentredString(width / 2.0, height - 1.5 * cm, sr)
    canvas.drawRightString(width - 1 * cm, height - 1.9 * cm, f"г. {ms}")  # город
    canvas.drawString(0.8 * cm, height - 1.9 * cm, data_comp)  # дата начала
    # ==== текст судейская коллегия
    canvas.setFont("DejaVuSerif-Italic", 11)
    canvas.setFillColor(blue)  # меняет цвет шрифта списка судейской коллеги
    if pv == landscape(A4):
        main_referee_collegia = f"Гл. судья: судья {title.kat_ref}______________ {title.referee}   " \
                                f"Гл. секретарь: судья {title.kat_sec} ______________{title.secretary}"
        if current_button == "3":
            # текста титула по основным
            canvas.drawCentredString(width / 2.0, height - 15 * cm, main_referee_collegia)
        else:
            # текста титула по основным
            canvas.drawCentredString(width / 2.0, height - 20 * cm, main_referee_collegia)
    else:
        main_referee = f"Гл. судья: судья {title.kat_ref} ______________{title.referee}"
        if current_button == "1":
            # подпись главного судьи
            canvas.drawString(2 * cm, 20 * cm, main_referee)
        elif current_button == "2":
            regions_list = [] # сдвигает подпись судьи относительно кол-во регионов
            regions = Player.select().where(Player.title_id == title_id())
            for k in regions:
                region = k.region
                regions_list.append(region)
            regions_set = set(regions_list)
            count = len(regions_set)
            # подпись главного судьи
            canvas.drawString(2 * cm, ((29 - count ) * cm), main_referee)
        else:
            main_secretary = f"Гл. секретарь: судья {title.kat_sec} ______________{title.secretary} "
            # подпись главного судьи
            canvas.drawString(2 * cm, 1.8 * cm, main_referee)
            # подпись главного секретаря
            canvas.drawString(2 * cm, 0.8 * cm, main_secretary)
    canvas.restoreState()
    return func_zagolovok


def tbl(stage, kg, ts, zagolovok, cW, rH):
    """данные таблицы и применение стиля и добавления заголовка столбцов
    tdt_new - [[[участник],[регион счет в партиях]]]"""
    from reportlab.platypus import Table
    group_dict = {}
    cell_list_tmp = []
    cell_list = [] # список номеров ячеек где большие фамилиии
    tdt_temp = []
    tdt_new_tmp = []
    dict_tbl = {}
    tdt_all = table_data(stage, kg)  # данные результатов в группах
    # данные результатов победителей в группах для окрашивания очков в красный цвет
    tdt_new = tdt_all[0]
    # убирает id от фамилии и перезаписывает tdt_new
    g = 0
    for group in tdt_new:
        l = 0
        for z in group:
            if l % 2 == 0:
                fam_id = z[1]
                znak = fam_id.find("/")
                if znak != -1:
                    family = fam_id[:znak]
                else:
                    family = fam_id
                z[1] = family 
            # === определяет большие фамилии
                length_of_last_name = len(z[1])
                if length_of_last_name > 17:
                    cell_list_tmp.append(l)
            l += 1
        cell_list = cell_list_tmp.copy()
        group_dict[g] = cell_list
        cell_list_tmp.clear()
        g += 1

    for k in tdt_new:
        tdt_temp = k.copy()
        k.clear()
        tdt_new_temp = tdt_temp.copy()
        tdt_new_tmp.append(tdt_new_temp)
        tdt_temp.clear()
    tdt_new.clear()
    tdt_new = tdt_new_tmp.copy()
    # ===========================
    for i in range(0, kg):
        tdt_new[i].insert(0, zagolovok)       
        dict_tbl[i] = Table(tdt_new[i], colWidths=cW, rowHeights=rH)
        list_cells = group_dict[i]
        ts.add('TEXTCOLOR', (0, 0), (-1, -1), colors.darkblue)
        for m in list_cells:
            ts.add('FONTSIZE', (1, m + 1), (1, m + 1), 5)
            # ts.add('TEXTCOLOR', (1, m + 1), (1, m + 1), colors.green)
        # ставит всю таблицу в синий цвет
        for k in tdt_all[1][i]:
            col = k[0]  # столбец очков победителя
            row = k[1]  # ряд очков победителя
            ts.add('TEXTCOLOR', (col, row + 1), (col, row + 1), colors.red)  # красный цвет очков победителя
        dict_tbl[i].setStyle(ts)  # применяет стиль к таблице данных
    return dict_tbl


def tbl_begunki(ts, stage, number_group, tours, list_tours):
    """данные таблицы и применение стиля и добавления заголовка столбцов
    tdt_new - [[[участник],[регион счет в партиях]]]"""
    msgBox = QMessageBox
    stiker = []
    group_list = ["Предварительный", "1-й полуфинал", "2-й полуфинал"]
    final_type = "круг"
    from reportlab.platypus import Table
    systems = System.select().where(System.title_id == title_id())
    result = Result.select().where(Result.title_id == title_id())
    titles = Title.select().where(Title.id == title_id()).get()
    gamer_txt = titles.gamer
    gm = gamer_txt[:1]
    # ==== новый вариант с использованием system id   
    id_system = system_id(stage=number_group) if stage == "Финальный" else system_id(stage)
    # ========
    system = systems.select().where(System.id == id_system).get()
    final_type = system.type_table
     # # кол-во столбцов в таблице и их ширина
    cW = (1.6 * cm)
    rH = (0.6 * cm, 0.9 * cm, 1 * cm, 0.6 * cm, 0.6 * cm, 0.6 * cm, 0.6 * cm, 0.6 * cm,
           0.5 * cm, 0.5 * cm)
    dict_tbl = {}
    tdt_new_tmp = []

    if final_type == "сетка":
        result_setka = result.select().where(Result.number_group == number_group)
        result_all = result_setka.select().where((Result.player1 != "") & (Result.player2 != ""))
        result_group = result_all.select().where(Result.winner.is_null())
    else:    
        if number_group == "все" and tours == "все":
            result_group = result.select().where(Result.system_id == id_system)
        elif number_group == "все" and tours == "диапазон":
            result_group = result.select().where((Result.system_id == id_system) & (Result.round.in_(list_tours)))
        elif number_group == "все" and tours == "несыгранные":
            result_group = result.select().where((Result.system_id == id_system) & (Result.winner.is_null()))
        elif number_group != "все" and tours == "все":
            if stage in group_list:
                result_group = result.select().where((Result.system_id == id_system) & (Result.number_group == number_group))
            else:
                result_group = result.select().where(Result.system_id == id_system)
        elif number_group != "все" and tours == "диапазон":
            result_group = result.select().where((Result.number_group == number_group) & (Result.round.in_(list_tours)))
 
    shot_stage = ""
    count = len(result_group)
    if count == 0:
        msgBox.information(my_win, "Уведомление", "Нет не сыгранных встреч,\nв печати бегунков\nнет необходимости.")
        return
    for res_id in result_group:
        tours = res_id.tours # номера игроков в туре
        pl1 = res_id.player1 # 1-й игроков и его город в туре
        pl2 = res_id.player2 # 2-й игроков и его город в туре
        st = res_id.number_group # этап
        n_gr = ""
        if stage == "Предварительный":
            shot_stage = "ПР"
            mark = st.find(" ")
            gr = st[:mark]
            sys_stage = f"{shot_stage}"
            n_gr = f"{gr}гр"
        elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
            shot_stage = "ПФ1" if stage == "1-й полуфинал" else "ПФ2"
            mark = st.find(" ")
            gr = st[:mark]
            sys_stage = f"{shot_stage}"
            n_gr = f"{gr}гр"
        elif stage == "Финальный":
            n_gr = gm
            shot_stage = "Ф"
            mark = st.find("-")
            sys_stage = f"{st[:mark]}{shot_stage}"
        elif stage == "Одна таблица":
            n_gr = gm
            sys_stage = "Ф"
        round = res_id.round # раунд
        s1 = pl1.find("/")  
        s2 = pl2.find("/")   
        player1 = pl1[:s1]
        dlina_pl1 = len(player1)
        if dlina_pl1 >= 16:
            player1 = short_player_begunki(fam=player1)
        city1 = pl1[s1 + 1:]
        player2 = pl2[:s2]
        dlina_pl2 = len(player2)
        if dlina_pl2 >= 16:
            player2 = short_player_begunki(fam=player2)
        city2 = pl2[s2 + 1:]
        pl1 = f"{player1}\n{city1}" # делает фамилия и город на разнызх строчках
        pl2 = f"{player2}\n{city2}"
            # список строк бегунка
        d_tmp = [[n_gr, 'тур', 'вст', 'стол'],
                [sys_stage, round, tours, ''],
                [pl1, '', pl2, ''],
                ['', '', '', ''],
                ['', '', '', ''],
                ['', '', '', ''],
                ['', '', '', ''],
                ['', '', '', ''],
                ['общ счет:', '', '', ''],
                ['Победитель', '', '', '']]
        tdt_temp = d_tmp.copy()
        d_tmp.clear()
        tdt_new_temp = tdt_temp.copy()
        tdt_new_tmp.append(tdt_new_temp)
        tdt_temp.clear()
    game = len(tdt_new_tmp)
        # ===========================
    for i in range(0, game):      
        dict_tbl[i] = Table(tdt_new_tmp[i], colWidths=cW, rowHeights=rH)
        dict_tbl[i].setStyle(ts)  # применяет стиль к таблице данных
    stiker.append(dict_tbl)
    stiker.append(game)
    return stiker


def short_player_begunki(fam):
    """Сокращает имя игрока до одной буквы, если длина фамилии и имя более 15 символов"""
    mark = fam.find(" ")
    family = fam[:mark]
    name = fam[mark +1:mark + 2]
    full_name = f"{family} {name}."
    return full_name


def begunki_made():
    """создание бегунков"""
    from sys import platform
    from reportlab.platypus import Table
    msgBox = QMessageBox
    system = System.select().where(System.title_id == title_id())  # находит system id последнего
    result = Result.select().where(Result.title_id == title_id())
    number_group = my_win.comboBox_select_group_begunki.currentText()
    stage = my_win.comboBox_select_stage_begunki.currentText()
    tours = my_win.comboBox_select_tours.currentText()

    id_system = system_id(stage=number_group) if stage == "Финальный" else system_id(stage)
    elements = []
    ts = []
    tblstyle = []
    for p in range(0, 8):
        fn = ('SPAN',(0, 2 + p), (1, 2 + p))
        tblstyle.append(fn)
        fn = ('SPAN',(2, 2 + p), (3, 2 + p))
        tblstyle.append(fn)

    ts.append(tblstyle)
    ts = TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
                        ('BOX', (0,0), (-1,-1), 1, colors.black)]
                        + tblstyle +
                        [('FONTSIZE', (0, 1), (0, 1), 20),
                        ('VALIGN', (0, 1), (0, 1), 'TOP'),
                        ('ALIGN',(0, 1), (0, 1),'CENTER'),
                        ('FONTSIZE', (0, 2), (3, 2), 7), 
                        ('VALIGN', (1, 0), (3, 0), 'MIDDLE'),
                        ('FONTSIZE', (1, 1), (3, 1), 12), 
                        ('VALIGN', (1, 1), (3, 1), 'MIDDLE'),
                        ('ALIGN',(1, 1), (3, 1),'CENTER'),
                        ('FONTSIZE', (0, 0), (0, 0), 12), 
                        ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                        ('ALIGN',(0, 0), (0, 0),'CENTER')])
    #  ========= формирование диапазона печати бегунков ==========
    sys = system.select().where(System.id == id_system).get()
    final_type = sys.type_table
    list_tours = []
    if final_type == "сетка":
        # list_tours.append("несыгранные")
        result_setka = result.select().where(Result.number_group == number_group)
        result_all = result_setka.select().where((Result.player1 != "") & (Result.player2 != ""))
        result_fin = result_all.select().where(Result.winner.is_null())
        if len(result_fin) == 0:
            msgBox.information(my_win, "Уведомление", "Все встречи сыграны,\n в печати бегунков нет необходимости.")
            return
    elif final_type == "круг" or final_type == "группы":
        if tours != "все":
            range_tours_str = my_win.lineEdit_range_tours.text()
            txt = range_tours_str.replace(" ", "")
            range_tours_list = list(txt)
            if "-" in range_tours_list:
                range_tours_list.remove("-")
                result_int = [int(item) for item in range_tours_list] # преобразовывает список строковых данных в числовой тип
                for b in range (result_int[0], result_int[1] + 1):
                    b = int(b)
                    list_tours.append(b)
            else:
                tours_list = range_tours_list
                for b in tours_list:
                    if b != ",":
                        b = int(b)
                        list_tours.append(b)
        else:
            if number_group != "все":
                if stage == "1-й полуфинал" or stage == "2-й полуфинал":
                    result_group = result.select().where((Result.system_id == id_system) & (Result.number_group == number_group))
                else:   
                    result_group = result.select().where(Result.system_id == id_system)
                for i in result_group:
                    r = int(i.round)
                    if r not in list_tours:
                        list_tours.append(r)
        
    stiker = tbl_begunki(ts, stage, number_group, tours, list_tours) # здесь надо менять данные бегунков
    if stiker == None: # если все встречи сыграны нет смысла печать бегунков выход из функции
        return
    dict_table = stiker[0]
    game = stiker[1]

    data_tmp = []
    data_temp = []
    tmp = []
    temp = []
    data = []
    celoe = game // 3
    ostatok = game % 3
    end = 0
    row = 3
    if ostatok == 0:
        end = celoe + 1
    else:
        end = celoe + 2
    a = 0
    for k in range(1, end):
        if ostatok !=0 and k == end - 1:
            row = ostatok
        for i in range(0, row): # кол-во бегунков в 
            data_tmp.append(dict_table[a])
            a += 1
        tmp = data_tmp.copy()
        data_temp.append(tmp) 
        temp = data_temp.copy()
        data.append(temp)
        data_tmp.clear()
        data_temp.clear()
    shell_table = []
    count_data = len(data)
    s_tmp = []
    for l in range(0, count_data): 
        shell_tmp = Table(data[l], colWidths=["*"])
        s_tmp.append(shell_tmp)
        tmp_copy = s_tmp.copy()
        shell_table.append(tmp_copy)
        s_tmp.clear()
        elements.append(shell_table[l][0])
 
    name_table = "begunki.pdf"
    # устанавливает поля на странице pdf
    doc = SimpleDocTemplate(name_table, pagesize=A4, rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(elements)
    my_win.lineEdit_range_tours.clear()
    my_win.lineEdit_range_tours.hide()
    view_file = name_table
    if platform == "darwin":  # OS X
        os.system(f"open {view_file}")
    elif platform == "win32":  # Windows...
        os.system(f"{view_file}")
    os.chdir("..") # возврат на предыдущий уровень


def select_stage_for_begunki():
    """выбор финалов или номеров групп для печати бегунков"""
    group_etap_list = ["Предварительный", "1-й полуфинал", "2-й полуфинал"]
    my_win.comboBox_select_group_begunki.clear()
    if my_win.comboBox_select_stage_begunki.currentIndex() != 0:
        my_win.Button_print_begunki.setEnabled(True)
    systems = System.select().where(System.title_id == title_id())
    group_list = ["все"]
    stage = my_win.comboBox_select_stage_begunki.currentText()

    if stage in group_etap_list:
        id_system = system_id(stage)
        sys_id = systems.select().where(System.id == id_system).get()
        group = sys_id.total_group
        group_list = [f"{i} группа" for i in range(1, group + 1)] # генератор списка
        group_list.insert(0, "все")
    elif stage == "Одна таблица":
        pass
    else:
        for k in systems:
            if k.stage not in group_etap_list:
                group_list.append(k.stage)
    my_win.comboBox_select_group_begunki.addItems(group_list)
    if  stage in group_etap_list or stage == "-Выбор спортсменов-":
        my_win.comboBox_select_group_begunki.setCurrentIndex(0)
    else:
        my_win.comboBox_select_group_begunki.setCurrentIndex(1)

        
def select_tour_for_begunki():
    """выбор номеров тура или диапазона туров""" 
    my_win.comboBox_select_tours.clear()
    tour_list = ["все", "несыгранные", "диапазон"]
    my_win.comboBox_select_tours.addItems(tour_list)
    index = my_win.comboBox_select_tours.currentIndex()
    if index != 0:
        my_win.lineEdit_range_tours.show()


def select_diapazon():
    """показывает поле для ввода дмапазона туров"""
    my_win.lineEdit_range_tours.clear()
    index = my_win.comboBox_select_tours.currentIndex()
    if index == 0 or index == 1:
        my_win.lineEdit_range_tours.hide()
    else:
        my_win.lineEdit_range_tours.show()
        my_win.lineEdit_range_tours.setFocus()


def enter_print_begunki():
    """Печать бегунков при нажатии энтер на поле диапазона"""
    sender = my_win.sender()
    if sender == my_win.lineEdit_range_tours:
        begunki_made()


def merdge_pdf_files():
    """Слияние все таблиц соревнований в один файл"""
    pdf_merger = PdfMerger()

    title = Title.get(Title.id == title_id())
    pdf_files_list = []
    short_name = title.short_name_comp
 
    count = my_win.tableWidget.rowCount()
    for k in range(0, count):
        name_files = my_win.tableWidget.item(k, 1).text()
        pdf_files_list.append(name_files)

    my_win.tableWidget.setColumnCount(2) # устанавливает колво столбцов
    my_win.tableWidget.setRowCount(count)
    column_label = ["№", "Этапы"]

    my_win.tableWidget.setHorizontalHeaderLabels(column_label) # заголовки столбцов в tableWidget
    my_win.tableWidget.setDragDropOverwriteMode(True)
    my_win.tableWidget.setSelectionMode(QAbstractItemView.SingleSelection)
    my_win.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
    my_win.tableWidget.show()
    catalog = 1 # переходит в каталог /table_pdf и считывает все файлы этого соревнования
    change_dir(catalog)
    with contextlib.ExitStack() as stack:
        files = [stack.enter_context(open(pdf, 'rb')) for pdf in pdf_files_list]
        for f in files:
            pdf_merger.append(f)
        os.chdir("..")
        catalog = 2
        change_dir(catalog)
        with open(f'{short_name}.pdf', 'wb') as f:
            pdf_merger.write(f)
            title.pdf_comp = f'{short_name}.pdf'
            title.save()
    my_win.tableWidget.show()


# def sign_referee_on_page(f):
#     """создание подписи судей на листах соревнований в PDF"""
#     title = Title.select().where(Title.id == title_id()).get()
#     # short_name = title.short_name_comp
#     name_f = f.name

#     can = Canvas(name_f)
 
#     from PyPDF2 import PdfFileWriter, PdfFileReader
#     import io
 
#     in_pdf_file = name_f
#     out_pdf_file = 'with_image.pdf'
#     fname = QFileDialog.getOpenFileName(
#         my_win, "Выбрать файл подписи", "", "Image files(*.jpg)")
#     # canvas = Canvas(f"{name_f}", pagesize=A4)
#     filepatch = str(fname[0])
#     # img_file = '../../static/img/code_maven_440x440.png'
 
#     packet = io.BytesIO()
#     can = Canvas(packet)
#     #can.drawString(10, 100, "Hello world")
#     x_start = 0
#     y_start = 0
#     can.drawImage(filepatch, x_start, y_start, width=120, preserveAspectRatio=True, mask='auto')
#     can.showPage()
#     # can.showPage()
#     # can.showPage()
#     can.save()
 
#     # #move to the beginning of the StringIO buffer
#     # packet.seek(0)
 
#     # # new_pdf = PdfFileReader(packet)
 
#     # # read the existing PDF
#     # existing_pdf = PdfFileReader(open(in_pdf_file, "rb"))
#     # output = PdfFileWriter()
 
#     # # for i in range(len(existing_pdf.pages)):
#     # #     page = existing_pdf.getPage(i)
#     # #     page.mergePage(new_pdf.getPage(i))
#     # #     output.addPage(page)
 
#     # outputStream = open(out_pdf_file, "wb")
#     # output.write(outputStream)
#     # outputStream.close()
 
 
# # create_pdf(f)
# # add_image()
# # ==========



def load_name_net_after_choice_for_wiev(fin):
    """загружает список сетки после жеребьевки для ее просмотра"""
    system = System.select().where(System.title_id == title_id())
    for k in system:
        stage = k.stage
        if stage == fin:
            vid_net = k.label_string
            break
    if vid_net == 'Сетка (с розыгрышем всех мест) на 8 участников':
        setka_8_full_made(fin)
    elif vid_net == 'Сетка (-2) на 8 участников':
        setka_8_2_made(fin)
    elif vid_net == 'Сетка (с розыгрышем всех мест) на 16 участников':
        setka_16_full_made(fin)
    elif vid_net == 'Сетка (-2) на 16 участников':
        setka_16_2_made(fin)
    elif vid_net == 'Сетка (с розыгрышем всех мест) на 32 участников':
        setka_32_full_made(fin)
    elif vid_net == 'Сетка (-2) на 32 участников':
        setka_32_2_made(fin)
    elif vid_net == 'Сетка (1-3 место) на 32 участников':
        setka_32_made(fin)


def table_made(pv, stage):
    """создание таблиц kg - количество групп(таблиц), g2 - наибольшое кол-во участников в группе
     pv - ориентация страницы, е - если участников четно группам, т - их количество"""
    sender = my_win.sender()
    stage_list_sf = ["1-й полуфинал", "2-й полуфинал"]
    from reportlab.platypus import Table
     # ==== новый вариант с использованием system id
    id_system = system_id(stage)
    system = System.select().where((System.title_id == title_id()) & (System.id == id_system)).get()  # находит system id последнего
    titles = Title.select().where(Title.id == title_id()).get()
    sex = titles.gamer 

    if stage in stage_list_sf: # если этап полуфинал
        kg = system.total_group  # кол-во групп
        max_pl = system.max_player // kg 
    elif stage == "Предварительный":
        kg = system.total_group  # кол-во групп
        max_pl = system.max_player
    else: # игры в финале по кругу или одна круговая таблица
        kg = 1
        max_pl = system.max_player
        
    family_col = 3.2
    if pv == "альбомная":  # альбомная ориентация стр
        pv = landscape(A4)
        center_stage = 210 # откуда начинается надпись -предварительный этап-
        if kg == 1 or max_pl in [10, 11, 12, 13, 14, 15, 16, 17]:
            # ширина столбцов таблицы в зависимости от кол-во чел (1 таблица)
            wcells = 21.4 / max_pl
        else:
            # ширина столбцов таблицы в зависимости от кол-во чел (2-ух в ряд)
            wcells = 7.4 / max_pl
    else:  # книжная ориентация стр
        pv = A4
        center_stage = 140 # откуда начинается надпись -предварительный этап-
        if max_pl < 7:
            family_col = 3.8
            wcells = 12.0 / max_pl  # ширина столбцов таблицы в зависимости от кол-во чел
            # wcells = round(wcells, 2)
        else:
            family_col = 3.8
            wcells = 12.8 / max_pl  # ширина столбцов таблицы в зависимости от кол-во чел
        wcells = round(wcells, 2)

    col = ((wcells * cm,) * max_pl)
    elements = []

    # кол-во столбцов в таблице и их ширина
    cW = ((0.4 * cm, family_col * cm) + col + (0.8 * cm, 1 * cm, 1 * cm))

    if kg == 1:
        if max_pl > 16:
            rH = (0.42 * cm)  # высота строки
        else:
            rH = (0.45 * cm)  # высота строки
    else:
        if max_pl < 5:
            rH = (0.34 * cm)  # высота строки
        else:
            rH = (0.3 * cm)  # высота строки
    num_columns = []  # заголовки столбцов и их нумерация в зависимости от кол-во участников

    for i in range(max_pl):
        i += 1
        i = str(i)
        num_columns.append(i)
    zagolovok = (['№', 'Участники/ Город'] + num_columns + ['Очки', 'Соот', 'Место'])

    tblstyle = []
    # =========  цикл создания стиля таблицы ================
    for q in range(1, max_pl + 1):  # город участника делает курсивом
        # город участника делает курсивом
        fn = ('FONTNAME', (1, q * 2), (1, q * 2), "DejaVuSerif-Italic")
        tblstyle.append(fn)
        fn = ('FONTNAME', (1, q * 2 - 1), (1, q * 2 - 1), "DejaVuSerif-Bold")  # участника делает жирным шрифтом
        tblstyle.append(fn)
        # центрирование текста в ячейках)
        fn = ('ALIGN', (1, q * 2 - 1), (1, q * 2 - 1), 'LEFT')
        tblstyle.append(fn)
        # объединяет 1-2, 3-4, 5-6, 7-8 ячейки 1 столбца
        fn = ('SPAN', (0, q * 2 - 1), (0, q * 2))
        tblstyle.append(fn)
        # объединяет клетки очки
        fn = ('SPAN', (max_pl + 2, q * 2 - 1), (max_pl + 2, q * 2))
        tblstyle.append(fn)
        # объединяет клетки соот
        fn = ('SPAN', (max_pl + 3, q * 2 - 1), (max_pl + 3, q * 2))
        tblstyle.append(fn)
        # объединяет клетки  место
        fn = ('SPAN', (max_pl + 4, q * 2 - 1), (max_pl + 4, q * 2))
        tblstyle.append(fn)
        # объединяет диагональные клетки
        fn = ('SPAN', (q + 1, q * 2 - 1), (q + 1, q * 2))
        tblstyle.append(fn)
        fn = ('BACKGROUND', (q + 1, q * 2 - 1), (q + 1, q * 2), colors.lightgreen)  # заливает диагональные клетки
        tblstyle.append(fn)

    ts = []
    ts_grid = []
    # создание внутренней сетки таблицы (столбец, строка), (столбец, строка) (2, 2) - вид пунктирной линии
    for p in range(0, max_pl * 2 + 1):
        if p % 2 == 0:
            tsg = ('LINEBELOW', (1, p + 2), (-1, p + 2), 0.25, colors.black)
        else:
            tsg = ('LINEBELOW', (2, p), (-1, p), 0.25, colors.grey, None, (1, 1)) # пунктирная линия под ячейкой 2-й столбец, 3-я строка (начинается с 0)
        ts_grid.append(tsg)


    ts.append(tblstyle)
    ts.append(ts_grid)
    # ============= полный стиль таблицы ======================
    ts = TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                     ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                     ('FONTSIZE', (0, 0), (-1, -1), 6),
                     # вставить размер шрифта конкретной ячейки под длинную фамилию
                     ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                     ('FONTNAME', (0, 0), (max_pl + 5, 0), "DejaVuSerif-Bold"),
                     ('VALIGN', (0, 0), (max_pl + 5, 0), 'MIDDLE'), # центрирование текста в ячейках вертикальное
                     ('BOTTOMPADDING', (0, 0), (-1, -1), 0)]
                    + tblstyle +
                    [('BACKGROUND', (0, 0), (max_pl + 5, 0), colors.yellow),
                     # цвет шрифта в ячейках
                     ('TEXTCOLOR', (0, 0), (-1, -1), colors.darkblue),
                     ('LINEABOVE', (0, 0), (-1, 1), 1, colors.black),  # цвет линий нижней
                     # цвет и толщину внутренних линий
                     ('INNERGRID', (0, 0), (1, -1), 0.25, colors.black),

                     ('LINEAFTER', (1, 0), (-1, -1), 0.25, colors.black)] # линия справа 1-я ячейка (1-й столбец, 0-я строка), 2-я ячейка (1-й столбец, строка до конца) (начинается с 0)
                    + ts_grid +
                    [('BOX', (0, 0), (-1, -1), 2, colors.black)])  # внешние границы таблицы

    #  ============ создание таблиц и вставка данных =================
    h1 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic",
            leftIndent=center_stage, spacebefore=10, textColor="green")  # стиль параграфа ()

    h2 = PS("normal", fontSize=10, fontName="DejaVuSerif-Italic",
            leftIndent=290, spacebefore=20, textColor="brown")  # стиль параграфа (номера таблиц)
            #========
    
    dict_table = tbl(stage, kg, ts, zagolovok, cW, rH)
    if kg == 1:  # одна таблицу
        data = [[dict_table[0]]]
        # shell_table = Table(data, colWidths=["*"])
        shell_table = Table(data, colWidths=[28 * cm])
        text = ""
        elements.append(Paragraph(text, h2))
        elements.append(shell_table)
    else:
        data_tmp = []
        data_temp = []
        tmp = []
        temp = []
        data = []
        if pv == landscape(A4):  # страница альбомная, то таблицы размещаются обе в ряд
            for k in range(1, kg // 2 + 1):
                data_tmp = [dict_table[(k * 2 - 2) + 0], dict_table[(k * 2 - 2) + 1]]
                tmp = data_tmp.copy()
                data_temp.append(tmp) 
                temp = data_temp.copy()
                data.append(temp)
                data_tmp.clear()
                data_temp.clear()
            shell_table = []
            s_tmp = []
            for l in range(0, kg // 2): 
                shell_tmp = Table(data[l], [410, 410])
                shell_tmp.setStyle(TableStyle([('VALIGN',(0, 0), (-1, -1), 'TOP')]))
                gr_1 = f'группа {l * 2 + 1}'
                gr_2 = f'группа {l * 2 + 2}'
                s_tmp.append(shell_tmp)
                tmp_copy = s_tmp.copy()
                shell_table.append(tmp_copy)
                s_tmp.clear()
                text = gr_1 + (' ' * 10) + gr_2
                elements.append(Paragraph(text, h2))
                elements.append(shell_table[l][0])
        else:  # страница книжная, то таблицы размещаются в столбец
            for k in range(1, kg // 2 + 1):
                for i in range(0, kg):
                    data_tmp.append(dict_table[i])  
                    tmp = data_tmp.copy()
                    data_temp.append(tmp) 
                    temp = data_temp.copy()
                    data.append(temp)
                    data_tmp.clear()
                    data_temp.clear()
            shell_table = []
            s_tmp = []
            for l in range(0, kg): 
                shell_tmp = Table(data[l], colWidths=["*"])
                s_tmp.append(shell_tmp)
                tmp_copy = s_tmp.copy()
                shell_table.append(tmp_copy)
                s_tmp.clear()
                elements.append(Paragraph(f'группа {l + 1}', h2))
                elements.append(shell_table[l][0])

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    short_name = t_id.short_name_comp

    if stage == "Одна таблица":
        title = "Финальные соревнования. Одиночный разряд"
        name_table = f"{short_name}_one_table.pdf"
    elif stage == "Предварительный":
        title = "Квалификационные соревнования"
        name_table = f"{short_name}_table_group.pdf"
    elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
        txt = stage.rfind("-")
        number_fin = stage[:txt]
        title = stage
        name_table = f"{short_name}_{number_fin}-semifinal.pdf"
    else:
        txt = stage.rfind("-")
        number_fin = stage[:txt]
        sys = System.select().where(System.id == id_system).get()
        max_pl = sys.max_player # максимальное число игроков в сетке  
        first_mesto = mesto_in_final(fin=stage)
        last_mesto = max_pl if stage == "1-й финал" else first_mesto + max_pl - 1
        title = f'Финальные соревнования.({first_mesto}-{last_mesto} место). Одиночный разряд' # титул на таблице
        name_table = f"{short_name}_{number_fin}-final.pdf"
    doc = SimpleDocTemplate(name_table, pagesize=pv)
    catalog = 1
    change_dir(catalog)
    doc.topMargin = 1.8 * cm # высота отступа от верха листа pdf
    if sender == my_win.indent_edit_Action:
        indent = change_indent_page()
        doc.leftMargin = indent * cm
    # doc.leftMargin = 0 * cm
    elements.insert(0, (Paragraph(f"{title}. {sex}", h1)))
    doc.build(elements, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")


def list_regions_pdf():
    """список субъектов РФ"""
    from reportlab.platypus import Table
    story = []  # Список данных таблицы участников
    elements = []  # Список Заголовки столбцов таблицы
    region_list = []
    tit = Title.get(Title.id == title_id())
    short_name = tit.short_name_comp
    regions = Player.select().where((Player.title_id == title_id()) & (Player.player != "x"))

    for k in regions:
        reg = k.region
        if reg not in region_list:
            region_list.append(reg)

    kp = len(region_list)
    region_list.sort()
    n = 0
    for reg in region_list:
        n += 1
        num = n
        data = [num, reg]
        elements.append(data)
    elements.insert(0, ["№", "Субъекты РФ"])
    t = Table(elements,
              colWidths=(1.0 * cm, 10.0 * cm),
              rowHeights=None, repeatRows=1)  # ширина столбцов, если None-автоматическая
    t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                           ('FONTSIZE', (0, 0), (-1, -1), 10),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 2),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (-1, -1), 'CENTER'), # (1-я ячейка столб, ряд)  (2-я ячейка столб, ряд)
                        #    ('ALIGN', (0, 0), (0, kp), 'CENTER'), # (1-я ячейка столб, ряд)  (2-я ячейка столб, ряд)
                           ('ALIGN', (1, 1), (1, kp), 'LEFT'),
                           ('BACKGROUND', (0, 0), (1, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (1, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.8, colors.black)
                           ]))


    h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=150, textColor="green",
            firstLineIndent=-20)  # стиль параграфа
    h3.spaceAfter = 10  # промежуток после заголовка
    story.append(Paragraph(f'Список субъектов РФ', h3))
    story.append(t)

    doc = SimpleDocTemplate(f"{short_name}_regions_list.pdf", pagesize=A4)
    catalog = 1
    change_dir(catalog)
    doc.build(story, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")


def list_winners_pdf():
    """список призеров"""
    from reportlab.platypus import Table
    story = []  # Список данных таблицы участников
    elements = []  # Список Заголовки столбцов таблицы
    tit = Title.get(Title.id == title_id())
    short_name = tit.short_name_comp
   
    count_Row = my_win.tableWidget.rowCount()
    kp = count_Row + 1
    n = 0
    for l in range(0, count_Row):
        mesto = my_win.tableWidget.item(l, 0).text()
        player = my_win.tableWidget.item(l, 1).text()
        bday = my_win.tableWidget.item(l, 2).text()
        rank = my_win.tableWidget.item(l, 3).text()
        city = my_win.tableWidget.item(l, 4).text()
        region = my_win.tableWidget.item(l, 5).text()
        razryad = my_win.tableWidget.item(l, 6).text()
        coach = my_win.tableWidget.item(l, 7).text()
        coach = chop_line(t=coach)
        n += 1

        data = [mesto, player, bday, rank, city, region, razryad, coach]
        elements.append(data)
    elements.insert(0, ["Место", "Фамилия, Имя", "Дата рождения", "Рейтинг", "Город", "Регион", "Разряд", "Тренеры"])
    t = Table(elements,
              colWidths=(2.3 * cm, 5.6 * cm, 3.0 * cm, 2.0 * cm, 3.0 * cm, 4.0 * cm, 2.0 * cm, 5.5 * cm,),
              rowHeights=None, repeatRows=1)  # ширина столбцов, если None-автоматическая
    t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                           ('FONTSIZE', (0, 0), (-1, -1), 9),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 1),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (0, 7), 'CENTER'),
                           ('ALIGN', (1, 0), (1, kp), 'LEFT'),
                           ('BACKGROUND', (0, 0), (8, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (8, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, kp * -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.8, colors.black)
                           ]))


    h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=200, textColor="green",
            firstLineIndent=-20)  # стиль параграфа
    h3.spaceAfter = 10  # промежуток после заголовка
    story.append(Paragraph(f'Список победителей и призеров', h3))
    story.append(t)

    doc = SimpleDocTemplate(f"{short_name}_winners_list.pdf", pagesize=landscape(A4))
    catalog = 1
    change_dir(catalog)
    doc.build(story, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")


def protokol_pdf():
    """Судейский протокол для таблицы в круг"""
    from sys import platform
    from reportlab.platypus import Table
    fin_list = []
    stage_list = ["Предварительный", "1-й полуфинал", "2-й полуфинал"]
    systems = System.select().where(System.title_id == title_id())
    for k in systems:
        stage_system = k.stage
        system_string = k.label_string
        if stage_system not in stage_list and system_string == "Круговая таблица на 16 участников":
            fin_list.append(stage_system)

    stage, ok = QInputDialog.getItem(my_win, "Финалы", "Выберите финал по кругу для\n"
                                        "создания судейского протокола", fin_list)
    fin = f'{stage[:1]}-fin'
    id_system = system_id(stage)
    #========
    story = []  # Список данных таблицы участников
    elements = []  # Список Заголовки столбцов таблицы
    result_list = Result.select().where((Result.title_id == title_id()) & (Result.system_id == id_system))
    titles = Title.get(Title.id == title_id())
    short_name = titles.short_name_comp
    gamer = titles.gamer
    count = len(result_list)  # количество записей в базе
    kp = count + 1
    n = 0
    for l in result_list:
        n += 1
        round = l.round
        randevy = l.tours
        player_1 = l.player1
        player_2 = l.player2
        win_pl = l.winner
  
        data = [n, round, randevy, player_1, player_2, win_pl]

        elements.append(data)
    elements.insert(0, ["№", "Тур", "Встреча", "Игрок-1", "Игрок-2", "Победитель", "Счет"])
    t = Table(elements,
              colWidths=(1.0 * cm, 0.9 * cm, 1.2 * cm, 5.7 * cm, 5.7 * cm, 3.5 * cm, 1.5 * cm),
              rowHeights=(0.43 * cm), repeatRows=1)  # ширина столбцов, если None-автоматическая
    t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 1),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (-1, kp * -1), 'CENTER'),
                           ('BACKGROUND', (0, 0), (8, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (8, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, kp * -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.5, colors.black)
                           ]))

    h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=150,
            firstLineIndent=-20, textColor="green")  # стиль параграфа
    h3.spaceAfter = 5  # промежуток после заголовка
    story.append(Paragraph(f'Протокол. {gamer}-{stage}', h3))
    story.append(t)
    
    doc = SimpleDocTemplate(f"{short_name}_protokol_{fin}.pdf", pagesize=A4, rightMargin=1*cm, leftMargin=1*cm, topMargin=0.8*cm, bottomMargin=0.8*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(story)
    view_file = f"{short_name}_protokol_{fin}.pdf"
    if platform == "darwin":  # OS X
        os.system(f"open {view_file}")
    elif platform == "win32":  # Windows...
        os.system(f"{view_file}")
    os.chdir("..")


def list_referee_pdf():
    """список судейской коллегии"""
    from reportlab.platypus import Table
    story = []  # Список данных таблицы участников
    elements = []  # Список Заголовки столбцов таблицы
    tit = Title.get(Title.id == title_id())
    short_name = tit.short_name_comp

    count_Row = my_win.tableWidget.rowCount()
    kp = count_Row + 1
    n = 0
    for l in range(0, count_Row):
        if n < 2: 
            num = my_win.tableWidget.item(l, 0).text()
            post = my_win.tableWidget.item(l, 1).text()
            fam_city = my_win.tableWidget.item(l, 2).text()
            category = my_win.tableWidget.item(l, 3).text()
        else:
            post_combobox = my_win.tableWidget.cellWidget(l, 1)
            post = post_combobox.currentText()
            fam_city_comboBox = my_win.tableWidget.cellWidget(l, 2)
            fam_city = fam_city_comboBox.currentText()
            category_combobox = my_win.tableWidget.cellWidget(l, 3)
            category = category_combobox.currentText()
        num = my_win.tableWidget.item(l, 0).text()
        n += 1
        data = [num, post, fam_city, category]
        elements.append(data)
    elements.insert(0, ["№", "Должность", "Фамилия Имя/ город", "Категория"])
    t = Table(elements,
              colWidths=(0.6 * cm, 7.0 * cm, 7.7 * cm, 3.2 * cm),
              rowHeights=None, repeatRows=1)  # ширина столбцов, если None-автоматическая
    t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                           ('FONTSIZE', (0, 0), (-1, -1), 10),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 1),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (-1, kp * -1), 'CENTER'),
                           ('BACKGROUND', (0, 0), (8, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (8, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, kp * -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.9, colors.black)
                           ]))


    h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=150, textColor="green",
            firstLineIndent=-20)  # стиль параграфа
    h3.spaceAfter = 10  # промежуток после заголовка
    story.append(Paragraph(f'Список главной судейской коллегии', h3))
    story.append(t)

    doc = SimpleDocTemplate(f"{short_name}_referee_list.pdf", pagesize=A4)
    catalog = 1
    change_dir(catalog)
    doc.build(story, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")


def setka_8_superfinal(fin):
    """сетка на 8 суперфинал в pdf"""
    sender = my_win.sender()
    from reportlab.platypus import Table
    table = "setka_8_superfinal"
    elements = []
    data = []
    style = []
    column = ['']
    column_count = column * 10
    # добавить в аргументы функции
    final = fin
    titles = Title.select().where(Title.id == title_id()).get()
    gamer = titles.gamer
    id_system = system_id(fin)
    sys = System.select().where(System.id == id_system).get()
    max_pl = sys.max_player # максимальное число игроков в сетке
    first_mesto = mesto_in_final(fin) if sender != my_win.clear_s8_full_Action else 1
    last_mesto = max_pl if fin == "1-й финал" else first_mesto + max_pl - 1
    fin_title = f'Финальные соревнования.({first_mesto}-{last_mesto} место)' # титул на таблице
    for i in range(0, 40):
        # column_count[9] = i  # нумерация 10 столбца для удобного просмотра таблицы
        list_tmp = column_count.copy()
        data.append(list_tmp)
    # ========= места ==========
    y = 0
    for i in range(0, 16, 2):
        y += 1
        data[i][0] = str(y)  # рисует начальные номера таблицы 1-16
    # ========= нумерация встреч сетки ==========
    draw_num(row_n=1, row_step=2, col_n=2, number_of_columns=3, number_of_game=1, player=8, data=data) # рисует номера встреч 1-32
    draw_num(row_n=16, row_step=2, col_n=6, number_of_columns=2, number_of_game=8, player=2, data=data) # рисует номера встреч 1-32
    draw_num(row_n=20, row_step=2, col_n=4, number_of_columns=2, number_of_game=9, player=4, data=data) # рисует номера встреч 1-32
    draw_num_lost(row_n=16, row_step=2, col_n=4, number_of_game=5, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=20, row_step=2, col_n=2, number_of_game=1, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=28, row_step=2, col_n=4, number_of_game=9, player=2, data=data) # номера минус проигравшие встречи -1 -16
   
    data[28][6] = str(12)  # создание номеров встреч 15
    data[13][6] = str(-7)
    data[18][6] = str(-8)
    data[25][6] = str(-11)
    data[30][6] = str(-12)
    # ============= данные игроков и встреч и размещение по сетке =============
    tds = write_in_setka(data, fin, first_mesto, table)
    #===============
    cw = ((0.3 * cm, 4.6 * cm, 0.4 * cm, 3.0 * cm, 0.4 * cm, 3.0 * cm, 0.4 * cm, 4.8 * cm, 1.5 * cm, 0.4 * cm))
    # основа сетки на чем чертить таблицу (ширина столбцов и рядов, их кол-во)
    color_mesta(data, first_mesto, table, fin) # раскрашивает места участников красным цветом
    t = Table(data, cw, 40 * [0.6 * cm])
    # =========  цикл создания стиля таблицы ================
    # ==== рисует основной столбец сетки 
    style = draw_setka(1, 1, 8, style) # рисует кусок сетки(номер столбца, номер строки на 8 человека)
    style = draw_setka(3, 20, 4, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(5, 28, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(5, 16, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    # ======= встречи за места =====
    for q in range(0, 7, 6):
        fn = ('LINEABOVE', (7, q + 8), (8, q + 8),
              1, colors.darkblue)  # за 1-2 место
        style.append(fn)
    for q in range(0, 3, 2):
        fn = ('LINEABOVE', (7, q + 17), (8, q + 17),
              1, colors.darkblue)  # за 3-4 место
        style.append(fn)
        fn = ('LINEABOVE', (7, q + 29), (8, q + 29),
              1, colors.darkblue)  # за 7-8 место
        style.append(fn)
    for q in range(0, 4, 3):
        fn = ('LINEABOVE', (7, q + 23), (8, q + 23),
              1, colors.darkblue)  # за 5-6 место
        style.append(fn)

    for i in range(1, 6, 2):
        fn = ('TEXTCOLOR', (i, 0), (i, 39), colors.black)  # цвет шрифта игроков
        style.append(fn)
        fn = ('TEXTCOLOR', (i + 1, 0), (i + 1, 39), colors.green)  # цвет шрифта номеров встреч
        style.append(fn)
        # выравнивание фамилий игроков по левому краю
        fn = ('ALIGN', (i, 0), (i, 39), 'LEFT') 
        style.append(fn)
        # центрирование номеров встреч
        fn = ('ALIGN', (i + 1, 0), (i + 1, 39), 'CENTER')
        style.append(fn)
    fn = ('INNERGRID', (0, 0), (-1, -1), 0.01, colors.grey)  # временное отображение сетки
    style.append(fn)

    ts = style   # стиль таблицы (список оформления строк и шрифта)
    t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                           ('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           ('FONTNAME', (1, 0), (1, 16), "DejaVuSerif-Bold"),
                           ('FONTSIZE', (1, 0), (1, 16), 7),
                           # 10 столбец с 0 по 68 ряд (цвет места)
                           ('TEXTCOLOR', (8, 0), (8, 39), colors.red),
                           ('ALIGN', (8, 0), (8, 39), 'RIGHT'),
                           ('ALIGN', (7, 0), (7, 39), 'LEFT'),
                           # цвет шрифта игроков 1 ого тура
                           ('TEXTCOLOR', (0, 0), (0, 39), colors.blue),
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                           ] + ts))
# === надпись финала
    h2 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic",
            leftIndent=50, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
    elements.append(Paragraph(f"{fin_title}. Одиночный разряд. {gamer}", h2))
# ====
    elements.append(t)
    pv = A4
    znak = final.rfind("-")
    if znak == -1:
        f = "superfinal"
    else:
        f = final[:znak]

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    if tds is not None:
        short_name = t_id.short_name_comp
        if fin == "Одна таблица":
            name_table_final = f"{short_name}_one_table.pdf"
        elif fin != "Суперфинал":
            name_table_final = f"{short_name}_{f}-final.pdf"
        else:
            name_table_final = f"{short_name}_{f}.pdf"
    else:
        short_name = "clear_8_full_net"  # имя для чистой сетки
        name_table_final = f"{short_name}.pdf"
    doc = SimpleDocTemplate(name_table_final, pagesize=pv, rightMargin=1*cm, leftMargin=1*cm, topMargin=3*cm, bottomMargin=1*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(elements, onFirstPage=func_zagolovok)
    os.chdir("..") # переходит на один уровень на верх
    return tds


def setka_8_full_made(fin):
    """сетка на 8 в pdf"""
    sender = my_win.sender()
    from reportlab.platypus import Table
    table = "setka_8_full"
    elements = []
    data = []
    style = []
    column = ['']
    column_count = column * 10
    # добавить в аргументы функции
    final = fin
    titles = Title.select().where(Title.id == title_id()).get()
    gamer = titles.gamer
    id_system = system_id(fin)
    sys = System.select().where(System.id == id_system).get()
    max_pl = sys.max_player # максимальное число игроков в сетке
    first_mesto = mesto_in_final(fin) if sender != my_win.clear_s8_full_Action else 1
    last_mesto = max_pl if fin == "1-й финал" else first_mesto + max_pl - 1
    fin_title = f'Финальные соревнования.({first_mesto}-{last_mesto} место)' # титул на таблице
    for i in range(0, 40):
        # column_count[9] = i  # нумерация 10 столбца для удобного просмотра таблицы
        list_tmp = column_count.copy()
        data.append(list_tmp)
    # ========= места ==========
    y = 0
    for i in range(0, 16, 2):
        y += 1
        data[i][0] = str(y)  # рисует начальные номера таблицы 1-16
    # ========= нумерация встреч сетки ==========
    draw_num(row_n=1, row_step=2, col_n=2, number_of_columns=3, number_of_game=1, player=8, data=data) # рисует номера встреч 1-32
    draw_num(row_n=16, row_step=2, col_n=6, number_of_columns=2, number_of_game=8, player=2, data=data) # рисует номера встреч 1-32
    draw_num(row_n=20, row_step=2, col_n=4, number_of_columns=2, number_of_game=9, player=4, data=data) # рисует номера встреч 1-32
    draw_num_lost(row_n=16, row_step=2, col_n=4, number_of_game=5, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=20, row_step=2, col_n=2, number_of_game=1, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=28, row_step=2, col_n=4, number_of_game=9, player=2, data=data) # номера минус проигравшие встречи -1 -16
   
    data[28][6] = str(12)  # создание номеров встреч 15
    data[13][6] = str(-7)
    data[18][6] = str(-8)
    data[25][6] = str(-11)
    data[30][6] = str(-12)
    # ============= данные игроков и встреч и размещение по сетке =============
    tds = write_in_setka(data, fin, first_mesto, table)
    #===============
    cw = ((0.3 * cm, 4.6 * cm, 0.4 * cm, 3.0 * cm, 0.4 * cm, 3.0 * cm, 0.4 * cm, 4.8 * cm, 1.5 * cm, 0.4 * cm))
    # основа сетки на чем чертить таблицу (ширина столбцов и рядов, их кол-во)
    color_mesta(data, first_mesto, table, fin) # раскрашивает места участников красным цветом
    t = Table(data, cw, 40 * [0.6 * cm])
    # =========  цикл создания стиля таблицы ================
    # ==== рисует основной столбец сетки 
    style = draw_setka(1, 1, 8, style) # рисует кусок сетки(номер столбца, номер строки на 8 человека)
    style = draw_setka(3, 20, 4, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(5, 28, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(5, 16, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    # ======= встречи за места =====
    for q in range(0, 7, 6):
        fn = ('LINEABOVE', (7, q + 8), (8, q + 8),
              1, colors.darkblue)  # за 1-2 место
        style.append(fn)
    for q in range(0, 3, 2):
        fn = ('LINEABOVE', (7, q + 17), (8, q + 17),
              1, colors.darkblue)  # за 3-4 место
        style.append(fn)
        fn = ('LINEABOVE', (7, q + 29), (8, q + 29),
              1, colors.darkblue)  # за 7-8 место
        style.append(fn)
    for q in range(0, 4, 3):
        fn = ('LINEABOVE', (7, q + 23), (8, q + 23),
              1, colors.darkblue)  # за 5-6 место
        style.append(fn)

    for i in range(1, 6, 2):
        fn = ('TEXTCOLOR', (i, 0), (i, 39), colors.black)  # цвет шрифта игроков
        style.append(fn)
        fn = ('TEXTCOLOR', (i + 1, 0), (i + 1, 39), colors.green)  # цвет шрифта номеров встреч
        style.append(fn)
        # выравнивание фамилий игроков по левому краю
        fn = ('ALIGN', (i, 0), (i, 39), 'LEFT') 
        style.append(fn)
        # центрирование номеров встреч
        fn = ('ALIGN', (i + 1, 0), (i + 1, 39), 'CENTER')
        style.append(fn)
    # fn = ('INNERGRID', (0, 0), (-1, -1), 0.01, colors.grey)  # временное отображение сетки
    # style.append(fn)

    ts = style   # стиль таблицы (список оформления строк и шрифта)
    t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                           ('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           ('FONTNAME', (1, 0), (1, 16), "DejaVuSerif-Bold"),
                           ('FONTSIZE', (1, 0), (1, 16), 7),
                           # 10 столбец с 0 по 68 ряд (цвет места)
                           ('TEXTCOLOR', (8, 0), (8, 39), colors.red),
                           ('ALIGN', (8, 0), (8, 39), 'RIGHT'),
                           ('ALIGN', (7, 0), (7, 39), 'LEFT'),
                           # цвет шрифта игроков 1 ого тура
                           ('TEXTCOLOR', (0, 0), (0, 39), colors.blue),
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                           ] + ts))
# === надпись финала
    h2 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic",
            leftIndent=50, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
    elements.append(Paragraph(f"{fin_title}. Одиночный разряд. {gamer}", h2))
# ====
    elements.append(t)
    pv = A4
    znak = final.rfind("-")
    if znak == -1:
        f = "superfinal"
    else:
        f = final[:znak]

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    if tds is not None:
        short_name = t_id.short_name_comp
        if fin == "Одна таблица":
            name_table_final = f"{short_name}_one_table.pdf"
        elif fin != "Суперфинал":
            name_table_final = f"{short_name}_{f}-final.pdf"
        else:
            name_table_final = f"{short_name}_{f}.pdf"
    else:
        short_name = "clear_8_full_net"  # имя для чистой сетки
        name_table_final = f"{short_name}.pdf"
    doc = SimpleDocTemplate(name_table_final, pagesize=pv, rightMargin=1*cm, leftMargin=1*cm, topMargin=3*cm, bottomMargin=1*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(elements, onFirstPage=func_zagolovok)
    os.chdir("..") # переходит на один уровень на верх
    return tds


def setka_8_2_made(fin):
    """сетка на 8 минус 2 в pdf"""
    sender = my_win.sender()
    from reportlab.platypus import Table
    table = "setka_8_2"
    elements = []
    data = []
    style = []
    column = ['']
    column_count = column * 10
    # добавить в аргументы функции
    final = fin
    titles = Title.select().where(Title.id == title_id()).get()
    gamer = titles.gamer
    id_system = system_id(fin)
    sys = System.select().where(System.id == id_system).get()
    max_pl = sys.max_player # максимальное число игроков в сетке
    first_mesto = mesto_in_final(fin) if sender != my_win.clear_s8_2_Action else 1
    last_mesto = max_pl if fin == "1-й финал" else first_mesto + max_pl - 1
    fin_title = f'Финальные соревнования.({first_mesto}-{last_mesto} место)' # титул на таблице
    for i in range(0, 40):
        # column_count[9] = i  # нумерация 10 столбца для удобного просмотра таблицы
        list_tmp = column_count.copy()
        data.append(list_tmp)
    # ========= места ==========
    y = 0
    for i in range(0, 16, 2):
        y += 1
        data[i][0] = str(y)  # рисует начальные номера таблицы 1-16
    # ========= нумерация встреч сетки ==========
    draw_num(row_n=1, row_step=2, col_n=2, number_of_columns=3, number_of_game=1, player=8, data=data) # рисует номера встреч 1-32
    draw_num_2(row_n=17, row_step=2, col_n=2, number_of_columns=2, number_of_game=8, player=4, data=data) # рисует номера встреч 33-47 
    draw_num_lost_2(row_n=15, row_step=2, col_n=2, revers_number=1, number_of_game=5, player=2, data=data) # номера минус проигравшие встречи -17-24
    draw_num_lost(row_n=17, row_step=2, col_n=0, number_of_game=1, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=25, row_step=2, col_n=4, number_of_game=10, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=31, row_step=2, col_n=4, number_of_game=8, player=2, data=data) # номера минус проигравшие встречи -1 -16
    
    data[13][6] = str(-7)
    data[17][6] = str(12)  # создание номеров встреч 12
    data[25][6] = str(13)
    data[31][6] = str(14)
    data[22][6] = str(-12)
    data[28][6] = str(-13)
    data[34][6] = str(-14)  # создание номеров встреч 27
    # ============= данные игроков и встреч и размещение по сетке =============
    tds = write_in_setka(data, fin, first_mesto, table)
    #===============
    cw = ((0.3 * cm, 4.6 * cm, 0.4 * cm, 3.0 * cm, 0.4 * cm, 3.0 * cm, 0.4 * cm, 4.8 * cm, 1.0 * cm, 0.2 * cm))
    # основа сетки на чем чертить таблицу (ширина столбцов и рядов, их кол-во)
    color_mesta(data, first_mesto, table, fin) # раскрашивает места участников красным цветом
    t = Table(data, cw, 40 * [0.6 * cm])
    # =========  цикл создания стиля таблицы ================
    # ==== рисует основной столбец сетки 
    style = draw_setka(1, 1, 8, style) # рисует кусок сетки(номер столбца, номер строки на 16 человека)
    style = draw_setka_2(1, 17, 4, style) # рисует кусок сетки(номер столбца, номер строки на 16 человека)
    style = draw_setka(5, 25, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_setka(5, 31, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    # ======= встречи за места =====
    for q in range(0, 7, 6):
        fn = ('LINEABOVE', (7, q + 8), (8, q + 8),
              1, colors.darkblue)  # за 1-2 место
        style.append(fn)
    for q in range(0, 5, 4):
        fn = ('LINEABOVE', (7, q + 19), (8, q + 19),
              1, colors.darkblue)  # за 3-4 место
        style.append(fn)
    for q in range(0, 4, 3):
        fn = ('LINEABOVE', (7, q + 26), (8, q + 26),
              1, colors.darkblue)  # за 5-6 место
        style.append(fn)
        fn = ('LINEABOVE', (7, q + 32), (8, q + 32),
              1, colors.darkblue)  # за 7-8 место
        style.append(fn)

    for i in range(1, 6, 2):
        fn = ('TEXTCOLOR', (i, 0), (i, 39), colors.black)  # цвет шрифта игроков
        style.append(fn)
        fn = ('TEXTCOLOR', (i + 1, 0), (i + 1, 39), colors.green)  # цвет шрифта номеров встреч
        style.append(fn)
        # выравнивание фамилий игроков по левому краю
        fn = ('ALIGN', (i, 0), (i, 39), 'LEFT') 
        style.append(fn)
        # центрирование номеров встреч
        fn = ('ALIGN', (i + 1, 0), (i + 1, 39), 'CENTER')
        style.append(fn)
    # fn = ('INNERGRID', (0, 0), (-1, -1), 0.01, colors.grey)  # временное отображение сетки
    # style.append(fn)

    ts = style   # стиль таблицы (список оформления строк и шрифта)
    t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                           ('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           ('FONTNAME', (1, 0), (1, 15), "DejaVuSerif-Bold"),
                        #    ('FONTSIZE', (1, 0), (1, 15), 7),
                           # 10 столбец с 0 по 68 ряд (цвет места)
                           ('TEXTCOLOR', (8, 0), (8, 39), colors.red),
                           # столбец с фамилиями за места выравнивает слева
                           ('ALIGN', (7, 0), (7, 39), 'LEFT'), 
                           # столбец с местами выравнивает справа
                           ('ALIGN', (8, 0), (8, 39), 'RIGHT'),
                           # цвет шрифта игроков 1 ого тура
                           ('TEXTCOLOR', (0, 0), (0, 39), colors.blue),
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                           ] + ts))
# === надпись финала
    h2 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic",
            leftIndent=50, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
    elements.append(Paragraph(f"{fin_title}. Одиночный разряд. {gamer}", h2))
# ====
    elements.append(t)
    pv = A4
    znak = final.rfind("-")
    f = final[:znak]

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    if tds is not None:
        short_name = t_id.short_name_comp
        if fin == "Одна таблица":
            name_table_final = f"{short_name}_one_table.pdf"
        else:
            name_table_final = f"{short_name}_{f}-final.pdf"
    else:
        short_name = "clear_8_2_net"  # имя для чистой сетки
        name_table_final = f"{short_name}.pdf"
    doc = SimpleDocTemplate(name_table_final, pagesize=pv, rightMargin=1*cm, leftMargin=1*cm, topMargin=3*cm, bottomMargin=1*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(elements, onFirstPage=func_zagolovok)
    os.chdir("..")
    return tds


def setka_16_full_made(fin):
    """сетка на 16 в pdf"""
    sender = my_win.sender()
    from reportlab.platypus import Table
    table = "setka_16_full"
    elements = []
    data = []
    style = []
    column = ['']
    column_count = column * 11
    # добавить в аргументы функции
    final = fin
    titles = Title.select().where(Title.id == title_id()).get()
    gamer = titles.gamer
    id_system = system_id(stage=fin)
    sys = System.select().where(System.id == id_system).get()
    max_pl = sys.max_player # максимальное число игроков в сетке
    first_mesto = mesto_in_final(fin) if sender != my_win.clear_s16_Action else 1
    last_mesto = max_pl if fin == "1-й финал" else first_mesto + max_pl - 1
    fin_title = f'Финальные соревнования.({first_mesto}-{last_mesto} место)' # титул на таблице
    for i in range(0, 69):
        # column_count[10] = i  # нумерация 10 столбца для удобного просмотра таблицы
        list_tmp = column_count.copy()
        data.append(list_tmp) # пустая основа сетки
    # ========= места ==========
    y = 0
    for i in range(0, 32, 2):
        y += 1
        data[i][0] = str(y)  # рисует начальные номера таблицы 1-16
    # ========= нумерация встреч сетки ==========
    draw_num(row_n=1, row_step=2, col_n=2, number_of_columns=4, number_of_game=1, player=16, data=data) # рисует номера встреч 1-32
    draw_num(row_n=32, row_step=2, col_n=6, number_of_columns=2, number_of_game=17, player=4, data=data) # рисует номера встреч 1-32
    draw_num(row_n=41, row_step=2, col_n=4, number_of_columns=3, number_of_game=21, player=8, data=data) # рисует номера встреч 1-32
    draw_num(row_n=58, row_step=2, col_n=6, number_of_columns=2, number_of_game=29, player=4, data=data) # рисует номера встреч 1-32
    draw_num_lost(row_n=29, row_step=2, col_n=6, number_of_game=13, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=32, row_step=2, col_n=4, number_of_game=9, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=39, row_step=2, col_n=6, number_of_game=17, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=41, row_step=2, col_n=2, number_of_game=1, player=8, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=58, row_step=2, col_n=4, number_of_game=21, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=55, row_step=2, col_n=6, number_of_game=25, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=65, row_step=2, col_n=6, number_of_game=29, player=2, data=data) # номера минус проигравшие встречи -1 -16
   
    data[8][8] = str(15)  # создание номеров встреч 15
    data[25][8] = str(-15)
    data[29][8] = str(16)  # создание номеров встреч 16
    data[31][8] = str(-16)
    data[37][8] = str(-19)
    data[39][8] = str(20)
    data[41][8] = str(-20)
    data[44][8] = str(27)  # создание номеров встреч 27
    data[52][8] = str(-27)
    data[55][8] = str(28)  # создание номеров встреч 28
    data[57][8] = str(-28)
    data[63][8] = str(-31)
    data[65][8] = str(32)  # создание номеров встреч 32
    data[67][8] = str(-32)

    # ============= данные игроков и встреч и размещение по сетке =============
    tds = write_in_setka(data, fin, first_mesto, table)
    #===============
    cw = ((0.3 * cm, 4.6 * cm, 0.4 * cm, 2.6 * cm, 0.4 * cm, 2.6 * cm, 0.4 * cm, 2.6 * cm,
           0.4 * cm, 4.4 * cm, 1.3 * cm))
    # основа сетки на чем чертить таблицу (ширина столбцов и рядов, их кол-во)
    color_mesta(data, first_mesto, table, fin) # раскрашивает места участников красным цветом
    t = Table(data, cw, 69 * [0.35 * cm])
    # =========  цикл создания стиля таблицы ================
    # ==== рисует основной столбец сетки 
    style = draw_setka(1, 1, 16, style) # рисует кусок сетки(номер столбца, номер строки на 16 человека)
    style = draw_setka(7, 29, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(5, 32, 4, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(7, 39, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(3, 41, 8, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(7, 55, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(5, 58, 4, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(7, 65, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    # ======= встречи за места =====
    for q in range(0, 11, 10):
        fn = ('LINEABOVE', (9, q + 16), (10, q + 16),
              1, colors.darkblue)  # за 1-2 место
        style.append(fn)
    for q in range(0, 3, 2):
        fn = ('LINEABOVE', (9, q + 30), (10, q + 30),
              1, colors.darkblue)  # за 3-4 место
        style.append(fn)
        fn = ('LINEABOVE', (9, q + 40), (10, q + 40),
              1, colors.darkblue)  # за 7-8 место
        style.append(fn)
        fn = ('LINEABOVE', (9, q + 56), (10, q + 56),
              1, colors.darkblue)  # за 11-12 место
        style.append(fn)
        fn = ('LINEABOVE', (9, q + 66), (10, q + 66),
              1, colors.darkblue)  # за 15-16 место
        style.append(fn)
    for q in range(0, 4, 3):
        fn = ('LINEABOVE', (9, q + 35), (10, q + 35),
              1, colors.darkblue)  # за 5-6 место
        style.append(fn)
        fn = ('LINEABOVE', (9, q + 61), (10, q + 61),
              1, colors.darkblue)  # за 13-14 место
        style.append(fn)
    for q in range(0, 6, 5):
        fn = ('LINEABOVE', (9, q + 48), (10, q + 48),
              1, colors.darkblue)  # за 9-10 место
        style.append(fn)

    for i in range(1, 8, 2):
        fn = ('TEXTCOLOR', (i, 0), (i, 68), colors.black)  # цвет шрифта игроков
        style.append(fn)
        fn = ('TEXTCOLOR', (i + 1, 0), (i + 1, 68), colors.green)  # цвет шрифта номеров встреч
        style.append(fn)
        # выравнивание фамилий игроков по левому краю
        fn = ('ALIGN', (i, 0), (i, 68), 'LEFT') 
        style.append(fn)
        # центрирование номеров встреч
        fn = ('ALIGN', (i + 1, 0), (i + 1, 68), 'CENTER')
        style.append(fn)
    # fn = ('INNERGRID', (0, 0), (-1, -1), 0.01, colors.grey)  # временное отображение сетки
    # style.append(fn)

    ts = style   # стиль таблицы (список оформления строк и шрифта)
    t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                           ('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           ('FONTNAME', (1, 0), (1, 32), "DejaVuSerif-Bold"),
                           ('FONTSIZE', (1, 0), (1, 32), 7),
                           # 10 столбец с 0 по 68 ряд (цвет места)
                           ('TEXTCOLOR', (10, 0), (10, 68), colors.red),
                        #    ('ALIGN', (10, 0), (10, 68), 'RIGHT'),
                           ('ALIGN', (9, 0), (9, 68), 'LEFT'),
                           # цвет шрифта игроков 1 ого тура
                           ('TEXTCOLOR', (0, 0), (0, 68), colors.blue),
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
                           ] + ts))
# === надпись финала
    h2 = PS("normal", fontSize=10, fontName="DejaVuSerif-Italic",
            leftIndent=50, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
    elements.append(Paragraph(f"{fin_title}. Одиночный разряд. {gamer}", h2))
# ====
    elements.append(t)
    pv = A4
    znak = final.rfind("-")
    f = final[:znak]

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    if tds is not None:
        short_name = t_id.short_name_comp
        if fin == "Одна таблица":
            name_table_final = f"{short_name}_one_table.pdf"
        else:
            name_table_final = f"{short_name}_{f}-final.pdf"
    else:
        short_name = "clear_16_full_net"  # имя для чистой сетки
        name_table_final = f"{short_name}.pdf"
    doc = SimpleDocTemplate(name_table_final, pagesize=pv, rightMargin=1*cm, leftMargin=1*cm, topMargin=2.2*cm, bottomMargin=1*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(elements, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")
    return tds


def setka_16_2_made(fin):
    """сетка на 16_2 в pdf"""
    from reportlab.platypus import Table
    sender = my_win.sender()
    table = "setka_16_2"
    elements = []
    data = []
    style = []
    column = ['']    
    column_count = column * 11
    # добавить в аргументы функции
    final = fin
    titles = Title.select().where(Title.id == title_id()).get()
    gamer = titles.gamer 
    id_system = system_id(fin)
    sys = System.select().where(System.id == id_system).get()
    max_pl = sys.max_player # максимальное число игроков в сетке  
    first_mesto = mesto_in_final(fin) if sender != my_win.clear_s16_2_Action else 1
    last_mesto = max_pl if fin == "1-й финал" else first_mesto + max_pl - 1
    fin_title = f'Финальные соревнования.({first_mesto}-{last_mesto} место)' # титул на таблице
    for i in range(0, 86):
        # column_count[10] = i  # нумерация 10 столбца для удобного просмотра таблицы
        list_tmp = column_count.copy()
        data.append(list_tmp)
    # ========= места ==========
    y = 0
    for i in range(2, 34, 2):
        y += 1
        data[i][0] = str(y)  # рисует начальные номера таблицы 1-16
    # ========= нумерация встреч сетки ==========
    draw_num(row_n=3, row_step=2, col_n=2, number_of_columns=4, number_of_game=1, player=16, data=data) # рисует номера встреч 1-32
    draw_num_lost_2(row_n=45, row_step=1, col_n=0, revers_number=0, number_of_game=1, player=8, data=data) # номера минус проигравшие встречи -17-24
    draw_num_lost_2(row_n=44, row_step=2, col_n=2, revers_number=1, number_of_game=9, player=4, data=data) # номера минус проигравшие встречи -17-24
    draw_num_lost_2(row_n=43, row_step=4, col_n=6, revers_number=0, number_of_game=13, player=2, data=data) # номера минус проигравшие встречи -17-24
    draw_num(row_n=62, row_step=2, col_n=2, number_of_columns=2, number_of_game=31, player=4, data=data) # рисует номера встреч 1-32
    draw_num(row_n=74, row_step=2, col_n=2, number_of_columns=2, number_of_game=35, player=4, data=data) # рисует номера встреч 1-32

    draw_num(row_n=46, row_step=2, col_n=2, number_of_columns=1, number_of_game=16, player=8, data=data) # рисует номера встреч 1-32
    draw_num(row_n=45, row_step=2, col_n=4, number_of_columns=2, number_of_game=20, player=8, data=data) # рисует номера встреч 1-32
    draw_num(row_n=44, row_step=4, col_n=8, number_of_columns=1, number_of_game=26, player=4, data=data) # рисует номера встреч 1-32
    draw_num_lost(row_n=62, row_step=2, col_n=0, number_of_game=20, player=4, data=data) # номера минус проигравшие встречи -20 -23
    draw_num_lost(row_n=74, row_step=2, col_n=0, number_of_game=16, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=61, row_step=2, col_n=6, number_of_game=26, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=67, row_step=2, col_n=6, number_of_game=24, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=73, row_step=2, col_n=6, number_of_game=31, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=79, row_step=2, col_n=6, number_of_game=35, player=2, data=data) # номера минус проигравшие встречи -1 -16
   
    data[46][10] = str(28)  # создание номеров встреч 15
    data[34][8] = str(-15)
    data[57][8] = str(-28)
    data[70][4] = str(-33)
    data[82][4] = str(-37)
    data[61][8] = str(29)  # создание номеров встреч 27
    data[64][8] = str(-29)
    data[67][8] = str(30)  # создание номеров встреч 28
    data[70][8] = str(-30)
    data[73][8] = str(34)  # создание номеров встреч 32
    data[76][8] = str(-34)
    data[79][8] = str(38)  # создание номеров встреч 32
    data[82][8] = str(-38)


    # ============= данные игроков и встреч и размещение по сетке =============
    tds = write_in_setka(data, fin, first_mesto, table) # сюда приходит чистая data
    #===============
    cw = ((0.3 * cm, 4.6 * cm, 0.4 * cm, 2.6 * cm, 0.4 * cm, 2.6 * cm, 0.4 * cm, 2.6 * cm,
           0.4 * cm, 4.4 * cm, 0.4 * cm))
    # основа сетки на чем чертить таблицу (ширина столбцов и рядов, их кол-во)
    style_color = color_mesta(data, first_mesto, table, fin) # раскрашивает места участников красным цветом
    t = Table(data, cw, 86 * [0.55 * cm])
    # =========  цикл создания стиля таблицы ================
    # ==== рисует основной столбец сетки 
    style = draw_setka(1, 3, 16, style) # рисует кусок сетки(номер столбца, номер строки на 16 человека)
    style = draw_setka(1, 62, 4, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(1, 74, 4, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka_2(1, 46, 8, style) # рисует кусок сетки(номер столбца, номер строки на 8 человека)
    style = draw_setka_made(9, 46, 2, 8, 1, style) # рисует кусок сетки(номер столбца, колво уч, шаг между линиями)
    style = draw_setka(7, 61, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(7, 67, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(7, 73, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    style = draw_setka(7, 79, 2, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
    # ======= встречи за места =====
    for q in range(0, 18, 17):
        fn = ('LINEABOVE', (9, q + 18), (10, q + 18),
              1, colors.darkblue)  # за 1-2 место
        style.append(fn)
    for q in range(0, 9, 8):
        fn = ('LINEABOVE', (9, q + 50), (10, q + 50),
              1, colors.darkblue)  # за 3-4 место
        style.append(fn)
 
    for q in range(0, 4, 3):
        fn = ('LINEABOVE', (9, q + 62), (10, q + 62),
              1, colors.darkblue)  # за 5-6 место
        style.append(fn)
        fn = ('LINEABOVE', (9, q + 68), (10, q + 68),
              1, colors.darkblue)  # за 5-6 место
        style.append(fn)
        fn = ('LINEABOVE', (9, q + 74), (10, q + 74),
              1, colors.darkblue)  # за 5-6 место
        style.append(fn)
        fn = ('LINEABOVE', (9, q + 80), (10, q + 80),
              1, colors.darkblue)  # за 5-6 место
        style.append(fn)

    for q in range(0, 7, 6):
        fn = ('LINEABOVE', (5, q + 65), (6, q + 65),
              1, colors.darkblue)  # за 7-8 место
        style.append(fn)
        fn = ('LINEABOVE', (5, q + 77), (6, q + 77),
              1, colors.darkblue)  # за 9-10 место
        style.append(fn)

    for i in range(1, 10, 2):
        fn = ('TEXTCOLOR', (i, 0), (i, 85), colors.black)  # цвет шрифта игроков
        style.append(fn)
        fn = ('TEXTCOLOR', (i + 1, 0), (i + 1, 85), colors.green)  # цвет шрифта номеров встреч
        style.append(fn)
        # выравнивание фамилий игроков по левому краю
        fn = ('ALIGN', (i, 0), (i, 85), 'LEFT') 
        style.append(fn)
        # центрирование номеров встреч
        fn = ('ALIGN', (i + 1, 0), (i + 1, 85), 'CENTER')
        style.append(fn)
   
    # fn = ('INNERGRID', (0, 0), (-1, -1), 0.01, colors.grey)  # временное отображение сетки
    # style.append(fn)

    ts = style   # стиль таблицы (список оформления строк и шрифта)

    for b in style_color: # цикл окрашивания мест красным цветом
        ts.append(b)

    t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                           ('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           ('FONTNAME', (1, 0), (1, 32), "DejaVuSerif-Bold"),
                           ('FONTSIZE', (1, 0), (1, 32), 7)] + ts 
                           + [
                           # цвет шрифта игроков 1 ого тура
                           ('TEXTCOLOR', (0, 0), (0, 40), colors.blue),
                           ('TEXTCOLOR', (0, 41), (0, 85), colors.green),
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
# === надпись финала
    h2 = PS("normal", fontSize=10, fontName="DejaVuSerif-Italic",
            leftIndent=50, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
    elements.append(Paragraph(f"{fin_title}. Одиночный разряд. {gamer}", h2))
# ====
    elements.append(t)
    pv = A4
    znak = final.rfind("-")
    f = final[:znak]

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    if tds is not None:
        short_name = t_id.short_name_comp
        if fin == "Одна таблица":
            name_table_final = f"{short_name}_one_table.pdf"
        else:
            name_table_final = f"{short_name}_{f}-final.pdf"
    else:
        short_name = "clear_16_2_net"  # имя для чистой сетки
        name_table_final = f"{short_name}.pdf"
    doc = SimpleDocTemplate(name_table_final, pagesize=pv)
    catalog = 1
    change_dir(catalog)
    doc.build(elements, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")
    return tds


def setka_32_made(fin):
    """сетка на 32 с розыгрышем 1-3 места"""
    from reportlab.platypus import Table
    sender = my_win.sender()
    table = "setka_32"
    elements = []
    style = []
    data = []
    column = ['']
    column_count = column * 13
    final = fin
    titles = Title.select().where(Title.id == title_id()).get()
    gamer = titles.gamer
    id_system = system_id(fin)
    sys = System.select().where(System.id == id_system).get()
    max_pl = sys.max_player # максимальное число игроков в сетке
    first_mesto = mesto_in_final(fin) if sender != my_win.clear_s32_Action else 1
    last_mesto = max_pl if fin == "1-й финал" else first_mesto + max_pl - 1
    fin_title = f'Финальные соревнования.({first_mesto}-{last_mesto} место)' # титул на таблице
    strok = 69
    for i in range(0, strok):
        # column_count[12] = i  # нумерация 10 столбца для удобного просмотра таблицы
        list_tmp = column_count.copy()
        data.append(list_tmp)
    # ========= нумерация встреч сетки ==========
    y = 0
    for i in range(1, 65, 2):
        y += 1
        data[i + 1][0] = str(y)  # рисует начальные номера таблицы 1-32
    number_of_game = draw_num(row_n=3, row_step=2, col_n=2, number_of_columns=5, number_of_game=1, player=32, data=data) # рисует номера встреч 1-32
    data[60][8] = str((number_of_game - 3) * -1)  # номера проигравших 29
    data[62][8] = str((number_of_game - 2) * -1)  # номера проигравших 30
    data[18][10] = str(number_of_game - 1)  # создание номеров встреч (31)
    data[55][10] = str((number_of_game - 1) * -1)  # номер проигравшего финал (-31)
    data[61][10] = str(number_of_game)  # создание номеров встреч 32
    data[66][10] = str((number_of_game) * -1)  # номер проигравшего финал (-32)

    # ============= данные игроков и встреч и размещение по сетке =============
    # ======= создать словарь  ключ - номер встречи, значение - номер ряда
    dict_num_game = {}
    for d in range(2, 11, 2):
        for r in range(0, 69):
            key = data[r][d]
            if key != "":
                dict_num_game[key] = r
 
    tds = write_in_setka(data, fin, first_mesto, table)
    cw = ((0.2 * cm, 3.8 * cm, 0.35 * cm, 2.7 * cm, 0.35 * cm, 2.7 * cm, 0.35 * cm, 2.7 * cm, 0.35 * cm,
        2.5 * cm, 0.35 * cm, 3.0 * cm, 0.3 * cm))
    # основа сетки на чем чертить таблицу (ширина столбцов и рядов, их кол-во)
    style_color = color_mesta(data, first_mesto, table, fin) # раскрашивает места участников красным цветом
    t = Table(data, cw, strok * [0.35 * cm])
    # =========  цикл создания стиля таблицы =======
    # ========= 1 страница =========
    style = draw_setka(1, 3, 32, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
   
    for l in range(34, 57, 22):
        fn = ('LINEABOVE', (11, l), (13, l), 1, colors.darkblue)  # рисует линии встреч за 1-2 места
        style.append(fn)
    for l in range(62, 68, 5):
        fn = ('LINEABOVE', (11, l), (13, l), 1, colors.darkblue)  # рисует линии встреч за 3-4 места
        style.append(fn)
    for l in range(61, 64, 2):
        fn = ('LINEABOVE', (9, l), (10, l), 1, colors.darkblue)  # рисует линии встреч за -29 -30
        style.append(fn)
    fn = ('BOX', (10, 61), (10, 62), 1, colors.darkblue)
    style.append(fn) 
    fn = ('SPAN', (10, 61), (10, 62))  # встреча 32
    style.append(fn)       
    fn = ('BACKGROUND', (10, 61), (10, 62), colors.lightyellow)  # встречи 32 за 3-4 место
    style.append(fn)

    for i in range(0, 11, 2):
        fn = ('TEXTCOLOR', (i + 1, 0), (i + 1, strok), colors.black)  # цвет шрифта игроков
        style.append(fn)
        fn = ('TEXTCOLOR', (i, 0), (i, strok), colors.green)  # цвет шрифта номеров встреч
        style.append(fn)
        # выравнивание фамилий игроков по левому краю
        fn = ('ALIGN', (i + 1, 0), (i + 1, strok), 'LEFT')
        style.append(fn)
        # центрирование номеров встреч
        fn = ('ALIGN', (i, 0), (i, strok), 'CENTER')
        style.append(fn)
    fn = ('INNERGRID', (0, 0), (-1, -1), 0.01, colors.grey)  # временное отображение сетки
    style.append(fn)
    ts = style   # стиль таблицы (список оформления строк и шрифта)
    for b in style_color:
        ts.append(b)

    t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                           ('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           ('FONTNAME', (1, 0), (1, 32), "DejaVuSerif-Bold"),
                           ('FONTSIZE', (1, 0), (1, 32), 7)] + ts 
                           + [
                           # цвет шрифта игроков 1 ого тура
                           ('TEXTCOLOR', (0, 0), (0, 68), colors.blue),
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')])) 
# === надпись финала
    h2 = PS("normal", fontSize=10, fontName="DejaVuSerif-Italic",
            leftIndent=50, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
    elements.append(Paragraph(f"{fin_title}. Одиночный разряд. {gamer}", h2))
# ====                         
    elements.append(t)
    pv = A4
    znak = final.rfind("-")
    f = final[:znak]

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    if tds is not None:
        short_name = t_id.short_name_comp
        name_table_final = f"{short_name}_{f}-final.pdf"
    else:
        short_name = "clear_32_net"
        name_table_final = f"{short_name}.pdf"
    doc = SimpleDocTemplate(name_table_final, pagesize=pv, rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(elements, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")
    return tds


def setka_32_full_made(fin):
    """сетка на 32 с розыгрышем всех мест"""
    from reportlab.platypus import Table
    sender = my_win.sender()
    table = "setka_32_full"
    elements = []
    style = []
    data = []
    column = ['']
    column_count = column * 13
    final = fin
    titles = Title.select().where(Title.id == title_id()).get()
    gamer = titles.gamer
    id_system = system_id(fin)
    sys = System.select().where(System.id == id_system).get()
    max_pl = sys.max_player # максимальное число игроков в сетке
    first_mesto = mesto_in_final(fin) if sender != my_win.clear_s32_full_Action else 1
    last_mesto = max_pl if fin == "1-й финал" else first_mesto + max_pl - 1
    fin_title = f'Финальные соревнования.({first_mesto}-{last_mesto} место)' # титул на таблице
    strok = 207
    for i in range(0, strok):
        # column_count[12] = i  # нумерация 10 столбца для удобного просмотра таблицы
        list_tmp = column_count.copy()
        data.append(list_tmp)
    # ========= нумерация встреч сетки ==========
    y = 0
    for i in range(1, 65, 2):
        y += 1
        data[i + 1][0] = str(y)  # рисует начальные номера таблицы 1-32
    number_of_game = draw_num(row_n=3, row_step=2, col_n=2, number_of_columns=5, number_of_game=1, player=32, data=data) # рисует номера встреч 1-32
    data[60][8] = str((number_of_game - 3) * -1)  # номера проигравших 29
    data[62][8] = str((number_of_game - 2) * -1)  # номера проигравших 30
    data[18][10] = str(number_of_game - 1)  # создание номеров встреч (31)
    data[55][10] = str((number_of_game - 1) * -1)  # номер проигравшего финал (-31)
    data[61][10] = str(number_of_game)  # создание номеров встреч 32
    data[66][10] = str((number_of_game) * -1)  # номер проигравшего финал (-32)
    #===== 2-я страница =========
    y = 0
    for i in range(78, 89, 10):
        y += 1
        data[i][8] = str((y + 34) * -1)  # номер проигравшего финал (-35, -36)
    y = 0
    for i in range(101, 112, 9):
        y += 1
        data[i][8] = str((y + 42) * -1)  # номер проигравшего финал (-43, -44)
    y = 0
    for i in range(120, 131, 10):
        y += 1
        data[i][8] = str((y + 46) * -1)  # номер проигравшего финал (-47, -48)
    y = 0
    for i in range(172, 187, 14):
        y += 4
        data[i][10] = str((y + 60) * -1)  # номер проигравшего финал (-64, -68)
    y = 0
    for i in range(198, 206, 7):
        y += 4
        data[i][10] = str((y + 72) * -1)  # номер проигравшего финал (-76, -80)
    data[178][8] = str(-67)  # номер проигравшего финал (-67)
    data[191][6] = str(-75)  # номер проигравшего финал (-75)
    data[164][8] = str(-63)  # номер проигравшего финал (-63)
    data[203][4] = str(-79)  # номер проигравшего финал (-79)

    number_of_game = draw_num(row_n=72, row_step=2, col_n=6, number_of_columns=2, number_of_game=33, player=4, data=data) # рисует номера встреч 1-32
    data[84][8] = str(number_of_game)  # создание номеров встреч 36
    number_of_game = draw_num(row_n=89, row_step=2, col_n=4, number_of_columns=3, number_of_game=37, player=8, data=data) # рисует номера встреч 1-32
    data[106][8] = str(number_of_game)
    number_of_game = draw_num(row_n=114, row_step=2, col_n=6, number_of_columns=2, number_of_game=45, player=4, data=data) # рисует номера встреч 1-32
    data[126][8] = str(number_of_game)  # создание номеров встреч 48
    draw_num_lost(row_n=72, row_step=2, col_n=4, number_of_game=25, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=84, row_step=2, col_n=6, number_of_game=33, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=89, row_step=2, col_n=2, number_of_game=17, player=8, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=106, row_step=2, col_n=6, number_of_game=41, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=114, row_step=2, col_n=4, number_of_game=37, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=126, row_step=2, col_n=6, number_of_game=45, player=2, data=data) # номера минус проигравшие встречи -1 -16
#========== 3-я страница ==============
    number_of_game = draw_num(row_n=140, row_step=2, col_n=2, number_of_columns=4, number_of_game=49, player=16, data=data) # рисует номера встреч 1-32
    data[168][10] = str(number_of_game)  # создание номеров встреч 64
    number_of_game = draw_num(row_n=172, row_step=2, col_n=6, number_of_columns=2, number_of_game=65, player=4, data=data) # рисует номера встреч 1-32
    data[182][10] = str(number_of_game)  # создание номеров встреч 68
    number_of_game = draw_num(row_n=179, row_step=2, col_n=2, number_of_columns=3, number_of_game=69, player=8, data=data) # рисует номера встреч 1-32
    data[194][10] = str(number_of_game)
    number_of_game = draw_num(row_n=197, row_step=2, col_n=2, number_of_columns=2, number_of_game=77, player=4, data=data) # рисует номера встреч 1-32
    data[201][10] = str(number_of_game)  # создание номеров встреч 68
    draw_num_lost(row_n=140, row_step=2, col_n=0, number_of_game=1, player=16, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=172, row_step=2, col_n=4, number_of_game=57, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=179, row_step=2, col_n=0, number_of_game=49, player=8, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=197, row_step=2, col_n=0, number_of_game=69, player=4, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=168, row_step=2, col_n=8, number_of_game=61, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=182, row_step=2, col_n=8, number_of_game=65, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=194, row_step=2, col_n=8, number_of_game=73, player=2, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=201, row_step=2, col_n=8, number_of_game=77, player=2, data=data) # номера минус проигравшие встречи -1 -16
    # ============= данные игроков и встреч и размещение по сетке =============
    tds = write_in_setka(data, fin, first_mesto, table)
    #===============
    cw = ((0.2 * cm, 3.8 * cm, 0.35 * cm, 2.7 * cm, 0.35 * cm, 2.7 * cm, 0.35 * cm, 2.7 * cm, 0.35 * cm,
        2.5 * cm, 0.35 * cm, 3.0 * cm, 0.3 * cm))
    # основа сетки на чем чертить таблицу (ширина столбцов и рядов, их кол-во)
    style_color = color_mesta(data, first_mesto, table, fin) # раскрашивает места участников красным цветом
    t = Table(data, cw, strok * [0.35 * cm])
    # =========  цикл создания стиля таблицы =======
    # ========= 1 страница =========
    style = draw_setka(1, 3, 32, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
   
    for l in range(34, 57, 22):
        fn = ('LINEABOVE', (11, l), (13, l), 1, colors.darkblue)  # рисует линии встреч за 1-2 места
        style.append(fn)
    for l in range(62, 68, 5):
        fn = ('LINEABOVE', (11, l), (13, l), 1, colors.darkblue)  # рисует линии встреч за 3-4 места
        style.append(fn)
    for l in range(61, 64, 2):
        fn = ('LINEABOVE', (9, l), (10, l), 1, colors.darkblue)  # рисует линии встреч за -29 -30
        style.append(fn)

    fn = ('BOX', (10, 61), (10, 62), 1, colors.darkblue)
    style.append(fn) 
    fn = ('SPAN', (10, 61), (10, 62))  # встреча 32
    style.append(fn)       
    fn = ('BACKGROUND', (10, 61), (10, 62), colors.lightyellow)  # встречи 32 за 3-4 место
    style.append(fn)
# =========== 2 страница ===================
    # ======= встречи (33-35) за 5-6 место =====
    style = draw_setka(5, 72, 4, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=75, col=9, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (36) за 7-8 место =====
    style = draw_setka(7, 84, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=85, col=9, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (37-43) за 9-10 место =====
    style = draw_setka(3, 89, 8, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=96, col=9, player=8, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (44) за 11-12 место =====
    style = draw_setka(7, 106, 2, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=107, col=9, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (45-47) за 13-14 место =====
    style = draw_setka(5, 114, 4, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=117, col=9, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (48) за 15-16 место =====
    style = draw_setka(7, 126, 2, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=127, col=9, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
# =========== 3 страница ==================
    # ======= встречи (49-56) за 17-18 место =====
    style = draw_setka(1, 140, 16, style) # рисует кусок сетки(номер столбца, номер строки на 16человека)
    style = draw_mesta(row=155, col=9, player=16, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (64) за 19-20 место =====
    style = draw_setka(9, 168, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=169, col=10, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (33-35) за 21-24 место =====
    style = draw_setka(5, 172, 4, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=175, col=8, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (68) за 25-26 место =====
    style = draw_setka(9, 182, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=183, col=11, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (69 - 75) за 25-26 место =====
    style = draw_setka(1, 179, 8, style) # рисует кусок сетки(номер столбца, номер строки на 8 человека)
    style = draw_mesta(row=186, col=7, player=8, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (64) за 25-26 место =====
    style = draw_setka(9, 194, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=195, col=10, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
     # ======= встречи (64) за 25-26 место =====
    style = draw_setka(1, 197, 4, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=200, col=5, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
     # ======= встречи (64) за 25-26 место =====
    style = draw_setka(9, 201, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=202, col=10, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    
# =========================================
    for i in range(0, 11, 2):
        fn = ('TEXTCOLOR', (i + 1, 0), (i + 1, 206), colors.black)  # цвет шрифта игроков
        style.append(fn)
        fn = ('TEXTCOLOR', (i, 0), (i, 206), colors.green)  # цвет шрифта номеров встреч
        style.append(fn)
        # выравнивание фамилий игроков по левому краю
        fn = ('ALIGN', (i + 1, 0), (i + 1, 206), 'LEFT')
        style.append(fn)
        # центрирование номеров встреч
        fn = ('ALIGN', (i, 0), (i, 206), 'CENTER')
        style.append(fn)
    # fn = ('INNERGRID', (0, 0), (-1, -1), 0.01, colors.grey)  # временное отображение сетки
    # style.append(fn)
    ts = style   # стиль таблицы (список оформления строк и шрифта)
    for b in style_color:
        ts.append(b)

    t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                           ('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           ('FONTNAME', (1, 0), (1, 64), "DejaVuSerif-Bold"),
                           ('FONTSIZE', (1, 0), (1, 64), 7)] + ts 
                           + [
                           # цвет шрифта игроков 1 ого тура
                           ('TEXTCOLOR', (0, 0), (0, 68), colors.blue),
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
                           
# === надпись финала
    h2 = PS("normal", fontSize=10, fontName="DejaVuSerif-Italic",
            leftIndent=50, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
    elements.append(Paragraph(f"{fin_title}. Одиночный разряд. {gamer}", h2))
# ====
    elements.append(t)
    pv = A4
    znak = final.rfind("-")
    f = final[:znak]

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    if tds is not None:
        short_name = t_id.short_name_comp
        name_table_final = f"{short_name}_{f}-final.pdf"
    else:
        short_name = "clear_32_full_net"
        name_table_final = f"{short_name}.pdf"
    doc = SimpleDocTemplate(name_table_final, pagesize=pv, rightMargin=1*cm, leftMargin=1*cm, topMargin=3.5*cm, bottomMargin=1.0*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(elements, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")
    return tds


def setka_32_2_made(fin):
    """сетка на 32 (-2) с розыгрышем всех мест"""
    from reportlab.platypus import Table
    sender = my_win.sender()
    table = "setka_32_2"
    elements = []
    style = []
    data = []
    column = ['']
    column_count = column * 15
    final = fin
    titles = Title.select().where(Title.id == title_id()).get()
    gamer = titles.gamer
    id_system = system_id(fin)
    sys = System.select().where(System.id == id_system).get()
    max_pl = sys.max_player # максимальное число игроков в сетке
    first_mesto = mesto_in_final(fin) if sender != my_win.clear_s32_2_Action else 1
    last_mesto = max_pl if fin == "1-й финал" else first_mesto + max_pl - 1
    last_mesto = max_pl if fin == "1-й финал" else first_mesto + max_pl - 1
    fin_title = f'Финальные соревнования.({first_mesto}-{last_mesto} место)' # титул на таблице
    strok = 207
    for i in range(0, strok):
        # column_count[14] = i  # нумерация 10 столбца для удобного просмотра таблицы
        list_tmp = column_count.copy()
        data.append(list_tmp)
    # ========= нумерация встреч сетки ==========
    y = 0
    for i in range(1, 64, 2):
        y += 1
        data[i + 1][0] = str(y)  # рисует начальные номера таблицы 1-32
    number_of_game = draw_num(row_n=3, row_step=2, col_n=2, number_of_columns=5, number_of_game=1, player=32, data=data) # рисует номера встреч 1-32 
    data[18][10] = str(number_of_game - 1)  # создание номеров встреч (31)
    data[55][10] = str((number_of_game - 1) * -1)  # номер проигравшего финал (-31)
 # ======= 2-я страница ===========
    draw_num_lost(row_n=74, row_step=2, col_n=0, number_of_game=1, player=16, data=data) # номера минус проигравшие встречи -1 -16
    draw_num_lost(row_n=102, row_step=2, col_n=10, number_of_game=58, player=2, data=data) # номера минус проигравшие встречи -58-59
    draw_num_lost(row_n=110, row_step=2, col_n=10, number_of_game=56, player=2, data=data) # номера минус проигравшие встречи -56-57
    draw_num_lost(row_n=112, row_step=2, col_n=0, number_of_game=52, player=4, data=data) # номера минус проигравшие встречи -52-55
    draw_num_lost(row_n=124, row_step=2, col_n=0, number_of_game=48, player=4, data=data) # номера минус проигравшие встречи -48-51
    draw_num_lost(row_n=120, row_step=2, col_n=8, number_of_game=63, player=2, data=data) # номера минус проигравшие встречи -63-64
    draw_num_lost(row_n=128, row_step=2, col_n=8, number_of_game=67, player=2, data=data) # номера минус проигравшие встречи -67-68
    draw_num_lost_2(row_n=72, row_step=2, col_n=2, revers_number=1, number_of_game=17, player=8, data=data) # номера минус проигравшие встречи -17-24
    draw_num_lost_2(row_n=71, row_step=4, col_n=6, revers_number=1, number_of_game=25, player=2, data=data) # номера минус проигравшие встречи -25-26
    draw_num_lost_2(row_n=87, row_step=4, col_n=6, revers_number=1, number_of_game=27, player=2, data=data) # номера минус проигравшие встречи -27-28
    draw_num_lost_2(row_n=71, row_step=8, col_n=10, revers_number=1, number_of_game=29, player=2, data=data) # номера минус проигравшие встречи -29-30

    number_of_game = draw_num_2(row_n=74, row_step=2, col_n=2, number_of_columns=2, number_of_game=32, player=16, data=data) # рисует номера встреч 33-47 
    number_of_game = draw_num_2(row_n=74, row_step=4, col_n=6, number_of_columns=2, number_of_game=48, player=16, data=data) # рисует номера встреч 48-55
    number_of_game = draw_num_2(row_n=74, row_step=8, col_n=10, number_of_columns=1, number_of_game=56, player=16, data=data) # рисует номера встреч 56-57
    number_of_game = draw_num_2(row_n=72, row_step=8, col_n=12, number_of_columns=1, number_of_game=58, player=16, data=data) # рисует номера встреч 58-59
    number_of_game = draw_num(row_n=112, row_step=2, col_n=2, number_of_columns=2, number_of_game=63, player=4, data=data) # рисует номера встреч 63-65
    number_of_game = draw_num(row_n=124, row_step=2, col_n=2, number_of_columns=2, number_of_game=67, player=4, data=data) # рисует номера встреч 67-69

    data[75][14] = str(number_of_game - 10)  # создание номеров встреч (60)
    data[98][12] = str((number_of_game - 10) * -1)  # номер проигравшего финал (-60)
    data[102][12] = str(number_of_game - 9)  # создание номеров встреч (61)
    data[106][12] = str((number_of_game - 9) * -1)  # номер проигравшего финал (-61)
    data[110][12] = str(number_of_game - 8)  # создание номеров встреч (62)
    data[114][12] = str((number_of_game - 8) * -1)  # номер проигравшего финал (-62)
    data[120][10] = str(number_of_game - 4)  # создание номеров встреч (66)
    data[124][10] = str((number_of_game - 4) * -1)  # номер проигравшего финал (-66)
    data[128][10] = str(number_of_game)  # создание номеров встреч (70)
    data[132][10] = str((number_of_game) * -1)  # номер проигравшего финал (-70)
    data[118][4] = str((number_of_game - 5) * -1)  # номер проигравшего финал (-65)
    data[130][4] = str((number_of_game - 1) * -1)  # номер проигравшего финал (-69)
# ======= 3-я страница ===========
    draw_num_lost(row_n=141, row_step=2, col_n=0, number_of_game=40, player=8, data=data) # номера минус проигравшие встречи
    draw_num_lost(row_n=156, row_step=2, col_n=8, number_of_game=75, player=2, data=data) # номера минус проигравшие встречи
    draw_num_lost(row_n=171, row_step=2, col_n=8, number_of_game=79, player=2, data=data) # номера минус проигравшие встречи
    draw_num_lost(row_n=186, row_step=2, col_n=8, number_of_game=87, player=2, data=data) # номера минус проигравшие встречи
    draw_num_lost(row_n=201, row_step=2, col_n=8, number_of_game=91, player=2, data=data) # номера минус проигравшие встречи
    draw_num_lost(row_n=160, row_step=2, col_n=2, number_of_game=71, player=4, data=data) # номера минус проигравшие встречи
    draw_num_lost(row_n=171, row_step=2, col_n=0, number_of_game=32, player=8, data=data) # номера минус проигравшие встречи
    draw_num_lost(row_n=192, row_step=2, col_n=2, number_of_game=83, player=4, data=data) # номера минус проигравшие встречи

    number_of_game = draw_num(row_n=141, row_step=2, col_n=2, number_of_columns=3, number_of_game=71, player=8, data=data) # рисует номера встреч 49
    data[153][6] = str((number_of_game - 1) * -1)  # создание номеров встреч -77
    data[156][10] = str(number_of_game)  # создание номеров встреч 78
    data[160][10] = str(number_of_game * -1)  # создание номеров встреч -78
    number_of_game = draw_num(row_n=160, row_step=2, col_n=4, number_of_columns=2, number_of_game=79, player=4, data=data) # рисует номера встреч 49
    data[166][6] = str((number_of_game - 1) * -1)  # создание номеров встреч -77
    data[171][10] = str(number_of_game)  # создание номеров встреч 68
    data[175][10] = str(number_of_game * -1)  # создание номеров встреч 68
    number_of_game = draw_num(row_n=171, row_step=2, col_n=2, number_of_columns=3, number_of_game=83, player=8, data=data) # рисует номера встреч 49
    data[183][6] = str((number_of_game - 1) * -1)  # создание номеров встреч -77
    data[186][10] = str(number_of_game)  # создание номеров встреч 90
    data[190][10] = str(number_of_game * -1)  # создание номеров встреч -90
    number_of_game = draw_num(row_n=192, row_step=2, col_n=4, number_of_columns=2, number_of_game=91, player=4, data=data) # рисует номера встреч 49
    data[198][6] = str((number_of_game - 1) * -1)  # создание номеров встреч -77
    data[201][10] = str(number_of_game)  # создание номеров встреч 94
    data[205][10] = str(number_of_game * -1)  # создание номеров встреч -94

    data[205][10] = str(number_of_game * -1)  # создание номеров встреч -94
    data[205][10] = str(number_of_game * -1)  # создание номеров встреч -94
    data[205][10] = str(number_of_game * -1)  # создание номеров встреч -94
    data[205][10] = str(number_of_game * -1)  # создание номеров встреч -94

    # ============= данные игроков и встреч и размещение по сетке =============
    # ======= создать словарь  ключ - номер встречи, значение - номер ряда
    dict_num_game = {}
    for d in range(2, 15, 2):
        for r in range(0, 69):
            key = data[r][d]
            if key != "":
                dict_num_game[key] = r
    # ===== добавить данные игроков и счета в data ==================
    tds = write_in_setka(data, fin, first_mesto, table)
    # ==============
    cw = ((0.2 * cm, 3.5 * cm, 0.35 * cm, 2.4 * cm, 0.35 * cm, 2.4 * cm, 0.35 * cm, 2.4 * cm, 0.35 * cm, 2.4 * cm, 0.35 * cm,
        2.4 * cm, 0.35 * cm, 2.6 * cm, 0.35 * cm))
    # основа сетки на чем чертить таблицу (ширина столбцов и рядов, их кол-во)
    style_color = color_mesta(data, first_mesto, table, fin) # раскрашивает места участников красным цветом
    t = Table(data, cw, strok * [0.35 * cm])
    # =========  цикл создания стиля таблицы =======
    # ========= 1 страница =========
    style = draw_setka(1, 3, 32, style) # рисует кусок сетки(номер столбца, номер строки на 32 человека)
   
    for l in range(34, 57, 22):
        fn = ('LINEABOVE', (11, l), (13, l), 1, colors.darkblue)  # рисует линии встреч за 1-2 места
        style.append(fn)
# =========== 2 страница ===================
    # # ======= встречи (33-35) за 3-4 место =====
    style = draw_setka_2(1, 74, 16, style) # рисует кусок сетки(номер столбца, номер строки на 16 человека)
    for k in range(0, 7, 6):
        for l in range(72 + k, 89 + k, 16):
            fn = ('LINEABOVE', (11, l), (12, l), 1, colors.darkblue)  # рисует линии встреч за 1-2 места
            style.append(fn)   
    for k in range(0, 17, 16):
        fn = ('BOX', (12, 72 + k), (12, 77 + k), 1, colors.darkblue)
        style.append(fn) 
        fn = ('SPAN', (12, 72 + k), (12, 77 + k))  # встреча 32
        style.append(fn)       
        fn = ('BACKGROUND', (12, 72 + k), (12, 77 + k), colors.lightyellow)  # встречи 32 за 3-4 место
        style.append(fn) 
           
    for l in range(75, 101, 8):
        fn = ('LINEABOVE', (13, l), (13, l), 1, colors.darkblue)  # рисует линии встреч за 3-4 места
        style.append(fn)
    fn = ('BOX', (14, 75), (14, 90), 1, colors.darkblue)
    style.append(fn) 
    fn = ('SPAN', (14, 75), (14, 90))  # встреча 32
    style.append(fn)       
    fn = ('BACKGROUND', (14, 75), (14, 90), colors.lightyellow)  # встречи 32 за 3-4 место
    style.append(fn)  
    # # ======= встречи (61) за 5-6 место =====
    style = draw_setka(11, 102, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=103, col=13, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (36) за 7-8 место =====
    style = draw_setka(11, 110, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=111, col=13, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (37-43) за 9-10 место =====
    style = draw_setka(1, 112, 4, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=115, col=5, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (44) за 11-12 место =====
    style = draw_setka(9, 120, 2, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=121, col=11, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
     # ======= встречи (37-43) за 13-14 место =====
    style = draw_setka(1, 124, 4, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=127, col=5, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (44) за 15-16 место =====
    style = draw_setka(9, 128, 2, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=129, col=11, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
# =========== 3 страница ==================
    # ======= встречи (49-56) за 17-18 место =====
    style = draw_setka(1, 141, 8, style) # рисует кусок сетки(номер столбца, номер строки на 8 человека)
    style = draw_mesta(row=148, col=7, player=8, style=style) # рисует линии сетки за места(номер строки, участники)
   # ======= встречи (64) за 19-20 место =====
    style = draw_setka(9, 156, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=157, col=10, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (33-35) за 21-24 место =====
    style = draw_setka(3, 160, 4, style) # рисует кусок сетки(номер столбца, номер строки на 4 человека)
    style = draw_mesta(row=163, col=6, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (68) за 25-26 место =====
    style = draw_setka(9, 171, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=172, col=10, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (69 - 75) за 25-26 место =====
    style = draw_setka(1, 171, 8, style) # рисует кусок сетки(номер столбца, номер строки на 8 человека)
    style = draw_mesta(row=178, col=7, player=8, style=style) # рисует линии сетки за места(номер строки, участники)
    # ======= встречи (64) за 25-26 место =====
    style = draw_setka(9, 186, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=187, col=10, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
     # ======= встречи (64) за 25-26 место =====
    style = draw_setka(3, 192, 4, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=195, col=6, player=4, style=style) # рисует линии сетки за места(номер строки, участники)
     # ======= встречи (64) за 25-26 место =====
    style = draw_setka(9, 201, 2, style) # рисует кусок сетки(номер столбца, номер строки на 2 человека)
    style = draw_mesta(row=202, col=10, player=2, style=style) # рисует линии сетки за места(номер строки, участники)
   
# =========================================
    for i in range(0, 15, 2):
        fn = ('TEXTCOLOR', (i + 1, 0), (i + 1, 206), colors.black)  # цвет шрифта игроков
        style.append(fn)
        fn = ('TEXTCOLOR', (i, 0), (i, 206), colors.green)  # цвет шрифта номеров встреч
        style.append(fn)
        # выравнивание фамилий игроков по левому краю
        fn = ('ALIGN', (i + 1, 0), (i + 1, 206), 'LEFT')
        style.append(fn)
        # центрирование номеров встреч
        fn = ('ALIGN', (i, 0), (i, 206), 'CENTER')
        style.append(fn)
    # fn = ('INNERGRID', (0, 0), (-1, -1), 0.01, colors.grey)  # временное отображение сетки
    # style.append(fn)
    ts = style   # стиль таблицы (список оформления строк и шрифта)
    for b in style_color:
        ts.append(b)

    t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                           ('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),
                           ('FONTSIZE', (0, 0), (-1, -1), 7),
                           ('FONTNAME', (1, 0), (1, 64), "DejaVuSerif-Bold"),
                           ('FONTSIZE', (1, 0), (1, 64), 7)] + ts 
                           + [
                           # цвет шрифта игроков 1 ого тура
                           ('TEXTCOLOR', (0, 0), (0, 68), colors.blue),
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
# === надпись финала
    h2 = PS("normal", fontSize=10, fontName="DejaVuSerif-Italic",
            leftIndent=50, textColor=Color(1, 0, 1, 1))  # стиль параграфа (номера таблиц)
    elements.append(Paragraph(f"{fin_title}. Одиночный разряд. {gamer}", h2))
# ====                       
    elements.append(t)
    pv = A4
    znak = final.rfind("-")
    f = final[:znak]

    if pv == A4:
        pv = A4
    else:
        pv = landscape(A4)
    t_id = Title.get(Title.id == title_id())
    if tds is not None:
        short_name = t_id.short_name_comp
        name_table_final = f"{short_name}_{f}-final.pdf"
    else:
        short_name = "clear_32_2_net"
        name_table_final = f"{short_name}.pdf"
    doc = SimpleDocTemplate(name_table_final, pagesize=pv, rightMargin=1*cm, leftMargin=1*cm, topMargin=3.4*cm, bottomMargin=1.0*cm)
    catalog = 1
    change_dir(catalog)
    doc.build(elements, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
    os.chdir("..")
    return tds


def mesto_in_final(fin):
    """с какого номера расставляются места в финале, в зависимости от его номера и кол-во участников fin - финал"""
    final = []
    mesto = {}

    system = System.select().where(System.title_id == title_id()) # находит system id последнего
    k = 0
    if fin == "Одна таблица" or fin == "1-й финал" or fin == "Суперфинал":
       mesto[fin] = 1 
    else:
        id_list = []
        fin_list = ["1-й финал", "2-й финал", "3-й финал", "4-й финал",
                            "5-й финал", "6-й финал", "7-й финал", "8-й финал",
                             "9-й финал", "10-й финал", "Суперфинал"]
        for l in system:
            sys_id = l.id
            stage_fin = l.stage
            if stage_fin in fin_list:
                id_list.append(sys_id)
        for k in id_list:
            sys = system.select().where(System.id == k).get()
            max_player = sys.max_player
            stage = sys.stage
            if stage != fin:
                final.append(max_player)
            else:   
                mesto[fin] = sum(final) + 1
                break
    first_mesto = mesto[fin]

    return first_mesto


def write_in_setka(data, stage, first_mesto, table):
    """функция заполнения сетки результатами встреч data поступает чистая только номера в сетке, дальше идет заполнение игроками и счетом"""
    "row_num_win - словарь, ключ - номер игры, значение - список(номер строки 1-ого игрока, номер строки 2-ого игрока) и записвает итоговые места в db"
    sender = my_win.sender()
    player = Player.select().where(Player.title_id == title_id())    
    system_flag = True
    if system_flag is True: 
        id_system = system_id(stage)
    row_num_los = {}
    row_end = 0  # кол-во строк для начальной расстоновки игроков в зависимости от таблицы
    flag_clear = False
    # уточнить кол-во столбцов
    if table == "setka_8_full":
        row_last = 33
        column_last = 8
        row_end = 15
        row_num_win = {1: [1], 2: [5], 3: [9], 4: [13], 5: [3], 6: [11], 7: [7], 8: [16], 9: [20], 10: [24],
                        11: [22], 12: [28]}
                 # ======= list mest
        mesta_dict = {7: 7, 8: 16, 11: 22, 12: 28}
    elif table == "setka_8_2":
        # это вариант при сетке минус 2
        # если встреча верху четная, то на встречу куда идет победитель ( список наоборот) 12: [20, 16]
        row_last = 39
        column_last = 9
        row_end = 15
        row_num_win = {1: [1], 2: [5], 3: [9], 4: [13], 5: [3], 6: [11], 7: [7], 8: [17], 9: [21],
                        10: [16], 11: [20], 12: [18], 13: [25], 14: [31]}
                 # ======= list mest
        mesta_dict = {7: 7, 12: 18, 13: 25, 14: 31}
    elif table == "setka_16_full":
        row_last = 69
        column_last = 11
        row_end = 31
        row_num_win = {1: [1], 2: [5], 3: [9], 4: [13], 5: [17], 6: [21], 7: [25], 8: [29], 9: [3], 10: [11], 11: [19], 12: [27], 13: [7], 14: [23], 
                       15: [15], 16: [29], 17: [32], 18: [36], 19: [34], 20: [39], 21: [41], 22: [45], 23: [49], 24: [53], 25: [43], 26: [51], 27: [47],
                       28: [55], 29: [58], 30: [62], 31: [60], 32: [65]}
                 # ======= list mest
        mesta_dict = {15: 15, 16: 29, 19: 34, 20: 39, 27: 47, 28: 55, 31: 60, 32: 65}
    elif table == "setka_16_2": # встречи, где играют победители и проигравший из основного тура  например 22: [54, 54] в списке одинаковые строки
        row_last = 85
        column_last = 10
        row_end = 33
        row_num_win = {1: [3], 2: [7], 3: [11], 4: [15], 5: [19], 6: [23], 7: [27], 8: [31], 9: [5], 10: [13], 11: [21], 12: [29], 13: [9], 14: [25], 15: [17], 
                       16: [46], 17: [50], 18: [54], 19: [58], 20: [45], 21: [49], 22: [53], 23: [57], 24: [47], 25: [55], 26: [45], 27: [53], 28: [49], 29: [61],
                        30: [67],  31: [62], 32: [66], 33: [64], 34: [73], 35: [74], 36: [78], 37: [76], 38: [79]} 
                 # ======= list mest
        mesta_dict = {15: 17, 28: 49, 29: 61, 33: 64, 30: 67, 34: 73, 37: 76, 38: 79} # номер встречи - номер строки
    elif table == "setka_32":
        row_last = 69
        column_last = 11
        row_end = 65
        row_num_win = {1: [3], 2: [7], 3: [11], 4: [15], 5: [19], 6: [23], 7: [27], 8: [31], 9: [35], 10: [39], 11: [43], 12: [47],
        13: [51], 14: [55], 15: [59], 16: [63], 17: [5], 18: [13], 19: [21], 20: [29], 21: [37], 22: [45], 23: [53], 24: [61],
        25: [9], 26: [25], 27: [41], 28: [57], 29: [17], 30:[49], 31: [33], 32: [61]}
        mesta_dict = {31: 33, 32: 61}
    elif table == "setka_32_2":
        # встреч, которые попадают на сноски (в сетке за 3 место) должно быть в row_num_win а список состоит из одного номера встречи куда идет победитель
        row_last = 207
        column_last = 15
        row_end = 65
        row_num_win = {1: [3], 2: [7], 3: [11], 4: [15], 5: [19], 6: [23], 7: [27], 8: [31], 9: [35], 10: [39], 11: [43], 12: [47],
        13: [51], 14: [55], 15: [59], 16: [63], 17: [5], 18: [13], 19: [21], 20: [29], 21: [37], 22: [45], 23: [53], 24: [61],
        25: [9], 26: [25], 27: [41], 28: [57], 29: [17], 30:[49], 31: [33], 32: [74], 33: [78], 34: [82], 35: [86], 36: [90],
        37: [94], 38: [98], 39: [102], 40: [73], 41: [77], 42: [81], 43: [85], 44:[89], 45: [93], 46: [97], 47: [101], 
        48: [75], 49: [83], 50: [91],  51: [99], 52: [73], 53: [81], 54: [89], 55: [97], 56: [77], 57: [93], 58: [74], 59: [90], 
        60: [81], 63: [112], 64: [116],  67: [124], 68: [128], 71: [141], 72: [145], 73: [149], 74: [153], 75: [143], 76: [151], 
        79: [160], 80: [164], 83: [171], 84: [175], 85: [179], 86: [183], 87: [173], 88: [181], 91: [192], 92: [196]}
                 # ======= dict mest
        mesta_dict = {31: 33, 60: 82, 61: 102, 62: 110, 65: 114, 66: 120, 69: 126, 70: 128, 77: 147,
                        78: 156, 81: 162, 82: 171, 89: 177, 90: 186, 93: 194, 94: 201}
    elif table == "setka_32_full":
        row_last = 207
        column_last = 11
        row_first = 0
        row_end = 65
        row_num_win = {1: [3], 2: [7], 3: [11], 4: [15], 5: [19], 6: [23], 7: [27], 8: [31], 9: [35], 10: [39], 11: [43], 12: [47],
        13: [51], 14: [55], 15: [59], 16: [63], 17: [5], 18: [13], 19: [21], 20: [29], 21: [37], 22: [45], 23: [53], 24: [61],
        25: [9], 26: [25], 27: [41], 28: [57], 29: [17], 30:[49], 31: [33], 32: [61], 33: [72], 34: [76], 35: [74], 36: [84], 37: [89],
        38: [93], 39: [97], 40: [101], 41: [91], 42: [99], 43: [95], 44: [106], 45: [114], 46: [118], 47: [116], 48: [126],  49: [140],
        50: [144], 51: [148], 52: [152], 53: [156], 54: [160], 55: [164], 56: [168], 57: [142], 58: [150], 59: [158], 60: [166], 61: [146],
        62: [162], 63: [154], 64: [168], 65: [172], 66: [176], 67: [174], 68: [182], 69: [179], 70: [183], 71: [187], 72: [191], 73: [181],
        74: [189], 75: [185], 76: [194], 77: [197], 78: [201], 79: [199]}
                 # ======= dict mest
        mesta_dict = {31: 33, 32: 61, 35: 74, 36: 84, 43: 95, 44: 106, 47: 116, 48: 126, 63: 154,
                        64: 168, 67: 174, 68: 182, 75: 185, 76: 194, 79: 199, 80: 201}
    
    if sender == my_win.clear_s32_Action or sender == my_win.clear_s32_full_Action or sender == my_win.clear_s32_2_Action:
        all_list = setka_data_clear(stage, table)  # печать чистой сетки
        col_first = 0
        row_first = 2
        flag_clear = True
    elif sender == my_win.clear_s16_Action:
        all_list = setka_data_clear(stage, table)  # печать чистой сетки
        col_first = 2
        row_first = 0
        flag_clear = True
    elif sender == my_win.clear_s16_2_Action:
        all_list = setka_data_clear(stage, table)  # печать чистой сетки
        col_first = 2
        row_first = 2
        flag_clear = True
    elif sender == my_win.clear_s8_full_Action:
        all_list = setka_data_clear(stage, table)  # печать чистой сетки
        col_first = 2
        row_first = 2
        flag_clear = True
    elif sender == my_win.clear_s8_2_Action:
        all_list = setka_data_clear(stage, table)  # печать чистой сетки
        col_first = 2
        row_first = 2
        flag_clear = True
    else:
        sys = System.select().where(System.title_id == title_id())
        system = sys.select().where(System.id == id_system).get()
        setka_string = system.label_string
        s_2 = 0
        if setka_string == "Сетка (с розыгрышем всех мест) на 8 участников":
            col_first = 0
            row_first = 0
            place_3rd = 8
        elif setka_string == "Сетка (-2) на 8 участников":
            col_first = 0
            row_first = 0
            place_3rd = 12
        elif setka_string == "Сетка (с розыгрышем всех мест) на 16 участников":
            col_first = 2
            row_first = 0
            place_3rd = 16
        elif setka_string == "Сетка (-2) на 16 участников":
            col_first = 0
            row_first = 2
            place_3rd = 28
            s_2  = 16
        elif setka_string == "Сетка (с розыгрышем всех мест) на 32 участников":
            col_first = 0
            row_first = 2
            place_3rd = 32
        elif setka_string == "Сетка (-2) на 32 участников":
            col_first = 0
            row_first = 2
            place_3rd = 60
            s_2  = 32
        posev_data = setka_player_after_choice(stage) # игроки 1-ого посева
        all_list = setka_data(stage, posev_data)
        id_sh_name = all_list[2] # словарь {Фамилия Имя: id}
    tds = []
    tds.append(all_list[0]) # список фамилия/ город 1-ого посева
    # ======
    if flag_clear is False:
        tds.append(id_sh_name)
 
    for d in range(col_first, column_last, 2):
        for r in range(row_first, row_last):
            key = data[r][d]
            if key != "":
                k = int(key)
            if key != "" and k < 0:
                row_num_los[key] = r # словарь номер игры, сноски - номер строки

    n = 0
    for t in range(row_first, row_end, 2):  # цикл расстановки игроков по своим номерам в 1-ом посеве (фамилия инциалы имени/ город)
        data[t][1] = tds[0][n]
        n += 1
    # ==============
    if flag_clear is False:
        # функция расстановки счетов и сносок игроков
        dict_setka = score_in_setka(stage, place_3rd) # список (номер, игрок, счет в партии, номер куда сносится проигравший, его фамилия)
        key_list = []
        mesta_list = []
        for k in dict_setka.keys():
            key_list.append(k) # список всех номеров встреч, которые сыграны
        for v in mesta_dict.keys():
            mesta_list.append(v) # список номеров встреч за места
        # ======
        # if my_win.checkBox_no_play_3.isChecked():
        #     key_list.append(place_3rd)
        key_list.sort()
        # ============
        for i in key_list: # спиисок встреч которые сыграны
            match = dict_setka[i]
            pl_win = match[1]
            pl_los = match[4]
            if pl_win != "X":
                id_win = id_sh_name[pl_win]
            if pl_los != "X":
                id_los = id_sh_name[pl_los]
            else:
                id_los = ""
            # вариант с двумя крестами ===
            if pl_win == "X" and pl_los == "X":
                id_win = ""
                id_los = ""
            r = str(match[3]) # сноска проигравшего
            # ===== определение итоговых мест и запись в db
            if i in mesta_list: # i - номер данной встречи
                index = mesta_list.index(i)
                mesto = first_mesto + (index * 2)
                # записывает места в таблицу -Player-
                if my_win.checkBox_no_play_3.isChecked() and i == place_3rd:
                    for n in [id_win, id_los]: # записывает место в сетке в таблицу -choice- и итоговое место игроку в -player-
                        choice_pl = Choice.get(Choice.player_choice_id == n)
                        choice_pl.mesto_final = mesto
                        choice_pl.save()
                        player = Player.get(Player.id == n)
                        player.mesto = mesto
                        player.save()
                else:
                    m = 0
                    for n in [id_win, id_los]: # записывает место в сетке в таблицу -choice-
                        if n != "":
                            choice_pl = Choice.get(Choice.player_choice_id == n)
                            player = Player.get(Player.id == n)
                            if stage == "Суперфинал":
                                pl = Player.update(mesto=mesto+m).where(Player.id == n).execute()
                            else:
                                choice_pl.mesto_final = mesto + m
                                choice_pl.save()
                                player.mesto = mesto + m
                                player.save()
                            
                            if n == id_win:
                                win = f"{player.player}/{player.city}" 
                            else:
                                los = f"{player.player}/{player.city}"
                            # player.mesto = mesto + m
                            # player.save()
                            m += 1
                    if id_los == "":
                        los = "X"
                    # вариант с двумя крестами
                    if id_win == "":
                        win = "X"
            c = match[0] # номер встречи, куда попадают победитель данной встречи (i)
            # ========== расстановка для сетки на 16
            if c != 0: #  номер встречи в сетке куда попадает победитель (кроме встреч за места)
                row_win = row_num_win[i][0] # номера строк данной встречи в сетке
                c1 = []
                c1_tmp = []
                win = match[1]
                los = match[4]
            elif c == 0:  # встречи за места
                row_win = mesta_dict[i]
                win = match[1]
                los = match[4]
            c = str(i)
           # цикл создания списков номеров встреч по столбцам новый
            column_dict = {}
            # if my_win.checkBox_no_play_3.not isChecked() and c != str(place_3rd):
            # check = my_win.checkBox_no_play_3.isChecked():
            # if check is False:
            for cd in range(2, column_last, 2):
                c1_tmp.clear()
                for rd in range(0, row_last):
                    d1 = data[rd][cd]
                    if d1 != "" and type(d1) == str and int(d1) > 0:
                        c1_tmp.append(d1)
                        c1 = c1_tmp.copy()
                column_dict[cd] = c1    # ключ -номер столбца, значение - список номеров встреч   
                            # =======
            for k in column_dict.keys():
                num_game_list = column_dict[k]  
                if str(i) in num_game_list:
                    if (i == place_3rd and s_2 == 16) or (i == place_3rd and s_2 == 32): # вариант у таблицы 16-2 или 32-2 встреча за 3 место номер столбца
                        col_win = k
                    else: 
                        col_win = k + 1
                    break   

            row_los = row_num_los[r]  # строка проигравшего
            score = match[2]  # счет во встречи
            row_list_los = data[row_los]  # получаем список строки, где ищет номер куда сносится проигравший
            col_los = row_list_los.index(r) # номер столбца проигравшего            
            data[row_win][col_win] = win
            data[row_win + 1][col_win] = score
            data[row_los][col_los + 1] = los
        return tds


def setka_data_clear(fin, table):
    """заполняет сетку для просмотра пустыми фамилиями"""
    all_list = []
    tmp = [""]
    if table == "setka_8_full" or table == "setka_8_2":
        max_pl = 8
    elif table == "setka_16_full" or table == "setka_16_2":
        max_pl = 16
    elif table == "setka_32" or table == "setka_32_full" or table == "setka_32_2":
        max_pl = 32
    tds = tmp * max_pl
    all_list.append(tds)
    return all_list
    

def kol_player(stage):
    """выводит максимальное количество человек в группе t если все группы равны, а g2 если разное количество"""
    id_system = system_id(stage)
    system = System.select().where((System.title_id == title_id()) & (System.id == id_system)).get()
    if stage == "Предварительный":
        all_players = system.total_athletes
        all_group = system.total_group
        flat = all_players % all_group  # если количество участников равно делится на группы
    # если количество участников не равно делится на группы, g2 наибольшое кол-во человек в группе
        player_flat = all_players // all_group
        if flat == 0:
            max_gamer = player_flat
        else:
            max_gamer = player_flat + 1
    else:
        max_gamer = system.max_player // system.total_group
    return max_gamer


def  table_data(stage, kg):
    """циклом создаем список участников каждой группы или финалов по кругу"""
    tdt_all = []  # список списков [tdt_new] и [tdt_color]
    tdt_color = []
    tdt_new = []
    tdt_new_id = []
    result = Result.select().where(Result.title_id == title_id())  # находит system id последнего
    id_system = system_id(stage)
    if kg == 1:  # система одна таблица круг или финалу по кругу
        # список словарей участник и его регион
        result_fin = result.select().where(Result.system_id == id_system)
        tr = len(result_fin)  # общее кол-во игр в финалах или одной таблице
        posev_data = player_choice_one_table(stage) # posev_data (фамилия/ id)
        count_player_group = len(posev_data)
        max_gamer = count_player_group
        num_gr = stage
        tdt_tmp = tdt_news(max_gamer, posev_data, count_player_group, tr, num_gr)
        tdt_new.append(tdt_tmp[0])
        tdt_color.append(tdt_tmp[1])
        tdt_all.append(tdt_new)
        tdt_all.append(tdt_color)
    else:
        max_gamer = kol_player(stage)
        result_stage = result.select().where(Result.system_id == id_system)
        tr = len(result_stage)  # общее кол-во игр в группах
        for p in range(0, kg):
            num_gr = f"{p + 1} группа"
            if stage == "Предварительный":
                posev_data = player_choice_in_group(num_gr) # словарь фамилия:игрок/id регион: область
            else:
                posev_data = player_choice_semifinal(stage, num_gr)
            count_player_group = len(posev_data)
            tdt_tmp = tdt_news(max_gamer, posev_data, count_player_group, tr, num_gr)
            tdt_new.append(tdt_tmp[0])
            tdt_color.append(tdt_tmp[1])
            tdt_new_id.append(tdt_tmp[2])

            tdt_all.append(tdt_new)
            tdt_all.append(tdt_color)
            tdt_all.append(tdt_new_id)
    return tdt_all


def tdt_news(max_gamer, posev_data, count_player_group, tr, num_gr):
    tdt_tmp = []
    tbl_tmp = []  # временный список группы tbl
    tbl_id_tmp = [] # временный список группы вместо фамилия id
    # цикл нумерации строк (по 2-е строки на каждого участника)
    for k in range(1, max_gamer * 2 + 1):
        st = ['']
        # получаем пустой список (номер, фамилия и регион, клетки (кол-во уч), оч, соот, место)
        s = (st * (max_gamer + 4))
        s.insert(0, str((k + 1) // 2))  # получаем нумерацию строк по порядку
        # === добавил вариант с id игрока ===
        s1 = (st * (max_gamer + 4))
        s1.insert(0, str((k + 1) // 2))  # получаем нумерацию строк по порядку
        tbl_tmp.append(s)
        tbl_id_tmp.append(s1)
    for i in range(1, count_player_group * 2 + 1, 2):
        posev = posev_data[((i + 1) // 2) - 1]
        fam_id = posev["фамилия"]
        znak = fam_id.find("/")
        #==== вариант вместо фамилия id игрока
        # if znak != -1:
        #     tbl_tmp[i - 1][1] = fam_id[znak + 1:]
        # else:
        #     tbl_tmp[i - 1][1] = posev["фамилия"]
        #=====================================
        fam = fam_id[:znak]
        id_fam = fam_id[znak + 1:]
        if znak != -1:
            tbl_tmp[i - 1][1] = fam
            tbl_id_tmp[i - 1][1] = id_fam
        else:
            tbl_tmp[i - 1][1] = posev["фамилия"]
        # =============================
        tbl_tmp[i][1] = posev["регион"]
 
    td = tbl_tmp.copy()  # cписок (номер, фамилия, город и пустые ячейки очков)
    td_id = tbl_id_tmp.copy()
    td_color = []

    if tr != 0:  # если еще не была жеребьевка, то пропуск счета в группе
        # список очки победителя красные (ряд, столбец) без заголовка
        td_color = score_in_table(td, num_gr)

    tdt_new = td
    tdt_tmp.append(tdt_new)
    tdt_tmp.append(td_color)
    tdt_tmp.append(td_id)

    return tdt_tmp


def setka_player_after_choice(stage):
    """список игроков сетки после жеребьевки"""
    p_data = {}
    posev_data = []
    id_system = system_id(stage)
    player = Player.select().where(Player.title_id == title_id())
    game_list = Game_list.select().where(Game_list.title_id == title_id())
    pl_list = game_list.select().where(Game_list.system_id == id_system).order_by(Game_list.rank_num_player)
    for i in pl_list:
        p_data['посев'] = i.rank_num_player
        txt = i.player_group_id
        if txt != "X":
            # ==== вариант новый с id игрока
            id_pl = i.player_group_id
            pl = player.select().where(Player.id == id_pl).get()
            p_data['фамилия'] = pl.full_name
        else:
            p_data['фамилия'] = "X"
        tmp = p_data.copy()
        posev_data.append(tmp)
        p_data.clear()
    return posev_data


def setka_data(stage, posev_data):
    """данные сетки"""
    id_full_name = {}
    id_name = {}
    tds = []
    fam_name_city = []
    id_system = system_id(stage)
    system = System.select().where((System.title_id == title_id()) & (System.id == id_system)).get()  # находит system id последнего

    mp = system.max_player
    mp = full_net_player(player_in_final=mp)
    for i in range(1, mp * 2 + 1, 2):
        posev = posev_data[((i + 1) // 2) - 1]
        family = posev['фамилия'] # фамилия имя / город
        name_list = full_player_id(family) # словарь {name: фамилия/город, id: номер игрока}, {name: фамилия, id: номер мгрока}
        id_f_n = name_list[0] # словарь name: фамилия/город, id: номер игрока
        id_s_n = name_list[1] # {name: фамилия, id: номер игрока}
            # словарь ключ - полное имя/ город, значение - id
        id_full_name[id_f_n["name"]] = id_f_n["id"]
        id_name[id_s_n["name"]] = id_s_n["id"]
        if family != "X":
            # находит пробел отделяющий имя от фамилии
            space = family.find(" ")
            line = family.find("/")  # находит черту отделяющий имя от города
            city_slice = family[line:]  # получает отдельно город
            # получает отдельно фамилия и первую букву имени
            family_slice = family[:space + 2]
            family_city = f'{family_slice}.{city_slice}'   # все это соединяет
        tds.append(family)
        fam_name_city.append(family)
    all_list = [tds, id_full_name, id_name, fam_name_city]

    return all_list


def full_player_id(family):
    """получает словарь -фамилия игрока и его город и соответствующий ему id в таблице Players"""
    full_name = {}
    short_name = {}   
    player = Player.select().where(Player.title_id == title_id())
    if family != "X":
        pl_id = player.select().where(Player.full_name == family).get()
        player_id = pl_id.id # ид игрока
        f_name = pl_id.full_name
        s_name = pl_id.player
        full_name["name"] = f_name
        full_name["id"] = player_id 
        short_name["name"] = s_name
        short_name["id"] = player_id
    else:
        full_name["name"] = "X"
        full_name["id"] = 0
        short_name["name"] = "X"
        short_name["id"] = 0
    name_list = []
    name_list.append(full_name)
    name_list.append(short_name)

    return name_list


def score_in_table(td, num_gr):
    """заносит счет и места в таблицу группы или таблицу по кругу pdf
    -td- список строки таблицы, куда пишут счет"""
    td_color = []
    total_score = {}  # словарь, где ключ - номер участника группы, а значение - очки
    sender = my_win.sender()
    tab = my_win.tabWidget.currentIndex()
    if tab == 0:
        my_win.tabWidget.setCurrentIndex(3)
        tab = my_win.tabWidget.currentIndex()
    tab_etap = my_win.tabWidget_stage.currentIndex()
    result = Result.select().where(Result.title_id == title_id())
    choice = Choice.select().where(Choice.title_id == title_id())
    gamelist = Game_list.select().where(Game_list.title_id == title_id())
    if tab == 7: # открыта вкладка для редактирования групп
        stage = my_win.comboBox_edit_etap1.currentText()
        id_system = system_id(stage)
        mp = len(gamelist.select().where((Game_list.system_id == id_system) & (Game_list.number_group == num_gr)))
        results = result.select().where((Result.system_stage_id == id_system) & (Result.number_group == num_gr))
        ch = choice.select().where((Choice.semi_final == stage) & (Choice.sf_group == num_gr))  # фильтрует по группе
    elif tab == 3: # Результаты
        if tab_etap == 0: # группы
            stage = "Предварительный"
            id_system = system_id(stage)   
            ta = System.select().where(System.id == id_system).get()  # находит system id последнего
            results = result.select().where((Result.system_id== id_system) & (Result.number_group == num_gr))
            ch = choice.select().where(Choice.group == num_gr)  # фильтрует по группе
            mp = ta.max_player
            stage = ta.stage
        elif tab_etap == 1: # полуфиналы
            if sender == my_win.view_pf1_Action:
                stage = "1-й полуфинал"
            elif sender == my_win.view_pf2_Action:
                stage = "2-й полуфинал"
            else:
                stage = my_win.comboBox_filter_semifinal.currentText()
            id_system = system_id(stage) # получает id системы из комбобокса
            # вариант если в ПФ не полные группы
            ta = System.select().where(System.id == id_system).get()  # находит system id последнего
            max_pl = ta.max_player # общее кол-во игрков в ПФ
            group_in_sf = ta.total_group # общее кол-во групп в ПФ
            mp = max_pl // group_in_sf
            # =======
            # mp = len(gamelist.select().where((Game_list.system_id == id_system) & (Game_list.number_group == num_gr)))
            results = result.select().where((Result.system_id == id_system) & (Result.number_group == num_gr))
            ch = choice.select().where((Choice.semi_final == stage) & (Choice.sf_group == num_gr))  # фильтрует по группе
        elif tab_etap == 2 or my_win.choice_fin_Action: # финалы
            stage = num_gr
            id_system = system_id(stage)
            systems = System.select().where(System.id == id_system).get()
            etap_exit = systems.stage_exit # этап откуда выходят в данный финал
            results = result.select().where(Result.system_id == id_system)
            if num_gr == "Одна таблица":
                stage = "Одна таблица"
                ch = choice.select().where(Choice.basic == "Одна таблица")  # фильтрует по одной таблице
            else: # игры в финале по кругу
                if etap_exit == "Предварительный":
                    ch = choice.select().where(Choice.final == num_gr)
                else:
                    # ch = choice.select().where((Choice.final == num_gr) & (Choice.semi_final == etap_exit))
                    ch = choice.select().where(Choice.final == num_gr)
            mp = len(gamelist.select().where(Game_list.system_id == id_system))
    
    count = len(results)  # сколько игр в группе
    count_player = len(ch)  # определяет сколько игроков в группе
    result_list = results.dicts().execute()
    for s in range(1, count_player + 1):
        total_score[s] = 0
    for i in range(0, count):
        sc_game = str(list(result_list[i].values())[9])  # счет в партиях
        if sc_game != "" or sc_game != "None":
            scg = 9
        else:  # номер столбца
            scg = 8
        tours = str(list(result_list[i].values())[3])  # номера игроков в туре
        znak = tours.find("-")
        p1 = int(tours[:znak])  # игрок под номером в группе
        p2 = int(tours[znak + 1:])  # игрок под номером в группе

        win = str(list(result_list[i].values())[6])
        player1 = str(list(result_list[i].values())[4])

        if win != "" and win != "None":  # если нет сыгранной встречи данного тура
            if win == player1:  # если победитель игрок под первым номером в туре
                # очки 1-ого игрока
                td[p1 * 2 - 2][p2 + 1] = str(list(result_list[i].values())[7])  # ячейка в таблице очки
                # счет 1-ого игрока
                td[p1 * 2 - 1][p2 + 1] = str(list(result_list[i].values())[scg])   # ячейка в таблице счет впартии
                # очки 2-ого игрока
                td[p2 * 2 - 2][p1 + 1] = str(list(result_list[i].values())[11])  # ячейка в таблице очки
                # счет 2-ого игрока
                td[p2 * 2 - 1][p1 + 1] = str(list(result_list[i].values())[12])  # ячейка в таблице счет впартии
                # очки 1-ого игрока
                tp1 = str(list(result_list[i].values())[7])
                # очки 2-ого игрока
                tp2 = str(list(result_list[i].values())[11])
                tp1 = 0 if tp1 == "" else str(list(result_list[i].values())[7])
                tp2 = 0 if tp2 == "" else str(list(result_list[i].values())[11])
                # считывает из словаря 1-ого игрока всего очков
                plr1 = total_score[p1]
                # считывает из словаря 2-ого игрока всего очков
                plr2 = total_score[p2]
                plr1 = plr1 + int(tp1)  # прибавляет очки 1-ого игрока
                plr2 = plr2 + int(tp2)  # прибавляет очки 2-ого игрока
                total_score[p1] = plr1  # записывает сумму очков 1-му игроку
                total_score[p2] = plr2  # записывает сумму очков 2-му игроку
                col = p1 * 2 - 2
                row = p2 + 1
            else:  # если победитель игрок под вторым номером в туре
                # очки 1-ого игрока
                td[p1 * 2 - 2][p2 + 1] = str(list(result_list[i].values())[11])
                # счет 1-ого игрока
                td[p1 * 2 - 1][p2 + 1] = str(list(result_list[i].values())[12])
                # очки 2-ого игрока
                td[p2 * 2 - 2][p1 + 1] = str(list(result_list[i].values())[7])
                # счет 2-ого игрока
                td[p2 * 2 - 1][p1 + 1] = str(list(result_list[i].values())[scg])
                # очки 1-ого игрока
                tp1 = str(list(result_list[i].values())[11])
                # очки 2-ого игрока
                tp2 = str(list(result_list[i].values())[7])
                tp1 = 0 if tp1 == "" else str(list(result_list[i].values())[11])
                tp2 = 0 if tp2 == "" else str(list(result_list[i].values())[7])
                # считывает из словаря 1-ого игрока очки
                plr1 = total_score[p1]
                # считывает из словаря 2-ого игрока очки
                plr2 = total_score[p2]
                plr1 = plr1 + int(tp1)  # прибавляет очки 1-ого игрока
                plr2 = plr2 + int(tp2)  # прибавляет очки 2-ого игрока
                total_score[p1] = plr1  # записывает сумму очков 1-му игроку
                total_score[p2] = plr2  # записывает сумму очков 2-му игроку
                col = p2 * 2 - 2
                row = p1 + 1
            # список ряд столбец, где очки надо красить в красный
            td_tmp = [row, col]
            td_color.append(td_tmp)
    for t in range(0, count_player):  # записывает очки в зависимости от кол-во игроков в группе
        # записывает каждому игроку сумму очков
        td[t * 2][mp + 2] = total_score[t + 1]
    # ===== если сыграны все игры группе то выставляет места =========
    count_game = (count_player * (count_player - 1)) // 2 # сколько всего игр в группе

    results_playing = results.select().where((Result.points_win == 2) | (Result.points_win == 0))
    a = len(results_playing) # кол-во сыгранных игр

    if a == count_game:
        rank_in_group(total_score, td, num_gr, stage)  # определяет места в группе
    return td_color


def numer_game(num_game, vid_setki):
    """определяет куда записывать победителя и проигравшего по сноске в сетке, номера встреч"""
    snoska = []
    num_game = int(num_game)
    if vid_setki == 'Сетка (с розыгрышем всех мест) на 8 участников':
        dict_winner = {1:5, 2:5, 3:6, 4:6, 5:7, 6:7, 9:11, 10:11}
        dict_loser = {1:9, 2:9, 3:10, 4:10, 5:8, 6:8, 9:12, 10:12}
        dict_loser_pdf = {1: -1, 2: -2, 3: -3, 4: -4, 5: -5, 6: -6, 7: -7, 8: -8, 9: -9, 10: -10, 11: -11, 12: -12}
        dict_mesta = [7, 8, 11, 12]
    elif vid_setki == 'Сетка (-2) на 8 участников':
        dict_winner = {1:5, 2:5, 3:6, 4:6, 5:7, 6:7, 8:10, 9:11, 10:12, 11:12}
        dict_loser = {1:8, 2:8, 3:9, 4:9, 5:11, 6:10, 10:13, 11:13, 8:14, 9:14}
        dict_loser_pdf = {1: -1, 2: -2, 3: -3, 4: -4, 5: -5, 6: -6, 7: -7, 8: -8, 9: -9, 10: -10, 11: -11, 12: -12, 13: -13,
                      14: -14}
        dict_mesta = [7, 12, 13, 14]
    elif vid_setki == 'Сетка (с розыгрышем всех мест) на 16 участников':
        dict_winner = {1: 9, 2: 9, 3: 10, 4: 10, 5: 11, 6: 11, 7: 12, 8: 12, 9: 13, 10: 13, 11: 14, 12: 14, 13: 15, 14: 15,
                   17: 19, 18: 19, 21: 25, 22: 25, 23: 26, 24: 26, 25: 27, 26: 27, 29: 31, 30: 31}
        dict_loser = {1: 21, 2: 21, 3: 22, 4: 22, 5: 23, 6: 23, 7: 24, 8: 24, 9: 17, 10: 17, 11: 18, 12: 18, 13: 16, 14: 16,
                  17: 20, 18: 20, 21: 29, 22: 29, 23: 30, 24: 30, 25: 28, 26: 28, 29: 32, 30: 32}
        dict_loser_pdf = {1: -1, 2: -2, 3: -3, 4: -4, 5: -5, 6: -6, 7: -7, 8: -8, 9: -9, 10: -10, 11: -11, 12: -12, 13: -13,
                      14: -14, 17: -17, 18: -18, 21: -21, 22: -22, 23: -23, 24: -24, 25: -25, 26: -26, 29: -29, 30: -30}
        dict_mesta = [15, 16, 19, 20, 27, 28, 31, 32]
    elif vid_setki == 'Сетка (-2) на 16 участников':
        dict_winner = {1:9, 2:9, 3:10, 4:10, 5:11, 6:11, 7:12, 8:12, 9:13, 10:13, 11:14, 12:14, 13:15, 14:15,
                   16:20, 17:21, 18:22, 19:23, 20:24, 21:24, 22:25, 23:25, 24:26, 25:27, 26:28, 27:28, 31:33, 32:33, 35:37, 36:37}
        dict_loser = {1:16, 2:16, 3:17, 4:17, 5:18, 6:18, 7:19, 8:19, 9:23, 10:22, 11:21, 12:20, 13:26, 14:27,
                  16:35, 17:35, 18:36, 19:36, 20:31, 21:31, 22:32, 23:32, 24:30, 25:30, 26:29, 27:29, 31:34, 32:34, 35:38, 36:38}
        dict_loser_pdf = {1: -1, 2: -2, 3: -3, 4: -4, 5: -5, 6: -6, 7: -7, 8: -8, 9: -9, 10: -10, 11: -11, 12: -12, 13: -13,
                      14: -14, 16: -16, 17: -17, 18: -18, 19: -19, 20: -20, 21: -21, 22: -22, 23: -23, 24: -24, 25: -25, 26: -26, 
                      27: -27, 31: -31, 32: -32, 35: -35, 36: -36}
        dict_mesta = [15, 28, 29, 30, 33, 34, 37, 38]
    elif vid_setki == 'Сетка (с розыгрышем всех мест) на 32 участников':
        dict_winner = {1: 17, 2: 17, 3: 18, 4: 18, 5: 19, 6: 19, 7: 20, 8: 20, 9: 21, 10: 21, 11: 22, 12: 22, 13: 23, 14: 23,
                   15: 24, 16: 24, 17: 25, 18: 25, 19: 26, 20: 26, 21: 27, 22: 27, 23: 28, 24: 28, 25: 29, 26: 29, 27: 30, 28: 30, 
                   29: 31, 30: 31, 33: 35, 34: 35, 37: 41, 38: 41, 39: 42, 40: 42, 41: 43, 42: 43, 45: 47, 46: 47, 49: 57, 50: 57,
                   51: 58, 52: 58, 53: 59, 54: 59, 55: 60, 56: 60, 57: 61, 58: 61, 59: 62, 60: 62, 61: 63, 62: 63, 65: 67, 66: 67,
                   69: 73, 70: 73, 71: 74, 72: 74, 73: 75, 74: 75, 77: 79, 78: 79}
        dict_loser = {1: 49, 2: 49, 3: 50, 4: 50, 5: 51, 6: 51, 7: 52, 8: 52, 9: 53, 10: 53, 11: 54, 12: 54, 13: 55, 14: 55, 15: 56, 16: 56,
                  17: 37, 18: 37, 19: 38, 20: 38, 21: 39, 22: 39, 23: 40, 24: 40, 25: 33, 26: 33, 27: 34, 28: 34, 29: 32, 30: 32,
                  33: 36, 34: 36, 37: 45, 38: 45, 39: 46, 40: 46, 41: 44, 42: 44, 45: 48, 46: 48, 49: 69, 50: 69, 51: 70, 52: 70, 53: 71, 54: 71, 55: 72, 56: 72,
                  57: 65, 58: 65, 59: 66, 60: 66, 61: 64, 62: 64, 65: 68, 66: 68, 69: 77, 70: 77, 71: 78, 72: 78, 73: 76, 74: 76, 77: 80, 78: 80}
        dict_loser_pdf = {1: -1, 2: -2, 3: -3, 4: -4, 5: -5, 6: -6, 7: -7, 8: -8, 9: -9, 10: -10, 11: -11, 12: -12, 13: -13,
                    14: -14, 15: -15, 16: -16, 17: -17, 18: -18, 19: -19, 20: -20, 21: -21, 22: -22, 23: -23, 24: -24, 25: -25, 26: -26, 
                    27: -27, 28: -28, 29: -29, 30: -30, 33: -33, 34: -34, 37: -37, 38: -38, 39: -39, 40: -40, 41: -41, 42: -42, 45: -45, 46: -46,
                    49: -49, 50: -50, 51: -51, 52: -52, 53: -53, 54: -54, 55: -55, 56: -56, 57: -57, 58: -58, 59: -59, 60: -60, 61: -61, 62: -62, 65: -65, 66: -66,
                    69: -69, 70: -70, 71: -71, 72: -72, 73: -73, 74: -74, 75: -75, 77: -77, 78: -78, 79: -79}
        dict_mesta = [31, 32, 35, 36, 43, 44, 47, 48, 63, 64, 67, 68, 75, 76, 79, 80]
    elif vid_setki == 'Сетка (-2) на 32 участников':
        dict_winner = {1: 17, 2: 17, 3: 18, 4: 18, 5: 19, 6: 19, 7: 20, 8: 20, 9: 21, 10: 21, 11: 22, 12: 22, 13: 23, 14: 23,
                   15: 24, 16: 24, 17: 25, 18: 25, 19: 26, 20: 26, 21: 27, 22: 27, 23: 28, 24: 28, 25: 29, 26: 29, 27: 30, 28: 30, 
                   29: 31, 30: 31, 32: 40, 33: 41, 34: 42, 35: 43, 36: 44, 37: 45, 38: 46, 39: 47, 40: 48, 41: 48, 42: 49, 43: 49, 44: 50, 45: 50, 46: 51,
                   47: 51, 48: 52, 49: 53, 50: 54, 51: 55, 52: 56, 53: 56, 54: 57, 55: 57, 56: 58, 57: 59, 58: 60, 59: 60, 63: 65, 64: 65, 67: 69, 68: 69, 71: 75, 72: 75, 73: 76, 74: 76, 
                   75: 77, 76: 77, 79: 81, 80: 81, 83: 87, 84: 87, 85: 88, 86: 88, 87: 89, 88: 89, 91: 93, 92: 93}
        
        dict_loser = {1: 32, 2: 32, 3: 33, 4: 33, 5: 34, 6: 34, 7: 35, 8: 35, 9: 36, 10: 36, 11: 37, 12: 37, 13: 38, 14: 38, 15: 39, 16: 39,
                  17: 47, 18: 46, 19: 45, 20: 44, 21: 43, 22: 42, 23: 41, 24: 40, 25: 53, 26: 52, 27: 55, 28: 54, 29: 59, 30: 58,
                  32: 83, 33: 83, 34: 84, 35: 84, 36: 85, 37: 85, 38: 86, 39: 86, 40: 71, 41: 71, 42: 72, 43: 72, 44: 73, 45: 73, 46: 74, 47: 74, 48: 67, 49: 67, 
                  50: 68, 51: 68, 52: 63, 53: 63, 54: 64, 55: 64, 56: 62, 57: 62, 58: 61, 59: 61, 63: 66, 64: 66, 67: 70, 68: 70, 71: 79, 72: 79, 73: 80, 74: 80, 
                  79: 82, 80: 82, 75: 78, 76: 78, 83: 91, 84: 91, 85: 92, 86: 92, 87: 90, 88: 90, 91: 94, 92: 94}
        
        dict_loser_pdf = {1: -1, 2: -2, 3: -3, 4: -4, 5: -5, 6: -6, 7: -7, 8: -8, 9: -9, 10: -10, 11: -11, 12: -12, 13: -13,
                    14: -14, 15: -15, 16: -16, 17: -17, 18: -18, 19: -19, 20: -20, 21: -21, 22: -22, 23: -23, 24: -24, 25: -25, 26: -26, 
                    27: -27, 28: -28, 29: -29, 30: -30, 32: -32, 33: -33, 34: -34, 35: -35, 36: -36, 37: -37, 38: -38, 39: -39, 40: -40, 
                    41: -41, 42: -42, 43: -43, 44: -44, 45: -45, 46: -46, 47: -47, 48: -48, 49: -49, 50: -50, 51: -51, 52: -52, 53: -53, 54: -54, 
                    55: -55, 56: -56, 57: -57, 58: -58, 59: -59, 63: -63, 64: -64, 67: -67, 68: -68,
                    71: -71, 72: -72, 73: -73, 74: -74, 75: -75, 76: -76, 79: -79, 80: -80, 83: -83, 84: -84, 85: -85, 86: -86, 87: -87,
                    88: -88, 91: -91, 92: -92}
        dict_mesta = [31, 60, 61, 62, 65, 66, 69, 70, 77, 78, 81, 82, 89, 90, 93, 94]
    elif vid_setki == 'Сетка (1-3 место) на 32 участников': # поправить
        dict_winner = {1: 17, 2: 17, 3: 18, 4: 18, 5: 19, 6: 19, 7: 20, 8: 20, 9: 21, 10: 21, 11: 22, 12: 22, 13: 23, 14: 23,
                   15: 24, 16: 24, 17: 25, 18: 25, 19: 26, 20: 26, 21: 27, 22: 27, 23: 28, 24: 28, 25: 29, 26: 29, 27: 30, 28: 30, 
                   29: 31, 30: 31, 33: 35, 34: 35, 37: 41, 38: 41, 39: 42, 40: 42, 41: 43, 42: 43, 45: 47, 46: 47, 49: 57, 50: 57,
                   51: 58, 52: 58, 53: 59, 54: 59, 55: 60, 56: 60, 57: 61, 58: 61, 59: 62, 60: 62, 61: 63, 62: 63, 65: 67, 66: 67,
                   69: 73, 70: 73, 71: 74, 72: 74, 73: 75, 74: 75, 77: 79, 78: 79}
        dict_loser = {1: 49, 2: 49, 3: 50, 4: 50, 5: 51, 6: 51, 7: 52, 8: 52, 9: 53, 10: 53, 11: 54, 12: 54, 13: 55, 14: 55, 15: 56, 16: 56,
                  17: 37, 18: 37, 19: 38, 20: 38, 21: 39, 22: 39, 23: 40, 24: 40, 25: 33, 26: 33, 27: 34, 28: 34, 29: 32, 30: 32,
                  33: 36, 34: 36, 37: 45, 38: 45, 39: 46, 40: 46, 41: 44, 42: 44, 45: 48, 46: 48, 49: 69, 50: 69, 51: 70, 52: 70, 53: 71, 54: 71, 55: 72, 56: 72,
                  57: 65, 58: 65, 59: 66, 60: 66, 61: 64, 62: 64, 65: 68, 66: 68, 69: 77, 70: 77, 71: 78, 72: 78, 73: 76, 74: 76, 77: 80, 78: 80}
        dict_loser_pdf = {1: -1, 2: -2, 3: -3, 4: -4, 5: -5, 6: -6, 7: -7, 8: -8, 9: -9, 10: -10, 11: -11, 12: -12, 13: -13,
                    14: -14, 15: -15, 16: -16, 17: -17, 18: -18, 19: -19, 20: -20, 21: -21, 22: -22, 23: -23, 24: -24, 25: -25, 26: -26, 
                    27: -27, 28: -28, 29: -29, 30: -30, 33: -33, 34: -34, 37: -37, 38: -38, 39: -39, 40: -40, 41: -41, 42: -42, 45: -45, 46: -46,
                    49: -49, 50: -50, 51: -51, 52: -52, 53: -53, 54: -54, 55: -55, 56: -56, 57: -57, 58: -58, 59: -59, 60: -60, 61: -61, 62: -62, 65: -65, 66: -66,
                    69: -69, 70: -70, 71: -71, 72: -72, 73: -73, 74: -74, 75: -75, 77: -77, 78: -78, 79: -79}
        dict_mesta = [31, 32, 35, 36, 43, 44, 47, 48, 63, 64, 67, 68, 75, 76, 79, 80]

    if num_game in dict_mesta: # если встреча за места
        snoska = [0, 0, (num_game * -1)]
    else:
        game_winner = dict_winner[num_game]  # номер игры победителя
        snoska.append(game_winner)
        game_loser = dict_loser[num_game]  # номер игры проигравшего
        snoska.append(game_loser)
        # для отображения в pdf (встречи с минусом)
        game_loser = dict_loser_pdf[num_game]
        snoska.append(game_loser) # список: номер встречи победителя, номер - проигравшего и куда снести проигравшего
    return snoska


def score_in_setka(stage, place_3rd):
    """ выставляет счет победителя и сносит на свои места в сетке"""
    dict_setka = {}
    match = []
    tmp_match = []
    id_system = system_id(stage)
    system = System.select().where(System.id == id_system).get()
    vid_setki = system.label_string
    visible_game = system.visible_game
    # получение id последнего соревнования
    player = Player.select().where(Player.title_id == title_id())
    result = Result.select().where(Result.system_id == id_system)
    for res in result:
        num_game = int(res.tours)
        
        if res.winner is not None and res.winner != "": # значит встреча сыграна
            if num_game == place_3rd and my_win.checkBox_no_play_3.isChecked(): # если два 3-х места
                if res.player1 != "" and res.player2 != "":
                    res = result.select().where(Result.tours == place_3rd).get()
                    id_pl1 = player.select().where(Player.full_name == res.player1).get()
                    id_pl2 = player.select().where(Player.full_name == res.player2).get()
                    short_name_win = id_pl1.player
                    short_name_los = id_pl2.player
                    match = [0, short_name_win, '', '', short_name_los]
                    dict_setka[num_game] = match
            elif res.winner != "X":
                id_pl_win = player.select().where(Player.full_name == res.winner).get()
                short_name_win = id_pl_win.player
                if res.loser == "X":
                    short_name_los = "X"
                else: 
                    id_pl_los = player.select().where(Player.full_name == res.loser).get()
                    short_name_los = id_pl_los.player
            else:
                short_name_win = "X"
                short_name_los = "X"

            snoska = numer_game(num_game, vid_setki) # список (номер встречи победителя, номер встречи проигравшего и минус куда идет проигравший в сетке)
            tmp_match.append(snoska[0]) # номер на сетке куда идет победитель
            tmp_match.append(short_name_win)
            if num_game == place_3rd and my_win.checkBox_no_play_3.isChecked():
                tmp_match.append('') # при два 3 места не писать счет в партии
            else:
                if visible_game == 1: # если счет в партиии
                    tmp_match.append(f'{res.score_in_game} {res.score_win}')
                else:
                    tmp_match.append(f'{res.score_in_game}')
            tmp_match.append(snoska[2])
            tmp_match.append(short_name_los)
            match = tmp_match.copy()
            tmp_match.clear()
            dict_setka[num_game] = match

    return dict_setka


def result_rank_group_in_choice(num_gr, player_rank_group, stage):
    """записывает места из группы в таблицу -Choice-, а если одна таблица в финале по кругу то в список
    player_rank_group список списков 1-е число номер игрок в группе, 2-е его место"""
    # tab = my_win.tabWidget.currentIndex()
    tab_etap = my_win.tabWidget_stage.currentIndex()

    chc = Choice.select().where(Choice.title_id == title_id())
    if len(player_rank_group) > 0:
        if tab_etap == 0:
            choice = chc.select().where(Choice.group == num_gr)
        elif tab_etap == 1:
            choice = chc.select().where((Choice.semi_final == stage) & (Choice.sf_group == num_gr))
        else:
            if num_gr == "Одна таблица":
                choice = chc.select().where(Choice.basic == "Одна таблица")
            else:
                choice = chc.select().where(Choice.final == num_gr).order_by(Choice.posev_final)
        count = len(choice)
        n = 0
        for ch in choice:
            if tab_etap == 0:
                for i in range(0, count):  # цикл по номерам посева в группе
                    # если есть совпадение, то место в списке
                    if ch.posev_group == player_rank_group[i][0]:
                        with db:
                            # записывает в таблицу -Choice- места в группе
                            ch.mesto_group = player_rank_group[i][1]
                            ch.save()
            elif tab_etap == 1:
                for i in range(0, count):  # цикл по номерам посева в группе
                    # если есть совпадение, то место в списке
                    if ch.posev_sf == player_rank_group[i][0]:
                        with db:
                            # записывает в таблицу -Choice- места в группе
                            ch.mesto_semi_final = player_rank_group[i][1]
                            ch.save()
            else:
                player_rank_group.sort()
                ch.mesto_final = player_rank_group[n][1]
                player_id = ch.player_choice_id
                ch.save()
                player = Player.get(Player.id == player_id)
                player.mesto = player_rank_group[n][1]
                player.save()
                n += 1


def rank_in_group(total_score, td, num_gr, stage):
    """выставляет места в группах соответственно очкам 
    -men_of_circle - кол-во человек в крутиловке
    -player_rank_group - список списков номер игрока - место 
    -num_player -player_rank - список списков участник - место
    -player_group - кол-во участников в группе"""
    tr_all = []
    pps = []
    group_list = ["Предварительный", "1-й полуфинал", "2-й полуфинал"]
    no_final_list = ["Одна таблица", "Предварительный", "1-й полуфинал", "2-й полуфинал"]
    pp = {}  # ключ - игрок, значение его очки
    pg_win = {}
    pg_los = {}
    tr = []
    player_rank_tmp = []

    rev_dict = {}  # словарь, где в качестве ключа очки, а значения - номера групп
    player_rank_group = []    
    result = Result.select().where(Result.title_id == title_id())
    game_list = Game_list.select().where(Game_list.title_id == title_id())
    # ========
    id_system = system_id(stage)    
    game_list_group = game_list.select().where((Game_list.system_id == id_system) & (Game_list.number_group == num_gr))
    if stage == "Одна таблица" or stage not in group_list:
        game_max = result.select().where(Result.system_id == id_system)  # сколько всего игр в группе
        game_list_group = game_list.select().where(Game_list.system_id == id_system)
        max_person = len(game_list_group)
    else:
        if stage == "Предварительный":
            systems = System.get(System.id == id_system)
            max_person = systems.max_player
        elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
            # ====
            systems = System.get(System.id == id_system)
            total_player = systems.max_player
            total_group = systems.total_group
            max_person = total_player // total_group

            # =====
            # game_list_group = game_list.select().where((Game_list.system_id == id_system) & (Game_list.number_group == num_gr))
            # max_person = len(game_list_group)
        game_max = result.select().where((Result.system_id == id_system) & (Result.number_group == num_gr))  # сколько всего игр в группе
    # ======== проверка на неявку ======
    fio_no_player = []
    game_not_player = game_max.select().where(Result.points_loser == 0)
    count_not_player = len(game_not_player)
    if count_not_player != 0:
        for k in game_not_player:
            pl = k.points_loser
            if pl == 0:
                player_no = k.loser
                if player_no not in fio_no_player:
                    fio_no_player.append(player_no)
                    
        for fio in fio_no_player:           
            fio_loser = game_not_player.select().where(Result.loser == fio)
            count_fio_loser = len(fio_loser)

            game_one_person = max_person // 2
            if count_fio_loser >= game_one_person: # игры по неявке более 50%
                result_no = result.select().where(Result.system_id == id_system)
                game_id_not_player = result_no.select().where((Result.player1 == fio) | (Result.player2 == fio))
                cou = len(game_id_not_player)
                for game_id in game_id_not_player:
                    game_id.points_win = 0
                    game_id.points_loser = 0
                    game_id.save()
     # ===========================================
    # 1-й запрос на выборку с группой
    game_played = game_max.select().where((Result.winner is None) | (Result.winner != ""))  # 2-й запрос на выборку
    # с победителями из 1-ого запроса
    kol_tours_played = len(game_played)  # сколько игр сыгранных
    kol_tours_in_group = len(game_max)  # кол-во всего игр в группе

    for key, value in total_score.items():
        rev_dict.setdefault(value, set()).add(key) # словарь (число очков, номера участников группы у которых они есть)
    res = [key for key, values in rev_dict.items() if len(values) > 1] # список очки, которых более чем у одного участника
    # сортирует словарь по убыванию очков
    sorted_tuple = {k: total_score[k] for k in sorted(total_score, key=total_score.get, reverse=True)}
    valuesList = list(sorted_tuple.values())  # список очков по убыванию
    unique_numbers = list(set(valuesList))  # множество уникальных очков
    unique_numbers.sort(reverse=True)  # список уникальных очков по убыванию
    mesto = 1
    # +++++ вариант начального места в финале
    if stage in no_final_list:
        mesto = 1
    else:
        mesta_list = []
        systems = System.select().where(System.title_id == title_id())
        for k in systems:
            if k.stage not in no_final_list:
                player_in_final = k.max_player
                if k.stage != stage:
                    mesta_list.append(player_in_final)
                else:
                    break
        mesto = 1 + sum(mesta_list)

    for f in unique_numbers:  # проходим циклом по уник. значениям
        num_player = rev_dict.get(f) # номера участников в таблице с одинаковым кол-вом очков
        for pl in num_player:
            tr.append(str(pl))  # создает список (встречи игроков)
        m_new = valuesList.count(f)  # подсчитываем сколько раз оно встречается (у скольких крутиловка)

        if m_new == 1:  # если кол-во очков у одного спортсмена
            # записывает место победителю
            td[pl * 2 - 2][max_person + 4] = mesto
            player_rank_tmp.append([pl, mesto])
        elif m_new == 2:  # если кол-во очков у двух спортсмена (определение мест по игре между собой)
            player_rank_tmp = circle_2_player(tr, td, max_person, mesto, num_gr, id_system)
        elif m_new == 3: # если кол-во очков у трех спортсмена
            men_of_circle = m_new
            # получает список 1-й уникальные
            u = summa_points_person(tr, tr_all, num_gr, pg_win, pg_los, pp, stage, id_system)
            # значения очков и список значения очков и у скольких спортсменов они есть
            z = u[1]  # список списков кол-во очков и у сколько игроков они есть
            points_person = z[0] # список [колво очко, у скольки игроков они есть]
            player_rank_tmp = circle_3_player(men_of_circle, points_person, tr, td, max_person, mesto, player_rank_tmp, num_gr, tr_all,
                    pg_win, pg_los, pp, pps, stage)
        elif m_new > 3:  # если кол-во очков у более трех спортсменов (крутиловка)
            player_rank_tmp = circle(tr, num_gr, td, max_person, mesto, stage, id_system)
        tr.clear()

        for i in player_rank_tmp:
            # список участников в группе и его место
            player_rank_group.append(i)

        mesto = mesto + m_new
        player_rank_tmp.clear()
    if kol_tours_played == kol_tours_in_group:  # когда все встречи сыграны
        # функция простановки мест из группы в -Choice-
        result_rank_group_in_choice(num_gr, player_rank_group, stage)


def get_unique_numbers(pp_all):
    """получение списка уникальных значений"""
    unique = []
    for number in pp_all:
        if number not in unique:
            unique.append(number)
    return unique


def circle(tr, num_gr, td, max_person, mesto, stage, id_system):
    """выставляет места в крутиловке -tour- встречи игроков, p1, p2 фамилии, num_gr номер группы
    -tr- список всех туров (номеров) участников в крутиловке men_of_circle кол-во игроков с одинаковым кол-вом очков,
    max_person общее кол-во игроков в группе player_rank - список (номер игроков и их места)"""
    pl_rank_tmp = []  # список списков (игрок и его место)
    player_rank_tmp = []
    tr_all = []
    ps = []
    pps = []
    rev_dict = {}  # словарь, где в качестве ключа очки, а значения - номера групп
    pp = {}  # ключ - игрок, значение его очки
    pg_win = {}
    pg_los = {}

    # получает список 1-й уникальные
    u = summa_points_person(tr, tr_all, num_gr, pg_win, pg_los, pp, stage, id_system)
    # значения очков и список значения очков и у скольких спортсменов они есть
    unique_numbers = u[0]
    tr.clear()
    # ====
    for key, value in pp.items(): # сортируем словарь PP по уменьшению значений (очков)
        rev_dict.setdefault(value, set()).add(key)
    # ====
    for f in unique_numbers:  # проходим циклом по уник. значениям, очки в крутиловке
        m_new = 0
        num_player = rev_dict.get(f)
        count_point = len(num_player)

        if count_point == 1:
            for x in num_player:
                p1 = x
            # записывает место победителю
            td[p1 * 2 - 2][max_person + 4] = mesto
            td[p1 * 2 - 2][max_person + 3] = f  # записывает место победителю
            player_rank_tmp.append([p1, mesto])
            m_new += 1
        elif count_point == 2:
            for x in num_player:
                tr.append(str(x))  # создает список (встречи игроков)
                m_new += 1
            player_rank_tmp = circle_2_player(tr, td, max_person, mesto, num_gr, id_system)
        else:
            point = 0
            for x in num_player:
                tr.append(str(x))  # создает список (встречи игроков)
                m_new += 1
            player_rank_tmp = circle_in_circle(m_new, td, max_person, mesto, tr, num_gr, point,
                                               player_rank_tmp, tr_all, pp, pg_win, pg_los, x, pps, ps, stage)
        mesto = mesto + m_new
        tr.clear()
        # заменяет список (места еще не проставлены) на новый с правильными местами
        for i in player_rank_tmp:
            pl_rank_tmp.append(i)
        player_rank_tmp.clear()
    player_rank_tmp = pl_rank_tmp
    return player_rank_tmp


def circle_in_circle(m_new, td, max_person, mesto, tr, num_gr, point, player_rank_tmp,
                     tr_all, pp, pg_win, pg_los, x, pps, ps, stage):
    """крутиловка в крутиловке
    m_new - число спортсменов в крутиловке
    td - группа
    tr - номера игроков в группе
    tr_all - все варианты встреч
    pp - очки
    pg_win - список выигранных партий каждого игрока
    pg_los - список проигранных партий каждого игрока"""
    num_player = []
    id_system = system_id(stage)
    if m_new == 1:
        p1 = x
        td[p1 * 2 - 2][max_person + 4] = mesto  # записывает место победителю
        td[p1 * 2 - 2][max_person + 3] = point  # очки во встрече победителя
        player_rank_tmp.append([p1, mesto])
    elif m_new == 2:
        player_rank_tmp = circle_2_player(tr, td, max_person, mesto, num_gr, id_system)
    elif m_new == 3:
        # получает список 1-й уникальные
        u = summa_points_person(tr, tr_all, num_gr, pg_win, pg_los, pp, stage, id_system)
        # значения очков и список значения очков и у скольких спортсменов они есть
        z = u[1]
        points_person = z[0]
        men_of_circle = m_new
        player_rank_tmp = circle_3_player(men_of_circle, points_person, tr, td, max_person, mesto, player_rank_tmp, num_gr, tr_all,
                                            pg_win, pg_los, pp, pps, stage)
    elif m_new > 3:
        dict_ratio = {}
        keys_list = list(pg_win.keys())
        # for k in range(1, m_new + 1):
        for k in keys_list:
            pg_win[k] = sum(pg_win[k])  # сумма выигранных партий
            pg_los[k] = sum(pg_los[k])  # сумма проигранных партий
            x = pg_win[k] / pg_los[k]
            x = float('{:.3f}'.format(x)) # соотношение выйгранных партий к проигранным
            dict_ratio[k] = x
        sorted_ratio = {k: dict_ratio[k] for k in
                            sorted(dict_ratio, key=dict_ratio.get, reverse=True)}  # сортирует словарь по убыванию соот 
        k_list = list(sorted_ratio.keys())  # отдельно составляет список ключей (группы)
        v_list = list(sorted_ratio.values())  # отдельно составляет список значений (соотношение)
        ratio_person = get_unique_numbers(v_list)  
        list_uniq = []  # список списков соотношение (выигранный партии к проигравшем) и кол-во игроков их имеющих
        list_tmp = []
        u = []

        for p in ratio_person:
            a = v_list.count(p)
            list_tmp.append(p)
            list_tmp.append(a)
            # список (очки и скольких игроков они встречаются)
            list_uniq.append(list_tmp.copy())
            list_tmp.clear()

        for m in list_uniq:
            point = m[0] # соотношение
            total_uniq = m[1] # сколько раз встречается
            index = v_list.index(m[0])
            p1 = k_list[index] # номер игрока в группе 
            if total_uniq == 1:
                td[p1 * 2 - 2][max_person + 4] = mesto  # записывает место победителю
                td[p1 * 2 - 2][max_person + 3] = point  # очки во встрече победителя
                player_rank_tmp.append([p1, mesto])   
                mesto += 1
            elif total_uniq == 2:
                for i in range(len(v_list)):
                    if v_list[i] == point:
                        num_pl = k_list[i]
                        num_player.append(num_pl)
                tr.clear()
                for x in num_player:
                    tr.append(str(x))  # создает список (встречи игроков)
                    m_new += 1
                player_rank_temp = circle_2_player(tr, td, max_person, mesto, num_gr, id_system)
                player_rank_tmp = player_rank_tmp + player_rank_temp
                mesto = mesto + m_new # новое место
            elif total_uniq == 3:
                for i in range(len(v_list)):
                    if v_list[i] == point:
                        num_pl = k_list[i]
                        num_player.append(num_pl)
                tr.clear()
                m_new = 0
                for x in num_player:
                    tr.append(str(x))  # создает список (встречи игроков)
                    m_new += 1
                # получает список 1-й уникальные
                u = summa_points_person(tr, tr_all, num_gr, pg_win, pg_los, pp, stage, id_system)
                # значения очков и список значения очков и у скольких спортсменов они есть
                z = u[1]
                points_person = z[0]
                men_of_circle = m_new
                player_rank_temp = circle_3_player(men_of_circle, points_person, tr, td, max_person, mesto, player_rank_tmp, num_gr, tr_all,
                    pg_win, pg_los, pp, pps, stage)
                player_rank_tmp = player_rank_tmp + player_rank_temp
                mesto = mesto + m_new # новое место

    tr_all.clear()
    tr.clear()
    return player_rank_tmp


def tour_circle(pp, per_circ, circ):
    tr_new = []
    k_list = list(pp.keys())  # отдельно составляет список ключей (группы)
    v_list = list(pp.values())  # отдельно составляет список значений (очки)
    y = 0
    for s in range(0, circ):
        index = v_list.index(per_circ, y)
        per = str(k_list[index])
        y = index + 1
        tr_new.append(per)
    return tr_new


def summa_points_person(tr, tr_all, num_gr, pg_win, pg_los, pp, stage, id_system):
    """подсчитывает сумму очков у спортсменов в крутиловке 
    -tr- номера игроков в группе, у которых крутиловка
    -tr_all- все варианты встреч в крутиловке
    -pg_los- словарь (номер игрока: список (кол-во проигранных партий)
    -pg_win- словарь (номер игрока: список (кол-во выйгранных партий)"""
    pp_all = []
    u = []
    tr_all.clear()
    pg_win.clear()
    pg_los.clear()
    pp.clear()
    for r in tr:
        r = int(r)
        pp[r] = []  # словарь (игрок - сумма очков)
        pg_win[r] = []
        pg_los[r] = []

    for i in combinations(tr, 2):  # получает список с парами игроков в крутиловке
        i = list(i)
        tr_all.append(i)
    count_game_circle = len(tr_all) # число игр в крутиловки

    for n in range(0, count_game_circle):
        tour = "-".join(tr_all[n])  # получает строку встреча в туре
        ki1 = int(tr_all[n][0])  # 1-й игрок в туре
        ki2 = int(tr_all[n][1])  # 2-й игрок в туре

        sum_points_circle(num_gr, tour, ki1, ki2, pg_win, pg_los, pp, stage, id_system)  # сумма очков игрока

    for i in tr:  # суммирует очки каждого игрока
        i = int(i)
        s = sum(pp[i])
        pp[i] = s  # словарь (участник - его очки)
        pp_all.append(s)

    list_uniq = []  # список списков сумма очков и кол-во игроков их имеющих
    list_tmp = []
    points_person = get_unique_numbers(pp_all)
    points_person.sort(reverse=True)

    for p in points_person:
        a = pp_all.count(p)
        list_tmp.append(p)
        list_tmp.append(a)
        # список (очки и скольких игроков они встречаются)
        list_uniq.append(list_tmp.copy())
        list_tmp.clear()
    u.append(points_person)
    u.append(list_uniq)
    return u # список списков 1-й кол-во игроков 2-й очки выйигранные и проигранные


def circle_2_player(tr, td, max_person, mesto, num_gr, id_system):
    """крутиловка из 2-ух человек"""
    result = Result.select().where((Result.title_id == title_id()) & (Result.system_id == id_system))
    player_rank_tmp = []
    tour = "-".join(tr)  # делает строку встреча в туре
    # =====приводит туры к читаемому виду (1-й игрок меньше 2-ого)
    znak = tour.find("-")
    p1 = int(tour[:znak])  # игрок под номером в группе
    p2 = int(tour[znak + 1:])  # игрок под номером в группе
    if p1 > p2:  # меняет последовательность игроков в туре на обратную, чтоб у 1-ого игрока был номер меньше
        p_tmp = p1
        p1 = p2
        p2 = p_tmp
        tour = f"{p1}-{p2}"
    if num_gr != "Одна таблица":
        c = result.select().where((Result.number_group == num_gr) &
                                  (Result.tours == tour)).get()  # ищет в базе
    # строчку номер группы и тур по двум столбцам
    else:
        c = result.select().where((Result.system_stage == num_gr) &
                                  (Result.tours == tour)).get()  # ищет в базе
        # строчку номер группы и тур по двум столбцам
    if c.winner == c.player1:
        points_p1 = c.points_win  # очки во встрече победителя
        points_p2 = c.points_loser  # очки во встрече проигравшего
        td[p1 * 2 - 2][max_person + 4] = mesto  # записывает место победителю
        td[p2 * 2 - 2][max_person + 4] = mesto + 1  # записывает место проигравшему
        player_rank_tmp.append([p1, mesto])
        player_rank_tmp.append([p2, mesto + 1])
    else:
        points_p1 = c.points_loser
        points_p2 = c.points_win
        td[p1 * 2 - 2][max_person + 4] = mesto + 1  # записывает место победителю
        td[p2 * 2 - 2][max_person + 4] = mesto  # записывает место проигравшему
        player_rank_tmp.append([p1, mesto + 1])
        player_rank_tmp.append([p2, mesto])
    td[p1 * 2 - 2][max_person + 3] = points_p1  # очки во встрече победителя
    td[p2 * 2 - 2][max_person + 3] = points_p2  # очки во встрече проигравшего

    return player_rank_tmp


def circle_3_player(men_of_circle, points_person, tr, td, max_person, mesto, player_rank_tmp, num_gr, tr_all,
                    pg_win, pg_los, pp, pps, stage):
    """в крутиловке 3-и спортсмена
    -pp- словарь (номер игрока, очки)
    -ps- список коэфициентов
    -points_person - список [1-е значение колво очков, 2-е у скольки участников оное есть"""
    id_system = system_id(stage)
    ps = []
    if points_person[1] == 3:  # у всех трех участников равное кол-во очков   
        for k in tr:  # суммирует выигранные и проигранные партии каждого игрока
            k = int(k)
            pg_win[k] = sum(pg_win[k])  # сумма выигранных партий
            pg_los[k] = sum(pg_los[k])  # сумма проигранных партий
            x = pg_win[k] / pg_los[k]
            x = float('{:.3f}'.format(x)) # соотношение выйгранных партий к проигранным
            ps.append(x) # коэфициент
            pps.append(pp[k])
        # получает словарь(ключ, номер участника)
        d = {index: value for index, value in enumerate(tr)}
        # получает словарь(ключ, соотношение)
        ds = {index: value for index, value in enumerate(ps)}
        # сортирует словарь по убываню соот
        sorted_tuple = {k: ds[k] for k in sorted(ds, key=ds.get, reverse=True)}
        key_l = list(sorted_tuple.keys()) # присваивает ключ игрокам
        val_l = list(sorted_tuple.values())
        vls = set(val_l)  # группирует разные значения
        vl = len(vls)  # подсчитывает их количество
        m = 0
        if vl == 1:  # подсчитывает соотношения выигранных и проигранных мячей в партиях
            plr_ratio = score_in_circle(tr_all, men_of_circle, num_gr, tr, stage)
            sorted_ratio = {k: plr_ratio[k] for k in
                            sorted(plr_ratio, key=plr_ratio.get, reverse=True)}  # сортирует словарь по убыванию соот
            # получает список ключей отсортированного словаря
            key_ratio = list(sorted_ratio.keys())
            r = 0
            for i in key_ratio:
                ratio = sorted_ratio[i]  # соотношение в крутиловке
                person = int(i)  # номер игрока
                # записывает соотношение
                td[person * 2 - 2][max_person + 3] = str(ratio)
                td[person * 2 - 2][max_person + 4] = str(mesto + r)  # записывает место
                # добавляет в список группа, место, чтоб занести в таблицу Choice
                player_rank_tmp.append([person, mesto + r])
                r += 1
        else:
            for i in val_l:
                # получает ключ, по которому в списке ищет игрока
                w = key_l[val_l.index(i)]
                # получает номер участника, соответствующий
                wq = int(d.setdefault(w))
                # записывает соотношения игроку
                td[wq * 2 - 2][max_person + 3] = str(i)
                # записывает место
                td[wq * 2 - 2][max_person + 4] = str(m + mesto)
                # добавляет в список группа, место, чтоб занести в таблицу Choice
                player_rank_tmp.append([wq, m + mesto])
                m += 1
    elif points_person[1] == 2 or points_person[1] == 1:
        # получает словарь(ключ, номер участника)
        d = {index: value for index, value in enumerate(tr)}
        # сортирует словарь по убыванию соот
        sorted_tuple = {k: pp[k] for k in sorted(pp, key=pp.get, reverse=True)}
        key_l = list(sorted_tuple.keys()) # номера игроков по убыванию очков
        val_l = list(sorted_tuple.values()) # очки игроков по убыванию
        m = 0
        # вставить если в крутиловке игра по неявке
        if points_person[1] == 2:
            if val_l[0] == val_l[1]:
                tr = [str(key_l[0]), str(key_l[1])] 
                player_rank_tmp = circle_2_player(tr, td, max_person, mesto, num_gr, id_system)
                player_rank_tmp.append([key_l[2], mesto + 2])
            else:
                player_rank_tmp = ([key_l[0], mesto])
                tr = [str(key_l[1]), str(key_l[2])] 
                player_rank_temp = circle_2_player(tr, td, max_person, mesto, num_gr, id_system)
                player_rank_tmp.extend(player_rank_temp)
            for k in player_rank_tmp:
                td[k[0] * 2 - 2][max_person + 4] = str(k[1])  # записывает место
        else:
            for i in val_l:
                q = val_l.index(i) # индекс в списке
                wq = key_l[q] # получает номер участника группы, соответствующий
                # записывает соотношения игроку
                td[wq * 2 - 2][max_person + 3] = str(i)
                td[wq * 2 - 2][max_person + 4] = str(m + mesto)  # записывает место
                # добавляет в список группа, место, чтоб занести в таблицу Choice
                player_rank_tmp.append([wq, m + mesto])
                m += 1
    return player_rank_tmp


def sum_points_circle(num_gr, tour, ki1, ki2, pg_win, pg_los, pp, stage, id_system):
    """сумма очков каждого игрока в крутиловке"""
    
    # # =====приводит туры к читаемому виду (1-й игрок меньше 2-ого)
    znak = tour.find("-")
    p1 = int(tour[:znak])  # игрок под номером в группе
    p2 = int(tour[znak + 1:])  # игрок под номером в группе
    if p1 > p2:  # меняет последовательность игроков в туре на обратную, чтоб у 1-ого игрока был номер меньше
        # уточнить смену тура при p1>p2
        tour = f"{p2}-{p1}"
        ki1 = p2
        ki2 = p1
    result = Result.select().where(Result.title_id == title_id())
    if num_gr == "Одна таблица":
        res = result.select().where(Result.system_id == id_system)
    elif stage == "1-й полуфинал" or stage == "2-й полуфинал":
        res = result.select().where((Result.system_id == id_system) & (Result.number_group == num_gr))
    else:
        res = result.select().where((Result.system_id == id_system) & (Result.number_group == num_gr))
    c = res.select().where(Result.tours == tour).get()  # ищет в базе  данную встречу c - id - встречи в таблице Result
 
    if c.winner == c.player1:  # победил 1-й игрок
        points_p1 = c.points_win  # очки победителя
        points_p2 = c.points_loser  # очки проигравшего
        game_p1 = c.score_in_game  # счет во встречи (выигранные и проигранные партии) победителя
        game_p2 = c.score_loser # счет во встречи (выигранные и проигранные партии) проигравшего
        if game_p1 == "В : П":
            p1_game_win = 0
            p1_game_los = 0
            p2_game_win = 0
            p2_game_los = 0
        else:
            p1_game_win = int(game_p1[0]) # кол-во выигранных партий 1 игрока
            p1_game_los = int(game_p1[4])
            p2_game_win = int(game_p2[0])
            p2_game_los = int(game_p2[4])
    else: # победил 2-й игрок
        points_p1 = c.points_loser # очки 1-ого игрока проигранные
        points_p2 = c.points_win # очки 2-ого игрока выигранные
        game_p1 = c.score_loser # счет во встречи 1-ого игрока
        game_p2 = c.score_in_game # счет во встречи 2-ого игрока
        # ======= если победа по неявке исправить
        if game_p1 == "П : В":
            p1_game_win = 0
            p1_game_los = 0
            p2_game_win = 0
            p2_game_los = 0
        else:
            p1_game_win = int(game_p1[0]) # кол-во выигранных партий 1 игрока
            p1_game_los = int(game_p1[4])
            p2_game_win = int(game_p2[0])
            p2_game_los = int(game_p2[4])
    pp[ki1].append(points_p1)  # добавляет очки 1-ому игроку встречи
    pp[ki2].append(points_p2)  # добавляет очки 2-ому игроку встречи
# записывает в словарь счет во встречи 1-ого игрока
    pg_win[ki1].append(p1_game_win)
    # записывает в словарь счет во встречи 1-ого игрока
    pg_los[ki1].append(p1_game_los)
    # записывает в словарь счет во встречи 2-ого игрока
    pg_win[ki2].append(p2_game_win)
    # записывает в словарь счет во встречи 2-ого игрока
    pg_los[ki2].append(p2_game_los)


def _score_in_circle(tr_all, men_of_circle, num_gr, tr, stage):
    """подсчитывает счет по партиям в крутиловке"""
    id_system = system_id(stage)
    result = Result.select().where(Result.title_id == title_id())
    plr_win = {0: [], 1: [], 2: []}
    plr_los = {0: [], 1: [], 2: []}
    plr_ratio = {0: [], 1: [], 2: []}
    for n in range(0, men_of_circle):
        tour = "-".join(tr_all[n])  # получает строку встреча в туре
        znak = tour.find("-")
        p1 = int(tour[:znak])  # игрок под номером в группе
        p2 = int(tour[znak + 1:])  # игрок под номером в группе
        if p1 > p2:  # меняет последовательность игроков в туре на обратную, чтоб у 1-ого игрока был номер меньше
            tour = f"{p2}-{p1}"
            k1 = str(p2)  # 1-й игрок в туре
            k2 = str(p1) # 2-й игрок в туре
        else:
            k1 = str(p1)  # 1-й игрок в туре
            k2 = str(p2) # 2-й игрок в туре
        c_res = result.select().where((Result.system_id == id_system) & (Result.number_group == num_gr))
        c = c_res.select().where(Result.tours == tour).get()
        
        # k1 = tr_all[n][0]  # 1-й игрок в туре
        # k2 = tr_all[n][1]  # 2-й игрок в туре
        ki1 = tr.index(k1)  # получение индекса 1-й игрока
        ki2 = tr.index(k2)
        g = c.score_win # счет в партии на сколько он выиграл
        g_len = len(g)
        g = g[1:g_len - 1]
        sc_game = g.split(",")
 
        if c.winner == c.player1:  # победил 1-й игрок
            for i in sc_game:
                i = int(i)
                if i < 0:
                    plr_win[ki1].append(abs(i))
                    plr_los[ki2].append(abs(i))
                    if abs(i) < 10:
                        plr_los[ki1].append(11)
                        plr_win[ki2].append(11)
                    else:
                        plr_los[ki1].append(abs(i) + 2)
                        plr_win[ki2].append(abs(i) + 2)
                elif 0 <= i < 10:
                    plr_win[ki1].append(11)
                    plr_los[ki1].append(i)
                    plr_win[ki2].append(i)
                    plr_los[ki2].append(11)
                elif i >= 10:
                    plr_win[ki1].append(i + 2)
                    plr_los[ki1].append(i)
                    plr_win[ki2].append(i)
                    plr_los[ki2].append(i + 2)
        else:  # если победил 2-й игрок
            for i in sc_game:
                i = int(i)
                if i < 0:  # партию проиграл
                    plr_win[ki2].append(abs(i))
                    plr_los[ki1].append(abs(i))
                    if abs(i) < 10:
                        plr_los[ki2].append(11)
                        plr_win[ki1].append(11)
                    else:
                        plr_los[ki2].append(abs(i) + 2)
                        plr_win[ki1].append(abs(i) + 2)
                elif 0 <= i < 10:  # выиграл партию
                    plr_win[ki2].append(11)
                    plr_los[ki2].append(i)
                    plr_win[ki1].append(i)
                    plr_los[ki1].append(11)
                elif i >= 10:  # выиграл партию на больше меньше
                    plr_win[ki2].append(i + 2)
                    plr_los[ki2].append(i)
                    plr_win[ki1].append(i)
                    plr_los[ki1].append(i + 2)
    ratio_list = []
    new_ratio_list = []
    for n in range(0, men_of_circle): # ставит колво знаков после запятой при подсчете соотнешения мячей
        plr_win[n] = sum(plr_win[n])
        plr_los[n] = sum(plr_los[n])
        x = plr_win[n] / plr_los[n]
        ratio_list.append(x)
    for m in range(0, 6):
        flag_ratio = ratio(ratio_list, m)
        if flag_ratio is True:
            for x in ratio_list:
                form = '{:.'f"{m + 1}"'f}'
                x = float(form.format(x))
                new_ratio_list.append(x)
            n = 0
            for l in new_ratio_list:
                plr_ratio[n] = l
                n += 1
            return plr_ratio
        else:
            continue   


def score_in_circle(tr_all, men_of_circle, num_gr, tr, stage):
    """подсчитывает счет по партиям в крутиловке"""
    id_system = system_id(stage)
    result = Result.select().where(Result.title_id == title_id())
    plr_win = {}
    plr_los = {}
    plr_ratio = {}
    for r in tr:
        plr_win[r] = []
        plr_los[r] = []
        plr_ratio[r] = []
    for n in range(0, men_of_circle):
        tour = "-".join(tr_all[n])  # получает строку встреча в туре
        znak = tour.find("-")
        p1 = int(tour[:znak])  # игрок под номером в группе
        p2 = int(tour[znak + 1:])  # игрок под номером в группе
        if p1 > p2:  # меняет последовательность игроков в туре на обратную, чтоб у 1-ого игрока был номер меньше
            tour = f"{p2}-{p1}"
            k1 = str(p2)  # 1-й игрок в туре
            k2 = str(p1) # 2-й игрок в туре
        else:
            k1 = str(p1)  # 1-й игрок в туре
            k2 = str(p2) # 2-й игрок в туре
        c_res = result.select().where((Result.system_id == id_system) & (Result.number_group == num_gr))
        c = c_res.select().where(Result.tours == tour).get()

        g = c.score_win # счет в партии на сколько он выиграл
        g_len = len(g)
        g = g[1:g_len - 1]
        sc_game = g.split(",") # создает список из счета по партиям
 
        if c.winner == c.player1:  # победил 1-й игрок
            for i in sc_game:
                i = int(i)
                if i < 0:
                    plr_win[k1].append(abs(i))
                    plr_los[k2].append(abs(i))
                    if abs(i) < 10:
                        plr_los[k1].append(11)
                        plr_win[k2].append(11)
                    else:
                        plr_los[k1].append(abs(i) + 2)
                        plr_win[k2].append(abs(i) + 2)
                elif 0 <= i < 10:
                    plr_win[k1].append(11)
                    plr_los[k1].append(i)
                    plr_win[k2].append(i)
                    plr_los[k2].append(11)
                elif i >= 10:
                    plr_win[k1].append(i + 2)
                    plr_los[k1].append(i)
                    plr_win[k2].append(i)
                    plr_los[k2].append(i + 2)
        else:  # если победил 2-й игрок
            for i in sc_game:
                i = int(i)
                if i < 0:  # партию проиграл
                    plr_win[k2].append(abs(i))
                    plr_los[k1].append(abs(i))
                    if abs(i) < 10:
                        plr_los[k2].append(11)
                        plr_win[k1].append(11)
                    else:
                        plr_los[k2].append(abs(i) + 2)
                        plr_win[k1].append(abs(i) + 2)
                elif 0 <= i < 10:  # выиграл партию
                    plr_win[k2].append(11)
                    plr_los[k2].append(i)
                    plr_win[k1].append(i)
                    plr_los[k1].append(11)
                elif i >= 10:  # выиграл партию на больше меньше
                    plr_win[k2].append(i + 2)
                    plr_los[k2].append(i)
                    plr_win[k1].append(i)
                    plr_los[k1].append(i + 2)
    ratio_list = []
    new_ratio_list = []
    # for n in range(0, men_of_circle): # ставит колво знаков после запятой при подсчете соотнешения мячей
    for n in tr: # ставит колво знаков после запятой при подсчете соотнешения мячей
        plr_win[n] = sum(plr_win[n])
        plr_los[n] = sum(plr_los[n])
        x = plr_win[n] / plr_los[n]
        ratio_list.append(x)
    for m in range(0, 6):
        flag_ratio = ratio(ratio_list, m)
        if flag_ratio is True:
            for x in ratio_list:
                form = '{:.'f"{m + 1}"'f}'
                x = float(form.format(x))
                new_ratio_list.append(x)
            n = 0
            for l in new_ratio_list:
                ind = tr[n]
                plr_ratio[ind] = l
                n += 1
            return plr_ratio
        else:
            continue   



def ratio(ratio_list, m):

    """проверяет на одинаковые коэфицинты при подсчете соотношения счета мячей"""
    if m >= 1:
        m += 1 
    new_ratio_list = []
    for k in ratio_list:
        mark1 = str(k)[m]
        new_ratio_list.append(int(mark1))

    count_frequency = filter(lambda x: new_ratio_list.count(x) > 1, new_ratio_list)
    count_frequency = list(set(count_frequency))
    flag_ratio = True if len(count_frequency) == 0 else False
    return flag_ratio


def player_choice_in_group(num_gr):
    """распределяет спортсменов по группам согласно жеребьевке"""
    posev_data = []
    choice_group = Choice.select().where((Choice.title_id == title_id()) & (Choice.group == num_gr))
    players = Player.select().where((Player.title_id == title_id()) & (Player.bday != '0000-00-00'))
    for posev in choice_group:
        pl = players.select().where(Player.id == posev.player_choice_id).get()
        city = pl.city
        id_pl = posev.player_choice_id
        posev_data.append({
            'фамилия': f"{posev.family}/{id_pl}",
            'регион': city,
        })
    return posev_data


def player_choice_one_table(stage):
    """список спортсменов одной таблицы"""
    posev_data = []
    choices = Choice.select().where(Choice.title_id == title_id())
    players = Player.select().where(Player.title_id == title_id())
    if stage == "Одна таблица":
        choice = choices.select().where(Choice.basic == "Одна таблица")
    else:
        choice = choices.select().where(Choice.final == stage).order_by(Choice.posev_final)
        
    for posev in choice:
        pl = players.select().where(Player.id == posev.player_choice_id).get()
        city = pl.city
        posev_data.append({
            'фамилия': posev.family,
            'регион': city,
        })
    return posev_data


def player_choice_semifinal(stage, num_gr):
    """список спортсменов полуфиналов"""
    posev_data = []
    choices = Choice.select().where((Choice.title_id == title_id()) & (Choice.semi_final == stage))
    choice_group_pf = choices.select().where(Choice.sf_group == num_gr).order_by(Choice.posev_sf)
    players = Player.select().where(Player.title_id == title_id())
    for posev in choice_group_pf:
        pl = players.select().where(Player.id == posev.player_choice_id).get()
        city = pl.city
        id_pl = posev.player_choice_id
        posev_data.append({
            'фамилия': f"{posev.family}/{id_pl}",
            'регион': city,
        })
    return posev_data


def player_choice_in_setka(fin):
    """распределяет спортсменов в сетке согласно жеребьевке"""
    stage = fin
    id_system = system_id(stage)
    systems = System.select().where((System.title_id == title_id()) & (System.id == id_system)).get()
    count_exit = systems.mesta_exit # сколько игроков выходят в финал
 
    flag = selection_of_the_draw_mode() # выбор ручная или автоматическая жеребьевка
    posev = choice_setka_automat(fin, flag, count_exit)
   
    posev_data = []
    for key in posev.keys():
        posev_data.append({'посев': key, 'фамилия': posev[key]})  
    # сортировка (списка словарей) по ключу словаря -посев-
    posev_data = sorted(posev_data, key=lambda i: i['посев'])
    with db:  # записывает в db, что жеребьевка произведена
        systems.choice_flag = True
        systems.save()
    if flag == 3: # если ручная жеребьевка, то закрывает сетку 
        my_win.tableView_net.hide()
        my_win.resize(1270, 825)
        my_win.tableView.setGeometry(QtCore.QRect(260, 150, 1000, 626))
        my_win.tabWidget.setGeometry(QtCore.QRect(260, 0, 1000, 147))
    return posev_data


def change_choice_group():
    """Смена жеребьевки групп если в группе 2 и более одинаковых регион чтоб развести тренеров"""
    msg = QMessageBox
    sender = my_win.sender()
    if my_win.checkBox_repeat_regions.isChecked():
        reg = []
        reg_d = []
        gr_key = []
        reg_tmp = []
        double_reg = {}
        fg = my_win.comboBox_filter_choice_stage.currentText()
        choice = Choice.select().where(Choice.title_id == title_id())
        system = System.select().where(System.title_id == title_id())
        sys = system.select().where(System.stage == "Предварительный").get()
        total_gr = sys.total_group
        for i in range(1, total_gr + 1):
            m = 0
            group = choice.select().where(Choice.group == f"{i} группа")
            for k in group:
                m += 1
                reg_n = k.region
                if reg_n not in reg:
                    reg.append(reg_n)
                else:
                    reg_tmp.append(reg_n)
            reg_d = reg_tmp.copy()
            count = len(reg_d)
            if count > 0:
                double_reg[f"{i} группа"] = reg_d
            reg_tmp.clear()
            reg.clear()
        dr_count = len(double_reg)
        if dr_count != 0:
            for key in double_reg.keys():
                gr_key.append(key)   
            gr_key.insert(0, "-выберите группу-")  
            my_win.comboBox_filter_choice_stage.clear()
            my_win.comboBox_filter_number_group_final.clear()
            stage_list = ["-выбор этапа-", "Предварительный"]
            my_win.comboBox_filter_choice_stage.addItems(stage_list)
            my_win.comboBox_filter_number_group_final.addItems(gr_key)
            return double_reg
        else:
            msg.information(my_win, "Уведомление", "Нет групп с повторяющимися регионами.")
    else:
        my_win.comboBox_filter_choice_stage.setCurrentIndex(0)
        return


def change_page_vid():
    """Смена вида страницы с таблицами""" 
    msgBox = QMessageBox
    sys = []
    sys.append("")
    system = System.select().where(System.title_id == title_id()) 
    for i in system:
        stage = i.stage
        sys.append(stage)
    stage, ok = QInputDialog.getItem(my_win, "Таблицы", "Выберите таблицы из списка для\n"
                                        "смены ориентации страницы", sys)  
    if ok: 
        id_system = system_id(stage)                                  
        sys = system.select().where(System.id == id_system).get()
        vid = sys.page_vid
        vid_ed = "книжная" if vid == "альбомная" else "альбомная"
        ok = msgBox.question(my_win, "Таблицы", "Текущая ориентация страницы\n"
                                            f"-{stage}-: {vid},\n"
                                            "Хотите ее изменить на:" f"{vid_ed}?", msgBox.Ok, msgBox.Cancel)
        if ok:
            sys.page_vid = vid_ed
            sys.save()
        else:
            return
    else:
        return


def change_indent_page():
    """меняет отступ в PDF странице"""
    msgBox = QMessageBox
    indent, ok = QInputDialog.getInt(my_win, "Таблицы", "Выберите отступ от края таблицы,\n"
                                        "если она выходит за рамки страницы\n"
                                        "0 - без отступа") 
    return indent 


def change_dir(catalog):
    """смена директории, чтоб все pdf фалы сохранялися в папке table_pdf"""

    dir_path = pathlib.Path.cwd()
    parent_dir = str(dir_path)
    f1 = parent_dir.rfind("table_pdf")
    f2 = parent_dir.rfind("competition_pdf")
    if catalog == 1:
        if f1 == -1 :
            os.chdir("table_pdf") # переходит в каталог pdf       
    else:
        if f2 == -1 :
            os.chdir("competition_pdf") # переходит в каталог pdf       
 

def draw_setka_made(col, row, num, step, tur, style):
    """рисование сетки встреч игроков
    col - начальный столбец, row - начальный ряд, num - кол-во игроков"""
    style_set = []  
   
    col_fin = (col + 1) + (2 * (tur - 1)) # последний столбец
    row_fin = row + (num - 1) * step # последняя строка 
    for i in range (col, col_fin + 1, 2): # номер столбца 
        for k in range(row, row_fin + 1, step): # номер строки
            fn = ('LINEABOVE', (i, k), (i + 1, k), 1, colors.darkblue)  # рисует линии встреч
            style_set.append(fn)  
    for m in range(col + 1, col_fin + 1, 2):
        for q in range(row, row_fin, step):  # встречи 33-34
            fn = ('SPAN', (m, q), (m, q + step - 1 ))             
            style_set.append(fn)
            fn = ('BACKGROUND', (m, q), (m, q + step - 1 ), colors.lightyellow)  
            style_set.append(fn) 
            fn = ('BOX', (m, q), (m, q + step - 1), 1, colors.darkblue)
            style_set.append(fn) 
    for fn in style_set:
        style.append(fn)
    return style


def draw_setka(col, row, num, style):
    """рисование сетки встреч игроков
    col - начальный столбец, row - начальный ряд, num - кол-во игроков"""
    style_set = []  
    s = 1
    cf = 0  # кол-во туров
    if num == 2:  # кол-во игроков
        cf = 1
    elif num == 4:
        cf = 2
    elif num == 8:
        cf = 3
    elif num == 16:
        cf = 4
    elif num == 32:
        cf = 5
    row_b = row
    col_fin = col + cf * 2 # последний столбец
    row_fin = row + num * 2 - 1 # последняя строка 
    for i in range (col, col_fin, 2): # номер столбца 
        s *= 2
        for k in range(row, row_fin, s): # номер строки
            fn = ('LINEABOVE', (i, k), (i + 1, k), 1, colors.darkblue)  # рисует линии встреч
            style_set.append(fn)  
        row = row + s // 2
    s = 1
    for m in range(col + 1, col_fin + 1, 2):
        s *= 2
        for q in range(row_b, row_fin, s * 2):  # встречи 33-34
            fn = ('SPAN', (m, q), (m, q + s - 1 ))             
            style_set.append(fn)
            fn = ('BACKGROUND', (m, q), (m, q + s - 1 ), colors.lightyellow)  
            style_set.append(fn) 
            fn = ('BOX', (m, q), (m, q + s - 1), 1, colors.darkblue)
            style_set.append(fn)
        row_b = row_b + s // 2   
    for fn in style_set:
        style.append(fn)
    return style


def draw_setka_2(col, row, num, style):
    """рисование сетки встреч игроков
    col - начальный столбец, row - начальный ряд, num - кол-во игроков"""
    style_set = []  
    s = 1
    cf = 0  # кол-во туров
    if num == 2:  # кол-во игроков
        cf = 1
    elif num == 4:
        cf = 3
    elif num == 8:
        cf = 4
    elif num == 16:
        cf = 5
    
    row_b = row
    col_fin = col + cf * 2 # последний столбец
    row_fin = row + num * 2 - 1 # последняя строка 
    for i in range (col, col_fin, 2): # номер столбца 
        s *= 2
        for k in range(row, row_fin, s): # номер строки
            fn = ('LINEABOVE', (i, k), (i + 1, k), 1, colors.darkblue)  # рисует линии встреч
            style_set.append(fn)  
        if i == 1:
            row -= 1
            s = 1
        elif i == 3:
            row += 1
            s = 2
        elif i == 5:
            row -= 2
            row_fin -= 1
            s = 2
        elif i == 7:
            row += 2
            s = 4
 
    s = 1
    row_fin = row_b + num * 2 - 2 # последняя строка 
    for m in range(col + 1, col_fin + 1, 2):
        s *= 2
        for q in range(row_b, row_fin, s * 2):  # встречи 33-34
            fn = ('SPAN', (m, q), (m, q + s - 1 ))             
            style_set.append(fn)
            fn = ('BACKGROUND', (m, q), (m, q + s - 1 ), colors.lightyellow)  
            style_set.append(fn) 
            fn = ('BOX', (m, q), (m, q + s - 1), 1, colors.darkblue)
            style_set.append(fn)
        if m == 2:
            row_b -= 1
            s = 1
        elif m == 4:
            row_b += 1
            s = 2
        elif m == 6:
            row_b -= 2
            s = 2
        elif m == 8:
            row_b += 2
            s = 4

    for fn in style_set:
        style.append(fn)
    return style


def draw_mesta(row, col, player, style):
    """рисует линии встреч за место"""
    p = 0
    if player == 2:
        p = 4
    elif player == 4:
        p = 4
    elif player == 8:
        p = 6
    elif player == 16:
        p = 10

    col_f = 11

    if col == 9:
        col_f = col + 2
    else:
        col_f = col + 1

    for l in range(row, row + p + 1, p):
        fn = ('LINEABOVE', (col, l), (col_f, l), 1, colors.darkblue)  # рисует линии мест 5-6 места (4 чел)
        style.append(fn)
    return style


def draw_num(row_n, row_step, col_n, number_of_columns, number_of_game, player, data):
    """рисует номера встреч, row_n - начальный ряд, col_n - начальный столбец, 
    number_of_game - начальный номер встречи, player - кол-во участников, number_of_columns - кол-во столбцов """
    s = 1
    col_f = col_n + number_of_columns * 2 - 1
    row_f = row_n + (player - 2) * row_step 
    for k in range(col_n, col_f, 2):
        step = row_step * 2
        for i in range (row_n, row_f + 1, step):
            data[i][k] = str(number_of_game)
            number_of_game += 1
        row_step *= 2
        s *= 2
        row_n = row_n + s // 2
    return number_of_game


def draw_num_2(row_n, row_step, col_n, number_of_columns, number_of_game, data, player):
    """рисует номера встреч, row_n - начальный ряд, col_n - начальный столбец, 
    number_of_game - начальный номер встречи, player - кол-во участников, 
    number_of_columns - число столбцов"""
  
    col_f = col_n + number_of_columns * 2 - 1
    row_f = row_n + (player * 2 - 1)
    for k in range(col_n, col_f, 2): 
        for i in range (row_n, row_f, row_step * 2):
            data[i][k] = str(number_of_game)
            number_of_game += 1
        
        row_n -= int(row_step / 2)
        row_f -= int(row_step / 2)

    return number_of_game


def draw_num_lost(row_n, row_step, col_n, number_of_game, player, data):
    """нумерация встреч проигранных"""
    row_f = row_n + (player - 1) * 2 + 1
    for d in range(row_n, row_f, row_step):
            data[d - 1][col_n] = str(number_of_game * -1)
            number_of_game += 1


def draw_num_lost_2(row_n, row_step, col_n, revers_number, number_of_game, player, data):
    """нумерация встреч проигранных"""
    if revers_number == 0:
        row_n = row_n
        row_f = row_n + (player - 1) * row_step * 2 + 1
        step = row_step * 2
    else:
        row_n = row_n + (player - 1) * row_step * 2
        row_f = row_n - (player - 1) * row_step * 2 - 1
        step = row_step * 2 * -1

    for d in range(row_n, row_f, step):
            data[d][col_n] = str(number_of_game * -1)
            number_of_game += 1


def color_mesta(data, first_mesto, table, fin):
    """окрашивает места в красный цвет"""
    b = 0
    style_color = []
    id_system = system_id(stage=fin)
    system = System.select().where((System.title_id == title_id()) & (System.id == id_system)).get()
    flag = system.no_game

    ml = [] # столбец, ряд -1 ого места, ряд 2-ого места + 1, шаг между местами
    f = 0 # количество столбцов
    if table == "setka_32":
        f = 2
    elif table == "setka_32_full":
        f = 13
    elif table == "setka_32_2":
        f = 16
    elif table == "setka_16_full":
        f = 8
    elif table == "setka_16_2":
        f= 8
    elif table == "setka_8_full":
        f= 4
    elif table == "setka_8_2":
        f= 4
    k = 0
    for c in range(0, f):
        if c == 0: # 1-2 место
            if table == "setka_32_2":
                ml = [13, 31, 54, 22] 
            elif table == "setka_16_full":
                # ml = [10, 15, 26, 10] 
                ml = [10, 14, 25, 10]
            elif table == "setka_16_2":
                ml = [9, 15, 33, 17]
            elif table == "setka_8_full":
                ml = [8, 6, 13, 6]
            elif table == "setka_8_2":
                ml = [8, 6, 13, 6]
            elif table == "setka_32_full":
                ml = [11, 31, 54, 22] 
            elif table == "setka_32":
                ml = [11, 31, 54, 22]
        elif c == 1: # 3-4 место
            if table == "setka_32_2":
                ml = [13, 80, 97, 16]  
            elif table == "setka_16_full":
                # ml = [10, 29, 32, 2] 
                ml = [10, 28, 31, 2] 
            elif table == "setka_16_2":
                ml = [9, 48, 56, 7] 
            elif table == "setka_8_full":
                ml = [8, 15, 18, 2]
            elif table == "setka_8_2": 
                ml = [8, 17, 22, 4]
            elif table == "setka_32_full":               
                ml = [11, 59, 65, 5] 
            elif table == "setka_32":               
                ml = [11, 59, 65, 5]    
        elif c == 2: # 5-6 место
            if table == "setka_32_2":
                ml = [13, 101, 106, 4]  
            elif table == "setka_16_full":
                # ml = [10, 34, 38, 3] 
                ml = [10, 33, 37, 3] 
            elif table == "setka_16_2":
                ml = [9, 60, 64, 3] 
            elif table == "setka_8_full":
                ml = [8, 21, 25, 3]
            elif table == "setka_8_2": 
                ml = [8, 24, 28, 3] 
            else:
                ml = [11, 72, 92, 5]
        elif c == 3: # 7-8 место
            if table == "setka_32_2":
                ml = [13, 109, 114, 4]  
            elif table == "setka_16_full":
                # ml = [10, 39, 42, 2] 
                ml = [10, 38, 41, 2]
            elif table == "setka_16_2":
                ml = [9, 66, 70, 3]
            elif table == "setka_8_full":
                ml = [8, 27, 30, 2] 
            elif table == "setka_8_2": 
                ml = [8, 30, 34, 3]
            else:
                ml = [11, 94, 95, 1]
        elif c == 4: # 9-10 место
            if table == "setka_32_2":
                ml = [5, 113, 118, 4]  
            elif table == "setka_16_full":
                # ml = [10, 47, 53, 5] 
                ml = [10, 46, 52, 5] 
            elif table == "setka_16_2":
                ml = [5, 63, 70, 6] 
            else:
                ml = [11, 99, 133, 5]
        elif c == 5: # 11-12 место
            if table == "setka_32_2":
                ml = [11, 119, 124, 4]  
            elif table == "setka_16_full":
                # ml = [10, 55, 58, 2] 
                ml = [10, 54, 57, 2]
            elif table == "setka_16_2":
                ml = [9, 72, 76, 3] 
            else:
                ml = [11, 152, 163, 10]
        elif c == 6: # 13-14 место
            if table == "setka_32_2":
                ml = [5, 125, 130, 4]  
            elif table == "setka_16_full":
                # ml = [10, 60, 64, 3]
                ml = [10, 59, 63, 3]
            elif table == "setka_16_2":
                ml = [5, 75, 82, 6]  
            else:
                ml = [11, 167, 172, 4]
        elif c == 7: # 15-16 место
            if table == "setka_32_2":
                ml = [11, 127, 132, 4] 
            elif table == "setka_16_full":
                # ml = [10, 65, 68, 2]
                ml = [10, 64, 67, 2]
            elif table == "setka_16_2":
                ml = [9, 78, 82, 3]  
            else:
                ml = [9, 173, 178, 4]
        elif c == 8: # 17-18 место
            if table == "setka_32_2":
                ml = [7, 145, 153, 7]  
            else:
                ml = [11, 180, 186, 5]
        elif c == 9: # 19-20 место
            if table == "setka_32_2":
                ml = [11, 154, 160, 5]  
            else:
                ml = [7, 184, 191, 6]
        elif c == 10: # 21-22 место
            if table == "setka_32_2":
                ml = [7, 161, 166, 4]  
            else:
                ml = [11, 192, 198, 5]
        elif c == 11: # 23-24 место
            if table == "setka_32_2":
                ml = [11, 169, 175, 5]  
            else:
                ml = [5, 198, 203, 4]
        elif c == 12: # 25-26 место
            if table == "setka_32_2":
                ml = [7, 175, 183, 7] 
            else:
                ml = [11, 200, 205, 4]
        elif c == 13: # 27-28 место
            ml = [11, 184, 190, 5]  
        elif c == 14: # 29-30 место 
            ml = [7, 193, 198, 4] 
        elif c == 15: # 31-32 место
            ml = [11, 199, 205, 5]

        for i in range(ml[1], ml[2], ml[3]):
            data[i][ml[0]] = str(first_mesto + b) + " Место"
            fn = (('TEXTCOLOR', (ml[0], i), (ml[0], i), colors.red))
            style_color.append(fn)
            fn =  ('ALIGN', (ml[0], i), (ml[0], i), 'CENTER')
            style_color.append(fn)
            if b == 2 and flag == "3": # значит два 3-х места
                k += 1
                b = 4 if k == 2 else 2
            else:
                b += 1    
    return style_color   


def tours_list(cp):
    """туры таблиц по кругу в зависимости от кол-во участников (-cp- + 3) кол-во участников"""
    tour_list = []
    tr = [[['1-3'], ['1-2'], ['2-3']],
          [['1-3', '2-4'], ['1-2', '3-4'], ['2-3', '1-4']],
          [['2-4', '1-5'], ['1-4', '3-5'], ['1-3', '2-5'], ['2-3', '4-5'], ['1-2', '3-4']],
          [['2-4', '1-5', '3-6'], ['1-4', '2-6', '3-5'], ['1-3', '2-5', '4-6'], ['2-3', '1-6', '4-5'],
            ['1-2', '3-4', '5-6']],
          [['2-6', '3-5', '1-7'], ['2-5', '1-6', '4-7'], ['1-5', '4-6', '3-7'], ['4-5', '2-7', '3-6'],
            ['1-3', '2-4', '5-7'], ['1-4', '2-3', '6-7'], ['1-2', '3-4', '5-6']],
          [['2-6', '3-5', '1-7', '4-8'], ['2-5', '1-6', '3-8', '4-7'], ['1-5', '2-8', '4-6', '3-7'],
            ['1-8', '4-5', '2-7', '3-6'], ['1-3', '2-4', '5-7', '6-8'], ['1-4', '2-3', '6-7', '5-8'],
            ['1-2', '3-4', '5-6', '7-8']],
          [['1-9', '2-8', '3-7', '4-6'], ['5-9', '1-8', '2-7', '3-6'], ['4-9', '5-8', '1-7', '2-6'],
            ['3-9', '4-8', '5-7', '1-6'], ['2-4', '1-5', '3-8', '7-9'], ['4-1', '5-3', '9-2', '8-6'],
            ['1-3', '2-5', '4-7', '6-9'], ['3-2', '5-4', '8-9', '7-6'], ['1-2', '3-4', '5-6', '7-8']],
          [['1-9', '2-8', '3-7', '4-6', '5-10'], ['5-9', '1-8', '2-7', '3-6', '4-10'], ['4-9', '5-8', '1-7', '2-6', '3-10'],
            ['3-9', '4-8', '5-7', '1-6', '2-10'], ['2-4', '1-5', '3-8', '7-9', '6-10'], ['4-1', '5-3', '9-2', '8-6', '7-10'],
            ['1-3', '2-5', '4-7', '6-9', '8-10'], ['3-2', '5-4', '8-9', '7-6', '1-10'], ['1-2', '3-4', '5-6', '7-8', '9-10']],
          [['1-11', '2-10', '3-9', '4-8', '5-7'], ['6-11', '1-10', '2-9', '3-8', '4-7'], ['5-11', '6-10', '1-9', '2-8', '3-7'],
            ['4-11', '5-10', '6-9', '1-8', '2-7'], ['3-11', '4-10', '5-9', '6-8', '1-7'], ['2-11', '3-10', '4-9', '5-8', '6-7'],
            ['2-4', '1-5', '3-6', '7-10', '9-11'], ['1-4', '2-6', '3-5', '8-10', '7-11'], ['1-3', '2-5', '4-6', '7-9', '8-11'],
            ['2-3', '1-6', '4-5', '8-9', '10-11'], ['1-2', '3-4', '5-6', '7-8', '9-10']],
          [['1-11', '2-10', '3-9', '4-8', '5-7', '6-12'], ['6-11', '1-10', '2-9', '3-8', '4-7', '5-12'],
            ['5-11', '6-10', '1-9', '2-8', '3-7', '4-12'], ['4-11', '5-10', '6-9', '1-8', '2-7', '3-12'],
            ['3-11', '4-10', '5-9', '6-8', '1-7', '2-12'], ['2-11', '3-10', '4-9', '5-8', '6-7', '1-12'],
            ['2-4', '1-5', '3-6', '7-10', '9-11', '8-12'], ['1-4', '2-6', '3-5', '8-10', '7-11', '9-12'],
            ['1-3', '2-5', '4-6', '7-9', '8-11', '10-12'], ['2-3', '1-6', '4-5', '8-9', '10-11', '7-12'],
            ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12']],
          [['1-13', '2-12', '3-11', '4-10', '5-9', '6-8'], ['7-13', '1-12', '2-11', '3-10', '4-9', '5-8'],
            ['6-13', '7-12', '1-11', '2-10', '3-9', '4-8'], ['5-13', '6-12', '7-11', '1-10', '2-9', '3-8'],
            ['4-13', '5-12', '6-11', '7-10', '1-9', '2-8'], ['3-13', '4-12', '5-11', '6-10', '7-9', '1-8'],
            ['1-7', '2-6', '3-5', '4-11', '9-13', '10-12'], ['1-6', '2-5', '4-7', '3-12', '8-11', '10-13'],
            ['1-4', '2-7', '3-6', '5-10', '8-13', '9-12'], ['1-5', '3-7', '4-6', '2-13', '8-12', '9-11'],
            ['1-3', '2-4', '5-7', '6-9', '8-10', '11-13'], ['2-3', '4-5', '6-7', '8-9', '10-11', '12-13'],
            ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12']],
          [['1-13', '2-12', '3-11', '4-10', '5-9', '6-8', '7-14'], ['7-13', '1-12', '2-11', '3-10', '4-9', '5-8', '6-14'],
          ['6-13', '7-12', '1-11', '2-10', '3-9', '4-8', '5-14'], ['5-13',
                                                                   '6-12', '7-11', '1-10', '2-9', '3-8', '4-14'],
          ['4-13', '5-12', '6-11', '7-10', '1-9', '2-8', '3-14'], ['3-13',
                                                                   '4-12', '5-11', '6-10', '7-9', '1-8', '2-14'],
          ['1-7', '2-6', '3-5', '4-11', '9-13', '10-12', '8-14'], ['1-6',
                                                                   '2-5', '4-7', '3-12', '8-11', '10-13', '9-14'],
          ['1-4', '2-7', '3-6', '5-10', '8-13', '9-12', '11-14'], ['1-5',
                                                                   '3-7', '4-6', '2-13', '8-12', '9-11', '10-14'],
          ['1-3', '2-4', '5-7', '6-9', '8-10', '11-13', '12-14'], ['2-3',
                                                                   '4-5', '6-7', '8-9', '10-11', '12-13', '1-14'],
          ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12', '13-14']],
          [['1-15', '2-14', '3-13', '4-12', '5-11', '6-10', '7-9'], ['8-15', '1-14', '2-13', '3-12', '4-11', '5-10', '6-9'],
          ['8-15', '1-14', '2-13', '3-12', '4-11', '5-10', '6-9'], ['7-15',
                                                                    '8-14', '1-13', '2-12', '3-11', '4-10', '5-9'],
          ['6-15', '7-14', '8-13', '1-12', '2-11', '3-10', '4-9'], ['5-15',
                                                                    '6-14', '7-13', '8-12', '1-11', '2-10', '3-9'],
          ['4-15', '5-14', '6-13', '7-12', '8-11', '1-10', '2-9'], ['3-15',
                                                                    '4-14', '5-13', '6-12', '7-11', '8-10', '1-9'],
          ['2-15', '3-14', '4-13', '5-12', '6-11', '7-10', '8-9'], ['1-7',
                                                                    '2-6', '3-5', '4-8', '9-13', '12-14', '11-15'],
          ['1-6', '2-5', '3-8', '4-7', '9-14', '10-13', '12-15'], ['1-5',
                                                                   '2-8', '3-7', '4-6', '9-15', '10-14', '11-13'],
          ['1-4', '2-7', '3-6', '5-8', '9-12', '10-15', '11-14'], ['1-3',
                                                                   '2-4', '5-7', '6-8', '9-11', '10-12', '13-15'],
          ['1-8', '2-3', '4-5', '6-7', '10-11', '12-13', '14-15'], ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12', '13-14']],
          [['1-15', '2-14', '3-13', '4-12', '5-11', '6-10', '7-9', '8-16'],
          ['8-15', '1-14', '2-13', '3-12', '4-11', '5-10', '6-9', '7-16'],
          ['7-15', '8-14', '1-13', '2-12', '3-11', '4-10', '5-9', '6-16'],
          ['6-15', '7-14', '8-13', '1-12', '2-11', '3-10', '4-9', '5-16'],
          ['5-15', '6-14', '7-13', '8-12', '1-11', '2-10', '3-9', '4-16'],
          ['4-15', '5-14', '6-13', '7-12', '8-11', '1-10', '2-9', '3-16'],
          ['3-15', '4-14', '5-13', '6-12', '7-11', '8-10', '1-9', '2-16'],
          ['2-15', '3-14', '4-13', '5-12', '6-11', '7-10', '8-9', '1-16'],
          ['1-7', '2-6', '3-5', '4-8', '9-13', '12-14', '11-15', '10-16'],
          ['1-6', '2-5', '3-8', '4-7', '9-14', '10-13', '12-15', '11-16'],
          ['1-5', '2-8', '3-7', '4-6', '9-15', '10-14', '11-13', '12-16'],
          ['1-4', '2-7', '3-6', '5-8', '9-12', '10-15', '11-14', '13-16'],
          ['1-3', '2-4', '5-7', '6-8', '9-11', '10-12', '13-15', '14-16'],
          ['1-8', '2-3', '4-5', '6-7', '10-11', '12-13', '14-15', '9-16'],
          ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12', '13-14', '15-16']],
          [['2-17', '3-16', '4-15', '5-14', '6-13', '7-12', '8-11','9-10'],
            ['1-17', '2-15', '3-14', '4-13', '5-12', '6-11', '7-10','8-9'],
            ['1-16', '15-17', '2-13', '3-12', '4-11', '5-10', '6-9','7-8'],
            ['1-15', '14-16', '13-17', '2-11', '3-10', '4-9', '5-8','6-7'],
            ['1-14', '13-15', '12-16', '11-17', '2-9', '3-8', '4-7','5-6'],
            ['1-13', '12-14', '11-15', '10-16', '9-17', '2-7', '3-6','4-5'],
            ['1-12', '11-13', '10-14', '9-15', '8-16', '7-17', '2-5','3-4'],
            ['1-11', '10-12', '9-13', '8-14', '7-15', '6-16', '5-17', '2-3'],
            ['1-10', '9-11', '8-12', '7-13', '6-14', '5-15', '4-16', '3-17'],
            ['1-9', '8-10', '7-11', '6-12', '5-13', '4-14', '3-15', '2-16'],
            ['1-8', '7-9', '6-10', '5-11', '4-12', '3-13', '2-14', '16-17'],
            ['1-7', '6-8', '5-9', '4-10', '3-11', '2-12', '14-17','15-16'],
            ['1-6', '5-7', '4-8', '3-9', '2-10', '12-17', '13-16','14-15'],
            ['1-5', '4-6', '3-7', '2-8', '10-17', '11-16', '12-15','13-14'],
            ['1-4', '3-5', '2-6', '8-17', '9-16', '10-15', '11-14','12-13'],
            ['1-3', '2-4', '6-17', '7-16', '8-15', '9-14', '10-13','11-12'],
            ['1-2', '4-17', '5-16', '6-15', '7-14', '8-13', '9-12','10-11']]
]

    tour_list = tr[cp]
    return tour_list

def _tours_list(cp):
    """туры таблиц по кругу в зависимости от кол-во участников (-cp- + 3) кол-во участников"""
    tour_list = []
    tr = [[['1-3'], ['1-2'], ['2-3']],
          [['1-3', '2-4'], ['1-2', '3-4'], ['2-3', '1-4']],
          [['2-4', '1-5'], ['1-4', '3-5'], ['1-3', '2-5'], ['2-3', '4-5'], ['1-2', '3-4']],
          [['2-4', '1-5', '3-6'], ['1-4', '2-6', '3-5'], ['1-3', '2-5', '4-6'], ['2-3', '1-6', '4-5'],
            ['1-2', '3-4', '5-6']],
          [['2-6', '3-5', '1-7'], ['2-5', '1-6', '4-7'], ['1-5', '4-6', '3-7'], ['4-5', '2-7', '3-6'],
            ['1-3', '2-4', '5-7'], ['1-4', '2-3', '6-7'], ['1-2', '3-4', '5-6']],
          [['2-6', '3-5', '1-7', '4-8'], ['2-5', '1-6', '3-8', '4-7'], ['1-5', '2-8', '4-6', '3-7'],
            ['1-8', '4-5', '2-7', '3-6'], ['1-3', '2-4', '5-7', '6-8'], ['1-4', '2-3', '6-7', '5-8'],
            ['1-2', '3-4', '5-6', '7-8']],
          [['1-9', '2-8', '3-7', '4-6'], ['5-9', '1-8', '2-7', '3-6'], ['4-9', '5-8', '1-7', '2-6'],
            ['3-9', '4-8', '5-7', '1-6'], ['2-4', '1-5', '3-8', '7-9'], ['4-1', '5-3', '9-2', '8-6'],
            ['1-3', '2-5', '4-7', '6-9'], ['3-2', '5-4', '8-9', '7-6'], ['1-2', '3-4', '5-6', '7-8']],
          [['1-9', '2-8', '3-7', '4-6', '5-10'], ['5-9', '1-8', '2-7', '3-6', '4-10'], ['4-9', '5-8', '1-7', '2-6', '3-10'],
            ['3-9', '4-8', '5-7', '1-6', '2-10'], ['2-4', '1-5', '3-8', '7-9', '6-10'], ['4-1', '5-3', '9-2', '8-6', '7-10'],
            ['1-3', '2-5', '4-7', '6-9', '8-10'], ['3-2', '5-4', '8-9', '7-6', '1-10'], ['1-2', '3-4', '5-6', '7-8', '9-10']],
          [['1-11', '2-10', '3-9', '4-8', '5-7'], ['6-11', '1-10', '2-9', '3-8', '4-7'], ['5-11', '6-10', '1-9', '2-8', '3-7'],
            ['4-11', '5-10', '6-9', '1-8', '2-7'], ['3-11', '4-10', '5-9', '6-8', '1-7'], ['2-11', '3-10', '4-9', '5-8', '6-7'],
            ['2-4', '1-5', '3-6', '7-10', '9-11'], ['1-4', '2-6', '3-5', '8-10', '7-11'], ['1-3', '2-5', '4-6', '7-9', '8-11'],
            ['2-3', '1-6', '4-5', '8-9', '10-11'], ['1-2', '3-4', '5-6', '7-8', '9-10']],
          [['1-11', '2-10', '3-9', '4-8', '5-7', '6-12'], ['6-11', '1-10', '2-9', '3-8', '4-7', '5-12'],
            ['5-11', '6-10', '1-9', '2-8', '3-7', '4-12'], ['4-11', '5-10', '6-9', '1-8', '2-7', '3-12'],
            ['3-11', '4-10', '5-9', '6-8', '1-7', '2-12'], ['2-11', '3-10', '4-9', '5-8', '6-7', '1-12'],
            ['2-4', '1-5', '3-6', '7-10', '9-11', '8-12'], ['1-4', '2-6', '3-5', '8-10', '7-11', '9-12'],
            ['1-3', '2-5', '4-6', '7-9', '8-11', '10-12'], ['2-3', '1-6', '4-5', '8-9', '10-11', '7-12'],
            ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12']],
          [['1-13', '2-12', '3-11', '4-10', '5-9', '6-8'], ['7-13', '1-12', '2-11', '3-10', '4-9', '5-8'],
            ['6-13', '7-12', '1-11', '2-10', '3-9', '4-8'], ['5-13', '6-12', '7-11', '1-10', '2-9', '3-8'],
            ['4-13', '5-12', '6-11', '7-10', '1-9', '2-8'], ['3-13', '4-12', '5-11', '6-10', '7-9', '1-8'],
            ['1-7', '2-6', '3-5', '4-11', '9-13', '10-12'], ['1-6', '2-5', '4-7', '3-12', '8-11', '10-13'],
            ['1-4', '2-7', '3-6', '5-10', '8-13', '9-12'], ['1-5', '3-7', '4-6', '2-13', '8-12', '9-11'],
            ['1-3', '2-4', '5-7', '6-9', '8-10', '11-13'], ['2-3', '4-5', '6-7', '8-9', '10-11', '12-13'],
            ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12']],
          [['1-13', '2-12', '3-11', '4-10', '5-9', '6-8', '7-14'], ['7-13', '1-12', '2-11', '3-10', '4-9', '5-8', '6-14'],
          ['6-13', '7-12', '1-11', '2-10', '3-9', '4-8', '5-14'], ['5-13',
                                                                   '6-12', '7-11', '1-10', '2-9', '3-8', '4-14'],
          ['4-13', '5-12', '6-11', '7-10', '1-9', '2-8', '3-14'], ['3-13',
                                                                   '4-12', '5-11', '6-10', '7-9', '1-8', '2-14'],
          ['1-7', '2-6', '3-5', '4-11', '9-13', '10-12', '8-14'], ['1-6',
                                                                   '2-5', '4-7', '3-12', '8-11', '10-13', '9-14'],
          ['1-4', '2-7', '3-6', '5-10', '8-13', '9-12', '11-14'], ['1-5',
                                                                   '3-7', '4-6', '2-13', '8-12', '9-11', '10-14'],
          ['1-3', '2-4', '5-7', '6-9', '8-10', '11-13', '12-14'], ['2-3',
                                                                   '4-5', '6-7', '8-9', '10-11', '12-13', '1-14'],
          ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12', '13-14']],
          [['1-15', '2-14', '3-13', '4-12', '5-11', '6-10', '7-9'], ['8-15', '1-14', '2-13', '3-12', '4-11', '5-10', '6-9'],
          ['8-15', '1-14', '2-13', '3-12', '4-11', '5-10', '6-9'], ['7-15',
                                                                    '8-14', '1-13', '2-12', '3-11', '4-10', '5-9'],
          ['6-15', '7-14', '8-13', '1-12', '2-11', '3-10', '4-9'], ['5-15',
                                                                    '6-14', '7-13', '8-12', '1-11', '2-10', '3-9'],
          ['4-15', '5-14', '6-13', '7-12', '8-11', '1-10', '2-9'], ['3-15',
                                                                    '4-14', '5-13', '6-12', '7-11', '8-10', '1-9'],
          ['2-15', '3-14', '4-13', '5-12', '6-11', '7-10', '8-9'], ['1-7',
                                                                    '2-6', '3-5', '4-8', '9-13', '12-14', '11-15'],
          ['1-6', '2-5', '3-8', '4-7', '9-14', '10-13', '12-15'], ['1-5',
                                                                   '2-8', '3-7', '4-6', '9-15', '10-14', '11-13'],
          ['1-4', '2-7', '3-6', '5-8', '9-12', '10-15', '11-14'], ['1-3',
                                                                   '2-4', '5-7', '6-8', '9-11', '10-12', '13-15'],
          ['1-8', '2-3', '4-5', '6-7', '10-11', '12-13', '14-15'], ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12', '13-14']],
          [['1-15', '2-14', '3-13', '4-12', '5-11', '6-10', '7-9', '8-16'],
          ['8-15', '1-14', '2-13', '3-12', '4-11', '5-10', '6-9', '7-16'],
          ['7-15', '8-14', '1-13', '2-12', '3-11', '4-10', '5-9', '6-16'],
          ['6-15', '7-14', '8-13', '1-12', '2-11', '3-10', '4-9', '5-16'],
          ['5-15', '6-14', '7-13', '8-12', '1-11', '2-10', '3-9', '4-16'],
          ['4-15', '5-14', '6-13', '7-12', '8-11', '1-10', '2-9', '3-16'],
          ['3-15', '4-14', '5-13', '6-12', '7-11', '8-10', '1-9', '2-16'],
          ['2-15', '3-14', '4-13', '5-12', '6-11', '7-10', '8-9', '1-16'],
          ['1-7', '2-6', '3-5', '4-8', '9-13', '12-14', '11-15', '10-16'],
          ['1-6', '2-5', '3-8', '4-7', '9-14', '10-13', '12-15', '11-16'],
          ['1-5', '2-8', '3-7', '4-6', '9-15', '10-14', '11-13', '12-16'],
          ['1-4', '2-7', '3-6', '5-8', '9-12', '10-15', '11-14', '13-16'],
          ['1-3', '2-4', '5-7', '6-8', '9-11', '10-12', '13-15', '14-16'],
          ['1-8', '2-3', '4-5', '6-7', '10-11', '12-13', '14-15', '9-16'],
          ['1-2', '3-4', '5-6', '7-8', '9-10', '11-12', '13-14', '15-16']]]

    tour_list = tr[cp]
    return tour_list


def load_playing_game_in_table_for_semifinal(stage):
    """растановка в полуфинале игроков со встречей сыгранной в группе"""
    id_player_exit_out_gr = [] # список ид игроков попадающих в финал из группы в порядке занятых место по возрастанию
    posev_player_exit_out_gr = []
    player_exit = []    
    mesto_rank = 1 # начальное место с которого вышли в финал
    system = System.select().where(System.title_id == title_id())
    choice = Choice.select().where(Choice.title_id == title_id())
    results = Result.select().where(Result.title_id == title_id())
    sys = system.select().where(System.stage == "Предварительный").get()
    sys_semifin = system.select().where(System.stage == stage).get()
    kol_gr = sys.total_group
    if stage == "1-й полуфинал":
        mesto_rank = 1
    else:
        sys_fin_last = system.select().where(System.stage == stage).get()
        mesto_rank = sys_fin_last.mesta_exit + 1 # место, попадающих в финал из группы начало
    how_many_mest_exit = sys_semifin.mesta_exit # количество мест попадающих из предварительного этапа
    for i in range(1, kol_gr + 1): # цикл по группам
        posev_player_exit_out_gr.clear()
        id_player_exit_out_gr.clear()
        choice_group = choice.select().where(Choice.group == f"{i} группа") 
        kol_player = len(choice_group) # число участников в группе
        if mesto_rank + how_many_mest_exit <= kol_player:
            mesto_rank_end = mesto_rank + how_many_mest_exit
        else:
            mesto_rank_end = kol_player + 1
        n = 0
        for k in range(mesto_rank, mesto_rank_end): # цикл в группе начиная с места с которого выходят в финал (зависит скольк игроков выходят из группы)
            ch_mesto_exit = choice_group.select().where(Choice.mesto_group == k).get()
            pl_id = ch_mesto_exit.player_choice_id # id игрока, занявшего данное место
            pl_posev = ch_mesto_exit.posev_group
            id_player_exit_out_gr.append(pl_id)
            posev_player_exit_out_gr.append(pl_posev) # номера игроков в группе вышедших в финал
            n += 1

        posev_pl = []
        temp = []
        posev_id_pl = []
        all_posev_id_pl = []
        if n > 1:
            # получаем все варианты встреч, сыгранных в группе игроков которые попали в полуфинал
            for i in combinations(posev_player_exit_out_gr, 2):
                posev_player_exit = list(i)
                for v in posev_player_exit:
                    ind = posev_player_exit_out_gr.index(v)
                    id_player = id_player_exit_out_gr[ind]
                    temp.append(id_player)
                    posev_id_pl = temp.copy()
                temp.clear()
                posev_pl.append(posev_player_exit)
                all_posev_id_pl.append(posev_id_pl)

            result_pre = results.select().where(Result.system_stage == "Предварительный") # изменить откуда выходят из группы или пф
            for d in range(0, len(posev_pl)):
                posev_exit = posev_pl[d]
                id_player_exit = all_posev_id_pl[d]
                if posev_exit[0] > posev_exit[1]: # если спортсмены заняли места не по расстановки в табл меняем на номера встречи в правильном порядке по возр
                    id_player_exit.reverse()
                    
                player_exit.clear()
                posev_exit.clear()
                for l in id_player_exit:
                    players = Player.select().where(Player.id == l).get()
                    family_city = players.full_name
                    player_exit.append(family_city)  
                    # номер ид в таблице -Result- встречи игроков, попавших в полуфинал идущих по расстоновке в таблице   
                # result_gr = result_pre.select().where((Result.player1 == player_exit[0]) & (Result.player2 == player_exit[1])).get() 
                result_gr = result_pre.select().where((Result.player1 == player_exit[0]) & (Result.player2 == player_exit[1]))
                count = len(result_gr) # если была не полная группа и в ПФ вышел только один человек, то пропускает запись в таблицу Result
                if count == 0:
                    break
                else:
                    result_gr = result_pre.select().where((Result.player1 == player_exit[0]) & (Result.player2 == player_exit[1])).get()
                result_pre_fin = results.select().where(Result.system_stage == stage)
                result_semifin_player1 = result_pre_fin.select().where(Result.player1.in_(player_exit))
                result_semifin = result_semifin_player1.select().where(Result.player2.in_(player_exit)).get()

                with db:
                    result_semifin.winner = result_gr.winner
                    result_semifin.points_win = result_gr.points_win
                    result_semifin.score_in_game = result_gr.score_in_game
                    result_semifin.score_win = result_gr.score_win
                    result_semifin.loser = result_gr.loser
                    result_semifin.points_loser = result_gr.points_loser
                    result_semifin.score_loser = result_gr.score_loser
                    result_semifin.save()
    pv = sys_semifin.page_vid
    my_win.tabWidget.setCurrentIndex(4)


def load_playing_game_in_table_for_final(fin):
    """растановка в финале игроков со встречей сыгранной в группе"""
    id_player_exit_out_gr = [] # список ид игроков попадающих в финал из группы в порядке занятых место по возрастанию
    posev_player_exit_out_gr = []
    player_exit = []
    mesto_rank = 1 # начальное место с которого вышли в финал
    system = System.select().where(System.title_id == title_id())
    choice = Choice.select().where(Choice.title_id == title_id())
    results = Result.select().where(Result.title_id == title_id())
    # выбор выход в финал =======
    id_system = system_id(fin) # получаем id системы
    systems = system.select().where(System.id == id_system).get()
    stage = systems.stage_exit
    kol_gr = systems.total_group
    sys = system.select().where(System.stage == stage).get()
    kol_gr = sys.total_group
    if fin == "1-й финал":
        mesto_rank = 1
    else:
        sum_player = []
        etap_exit = stage
        for m in system:
            if m.stage == fin:
                break
            else:
                if m.stage_exit == etap_exit:
                    total_player_exit = m.mesta_exit
                    sum_player.append(total_player_exit)
        sum_player_exit = sum(sum_player) # сумма игроков вышедших в финал   
        mesto_rank = sum_player_exit + 1
    how_many_mest_exit = systems.mesta_exit # количество мест попадающих из предварительного этапа
    for i in range(1, kol_gr + 1): # цикл по группам
        posev_player_exit_out_gr.clear()
        id_player_exit_out_gr.clear()
        # =====
        if stage == "Предварительный":
            choice_group = choice.select().where(Choice.group == f"{i} группа")
        else:
            choice_group = choice.select().where((Choice.semi_final == stage) & (Choice.sf_group == f"{i} группа"))
        # =====
        kol_player = len(choice_group) # число участников в группе
        if mesto_rank + how_many_mest_exit <= kol_player:
            mesto_rank_end = mesto_rank + how_many_mest_exit
        else:
            mesto_rank_end = kol_player + 1
        n = 0
        for k in range(mesto_rank, mesto_rank_end): # цикл в группе начиная с места с которого выходят в финал (зависит скольк игроков выходят из группы)
            # =====
            if stage == "Предварительный":
                ch_mesto_exit = choice_group.select().where(Choice.mesto_group == k).get()
            else:
                ch_mesto_exit = choice_group.select().where(Choice.mesto_semi_final == k).get()
            #  =======
            pl_id = ch_mesto_exit.player_choice_id # id игрока, занявшего данное место
            pl_posev = ch_mesto_exit.posev_group  if stage == "Предварительный" else ch_mesto_exit.posev_sf 
            id_player_exit_out_gr.append(pl_id)
            posev_player_exit_out_gr.append(pl_posev) # номера игроков в группе вышедших в финал
            n += 1

        posev_pl = []
        temp = []
        posev_id_pl = []
        all_posev_id_pl = []
        if n > 1:
            # получаем все варианты встреч, сыгранных в группе игроков которые попали в финал
            for i in combinations(posev_player_exit_out_gr, 2):
                posev_player_exit = list(i)
                for v in posev_player_exit:
                    ind = posev_player_exit_out_gr.index(v)
                    id_player = id_player_exit_out_gr[ind]
                    temp.append(id_player)
                    posev_id_pl = temp.copy()
                temp.clear()
                posev_pl.append(posev_player_exit)
                all_posev_id_pl.append(posev_id_pl)

            result_pre = results.select().where(Result.system_stage == stage) # изменить откуда выходят из группы или пф
            for d in range(0, len(posev_pl)):
                posev_exit = posev_pl[d]
                id_player_exit = all_posev_id_pl[d]
                if posev_exit[0] > posev_exit[1]: # если спортсмены заняли места не по расстановки в табл меняем на номера встречи в правильном порядке по возр
                    id_player_exit.reverse()
                    
                player_exit.clear()
                posev_exit.clear()
                for l in id_player_exit:
                    players = Player.select().where(Player.id == l).get()
                    family_city = players.full_name
                    player_exit.append(family_city)  
                    # номер ид в таблице -Result- встречи игроков, попавших в финал идущих по расстоновке в таблице   
                result_gr = result_pre.select().where((Result.player1 == player_exit[0]) & (Result.player2 == player_exit[1])).get() 

                result_pre_fin = results.select().where(Result.number_group == fin)

                result_fin_1 = result_pre_fin.select().where((Result.player1 == player_exit[0]))
                result_fin = result_fin_1.select().where(Result.player2 == player_exit[1])
                count = len(result_fin)

                if count != 1:
                    result_fin = result_pre_fin.select().where((Result.player1 == player_exit[1]) & (Result.player2 == player_exit[0])).get()
                else:
                    result_fin = result_pre_fin.select().where((Result.player1 == player_exit[0]) & (Result.player2 == player_exit[1])).get()

                with db:
                    result_fin.winner = result_gr.winner
                    result_fin.points_win = result_gr.points_win
                    result_fin.score_in_game = result_gr.score_in_game
                    result_fin.score_win = result_gr.score_win
                    result_fin.loser = result_gr.loser
                    result_fin.points_loser = result_gr.points_loser
                    result_fin.score_loser = result_gr.score_loser
                    result_fin.save()
    stage = fin
    pv = systems.page_vid
    table_made(pv, stage)


def made_file_excel_for_rejting():
    """создание файла Excel для обсчета рейтинга"""
    result = Result.select().where(Result.title_id == title_id())
    players = Player.select().where(Player.title_id == title_id())
    player_result = result.select().where((Result.points_loser != 0) | (Result.score_in_game != "В : П")).order_by(Result.winner)
    book = op.Workbook()
    worksheet = book.active
    names_headers = ["Этап","Группа", "Победитель", "День рождения", "Проигравший", "День рождения", "Счет"]
    for m in range(1, 8):
        c =  worksheet.cell(row = 1, column = m)
        c.value = names_headers[m - 1]
    k = 2
    
    for l in player_result:
        id = l.id
        point_winner = l.points_win
        if point_winner == "":
            continue
        stage = l.system_stage
        group = l.number_group
        pl_win = l.winner       
        pl_los = l.loser
        id_win = players.select().where(Player.full_name == pl_win).get()
        pl_win = id_win.player
        b_day_win = id_win.bday
        bd_win = format_date_for_view(str_date=b_day_win)
        id_los =  players.select().where(Player.full_name == pl_los).get()
        pl_los = id_los.player
        b_day_los = id_los.bday
        bd_los = format_date_for_view(str_date=b_day_los)
        score = l.score_in_game

        c1 = worksheet.cell(row = k, column = 1)
        c1.value = stage
        c2 = worksheet.cell(row = k, column = 2)
        c2.value = group
        c3 = worksheet.cell(row = k, column = 3)
        c3.value = pl_win
        c4 = worksheet.cell(row = k, column = 4)
        c4.value = bd_win
        c5 = worksheet.cell(row = k, column = 5)
        c5.value = pl_los
        c6 = worksheet.cell(row = k, column = 6)
        c6.value = bd_los
        c7 = worksheet.cell(row = k, column = 7)
        c7.value = score
        k += 1

    t_id = Title.get(Title.id == title_id())
    short_name = t_id.short_name_comp 
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['b'].width = 15
    worksheet.column_dimensions['c'].width = 30
    worksheet.column_dimensions['d'].width = 15
    worksheet.column_dimensions['e'].width = 30
    worksheet.column_dimensions['g'].width = 15
    worksheet.column_dimensions['f'].width = 15
    f_name = f"{short_name}_report.xlsx"
    filename, filter = QtWidgets.QFileDialog.getSaveFileName(my_win, 'Save file', f'{f_name}','Excel files (*.xlsx)')
    book.save(filename)


def randevy_list():
    """Порядок встреч по кругу"""
    from openpyxl.styles import (
                        PatternFill, Border, Side, 
                        Alignment, Font, GradientFill
                        )
    group_list = ["Одна таблица", "Предварительный", "1-й полуфинал", "2-й полуфинал"]
    fin_circle_list = []
    t_id = Title.get(Title.id == title_id())  
    systems = System.select().where(System.title_id == title_id())
    name = t_id.name
    gamer = t_id.gamer
    for k in systems:
        stage_system = k.stage
        system_string = k.label_string
        if stage_system not in group_list and system_string == "Круговая таблица на 16 участников":
            fin_circle_list.append(stage_system)

    stage, ok = QInputDialog.getItem(my_win, "Финалы", "Выберите финал по кругу для\n"
                                        "создания порядка встреч", fin_circle_list)
    fin = f'{stage[:1]}-fin'
    id_system = system_id(stage)
    #========
    result_list = Result.select().where((Result.title_id == title_id()) & (Result.system_id == id_system))
    
    book = op.Workbook()
    worksheet = book.active
    thins = Side(border_style="medium", color="0000ff")
    double = Side(border_style="dashDot", color="ff0000")
    dDD = Side(border_style="double", color="0000ff")

    names_headers = ["Тур", "Встреча", "Спортсмен", "Спортсмен", "Тур", "Встреча", "Спортсмен", "Спортсмен"]

    for m in range(1, 9):
        c =  worksheet.cell(row = 4, column = m)
        с_title = worksheet.cell(row = 1, column = m)
        c.border  = Border(top=thins, bottom=thins, left=thins, right=thins)
        с_title.border  = Border(top=dDD, bottom=dDD, left=dDD, right=dDD)
        c.alignment = Alignment(horizontal='center')
        c.font = Font(italic = True, bold = True, name='Times New Roman', size=12)
        c.value = names_headers[m - 1]
    worksheet.merge_cells('A1:H2')
    worksheet.merge_cells('A3:H3')
    megre_cell_a1 = worksheet['A1']
    megre_cell_a3 = worksheet['A3']
    megre_cell_a1.alignment = Alignment(horizontal='center', vertical='center') 
    megre_cell_a3.alignment = Alignment(horizontal='center')
    megre_cell_a1.font = Font(bold = True, name='Times New Roman', size=18)
    megre_cell_a3.font = Font(italic = True, bold = True, name='Times New Roman', size=16)
    megre_cell_a1.value = f'{name}.{gamer}'
    megre_cell_a3.value =f'ПОРЯДОК ВСТРЕЧ {stage}'


    border_1 = Border(top=double, bottom=double, left=thins, right=thins)
    border_2 = Border(top=double, bottom=dDD, left=thins, right=thins)
    k = 0 
    l = 5 
    b = 1  
    for p in result_list:
        round = int(p.round)
        tour = p.tours
        pl1 = p.player1
        pl2 = p.player2
        zn1 = pl1.find("/")
        fio_1 = pl1[:zn1]
        zn2 = pl2.find("/")
        fio_2 = pl2[:zn2]

        t = 8 * round

        c1 = worksheet.cell(row = l, column = (1 + k))
        c1.alignment = Alignment(horizontal='center')
        c1.border = border_1 if b < t else border_2
        c1.value = round
        c2 = worksheet.cell(row = l, column = (2 + k))
        c2.alignment = Alignment(horizontal='center')
        c2.border = border_1 if b < t else border_2
        c2.value = tour
        c3 = worksheet.cell(row = l, column = (3 + k))
        c3.font = Font(name="Arial", size=12)
        c3.border = border_1 if b < t else border_2
        c3.value = fio_1
        c4 = worksheet.cell(row = l, column = (4 + k))
        c4.font = Font(name="Arial", size=12)
        c4.border = border_1 if b < t else border_2
        c4.value = fio_2
   
        l += 1
        b += 1
        if l > 68 and k == 0:
            k = 4
            l = 5
        
    t_id = Title.get(Title.id == title_id())

    worksheet.column_dimensions['A'].width = 5
    worksheet.column_dimensions['B'].width =8
    worksheet.column_dimensions['C'].width = 28
    worksheet.column_dimensions['D'].width = 28
    worksheet.column_dimensions['E'].width = 5
    worksheet.column_dimensions['F'].width = 8
    worksheet.column_dimensions['G'].width = 28
    worksheet.column_dimensions['H'].width = 28

    sex = t_id.gamer
    f_name = f"{sex}_порядок_встреч.xlsx"
    filename, filter = QtWidgets.QFileDialog.getSaveFileName(my_win, 'Save file', f'{f_name}','Excel files (*.xlsx)')
    book.save(filename)


def made_file_excel_list_player():
    """Создание списка участников в excel файле"""
    players = Player.select().where(Player.title_id == title_id())

    book = op.Workbook()
    worksheet = book.active
    names_headers = ["№", "Фамилия, Имя", "Дата рождения", "Рейтинг", "Город", "Субъект РФ", "Разряд", "Тренеры"]
    for m in range(1, 9):
        c =  worksheet.cell(row = 1, column = m)
        c.value = names_headers[m - 1]

    k = 2    
    for pl in players:
        fio = pl.player
        dr = pl.bday
        r = pl.rank
        gorod = pl.city
        obl = pl.region
        raz = pl.razryad
        id_coach = pl.coach_id
        coachs = Coach.get(Coach.id == id_coach)
        fio_coach = coachs.coach
        n = k - 1
        c1 = worksheet.cell(row = k, column = 1)
        c1.value = n
        c2 = worksheet.cell(row = k, column = 2)
        c2.value = fio
        c3 = worksheet.cell(row = k, column = 3)
        c3.value = dr
        c4 = worksheet.cell(row = k, column = 4)
        c4.value = r
        c5 = worksheet.cell(row = k, column = 5)
        c5.value = gorod
        c6 = worksheet.cell(row = k, column = 6)
        c6.value = obl
        c7 = worksheet.cell(row = k, column = 7)
        c7.value = raz
        c8 = worksheet.cell(row = k, column = 8)
        c8.value = fio_coach 
        k += 1

    t_id = Title.get(Title.id == title_id())

    worksheet.column_dimensions['A'].width = 8
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 12
    worksheet.column_dimensions['D'].width = 10
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    worksheet.column_dimensions['G'].width = 10
    worksheet.column_dimensions['H'].width = 40
    sex = t_id.gamer
    f_name = f"{sex}_списки.xlsx"
    filename, filter = QtWidgets.QFileDialog.getSaveFileName(my_win, 'Save file', f'{f_name}','Excel files (*.xlsx)')
    book.save(filename)


def button_move_enabled():
    """включает или выключает кнопки перемещения по таблице в зависимости от выделенной строки"""
    count = my_win.tableWidget.rowCount()
    row = my_win.tableWidget.currentRow()
    if row == 0:
        my_win.Button_down.setEnabled(True)
        my_win.Button_up.setEnabled(False)
    elif row == count - 1:
        my_win.Button_up.setEnabled(True)
        my_win.Button_down.setEnabled(False)
    else:
        my_win.Button_up.setEnabled(True)
        my_win.Button_down.setEnabled(True)


def move_row_in_tablewidget():
    """перемещяет выделенную строку по таблице вверх или вниз"""
    sender = my_win.sender()
    row_count = my_win.tableWidget.rowCount()
    row_cur = my_win.tableWidget.currentRow()
    if row_cur == 1:
        my_win.Button_up.setEnabled(False)
        my_win.Button_down.setEnabled(True)
    if row_cur == row_count:
        my_win.Button_down.setEnabled(False)
        my_win.Button_up.setEnabled(True)
    item_cur = my_win.tableWidget.item(row_cur, 1).text()
    item_cur_name = my_win.tableWidget.item(row_cur, 2).text()
    if sender == my_win.Button_down:
        item_tmp = my_win.tableWidget.item(row_cur + 1, 1).text()
        item_temp = my_win.tableWidget.item(row_cur + 1, 2).text()
        my_win.tableWidget.selectRow(row_cur + 1)
        my_win.tableWidget.setItem(row_cur + 1, 1, QTableWidgetItem(str(item_cur)))
        my_win.tableWidget.setItem(row_cur, 1, QTableWidgetItem(str(item_tmp)))
        my_win.tableWidget.setItem(row_cur + 1, 2, QTableWidgetItem(str(item_cur_name)))
        my_win.tableWidget.setItem(row_cur, 2, QTableWidgetItem(str(item_temp)))
    else:
        item_tmp = my_win.tableWidget.item(row_cur - 1, 1).text()
        item_temp = my_win.tableWidget.item(row_cur - 1, 2).text()
        my_win.tableWidget.selectRow(row_cur - 1)
        my_win.tableWidget.setItem(row_cur - 1, 1, QTableWidgetItem(str(item_cur)))
        my_win.tableWidget.setItem(row_cur, 1, QTableWidgetItem(str(item_tmp)))
        my_win.tableWidget.setItem(row_cur - 1, 2, QTableWidgetItem(str(item_cur_name)))
        my_win.tableWidget.setItem(row_cur, 2, QTableWidgetItem(str(item_temp)))


def made_list_GSK():
    """создание списка судейской коллегии"""
    my_win.tableWidget.clear()
    my_win.radioButton_GSK.setChecked(True)
    my_win.Button_made_page_pdf.setEnabled(True)
    my_win.tableWidget.show()
    number_of_referee, ok = QInputDialog.getInt(my_win, "Главная судейская коллегия", "Укажите число судей списка\n главной cудейской коллегии.", 4, 3, 10)
    if ok:
        title = Title.get(Title.id == title_id())
        referee = title.referee
        kat_referee = title.kat_ref
        secretary = title.secretary
        kat_secretary = title.kat_sec
        list_referee = [referee, secretary]
        list_kategory = [kat_referee, kat_secretary]

        my_win.tableWidget.setColumnCount(4) # устанавливает колво столбцов
        my_win.tableWidget.setRowCount(number_of_referee)

        column_label = ["№", "Должность", "Фамилия Имя Отчество/ Город", "Категория"]
        my_win.tableWidget.resizeColumnsToContents()
        for i in range(0, 4):  # закрашивает заголовки таблиц  рейтинга зеленым цветом
            my_win.tableWidget.showColumn(i)
            item = QtWidgets.QTableWidgetItem()
            brush = QtGui.QBrush(QtGui.QColor(76, 100, 255))
            brush.setStyle(QtCore.Qt.SolidPattern)
            item.setForeground(brush)
            my_win.tableWidget.setHorizontalHeaderItem(i, item)
        my_win.tableWidget.setHorizontalHeaderLabels(column_label) # заголовки столбцов в tableWidget
        referee_list = []
        post_list = ["", "ССВК", "1-й кат.", "2-й кат."]
        category_list = ["","Зам. Главного судьи", "Зам. Главного секретаря", "Ведущий судья"]
        my_win.tableWidget.setItem(0, 1, QTableWidgetItem("Гл. судья"))
        my_win.tableWidget.setItem(1, 1, QTableWidgetItem("Гл. секретарь"))
    else:
        return
    for k in range(0, 2):
        my_win.tableWidget.setItem(k, 2, QTableWidgetItem(str(list_referee[k])))
        my_win.tableWidget.setItem(k, 3, QTableWidgetItem(str(list_kategory[k])))
    for n in range(2, int(number_of_referee)): 
        comboBox_list_post = QComboBox()
        comboBox_list_category = QComboBox()  
        comboBox_family_city = QComboBox()
        referee_list = load_comboBox_referee()

        comboBox_family_city.setPlaceholderText("Введите фамилию судьи")
        comboBox_family_city.setCurrentIndex(-1)
        comboBox_family_city.setEditable(True)
        comboBox_list_category.addItems(category_list)
        comboBox_list_post.addItems(post_list) 
        comboBox_family_city.addItems(referee_list)

        my_win.tableWidget.setCellWidget(n, 1, comboBox_list_category)
        my_win.tableWidget.setCellWidget(n, 2, comboBox_family_city)
        my_win.tableWidget.setCellWidget(n, 3, comboBox_list_post)   
        for l in range(0, number_of_referee):
            my_win.tableWidget.setItem(l, 0, QTableWidgetItem(str(l + 1))) # (номер строки, номер столбца, значения)
        my_win.tableWidget.resizeColumnsToContents()
        comboBox_family_city.currentTextChanged.connect(change_on_comboBox_referee)   


def change_on_comboBox_referee(comboBox_family_city):
    """добавляет в базу данных судей если их там нет"""
    row_cur = my_win.tableWidget.currentRow()
    mark = comboBox_family_city.find("/") # если еще нет фамилии и города
    if mark != 0 and mark != -1:
        add_referee_to_db()
        family_referee = comboBox_family_city[:mark]
        family_referee = family_referee.title()
        referees = Referee.select().where(Referee.family == family_referee)
        if len(referees) > 0:
            for ref in referees:
                kategor = ref.category
                kat = my_win.tableWidget.cellWidget(row_cur, 3)
                kat.setCurrentText(kategor)
                return
    else:
        kat = my_win.tableWidget.cellWidget(row_cur, 3)
        kat.setCurrentText("")


def add_referee_to_db():
    """добавляет в базу данных новых судей"""
    sender = my_win.sender()
    count = my_win.tableWidget.rowCount()
    if sender == my_win.comboBox_kategor_ref:
        kat = my_win.comboBox_kategor_ref.currentText()
        item = my_win.comboBox_referee.currentText()
        whrite_referee_to_db(kat, item, k=0)
    elif sender == my_win.comboBox_kategor_sec:
        kat = my_win.comboBox_kategor_sec.currentText()
        item = my_win.comboBox_secretary.currentText()
        whrite_referee_to_db(kat, item, k=0)
    else:
        for k in range(2, count):
            item = my_win.tableWidget.cellWidget(k, 2).currentText()
            kat = my_win.tableWidget.cellWidget(k, 3).currentText()
            whrite_referee_to_db(kat, item, k)


def whrite_referee_to_db(kat, item, k=0):
    """запись рефери в db"""
    sender = my_win.sender()
    if kat != "":
        mark = item.find("/")
        family_referee = item[:mark]
        family_referee = family_referee.title()
        city_referee = item[mark + 2:]
        city_referee = city_referee.title()
        full_referee = f"{family_referee}/ {city_referee}"
        if sender == my_win.Button_made_page_pdf:
            f_referee = my_win.tableWidget.cellWidget(k, 2)
            f_referee.setCurrentText(full_referee)
        referees = Referee.select().where(Referee.family == family_referee)
        if len(referees) == 0:
            with db:
                ref = Referee(family=family_referee, city=city_referee, category=kat).save()


def view_all_page_pdf():    
    """просмотр все страниц соревнования pdf"""
    title = Title.get(Title.id == title_id())
    pdf_files_list = []
    rus_name_list = []
    pdf_file_canot_in_comp_list = ["player_list_payment.pdf", "player_list_debitor.pdf", "begunki", "player_list.pdf", 'player_list_duplicate.pdf', 'protokol_1-fin.pdf', 'protokol_2-fin.pdf'] # список, который игнорируется при составлении файла pdf соревнования
    stage_dict = {"table_group.pdf": "Предварительный",
                   "player_list_mesto.pdf": "Список участников по месту",
                   "winners_list.pdf": "Список победителей и призеров",
                   "player_list_alf.pdf": "Список участников по алф",
                   "title.pdf": "Титульный лист",
                   "referee_list.pdf": "ГСК",
                   "regions_list.pdf": "Список субъктов РФ",
                   "1-semifinal.pdf": "1-й полуфинал",
                   "2-semifinal.pdf": "2-й полуфинал",
                   "1-final.pdf": "1-й финал",
                   "2-final.pdf": "2-й финал",
                   "3-final.pdf": "3-й финал",
                   "4-final.pdf": "4-й финал",
                   "5-final.pdf": "5-й финал",
                   "6-final.pdf": "6-й финал",
                   "7-final.pdf": "7-й финал",
                   "8-final.pdf": "8-й финал",
                   "9-final.pdf": "9-й финал",
                   "10-final.pdf": "10-й финал",
                   "superfinal.pdf": "Суперфинал",
                   "one_table.pdf": "одна таблица"}
    short_name = title.short_name_comp
    count_mark = len(short_name)
    all_pdf_files_list = os.listdir("table_pdf")
    for name_files in all_pdf_files_list:
        text = name_files.find(short_name)
        text_stage = name_files[count_mark + 1:]
        if text == 0 and text_stage not in pdf_file_canot_in_comp_list:
            pdf_files_list.append(name_files)
            for latin_name in stage_dict.keys():
                if text_stage == latin_name:
                    rus_name = stage_dict[text_stage]
                    rus_name_list.append(rus_name)
                    break

    row = len(pdf_files_list)
    my_win.tableWidget.setColumnCount(3) # устанавливает колво столбцов
    my_win.tableWidget.setRowCount(row)
    column_label = ["№", "Файлы", "Этапы"]
    for c in range(0, 3):
        my_win.tableWidget.showColumn(c)

    my_win.tableWidget.setHorizontalHeaderLabels(column_label) # заголовки столбцов в tableWidget
    my_win.tableWidget.setDragDropOverwriteMode(True)
    my_win.tableWidget.setSelectionMode(QAbstractItemView.SingleSelection)
    my_win.tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)
    my_win.tableWidget.show()
    row_count = 0
    for item in pdf_files_list:
        item_name = rus_name_list[row_count]
        my_win.tableWidget.setItem(row_count, 0, (QTableWidgetItem(str(row_count + 1)))) # нумерация строк
        my_win.tableWidget.setItem(row_count, 1, (QTableWidgetItem(str(item)))) # файл на латинице
        my_win.tableWidget.setItem(row_count, 2, (QTableWidgetItem(str(item_name)))) # руское название файла
        row_count += 1
    my_win.Button_made_one_file_pdf.setEnabled(True)
    my_win.tableWidget.resizeColumnsToContents()


def made_list_regions():
    """создание списка регионов"""
    my_win.radioButton_regions.setChecked(True)
    my_win.Button_made_page_pdf.setEnabled(True)
    my_win.tableWidget.clear()
    region_list = []
    regions = Player.select().where(Player.title_id == title_id()) 
    
    for k in regions:
        reg = k.region
        if reg != "":
            if reg not in region_list:
                region_list.append(reg)
    count = len(region_list)
    region_list.sort()
    n = 0
    for l in region_list:
        my_win.tableWidget.setItem(n, 0, QTableWidgetItem(str(n + 1)))
        my_win.tableWidget.setItem(n, 1, QTableWidgetItem(str(l)))
        my_win.tableWidget.setColumnCount(2) # устанавливает колво столбцов
        my_win.tableWidget.setRowCount(count)
        column_label = ["№", "Субъекты РФ"]
        n += 1
    my_win.tableWidget.setHorizontalHeaderLabels(column_label) # заголовки столбцов в tableWidget
    for i in range(0, 2):  # закрашивает заголовки таблиц  рейтинга зеленым цветом
        my_win.tableWidget.showColumn(i)
        item = QtWidgets.QTableWidgetItem()
        brush = QtGui.QBrush(QtGui.QColor(76, 100, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        my_win.tableWidget.setHorizontalHeaderItem(i, item)


def made_list_players_for_pdf_file():
    """создание списка по алфавиту"""
    from reportlab.platypus import Table
    styles = getSampleStyleSheet()
    custom_style = styles['Normal'].fontName = 'DejaVuSerif'
    custom_style = styles['Normal'].fontSize = 6
    custom_style = styles["Normal"].clone("CustomStyle")
    custom_style.wordWrap = 'LTR' # Перенос слов (LTR - Left-To-Right)
    custom_style.leading = 6 # Межстрочный интервал
    story = []  # Список данных таблицы участников
    # elements = []  # Список Заголовки столбцов таблицы
    tit = Title.get(Title.id == title_id())
    for k in range(0, 2):
        elements = []  # Список Заголовки столбцов таблицы
        if k == 0:
            player_list_x = Player.select().where(Player.title_id == title_id()).order_by(Player.mesto)
        else:
            player_list_x = Player.select().where(Player.title_id == title_id()).order_by(Player.player)
        player_list = player_list_x.select().where(Player.player != "x")
        short_name = tit.short_name_comp
        gamer = tit.gamer
        otc = tit.otchestvo # если 1 значит в списках присутсвует отчество
        count = len(player_list)  # количество записей в базе
        kp = count + 1
        n = 0
        for l in player_list:
            n += 1
            if otc == 1:
                pat_id  = l.patronymic_id 
                patronymics = Patronymic.select().where(Patronymic.id == pat_id).get()  
                o = patronymics.patronymic
                p = l.player
                p = f"{p} {o}"
            else:
                p = l.player
            b = l.bday
            b = format_date_for_view(str_date=b)
            r = l.rank
            c = l.city
            g = l.region
            z = l.razryad
            coach_id = l.coach_id
            t = coach_id.coach
            m = l.mesto
            t = chop_line(t)
            # data = [n, p, b, r, c, g, z, t, m]
            data = [n, [Paragraph(p, custom_style)], b, r, c, g, z, [Paragraph(t, custom_style)], m]
            elements.append(data)
        elements.insert(0, ["№", "Фамилия, Имя", "Дата рожд.", "R", "Город", "Регион", "Разряд", "Тренер(ы)",
                            "Место"])
        t = Table(elements,
            colWidths=(0.8 * cm, 5.0 * cm, 1.6 * cm, 0.8 * cm, 2.5 * cm, 3.2 * cm, 1.1 * cm, 4.0 * cm, 1.0 * cm),
            # rowHeights=None, repeatRows=1)  # ширина столбцов, если None-автоматическая
              rowHeights=(0.45 * cm), repeatRows=1)  # ширина столбцов, если None-автоматическая
        t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                        #    ('FONTSIZE', (0, 0), (-1, -1), 7),
                            ('FONTSIZE', (0, 0), (-1, -1), 6),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 1),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (-1, kp * -1), 'CENTER'),
                           ('BACKGROUND', (0, 0), (8, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (8, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, kp * -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey, None, (1, 1)),
                        #    ('LINEBELOW', (0, 0), (-1, -1), 0.02, colors.grey, None, (1, 1)),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.5, colors.black)
                           ]))



        h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=150,
                firstLineIndent=-20, textColor="green")  # стиль параграфа
        h3.spaceAfter = 10  # промежуток после заголовка
        story.append(Paragraph(f'Список участников. {gamer}', h3))
        story.append(t)
        if k == 0:
            doc = SimpleDocTemplate(f"{short_name}_player_list_mesto.pdf", pagesize=A4)
        else:
            doc = SimpleDocTemplate(f"{short_name}_player_list_alf.pdf", pagesize=A4)
        catalog = 1
        change_dir(catalog)
        doc.build(story, onFirstPage=func_zagolovok, onLaterPages=func_zagolovok)
        os.chdir("..")


def made_list_winners():
    """создание списка победителей и призеров"""
    my_win.radioButton_winner.setChecked(True)
    my_win.Button_made_page_pdf.setEnabled(True)
    my_win.tableWidget.clear()
    players = Player.select().where(Player.title_id == title_id())
    winners = players.select().where((Player.mesto < 4) & (Player.bday != 0000-00-00)).order_by(Player.mesto)
    count = len(winners)
    if count == 0:
        return
    n = 0
    for l in winners:
        my_win.tableWidget.setColumnCount(8) # устанавливает колво столбцов
        my_win.tableWidget.setRowCount(count)       
        column_label = ["Место", "Фамилия, Имя", "Дата рождения", "Рейтинг", "Город", "Регион", "Разряд", "Тренеры"]
        coachs = Coach.select().where(Coach.id == l.coach_id).get()
        family_coach = coachs.coach
        bd = l.bday
        b_day = format_date_for_view(str_date=bd)
        my_win.tableWidget.setItem(n, 0, QTableWidgetItem(str(f"{l.mesto} место")))
        my_win.tableWidget.setItem(n, 1, QTableWidgetItem(str(l.player)))
        my_win.tableWidget.setItem(n, 2, QTableWidgetItem(b_day))
        my_win.tableWidget.setItem(n, 3, QTableWidgetItem(str(l.rank)))
        my_win.tableWidget.setItem(n, 4, QTableWidgetItem(str(l.city)))
        my_win.tableWidget.setItem(n, 5, QTableWidgetItem(str(l.region)))
        my_win.tableWidget.setItem(n, 6, QTableWidgetItem(str(l.razryad)))
        my_win.tableWidget.setItem(n, 7, QTableWidgetItem(str(family_coach)))

        for i in range(0, 8):  # закрашивает заголовки таблиц  рейтинга зеленым цветом
            my_win.tableWidget.showColumn(i)
            item = QtWidgets.QTableWidgetItem()
            brush = QtGui.QBrush(QtGui.QColor(76, 100, 255))
            brush.setStyle(QtCore.Qt.SolidPattern)
            item.setForeground(brush)
            my_win.tableWidget.setHorizontalHeaderItem(i, item)
        n += 1
    my_win.tableWidget.setHorizontalHeaderLabels(column_label) # заголовки столбцов в tableWidget


def made_pdf_list():
    """создание страниц PDF соревнования"""
    if my_win.radioButton_GSK.isChecked():
        add_referee_to_db()
        list_referee_pdf()
        my_win.view_referee_list_Action.setEnabled(True)
    elif my_win.radioButton_regions.isChecked():
        list_regions_pdf()
    elif my_win.radioButton_winner.isChecked():
        list_winners_pdf()
    my_win.Button_made_page_pdf.setEnabled(False)


def check_player_whitout_R():
    """Список участников для отметки кто оплатил лицензию за рейтинг"""
    from reportlab.platypus import Table
    from sys import platform
    sender = my_win.sender()
    elements = []
    story = []
    tit = Title.get(Title.id == title_id())
    if sender == my_win.print_list_nopay_R_Action:
        player_list = Player.select().where((Player.title_id == title_id()) & (Player.pay_rejting == "долг")).order_by(Player.city)  # сортировка по региону
    elif sender == my_win.print_list_pay_R_Action:
        player_list = Player.select().where((Player.title_id == title_id()) & (Player.pay_rejting == "оплачен")).order_by(Player.city)
    short_name = tit.short_name_comp
    gamer = tit.gamer
    count = len(player_list)  # количество записей в базе
    kp = count + 1
    n = 0
    for l in player_list:
        n += 1
        p = l.player
        c = l.city
        d = l.bday
        coach_id = l.coach_id
        t = coach_id.coach
        t = chop_line(t)
        app = l.comment
        data = [n, p, c, d, t, app]
        elements.append(data)

    elements.insert(0, ["№", "Фамилия, Имя", "Город", "Дата рожд", "Тренер(ы)", "Примечание"])
    t = Table(elements, colWidths=(0.7 * cm, 4.5 * cm, 3.2 * cm, 2.0 * cm, 6.0 * cm, 3.4 * cm), rowHeights=None, repeatRows=1)  # ширина столбцов, если None-автоматическая
    t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            # ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                           ('FONTSIZE', (0, 0), (-1, -1), 9),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 1),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (-1, kp * -1), 'CENTER'),
                           ('BACKGROUND', (0, 0), (8, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (8, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, kp * -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.5, colors.black)
                           ]))

    h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=150,
            firstLineIndent=-20, textColor="green")  # стиль параграфа
    h3.spaceAfter = 10  # промежуток после заголовка
    story.append(Paragraph(f'Список должников оплаты лицензий за R. {gamer}', h3))
    story.append(t)
    doc = SimpleDocTemplate(f"{short_name}_player_list_debitor.pdf", pagesize=A4, 
                            rightMargin=1*cm, leftMargin=1*cm, topMargin=1.5*cm, bottomMargin=1*cm) # название, вид страницы, размер полей
    view_file = f"{short_name}_player_list_debitor.pdf"
    catalog = 1
    change_dir(catalog)
    doc.build(story)
    if platform == "darwin":  # OS X
        os.system(f"open {view_file}")
    elif platform == "win32":  # Windows...
        os.system(f"{view_file}")
    change_dir(catalog)


def check_pay():
    """список для отметки оплаты взноса"""
    from reportlab.platypus import Table
    from sys import platform
    elements = []
    story = []
    tit = Title.get(Title.id == title_id())
    view_sort = ["По алфавиту", "По региону"]
    view_sort, ok = QInputDialog.getItem(
                            my_win, "Сортировка", "Выберите вид сортировки,\n просмотра списка участников.", view_sort, 0, False)
    if view_sort == "По алфавиту": 
        player_list = Player.select().where(Player.title_id == title_id()).order_by(Player.player)  # сортировка по алф
    elif view_sort == "По региону":
        player_list = Player.select().where(Player.title_id == title_id()).order_by(Player.region)  # сортировка по региону

    short_name = tit.short_name_comp
    gamer = tit.gamer
    count = len(player_list)  # количество записей в базе
    kp = count + 1
    n = 0
    for l in player_list:
        n += 1
        p = l.player
        c = l.city
        g = l.region
        coach_id = l.coach_id
        t = coach_id.coach

        data = [n, p, c, g, t]
        elements.append(data)

    elements.insert(0, ["№", "Фамилия, Имя", "Город", "Регион", "Тренер(ы)"])
    t = Table(elements, colWidths=(0.7 * cm, 5.0 * cm, 3.5 * cm, 4.5 * cm, 5.9 * cm), rowHeights=(0.6 * cm), repeatRows=1)  # ширина столбцов, если None-автоматическая
    t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            # ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                           ('FONTSIZE', (0, 0), (-1, -1), 10),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 1),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (-1, kp * -1), 'CENTER'),
                           ('BACKGROUND', (0, 0), (8, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (8, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, kp * -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.5, colors.black)
                           ]))

    h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=150,
            firstLineIndent=-20, textColor="green")  # стиль параграфа
    h3.spaceAfter = 10  # промежуток после заголовка
    story.append(Paragraph(f'Список участников. {gamer}', h3))
    story.append(t)
    doc = SimpleDocTemplate(f"{short_name}_player_list_payment.pdf", pagesize=A4, 
                            rightMargin=1*cm, leftMargin=1*cm, topMargin=1.5*cm, bottomMargin=1*cm) # название, вид страницы, размер полей
    view_file = f"{short_name}_player_list_payment.pdf"
    catalog = 1
    change_dir(catalog)
    doc.build(story, onFirstPage=func_zagolovok)
    if platform == "darwin":  # OS X
        os.system(f"open {view_file}")
    elif platform == "win32":  # Windows...
        os.system(f"{view_file}")
    change_dir(catalog)


def referee():
    """добавление судей в базу"""
    sender = my_win.sender()
    if sender == my_win.comboBox_referee: # комбобокс глав судьи
        text = my_win.comboBox_referee.currentText()
        index = my_win.comboBox_referee.findText(text)
        if index != -1:
            my_win.comboBox_referee.setCurrentIndex(index)
            my_win.comboBox_referee.lineEdit().setSelection(len(text), len(my_win.comboBox_referee.currentText()))
            category = find_referee_in_db(text)
            my_win.comboBox_kategor_ref.setCurrentText(category)
    elif sender == my_win.comboBox_secretary: # комбобокс глав секретарь:
        text = my_win.comboBox_secretary.currentText()
        index = my_win.comboBox_secretary.findText(text)
        if index != -1:
            my_win.comboBox_secretary.setCurrentIndex(index)
            my_win.comboBox_secretary.lineEdit().setSelection(len(text), len(my_win.comboBox_secretary.currentText()))
            category = find_referee_in_db(text)
            my_win.comboBox_kategor_sec.setCurrentText(category)
    elif sender == my_win.tableWidget.comboBox_family_city: # комбобокс выбора судей гск:
        text = my_win.tableWidget.comboBox_family_city.currentText()
        index = my_win.tableWidgetGSK.comboBox_family_city.findText(text)
        if index != -1:
            my_win.tableWidget.comboBox_family_city.setCurrentIndex(index)
            my_win.tableWidget.comboBox_family_city.lineEdit().setSelection(len(text), len(my_win.tableWidget.comboBox_family_city.currentText()))
            category = find_referee_in_db(text)
            my_win.comboBox_kategor_sec.setCurrentText(category)


def find_referee_in_db(text):
    """ищет фамилию судьи в базе данных и возвращает судейскую категорию"""
    mark = text.find("/")
    fio = text[:mark]
    referee = Referee.select().where(Referee.family == fio).get()
    category = referee.category
    return category


def open_close_file(view_file):
    # Проверить, существует
    if os.path.exists(view_file):
        flag = True
    else:
        flag = False
    return flag


def list_duplicate_family(double_id):
    """список двойных фамилий"""
    from reportlab.platypus import Table
    from sys import platform
    elements = []
    story = []
    n = 0
    tit = Title.get(Title.id == title_id())
    pl_list = Player.select().where(Player.title_id == title_id()) # сортировка по региону
    for pl in double_id:
        pl = pl_list.select().where(Player.id == pl).get() # сортировка по региону
        n += 1
        p = pl.player
        c = pl.city
        data = [n, p, c]
        elements.append(data)
        
    short_name = tit.short_name_comp
    gamer = tit.gamer
    count = len(double_id)  # количество записей в базе
    kp = count + 1
    elements.insert(0, ["№", "Фамилия, Имя", "Город"])
    t = Table(elements, colWidths=(0.7 * cm, 5.0 * cm, 5.0 * cm), rowHeights=(0.6 * cm), repeatRows=1)  # ширина столбцов, если None-автоматическая
    t.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), "DejaVuSerif"),  # Использую импортированный шрифт
                            # ('FONTNAME', (1, 1), (1, kp), "DejaVuSerif-Bold"),
                           # Использую импортированный шрифта размер
                           ('FONTSIZE', (0, 0), (-1, -1), 10),
                           # межстрочный верхний инервал
                           ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                           # межстрочный нижний инервал
                           ('TOPPADDING', (0, 0), (-1, -1), 1),
                           # вериткальное выравнивание в ячейке заголовка
                           ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                           # горизонтальное выравнивание в ячейке
                           ('ALIGN', (0, 0), (-1, kp * -1), 'CENTER'),
                           ('BACKGROUND', (0, 0), (8, 0), colors.yellow),
                           ('TEXTCOLOR', (0, 0), (8, 0), colors.darkblue),
                           ('LINEABOVE', (0, 0), (-1, kp * -1), 1, colors.blue),
                           # цвет и толщину внутренних линий
                           ('INNERGRID', (0, 0), (-1, -1), 0.02, colors.grey),
                           # внешние границы таблицы
                           ('BOX', (0, 0), (-1, -1), 0.5, colors.black)
                           ]))

    h3 = PS("normal", fontSize=12, fontName="DejaVuSerif-Italic", leftIndent=150,
            firstLineIndent=-20, textColor="green")  # стиль параграфа
    h3.spaceAfter = 10  # промежуток после заголовка
    story.append(Paragraph(f'Список участников. {gamer}', h3))
    story.append(t)
    doc = SimpleDocTemplate(f"{short_name}_player_list_duplicate.pdf", pagesize=A4, 
                            rightMargin=1*cm, leftMargin=1*cm, topMargin=1.5*cm, bottomMargin=1*cm) # название, вид страницы, размер полей
    view_file = f"{short_name}_player_list_duplicate.pdf"
    catalog = 1
    change_dir(catalog)
    doc.build(story, onFirstPage=func_zagolovok)
    if platform == "darwin":  # OS X
        os.system(f"open {view_file}")
    elif platform == "win32":  # Windows...
        os.system(f"{view_file}")
    change_dir(catalog)



def double_family():
    """создает список двойных фамилий"""
    double_pl_dict = {}
    players = Player.select().where(Player.title_id == title_id())
    for k in players:
        id_pl = k.id
        pl_fam_name = k.player
        mark = pl_fam_name.find(" ")
        if mark != -1:
            pl_fam = pl_fam_name[:mark]
            double_pl_dict[id_pl] = pl_fam
    double_id = find_duplicate_values(double_pl_dict) # список id игроков двойных фамилий
    list_duplicate_family(double_id)


def schedule_net():
     my_win.tabWidget_3.setTabEnabled(2, True)
    # book = op.Workbook()
    # worksheet = book.active
    # names_headers = ["№", "Фамилия, Имя", "Город"]
    # for m in range(1, 4):
    #     c =  worksheet.cell(row = 1, column = m)
    #     c.value = names_headers[m - 1]

    # k = 2    
    # for pl in players:
    #     fio = pl.player
    #     gorod = pl.city
    #     n = k - 1
    #     c1 = worksheet.cell(row = k, column = 1)
    #     c1.value = n
    #     c2 = worksheet.cell(row = k, column = 2)
    #     c2.value = fio
    #     c3 = worksheet.cell(row = k, column = 3)
    #     c3.value = gorod
 
    #     k += 1

    # t_id = Title.get(Title.id == title_id())

    # worksheet.column_dimensions['A'].width = 8
    # worksheet.column_dimensions['B'].width = 25
    # worksheet.column_dimensions['C'].width = 25
   
    # sex = t_id.gamer
    # f_name = f"{sex}_двойные _фамилии.xlsx"
    # filename, filter = QtWidgets.QFileDialog.getSaveFileName(my_win, 'Save file', f'{f_name}','Excel files (*.xlsx)')
    # book.save(filename)


def find_duplicate_values(double_pl_dict):
    double_id = []
    # Creating reverse dictionary to group keys by their values
    reverse_dict = {}
    for key, value in double_pl_dict.items():
        reverse_dict.setdefault(value, set()).add(key)

    # Finding values with more than one key to find duplicate values
    duplicate_values = [value for value, keys in reverse_dict.items() if len(keys) > 1]
    for dv in duplicate_values:
        id_player = list(reverse_dict[dv])
        for k in id_player:
            double_id.append(k)
    return double_id


def check_choice_net(fin):
    """Проверка после жеребьевки сетки на 1-ю встречи одних регионов или одинаковых тренеров"""
    msgBox = QMessageBox
    region_list = []
    coach_list = []
    coach_list_tmp = []
    id_system = system_id(stage=fin)
    gamelist_fin = Game_list.select().where(Game_list.system_id == id_system).order_by(Game_list.rank_num_player)
    c = 0
    g = 2
    for k in gamelist_fin:
        pl = k.player_group_id
        players = Player.select().where(Player.id == pl).get()
        region_pl = players.region
        coaches = Coach.select().where(Coach.id == players.coach_id).get()
        coach_str = coaches.coach
        coach_list_tmp = coach_str.split(", ")
        coach_list.append(coach_list_tmp.copy())
        coach_list_tmp.clear()
        region_list.append(region_pl)
        c += 1
        if c == 2:
            region_set = set(region_list)
            count = len(region_set)
            if count == 1:
                coach_union = set(coach_list[0]) & set(coach_list[1])    
                if len(coach_union) != 0:
                    msgBox.information(my_win, "Уведомление", f"Встреча № {g // 2},\nспортсмены одного тренера.\n{coach_union}")
                else:
                    msgBox.information(my_win, "Уведомление", f"Встреча № {g // 2},\nоба спортсмена из одного региона.")
            region_list.clear()
            coach_list.clear()
            c = 0
        g += 1


def mesto_3_no_play():
    """записывает в DB  изменения по разигрыванию 3 места"""
    if my_win.checkBox_no_play_3.isChecked():
        my_win.Button_3_mesta.setEnabled(True)
        n_g = "3"  
    else:
        n_g = ""
        my_win.Button_3_mesta.setEnabled(False)
    System.update(no_game=n_g).where((System.title_id == title_id()) & (System.stage == '1-й финал')).execute()
        

def two_3_place():
    """Когда не разигрывается 3-е место"""
    msgBox = QMessageBox()
    systems = System.select().where((System.title_id == title_id()) & (System.stage == '1-й финал')).get()
    results = Result.select().where((Result.title_id == title_id()) & (Result.number_group == '1-й финал'))
    system_table = systems.label_string
    if system_table == "Сетка (с розыгрышем всех мест) на 8 участников":
        number_game = 8
    elif system_table == "Сетка (-2) на 8 участников":
        number_game = 12
    elif system_table == "Сетка (с розыгрышем всех мест) на 16 участников":
        number_game = 16
    elif system_table == "Сетка (-2) на 16 участников":
        number_game = 28
    elif system_table == "Сетка (с розыгрышем всех мест) на 32 участников":
        number_game = 32
    elif system_table == "Сетка (-2) на 32 участников":
        number_game = 60
    elif system_table == "Сетка (1-3 место) на 32 участников":
        number_game = 32
    game = results.select().where(Result.tours == number_game).get()
    if game.player1 == "" or game.player2 == "":
        msgBox.information(my_win, "Уведомление", "Нет одного из игроков,\nзанявшего 3-е место.")
        return
    else: 
        win_3thd = game.winner
        if win_3thd is not None:
            question = msgBox.question(my_win, "Уведомление", "Вы уже занесли игроков, занявших 3-е место.\nЕсли хотите обновить данные. Нажмите - ОК-", msgBox.Ok, msgBox.Cancel)
            if question ==  msgBox.Cancel:
                return
        Result.update(winner=game.player1, loser=game.player2).where(Result.tours == number_game).execute()
    player_list = Result.select().where((Result.title_id == title_id()) & (Result.number_group == '1-й финал'))
    fill_table(player_list)


def add_double_player_to_list():
    """добавляет пару в списки"""
    pass


# def proba_pdf():
    # """проба пдф"""

# import itertools
# from random import randint
# from statistics import mean

# from reportlab.lib.pagesizes import A4
# from reportlab.pdfgen import canvas


# def grouper(iterable, n):
#     args = [iter(iterable)] * n
#     return itertools.zip_longest(*args)


# def export_to_pdf(data):
#     c = canvas.Canvas("grid-students.pdf", pagesize=A4)
#     w, h = landscape(A4)
#     max_rows_per_page = 45
#     # Margin.
#     x_offset = 50
#     y_offset = 50
#     # Space between rows.
#     padding = 15

#     xlist = [x + x_offset for x in [0, 200, 250, 300, 350, 400, 480]]
#     ylist = [h - y_offset - i*padding for i in range(max_rows_per_page + 1)]

    # for rows in grouper(data, max_rows_per_page):
    #     rows = tuple(filter(bool, rows))
    #     c.grid(xlist, ylist[:len(rows) + 1])
    #     for y, row in zip(ylist[:-1], rows):
    #         for x, cell in zip(xlist, row):
    #             c.drawString(x + 2, y - padding + 3, str(cell))
    #     c.showPage()

#     c.save()


# data = [("NAME", "GR. 1", "GR. 2", "GR. 3", "AVG", "STATUS")]

# for i in range(1, 101):
#     exams = [randint(0, 10) for _ in range(3)]
#     avg = round(mean(exams), 2)
#     state = "Approved" if avg >= 4 else "Disapproved"
#     data.append((f"Student {i}", *exams, avg, state))

# export_to_pdf(data)
# . =====

    # styles = getSampleStyleSheet()
    # styleN = styles['Normal']
    # styleH = styles['Heading1']
    # story = []
    #     #add some flowables
    # story.append(Paragraph("This is a Heading", styleH))
    # story.append(Paragraph("This is a paragraph in <i>Normal</i> style.", styleN))
    # c  = Canvas('mydoc.pdf', pagesize = landscape)
    # f = Frame(5* cm, 3 * cm, 6 * cm, 25 * cm, showBoundary=1) # высота прямоугольника  6 Х 25, showBoundary = 1, рамка 0- нет
    # f.addFromList(story, c)
#     # c.save()
# def proba():
#     choices = Choice.select()
#     for ch in choices:
#         ch_id = ch.id
#         reg = ch.region       
#         region_mod = reg.rstrip()
#         Choice.update(region=region_mod).where(Choice.id == ch_id).execute()
#     regions = Region.select()
#     # count = len(regions)
#     for r in regions:
#         r_id = r.id
#         rg = r.region
#         reg_mod = rg.rstrip()      
#         Region.update(region=reg_mod).where(Region.id == r_id).execute()
#     #     reg = str(reg.rstrip())
#     # players = Player.select()
#     # for p in players:
#     #     reg = p.region
#     #     reg = str(reg.rstrip())
#     #     # bd_new = format_date_for_db(str_date=bd)
#     #     # txt = str(bd_new)
#     #     Player.update(region=reg).execute()
#     print("Все записи обновлены")
# =======        
# def proba():
#     myconn = pymysql.connect(host = "localhost", user = "root", passwd = "db_pass", database = "mysql_db") 
#     # создать таблицу
    
#     class Players_full(BaseModel):
#         player = CharField(50)    
#         bday = DateField()
#         rank = IntegerField()
#         city = CharField()
#         region = CharField()
#         razryad = CharField()
#         coach_id = ForeignKeyField(Coach)
#         patronymic_id = ForeignKeyField(Patronymic)
    
#         class Meta:
#             db_table = "players_full"
#             order_by = "player"

#     db.create_tables([Players_full])
#     db.close()

 
# # #creating the cursor object 
# #     cur = myconn.cursor() 
# #     try: 
# #         #adding a column branch name to the table Employee 
# #         cur.execute("ALTER TABLE Game_list MODIFY COLUMN player_group_id VARCHAR(30) NULL;") 
# #     except: 
# #         myconn.rollback() 
    
# #     myconn.close() 


#     migrator = MySQLMigrator(db)
#     # # no_game = TextField(default="")
#     patronymic_id = IntegerField(Delete_player)  # новый столбец, его поле и значение по умолчанию
    
#     # # posev_super_final = ForeignKeyField(Choise, field=System.id, null=True)

#     # # with db:
#     # #     # migrate(migrator.drop_column('choices', 'posev_super_final')) # удаление столбца
#     # #     # migrate(migrator.alter_column_type('system', 'mesta_exit', IntegerField()))
#     # #     # migrate(migrator.rename_column('titles', 'kat_sek', 'kat_sec')) # Переименование столбца (таблица, старое название, новое название столбца)
#     migrate(migrator.add_column('delete_players', 'patronymic_id', patronymic_id)) # Добавление столбца (таблица, столбец, повтор название столбца)

# my_win.Button_proba.clicked.connect(proba) # запуск пробной функции

# ===== переводит фокус на поле ввода счета в партии вкладки -группа-
my_win.lineEdit_pl1_s1.returnPressed.connect(focus)
my_win.lineEdit_pl2_s1.returnPressed.connect(focus)
my_win.lineEdit_pl1_s2.returnPressed.connect(focus)
my_win.lineEdit_pl2_s2.returnPressed.connect(focus)
my_win.lineEdit_pl1_s3.returnPressed.connect(focus)
my_win.lineEdit_pl2_s3.returnPressed.connect(focus)
my_win.lineEdit_pl1_s4.returnPressed.connect(focus)
my_win.lineEdit_pl2_s4.returnPressed.connect(focus)
my_win.lineEdit_pl1_s5.returnPressed.connect(focus)
my_win.lineEdit_pl2_s5.returnPressed.connect(focus)
my_win.lineEdit_pl1_s6.returnPressed.connect(focus)
my_win.lineEdit_pl2_s6.returnPressed.connect(focus)
my_win.lineEdit_pl1_s7.returnPressed.connect(focus)
my_win.lineEdit_pl2_s7.returnPressed.connect(focus)
# ===== проверка правильность ввода цифр

my_win.lineEdit_range_tours.returnPressed.connect(enter_print_begunki)
my_win.lineEdit_num_game_fin.returnPressed.connect(filter_fin)

my_win.lineEdit_pl1_score_total.returnPressed.connect(enter_total_score)
my_win.lineEdit_pl2_score_total.returnPressed.connect(enter_total_score)
my_win.lineEdit_Family_name.returnPressed.connect(input_player)
my_win.lineEdit_bday.returnPressed.connect(next_field)
my_win.lineEdit_city_list.returnPressed.connect(add_city)
# ====== отслеживание изменения текста в полях ============
my_win.lineEdit_find_player_in_system.textChanged.connect(find_player_on_tab_system)
my_win.lineEdit_find_player_stat.textChanged.connect(find_player_on_tab_system)
my_win.lineEdit_Family_name.textChanged.connect(find_in_rlist)  # в поле поиска и вызов функции
my_win.lineEdit_find_player_in_R.textChanged.connect(find_in_player_rejting_list)
my_win.lineEdit_coach.textChanged.connect(find_coach)
my_win.lineEdit_city_list.textChanged.connect(find_city)
my_win.lineEdit_pl1_double.textChanged.connect(tab_double)
my_win.lineEdit_pl2_double.textChanged.connect(tab_double)
my_win.lineEdit_otchestvo.textChanged.connect(find_otchestvo)
my_win.lineEdit_otchestvo.returnPressed.connect(add_patronymic)
my_win.comboBox_region.currentTextChanged.connect(find_city)
# comboBox_family_city = QComboBox()
# comboBox_family_city.currentTextChanged.connect(referee)
# ============= двойной клик
# двойной клик по listWidget (рейтинг, тренеры, отчество)
my_win.listWidget.itemDoubleClicked.connect(dclick_in_listwidget)
my_win.listWidget_double.itemDoubleClicked.connect(dclick_in_listWidget_double)

# двойной клик по строке игроков в таблице -результаты-, -списки-
my_win.tableView.doubleClicked.connect(select_player_in_game)
my_win.tableWidget.doubleClicked.connect(move_row_in_tablewidget)

my_win.tabWidget.currentChanged.connect(tab)
my_win.tabWidget_stage.currentChanged.connect(tab_etap)
my_win.tabWidget_3.currentChanged.connect(tab_double)
# my_win.tabWidget_stage.currentChanged.connect(tab_result)

my_win.toolBox.currentChanged.connect(tool_page)
# ==================================
my_win.spinBox_kol_group.textChanged.connect(kol_player_in_group)
# ======== изменение индекса комбобоксов ===========

# fir_window.comboBox.currentTextChanged.connect(change_sroki)


my_win.comboBox_table_1.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_2.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_3.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_4.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_5.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_6.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_7.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_8.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_9.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_10.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_11.currentTextChanged.connect(kol_player_in_final)
my_win.comboBox_table_12.currentTextChanged.connect(kol_player_in_final)

my_win.comboBox_etap.currentTextChanged.connect(made_system_load_combobox_etap)

my_win.comboBox_page_vid.currentTextChanged.connect(page_vid)
my_win.comboBox_filter_number_group_final.currentTextChanged.connect(filter_player_on_system)
my_win.comboBox_filter_choice_stage.currentTextChanged.connect(choice_filter_on_system)
my_win.comboBox_fltr_region.currentTextChanged.connect(change_city_from_region)
my_win.comboBox_filter_region_in_R.currentTextChanged.connect(change_city_from_region_in_R)
my_win.comboBox_select_stage_begunki.currentTextChanged.connect(select_stage_for_begunki)
my_win.comboBox_select_group_begunki.currentTextChanged.connect(select_tour_for_begunki)
my_win.comboBox_select_tours.currentTextChanged.connect(select_diapazon)
my_win.comboBox_edit_etap1.currentTextChanged.connect(select_stage_for_edit)
my_win.comboBox_edit_etap2.currentTextChanged.connect(select_stage_for_edit)
my_win.comboBox_first_group.currentTextChanged.connect(add_item_listwidget)
my_win.comboBox_second_group.currentTextChanged.connect(add_item_listwidget)

my_win.comboBox_filter_final.currentTextChanged.connect(filter_fin)
my_win.comboBox_choice_R.currentTextChanged.connect(r_list_load_tableView)
# my_win.comboBox_filter_region_in_R.currentTextChanged.connect(filter_rejting_list)
# my_win.comboBox_filter_city_in_R.currentTextChanged.connect(filter_rejting_list)
# my_win.comboBox_filter_date_in_R.currentTextChanged.connect(filter_rejting_list)

my_win.comboBox_referee.currentTextChanged.connect(referee)
my_win.comboBox_secretary.currentTextChanged.connect(referee)
my_win.comboBox_kategor_ref.currentTextChanged.connect(add_referee_to_db)
my_win.comboBox_kategor_sec.currentTextChanged.connect(add_referee_to_db)


# =======  отслеживание переключение чекбоксов =========
my_win.radioButton_match_3.toggled.connect(change_status_visible_and_score_game)
my_win.radioButton_match_5.toggled.connect(change_status_visible_and_score_game)
my_win.radioButton_match_7.toggled.connect(change_status_visible_and_score_game)

my_win.checkBox_repeat_regions.stateChanged.connect(change_choice_group) 

# при изменении чекбокса активирует кнопку создать
my_win.checkBox.stateChanged.connect(button_title_made_enable)
# при изменении чекбокса активирует кнопку создать
my_win.checkBox_3.stateChanged.connect(button_system_made_enable)
# при изменении чекбокса показывает поля для ввода счета
my_win.checkBox_4.stateChanged.connect(change_status_visible_and_score_game)
# при изменении чекбокса показывает поля для ввода счета финала)
my_win.checkBox_visible_game.stateChanged.connect(change_status_visible_and_score_game)
# при изменении чекбокса показывает список удаленных игроков
my_win.checkBox_6.stateChanged.connect(del_player_table)
my_win.checkBox_7.stateChanged.connect(no_play)  # поражение по неявке игрок 1 группа
my_win.checkBox_8.stateChanged.connect(no_play)  # поражение по неявке игрок 2 группа

my_win.checkBox_11.stateChanged.connect(debitor_R) # должники рейтинга оплаты
my_win.checkBox_15.stateChanged.connect(filter_player_list)
my_win.checkBox_find_player.stateChanged.connect(find_player)
my_win.checkBox_double.stateChanged.connect(page_double)
my_win.checkBox_no_play_3.stateChanged.connect(mesto_3_no_play)
# =======  нажатие кнопок =========


my_win.Button_Ok.setAutoDefault(True)  # click on <Enter>

my_win.Button_3_mesta.clicked.connect(two_3_place)
my_win.Button_pay_R.clicked.connect(save_in_db_pay_R)
my_win.Button_clear_del.clicked.connect(clear_del_player)
my_win.Button_reset_filter_gr.clicked.connect(reset_filter)
my_win.Button_reset_filter_fin.clicked.connect(reset_filter)
my_win.Button_reset_filter_sf.clicked.connect(reset_filter)
my_win.Button_filter_fin.clicked.connect(filter_fin)
my_win.Button_filter_sf.clicked.connect(filter_sf)
my_win.Button_filter_gr.clicked.connect(filter_gr)
my_win.Button_app.clicked.connect(check_real_player) # отмечает что игрок по заявке

my_win.Button_etap_made.clicked.connect(etap_made) # кнопка создание этапов системы
my_win.Button_add_edit_player.clicked.connect(add_player)  # добавляет игроков в список и базу
# записывает в базу или редактирует титул
my_win.Button_title_made.clicked.connect(title_made) # создание титула
# записывает в базу счет в партии встречи
my_win.Button_Ok.clicked.connect(enter_score) # кнопка ввода счета
my_win.Button_del_player.clicked.connect(delete_player) # удаляет игроков
my_win.Button_print_begunki.clicked.connect(begunki_made)



my_win.Button_add_pl1.clicked.connect(list_player_in_group_after_draw)
my_win.Button_add_pl2.clicked.connect(list_player_in_group_after_draw)
my_win.Button_change_player.clicked.connect(change_player_between_group_after_draw)

my_win.Button_sort_mesto.clicked.connect(sort)
my_win.Button_sort_R.clicked.connect(sort)
my_win.Button_sort_Name.clicked.connect(sort)
my_win.Button_fltr_list.clicked.connect(filter_player_list)
my_win.Button_reset_fltr_list.clicked.connect(filter_player_list)
my_win.Button_reset_fltr_in_R.clicked.connect(clear_filter_rejting_list)
my_win.Button_sort_alf_R.clicked.connect(filter_rejting_list)
my_win.Button_sort_rejting_in_R.clicked.connect(filter_rejting_list)
my_win.Button_filter_R.clicked.connect(filter_rejting_list)
my_win.Button_made_R_file.clicked.connect(made_file_excel_for_rejting) # создание excel файла для рейтинга
my_win.Button_made_player_list_excel.clicked.connect(made_file_excel_list_player) # создание excel файла списка
my_win.Button_made_protokol.clicked.connect(protokol_pdf) # создание файла протокола
my_win.Button_made_one_file_pdf.clicked.connect(merdge_pdf_files)

my_win.Button_up.clicked.connect(move_row_in_tablewidget)
my_win.Button_down.clicked.connect(move_row_in_tablewidget)
my_win.tableWidget.cellClicked.connect(button_move_enabled)
my_win.Button_list_referee.clicked.connect(made_list_GSK)
my_win.Button_list_regions.clicked.connect(made_list_regions)
my_win.Button_list_winner.clicked.connect(made_list_winners)
my_win.Button_players_on_pdf_file.clicked.connect(made_list_players_for_pdf_file)
# my_win.Button_players_on_pdf_file.clicked.connect(list_player_pdf)
my_win.Button_made_page_pdf.clicked.connect(made_pdf_list)
my_win.Button_view_page_pdf.clicked.connect(view_all_page_pdf)
my_win.Button_randevy.clicked.connect(randevy_list)
my_win.Button_made_schedule.clicked.connect(schedule_net)
my_win.Button_add_double.clicked.connect(add_double_player_to_list)


my_win.Button_pay.clicked.connect(check_pay)
sys.exit(app.exec())